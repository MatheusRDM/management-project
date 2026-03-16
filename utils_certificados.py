import streamlit as st
import pandas as pd
import sqlite3
import os
import io
import sys
import logging
from datetime import datetime
import docx
import re
import glob
import unicodedata
from contextlib import contextmanager
from thefuzz import fuzz

# ======================================================================================
# FUNÇÃO AUXILIAR PARA PARSING DE DATAS COM FORMATO EXPLÍCITO
# ======================================================================================
def parse_date_safe(date_str, format='%d/%m/%Y'):
    """
    Parses date strings safely with explicit format to avoid Pandas warnings.
    Assumes dd/mm/yyyy format by default.
    """
    return pd.to_datetime(date_str, format=format, errors='coerce')

# Tenta importar bibliotecas do Google (modo nuvem vs local)
try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload
    GOOGLE_LIBS_AVAILABLE = True
except ImportError:
    GOOGLE_LIBS_AVAILABLE = False

# Configuração de Logs
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ======================================================================================
# 1. CONFIGURAÇÃO DE ARQUIVOS E DIRETÓRIOS
# ======================================================================================
FILES_CONFIG = {
    "certificados_2026": {
        "local_path": r"G:\.shortcut-targets-by-id\1NUJ7pNAqedohSrLjiwFErFrtVqVZkrjk\006 - Lab. Central\1.0 CONTROLES\00.CERTIFICADOS\FORM 067 - REV 00 - Controle de Certificados(2026)..xlsm",
        "tipo": "certificado", "ano": "2026"
    },
    "certificados_2025": {
        "local_path": r"G:\.shortcut-targets-by-id\1NUJ7pNAqedohSrLjiwFErFrtVqVZkrjk\006 - Lab. Central\1.0 CONTROLES\00.CERTIFICADOS\FORM 067 - REV 00 - Controle de Certificados(2025).xlsm",
        "tipo": "certificado", "ano": "2025"
    },
    "recebimento_2026": {
        "local_path": r"G:\.shortcut-targets-by-id\1NUJ7pNAqedohSrLjiwFErFrtVqVZkrjk\006 - Lab. Central\1.0 CONTROLES\03. RECEBIMENTO DE AMOSTRAS\FORM 022 A - REV 00 - Controle de recebimentos e descarte de amostras-AGIR - 2026.xlsm",
        "tipo": "recebimento", "ano": "2026"
    },
    "recebimento_2025": {
        "local_path": r"G:\.shortcut-targets-by-id\1NUJ7pNAqedohSrLjiwFErFrtVqVZkrjk\006 - Lab. Central\1.0 CONTROLES\03. RECEBIMENTO DE AMOSTRAS\FORM 022 A - REV 00 - Controle de recebimentos e descarte de amostras-AGIR 2025.xlsm",
        "tipo": "recebimento", "ano": "2025"
    },
    "propostas_comerciais": {
        "local_path": r"G:\.shortcut-targets-by-id\1JbWwLDR6PaShh0-_xJZLFAvEXQKn65V1\008 - Comercial\007 - Controle de Propostas\FORM 044 - REV 05 - Controle de Propostas E-VIAS.AFIRMA.xlsx",
        "tipo": "proposta", "ano": "multi-ano"
    },
}

LOCAL_PATH_RELATORIOS  = r"G:\.shortcut-targets-by-id\1NUJ7pNAqedohSrLjiwFErFrtVqVZkrjk\006 - Lab. Central\0.1 RELATÓRIOS TÉCNICOS\003-PROJETOS"
DIRETORIO_BASE_CAUQ    = r"G:\.shortcut-targets-by-id\1NUJ7pNAqedohSrLjiwFErFrtVqVZkrjk\006 - Lab. Central\0.2 PROJETOS CAUQ MARSHALL"
BASE_DIR_PROJETOS_MRAF = r"G:\.shortcut-targets-by-id\1NUJ7pNAqedohSrLjiwFErFrtVqVZkrjk\006 - Lab. Central\0.4 PROJETOS MRAF"
DB_NAME                = "lab_central_master.db"
BASE_DIR_PROPOSTAS     = r"G:\.shortcut-targets-by-id\1JbWwLDR6PaShh0-_xJZLFAvEXQKn65V1\008 - Comercial\000 - Comercial\01 - Propostas Comerciais"


# ======================================================================================
# CONFIGURAÇÃO DE TODOS OS TIPOS DE PROJETO COMPASA
# Cada tipo possui:
#   base_dir    → pasta raiz do tipo de projeto na rede
#   label       → nome de exibição no dashboard
#   tem_pioneiro→ se False, a fase PIONEIRO é omitida (cards e rastreamento)
#   keywords    → palavras-chave no MATERIAL para auto-detecção do tipo
#   icone       → emoji identificador
# ======================================================================================
_BASE_LAB = r"G:\.shortcut-targets-by-id\1NUJ7pNAqedohSrLjiwFErFrtVqVZkrjk\006 - Lab. Central"

TIPOS_PROJETO_CONFIG = {
    'CAUQ_MARSHALL': {
        'label':        'CAUQ Marshall',
        'base_dir':     _BASE_LAB + r"\0.2 PROJETOS CAUQ MARSHALL",
        'tem_pioneiro': True,
        'keywords':     ['CAUQ', 'MARSHALL'],
        'icone':        '🛣️',
    },
    'CAUQ_SUPERPAVE': {
        'label':        'CAUQ Superpave',
        'base_dir':     _BASE_LAB + r"\0.3 PROJETOS CAUQ SUPERPAVE",
        'tem_pioneiro': True,
        'keywords':     ['SUPERPAVE'],
        'icone':        '🔬',
    },
    'MRAF': {
        'label':        'MRAF',
        'base_dir':     _BASE_LAB + r"\0.4 PROJETOS MRAF",
        'tem_pioneiro': True,
        'keywords':     ['MRAF', 'RECICLAGEM'],
        'icone':        '♻️',
    },
    'SOLO_CIMENTO': {
        'label':        'Solo Cimento',
        'base_dir':     _BASE_LAB + r"\0.5 PROJETOS SOLO CIMENTO",
        'tem_pioneiro': False,
        'keywords':     ['SOLO', 'CIMENTO'],
        'icone':        '🏔️',
    },
    'CAMADAS_GRANULARES': {
        'label':        'Camadas Granulares',
        'base_dir':     _BASE_LAB + r"\0.6 PROJETOS CAMADAS GRANULARES",
        'tem_pioneiro': False,
        'keywords':     ['CAMADAS', 'GRANULAR', 'BRITA GRADUADA', 'MACADAME'],
        'icone':        '⛏️',
    },
    'BGS': {
        'label':        'BGS',
        'base_dir':     _BASE_LAB + r"\0.7 PROJETOS BGS",
        'tem_pioneiro': False,
        'keywords':     ['BGS'],
        'icone':        '🪨',
    },
}

# Símbolos exportados para uso nos dashboards
__all__ = [
    'FILES_CONFIG', 'LOCAL_PATH_RELATORIOS', 'DIRETORIO_BASE_CAUQ',
    'BASE_DIR_PROJETOS_MRAF', 'DB_NAME', 'TIPOS_PROJETO_CONFIG',
    'DataBridge', 'bridge', 'parse_date_safe',
    'sync_all_data', 'carregar_dados_consolidados_sql', 'exportar_dados_csv',
    'rastrear_projetos_compasa_completo', 'rastrear_projetos_externos',
    'escanear_todos_projetos', 'escanear_tipo_projeto',
    'carregar_projetos_form022a',
    'carregar_dados_epr_raw', 'consolidar_fas_totais',
    'gerar_quantitativos_empresas', 'carregar_empresa_finalidade_raw',
    'calcular_fas_total', 'carregar_form044', 'carregar_form045',
    'buscar_e_extrair_form045',
    'verificar_fase_mraf', 'localizar_pasta_mraf',
]

# ======================================================================================
# 2. GERENCIADOR DE DADOS
# ======================================================================================
class DataBridge:
    def __init__(self):
        from cloud_config import IS_CLOUD
        self.is_cloud = IS_CLOUD

    def get_file_content(self, config_key):
        if config_key not in FILES_CONFIG:
            return None
        config = FILES_CONFIG[config_key]
        local_path = config["local_path"]
        if os.path.exists(local_path):
            return local_path
        return None

    @contextmanager
    def get_db_conn(self):
        conn = sqlite3.connect(DB_NAME)
        try: yield conn
        finally: conn.close()
            
    def init_db(self):
        with self.get_db_conn() as conn:
            # --- ATUALIZAÇÃO DA TABELA: ADICIONADA COLUNA 'OBS_RECEBIMENTO' ---
            conn.execute('''CREATE TABLE IF NOT EXISTS recebimentos (
                id INTEGER PRIMARY KEY, NUMERO_PROPOSTA TEXT, CLIENTE TEXT, STATUS TEXT, 
                DATA_RECEBIMENTO TEXT, DATA_ENTREGA TEXT, DIAS_VENCIMENTO REAL, 
                STATUS_PRAZO TEXT, MATERIAL TEXT, PEDREIRA TEXT, TEM_MATERIAL INTEGER,
                TIPO_PROPOSTA TEXT, ANO TEXT, PT_COLUNA_A TEXT, FONTE TEXT, 
                DATA_ENTREGA_FAS TEXT, PROJETO_FAS TEXT, E_CONTRATO_CONTINUO INTEGER,
                OBS_RECEBIMENTO TEXT)''')
            
            conn.execute('''CREATE TABLE IF NOT EXISTS certificados_067 (
                id INTEGER PRIMARY KEY,
                PT TEXT,
                PT_NORMALIZADO TEXT,
                CLIENTE TEXT,
                ENSAIO TEXT,
                NORMA TEXT,
                QUANTIDADE REAL,
                ENSAIO_CONCLUIDO INTEGER,
                RELATORIO_VINCULADO TEXT,
                ANO TEXT,
                DATA_CRIACAO DATETIME DEFAULT CURRENT_TIMESTAMP
            )''')

            conn.execute('''CREATE TABLE IF NOT EXISTS propostas (
                NUMERO_PROPOSTA TEXT, ANO TEXT, CLIENTE TEXT,
                STATUS_PROPOSTA TEXT, DATA_ACEITE_PROPOSTA DATETIME)''')

            # Migration para adicionar coluna se não existir (evita erro em DB existente)
            try:
                conn.execute("ALTER TABLE recebimentos ADD COLUMN OBS_RECEBIMENTO TEXT")
            except sqlite3.OperationalError:
                pass # Coluna já existe

bridge = DataBridge()
if not bridge.is_cloud:
    bridge.init_db()

# ======================================================================================
# 3. FUNÇÕES AUXILIARES
# ======================================================================================

def normalizar_pc(pc_raw):
    if not pc_raw: return None
    pc_str = str(pc_raw).upper().strip()
    match = re.search(r'PC\s*(\d+)[\.\-](\d+)', pc_str, re.IGNORECASE)
    if match: return f"PC{match.group(1).zfill(3)}.{match.group(2)[:2]}"
    nums = re.findall(r'\d+\.\d+', pc_str)
    if nums: 
        parts = nums[0].split('.')
        if len(parts) == 2: return f"PC{parts[0].zfill(3)}.{parts[1][:2]}"
    return None

def normalizar_pt(pt_raw):
    if not pt_raw: return None
    pt_str = str(pt_raw).upper().strip()
    match = re.search(r'(\d+)', pt_str)
    return match.group(1) if match else None

def normalizar_pt_busca(texto):
    if not texto: return ""
    match = re.search(r'(\d+)', str(texto))
    return str(int(match.group(1))) if match else ""

def normalizar_status(status_raw):
    if not status_raw: return 'A DEFINIR'
    s = status_raw.upper().strip()
    if any(x in s for x in ['CONCLU', 'FINALIZ', 'ENTREGUE', 'EMITIDO', 'OK', 'ENVIADO', 'PRONTO', 'RELATÓRIO FEITO']):
        return 'FINALIZADO'
    if any(x in s for x in ['ANDAMENTO', 'EXECU', 'INICIAR', 'TESTE', 'ENSAIO']):
        return 'EM ANDAMENTO'
    if 'AGUARDANDO' in s:
        return 'AGUARDANDO MATERIAL'
    return 'A DEFINIR'

def normalizar_texto(texto):
    if not texto: return ''
    texto = str(texto).strip()
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII')
    return texto.upper()

def categorizar_material(material_raw):
    if not material_raw: return 'Outros'
    mat = str(material_raw).upper()
    if any(x in mat for x in ['BRITA', 'AREIA', 'PEDRA', 'BGS', 'RACH']): return 'Agregado'
    if any(x in mat for x in ['SOLO', 'TERRA', 'ATERRO']): return 'Solo'
    if any(x in mat for x in ['CONCRETO', 'CIMENTO', 'ARGAMASSA']): return 'Concreto'
    if any(x in mat for x in ['ASFALTO', 'CBUQ', 'CAP']): return 'Asfalto'
    return 'Outros'

# ======================================================================================
# 4. FUNÇÕES ESPECÍFICAS COMPASA (BUSCA INTELIGENTE POR PEDREIRA E PT)
# ======================================================================================

def verificar_existencia_pdf(pasta):
    try:
        if not os.path.exists(pasta): return False
        for f in os.listdir(pasta):
            if f.lower().endswith('.pdf'): return True
        return False
    except: return False

def buscar_pt_dentro_excel(pasta_busca, pt_alvo, tipo_fase):
    if not os.path.exists(pasta_busca): return None, None
    pt_limpo = normalizar_pt_busca(pt_alvo)
    if not pt_limpo: return None, None
    estrategias = []
    if tipo_fase == 'COMPOSICAO': estrategias.append((0, 7, 8))
    else:
        estrategias.append(('007 E1', 8, 11))
        estrategias.append((0, 6, 4))
        estrategias.append((0, 26, 4))
    for f in os.listdir(pasta_busca):
        if f.endswith('.xlsm') or f.endswith('.xlsx'):
            if f.startswith('~$'): continue
            caminho_completo = os.path.join(pasta_busca, f)
            try:
                xls = pd.ExcelFile(caminho_completo, engine='openpyxl')
                sheet_names = xls.sheet_names
                for sheet_target, row, col in estrategias:
                    sheet_to_read = None
                    if isinstance(sheet_target, int):
                        if len(sheet_names) > sheet_target: sheet_to_read = sheet_names[sheet_target]
                    elif isinstance(sheet_target, str):
                        for s in sheet_names:
                            if sheet_target.lower() in s.lower():
                                sheet_to_read = s
                                break
                    if sheet_to_read:
                        df = pd.read_excel(xls, sheet_name=sheet_to_read, header=None, nrows=40)
                        if len(df) > row and len(df.columns) > col:
                            val = str(df.iloc[row, col]).strip()
                            if pt_limpo in val or re.search(rf'\b{pt_limpo}\b', val):
                                return caminho_completo, df
            except: continue
    return None, None

def extrair_detalhes_tecnicos(df, tipo_fase):
    dados = {}
    try:
        pass
    except Exception:
        pass
    return dados


# ======================================================================================
# 4.1. CONSTANTES DE PADRÃO — usadas pelo scanner universal
# ======================================================================================
#
# ┌─────────────────────────────────────────────────────────────────────────────────┐
# │  MAPA DE CAMINHOS CONFIRMADOS                                                   │
# │  Base: G:\...shortcut...\006 - Lab. Central\                                    │
# │   0.2 PROJETOS CAUQ MARSHALL                                                    │
# │   0.3 PROJETOS CAUQ SUPERPAVE                                                   │
# │   0.4 PROJETOS MRAF                                                             │
# │   0.5 PROJETOS SOLO CIMENTO                                                     │
# │   0.6 PROJETOS CAMADAS GRANULARES                                               │
# │   0.7 PROJETOS BGS                                                              │
# └─────────────────────────────────────────────────────────────────────────────────┘
#
# ESTRUTURAS INTERNAS DETECTADAS AUTOMATICAMENTE:
#
# A) PADRÃO CAUQ  — Marshall, Superpave (e possivelmente demais)
#    BASE/
#      003-COMPOSIÇÕES/{ANO}/{PEDREIRA}/*.xlsm
#      004-TRAÇOS PIONEIROS/{ANO}/{PEDREIRA}/*.xlsm     (só se tem_pioneiro=True)
#      006-PROJETOS/{ANO}/{PT}_{PEDREIRA}/005-ENTREGA/*.xlsm
#    CLIENTE: NÃO embutido na pasta → vem do FORM 022A por cruzamento de PT
#
# B) PADRÃO MRAF  — confirmado em campo; pode aparecer em outros tipos futuramente
#    BASE/
#      004-PROJETOS/
#        _AAAA PROJETOS {TIPO}/
#          {NNN.M.AAAA} - {CLIENTE} - {PED.*} - {MISTURA}/
#            004-ENTREGA/
#              {TIPO} - N° {NNN.M.AAAA}_R{REV}.xlsm
#    CLIENTE: EMBUTIDO no nome da pasta (parte[1] após primeiro " - ")
#    CÓDIGO:  NNN.M.AAAA  (ex: 001.5.2026)
#
# C) PADRÃO NOME-COM-CLIENTE  — qualquer estrutura onde o nome da pasta de projeto
#    segue o formato "{COD} - {CLIENTE} - ..."  (universal, independe de A ou B)
#    → Cliente extraído diretamente do nome, sem banco de dados
#
# D) PADRÃO FLAT (fallback)
#    BASE/{ANO}/{PT}_{PEDREIRA}/*.xlsm
#
# DETECÇÃO: o scanner testa A → B → C → D em ordem e usa o que encontrar.
# Para cada projeto também tenta extração de cliente por nome antes de ir ao DB.
# ======================================================================================

# ── Sub-nomes de fases CAUQ (tentativas em ordem de prioridade) ─────────────────────
_SUBFASES_CAUQ = {
    'COMPOSICAO': [
        '003-COMPOSIÇÕES', '003-COMPOSICOES', '003 COMPOSIÇÕES', '003 COMPOSICOES',
        'COMPOSICOES', 'COMPOSIÇÕES', '002-COMPOSICOES', '002-COMPOSIÇÕES',
    ],
    'PIONEIRO': [
        '004-TRAÇOS PIONEIROS', '004-TRACOS PIONEIROS',
        '004 TRAÇOS PIONEIROS', '004 TRACOS PIONEIROS',
        'TRAÇOS PIONEIROS', 'TRACOS PIONEIROS', 'PIONEIROS',
        '005-TRAÇOS PIONEIROS',
    ],
    'PROJETO': [
        '006-PROJETOS', '005-PROJETOS', '006 PROJETOS', '005 PROJETOS',
        'PROJETOS', '006-PROJETO', '007-PROJETOS', '004-PROJETOS',
    ],
}

# ── Sub-nomes de entrega (dentro da pasta do projeto) ───────────────────────────────
_SUBPASTAS_ENTREGA = [
    '005-ENTREGA', '004-ENTREGA', '006-ENTREGA', '007-ENTREGA',
    '003-ENTREGA', 'ENTREGA', 'RELATORIO', '005 ENTREGA',
]

# ── Padrão de pasta de ano ───────────────────────────────────────────────────────────
_RE_ANO_PASTA = re.compile(r'(\d{4})')

# Ano ativo — apenas pastas deste ano são escaneadas nos painéis de projeto
_ANO_PROJETOS_ATIVO = '2026'

# ── Padrões de código por tipo ───────────────────────────────────────────────────────
#
# MRAF:  001.5.2026  →  NNN.M.AAAA  (3 blocos numéricos separados por ponto, ano 4 dígitos)
_RE_COD_MRAF = re.compile(r'(\d{3}\.\d+\.\d{4})')

# CAUQ / demais:  042/25  042.25  042_25  BGS-001.25  SC-08.25  CG-03/26
# Aceita prefixos alfanuméricos opcionais (SP, SC, CG, BGS, etc.) seguidos de NN/AA ou NN.AA
_RE_COD_CAUQ = re.compile(
    r'^(?:[A-Z]{1,6}[-_\s]?)?(\d{2,4})[/\._](\d{2,4})',
    re.IGNORECASE
)

# Genérico: primeiro número de 2–4 dígitos encontrado
_RE_PT_GENERICO = re.compile(r'(\d{2,4})')

# ── Regex para nome de pasta de projeto com cliente embutido ────────────────────────
# Formato confirmado MRAF:
#   {NNN.M.AAAA} - {CLIENTE} - {PED.*} - {MISTURA}         (4 partes)
#   {NNN.M.AAAA} - {CLIENTE} - {DESCRICAO}                  (3 partes)
#
# Extensão genérica para outros tipos que sigam o mesmo padrão:
#   {QUALQUER-COD} - {CLIENTE} - {RESTO}                    (≥ 3 partes separadas por " - ")
#
_RE_PASTA_NOME_COM_CLIENTE_4 = re.compile(
    r'^(.+?)\s+-\s+(.+?)\s+-\s+(PED[^-]*?)\s+-\s+(.+)$',
    re.IGNORECASE
)
_RE_PASTA_NOME_COM_CLIENTE_3 = re.compile(
    r'^(.+?)\s+-\s+(.+?)\s+-\s+(.+)$',
    re.IGNORECASE
)

# ── Clientes COMPASA e contratos contínuos ──────────────────────────────────────────
_NOMES_COMPASA = ['COMPASA', 'COMPASA DO BRASIL', 'COMPASA LTDA']

_NOMES_CC = [           # Empresas de contrato contínuo (não-Compasa)
    'EPR', 'CBB', 'ASFALTEC', 'STRATA', 'EIXO SP', 'EIXOSP',
]

# Órgãos e empresas tipicamente EXTERNOS
_NOMES_EXTERNOS_CONHECIDOS = [
    'DER', 'DNIT', 'PMSP', 'SETRAN', 'DAER', 'DEINFRA',
    'FX II', 'FX III', 'AUTOPISTA', 'ARTERIS', 'CCR',
]

def _e_compasa(nome):
    if not nome:
        return False
    n = str(nome).upper().strip()
    return any(c in n for c in _NOMES_COMPASA)

def _e_cliente_interno(nome):
    if not nome:
        return False
    n = str(nome).upper().strip()
    return _e_compasa(nome) or any(t in n for t in _NOMES_CC)

def _classificar_cliente_externo(nome):
    """Retorna 'COMPASA' | 'CC' | 'EXTERNO'"""
    if not nome:
        return 'EXTERNO'
    if _e_compasa(nome):
        return 'COMPASA'
    n = str(nome).upper().strip()
    if any(t in n for t in _NOMES_CC):
        return 'CC'
    return 'EXTERNO'



# ======================================================================================
# 4.2. LEITURA DO FORM 022A — FONTE PRIMÁRIA DE PROJETOS
# ======================================================================================
#
# COLUNAS CONFIRMADAS (imagens + código existente):
#   idx  0  → col A  = PROTOCOLO (PT number, ex: 62, 157, 191)   ← chave primária
#   idx  1  → col B  = CLIENTE (nome da empresa)
#   idx  2  → col C  = DATA RECEBIMENTO
#   idx  3  → col D  = IDENTIFICAÇÃO DO MATERIAL
#   idx  4  → col E  = PROCEDÊNCIA (pedreira ou local do projeto)
#   idx 11  → col L  = OBSERVAÇÃO RECEBIMENTO (pode conter código "N° 004/2026")
#   idx 12  → col M  = FINALIDADE (tipo do projeto: "PROJETO CAUQ", "PROJETO MRAF" etc.)
#   idx 17  → col R  = STATUS ("Em andamento", "Concluído", "Aguardando")
#
# AGRUPAMENTO: um PROJETO = (PROTOCOLO, FINALIDADE_NORMALIZADA)
#   → N materiais por projeto → status_geral = mínimo comum
#   → pedreira = a mais frequente do grupo
#   → código_extra = OBS quando contém "N°" (SUPERPAVE etc.)
# ======================================================================================

# ── Mapeamento FINALIDADE (col M) → chave de TIPOS_PROJETO_CONFIG ────────────────────
_FINALIDADE_PARA_TIPO = {
    'PROJETO CAUQ':           'CAUQ_MARSHALL',
    'PROJETO CAUQ MARSHALL':  'CAUQ_MARSHALL',
    'PROJETO MARSHALL':       'CAUQ_MARSHALL',
    'PROJETO SUPERPAVE':      'CAUQ_SUPERPAVE',
    'PROJETO CAUQ SUPERPAVE': 'CAUQ_SUPERPAVE',
    'PROJETO MRAF':           'MRAF',
    'PROJETO DE RECICLAGEM':  'MRAF',
    'PROJETO RECICLAGEM':     'MRAF',
    'RECICLAGEM':             'MRAF',
    'PROJETO BGS':            'BGS',
    'BGS':                    'BGS',
    'PROJETO SOLO CIMENTO':   'SOLO_CIMENTO',
    'SOLO CIMENTO':           'SOLO_CIMENTO',
    'CAMADAS GRANULARES':     'CAMADAS_GRANULARES',
    'PROJETO CAMADAS':        'CAMADAS_GRANULARES',
    'PROJETO CAMADAS GRANULARES': 'CAMADAS_GRANULARES',
}

# ── Regex para extrair código de projeto da col OBS ──────────────────────────────────
# "N° 004/2026"  "Nº004/26"  "N 004.2026"
_RE_OBS_CODIGO_PROJETO = re.compile(
    r'N[°º\s]*[\s:]*(\d{2,4})[/\._\-](\d{2,4})',
    re.IGNORECASE
)

# ── Status normalizados do FORM (col R) ──────────────────────────────────────────────
_STATUS_FORM_NORM = {
    'concluído': 'CONCLUIDO',
    'concluido': 'CONCLUIDO',
    'em andamento': 'ANDAMENTO',
    'andamento': 'ANDAMENTO',
    'aguardando': 'AGUARDANDO',
    'não iniciado': 'NAO_INICIADO',
}

def _normalizar_status_form(raw):
    if not raw or str(raw).lower() in ('nan', 'none', ''):
        return 'DESCONHECIDO'
    return _STATUS_FORM_NORM.get(str(raw).strip().lower(), str(raw).strip().upper()[:20])

def _finalidade_para_tipo(finalidade_raw):
    """Converte string de FINALIDADE (col M) para chave de TIPOS_PROJETO_CONFIG."""
    if not finalidade_raw or str(finalidade_raw).lower() in ('nan', 'none', ''):
        return None
    f = str(finalidade_raw).strip().upper()
    # Busca exata
    if f in _FINALIDADE_PARA_TIPO:
        return _FINALIDADE_PARA_TIPO[f]
    # Busca parcial
    for chave, tipo in _FINALIDADE_PARA_TIPO.items():
        if chave in f:
            return tipo
    return None

def _extrair_codigo_obs(obs_raw):
    """
    Extrai código de projeto da coluna OBS (col L).
    Ex: "N° 004/2026" → "004/2026"
    Ex: "N° 004/26"   → "004/26"
    Retorna string normalizada ou None.
    """
    if not obs_raw or str(obs_raw).lower() in ('nan', 'none', '-', 'n.i', ''):
        return None
    m = _RE_OBS_CODIGO_PROJETO.search(str(obs_raw))
    if m:
        return f"{m.group(1)}/{m.group(2)}"
    return None

def _status_geral_projeto(lista_status):
    """
    Calcula o status geral de um projeto a partir da lista de status dos materiais.
    Hierarquia: AGUARDANDO > ANDAMENTO > CONCLUIDO > DESCONHECIDO
    """
    if not lista_status:
        return 'DESCONHECIDO'
    s = set(lista_status)
    if 'AGUARDANDO' in s:
        return 'AGUARDANDO'
    if 'ANDAMENTO' in s:
        return 'ANDAMENTO'
    if all(x == 'CONCLUIDO' for x in lista_status):
        return 'CONCLUIDO'
    return 'ANDAMENTO'


@st.cache_data(ttl=600, show_spinner=False)
def carregar_projetos_form022a(ano='2026'):
    """
    Lê o FORM 022A do ano indicado e retorna um DataFrame com UM REGISTRO POR PROJETO.

    Um "projeto" = agrupamento por (PROTOCOLO × FINALIDADE_TIPO).
    Múltiplos materiais do mesmo projeto são consolidados numa lista.

    Colunas retornadas:
        PT              → número do protocolo (col A)
        CLIENTE         → nome da empresa (col B)
        CLASSIFICACAO   → 'COMPASA' | 'CC' | 'EXTERNO'
        TIPO_PROJETO    → chave de TIPOS_PROJETO_CONFIG
        TIPO_LABEL      → label de exibição
        TIPO_ICONE      → emoji
        FINALIDADE_RAW  → valor original da col M
        PEDREIRA        → procedência principal (col E)
        OBS             → observação principal (col L)
        CODIGO_OBS      → código extraído da OBS ("004/2026" etc.)
        STATUS_FORM     → status geral do lab ('CONCLUIDO'|'ANDAMENTO'|'AGUARDANDO')
        N_MATERIAIS     → quantidade de materiais do projeto
        MATERIAIS       → lista de materiais (col D)
        DATA_RECEBIMENTO→ data mais recente do grupo
    """
    # ── CLOUD GUARD ──────────────────────────────────────────────────────────
    if bridge.is_cloud:
        from cloud_config import carregar_parquet_cache
        df = carregar_parquet_cache("recebimento_form022a")
        if not df.empty and ano:
            # Filtrar por ano se houver coluna DATA_RECEBIMENTO ou PT
            pass  # Cache já contém todos os anos consolidados
        return df if not df.empty else pd.DataFrame()
    # ─────────────────────────────────────────────────────────────────────────
    key_form = f'recebimento_{ano}'
    if key_form not in FILES_CONFIG:
        return pd.DataFrame()

    source = bridge.get_file_content(key_form)
    if not source:
        return pd.DataFrame()

    try:
        df_raw = pd.read_excel(source, header=None, engine='openpyxl')
    except Exception:
        return pd.DataFrame()

    registros_brutos = []
    for i in range(6, len(df_raw)):       # dados começam na linha 7 (idx 6)
        try:
            row = df_raw.iloc[i]

            # ── Protocolo ──────────────────────────────────────────────────────────
            pt_raw = row.iloc[0] if len(row) > 0 else None
            try:
                pt = int(float(str(pt_raw).strip()))
            except Exception:
                continue
            if pt <= 0:
                continue

            # ── Cliente ────────────────────────────────────────────────────────────
            cliente_raw = str(row.iloc[1]).strip() if len(row) > 1 else ''
            if len(cliente_raw) < 3 or cliente_raw.upper() in ('CLIENTE', 'NAN', 'NONE', ''):
                continue

            # ── Normalização de cliente (mesma lógica do sync_recebimento) ─────────
            cl_up = cliente_raw.upper()
            obs_classif = str(row.iloc[16]).strip().upper() if len(row) > 16 else ''
            finalidade_r = str(row.iloc[12]).strip() if len(row) > 12 else ''
            if 'ASFALTEC' in obs_classif:
                cliente = 'CC ASFALTEC'
            elif 'CBB' in cl_up or 'CBB' in obs_classif:
                cliente = 'CBB ASFALTOS'
            elif 'COMPASA' in cl_up:
                cliente = 'COMPASA DO BRASIL'
            else:
                cliente = cliente_raw

            # ── Data recebimento ───────────────────────────────────────────────────
            dt = pd.to_datetime(row.iloc[2], errors='coerce') if len(row) > 2 else pd.NaT

            # ── Material (col D) ───────────────────────────────────────────────────
            material = str(row.iloc[3]).strip() if len(row) > 3 else ''

            # ── Procedência / Pedreira (col E) ─────────────────────────────────────
            pedreira = str(row.iloc[4]).strip() if len(row) > 4 else ''
            if pedreira.lower() in ('nan', 'none', 'ni', 'n.i', '-', ''):
                pedreira = ''

            # ── OBS Recebimento (col L, idx 11) ────────────────────────────────────
            obs = str(row.iloc[11]).strip() if len(row) > 11 else ''
            if obs.lower() in ('nan', 'none', ''):
                obs = ''

            # ── Finalidade (col M, idx 12) ─────────────────────────────────────────
            tipo_projeto = _finalidade_para_tipo(finalidade_r)
            if not tipo_projeto:
                continue   # linha sem tipo reconhecido → ignora

            # ── Status lab (col R, idx 17) ─────────────────────────────────────────
            status_raw = str(row.iloc[17]).strip() if len(row) > 17 else ''
            status_lab = _normalizar_status_form(status_raw)

            registros_brutos.append({
                '_pt':         pt,
                '_cliente':    cliente,
                '_tipo':       tipo_projeto,
                '_finalidade': finalidade_r,
                '_pedreira':   pedreira,
                '_obs':        obs,
                '_material':   material,
                '_status':     status_lab,
                '_dt':         dt,
            })
        except Exception:
            continue

    if not registros_brutos:
        return pd.DataFrame()

    df_bruto = pd.DataFrame(registros_brutos)

    # ── Agrupa por (PT × TIPO_PROJETO) — um projeto por combinação ───────────────────
    projetos = []
    for (pt, tipo), grp in df_bruto.groupby(['_pt', '_tipo']):
        cliente = grp['_cliente'].mode().iloc[0] if not grp['_cliente'].mode().empty else ''
        classif = _classificar_cliente_externo(cliente)
        cfg = TIPOS_PROJETO_CONFIG.get(tipo, {})

        # Pedreira principal: a que aparece mais vezes (excluindo vazio)
        peds = grp['_pedreira'][grp['_pedreira'] != '']
        pedreira_principal = peds.mode().iloc[0] if not peds.empty else ''

        # OBS principal: a que tem código extraível, ou a mais comum
        obs_list = grp['_obs'][grp['_obs'] != ''].tolist()
        codigo_obs = None
        obs_principal = ''
        for o in obs_list:
            c = _extrair_codigo_obs(o)
            if c:
                codigo_obs = c
                obs_principal = o
                break
        if not obs_principal and obs_list:
            obs_principal = obs_list[0]

        status_geral = _status_geral_projeto(grp['_status'].tolist())
        dt_max = grp['_dt'].dropna().max() if not grp['_dt'].dropna().empty else pd.NaT

        projetos.append({
            'PT':               str(pt),
            'PT_NUM':           str(pt),
            'CLIENTE':          cliente,
            'CLASSIFICACAO':    classif,
            'TIPO_PROJETO':     tipo,
            'TIPO_LABEL':       cfg.get('label', tipo),
            'TIPO_ICONE':       cfg.get('icone', '📁'),
            'FINALIDADE_RAW':   grp['_finalidade'].iloc[0],
            'PEDREIRA':         pedreira_principal,
            'OBS':              obs_principal,
            'CODIGO_OBS':       codigo_obs,
            'STATUS_FORM':      status_geral,
            'N_MATERIAIS':      len(grp),
            'MATERIAIS':        ', '.join(grp['_material'].tolist()),
            'DATA_RECEBIMENTO': dt_max,
            'ANO_PASTA':        ano,
        })

    if not projetos:
        return pd.DataFrame()

    df = pd.DataFrame(projetos)
    # Ordena: COMPASA primeiro, depois por tipo e PT
    _ord_map = {'COMPASA': 0, 'CC': 1, 'EXTERNO': 2}
    df['_ord'] = df['CLASSIFICACAO'].map(_ord_map).fillna(9)
    df = df.sort_values(['_ord', 'TIPO_PROJETO', 'PT']).drop(columns='_ord').reset_index(drop=True)
    return df


# ======================================================================================
# 4.3. SCANNER UNIVERSAL — detecta o padrão da base_dir e varre projetos
# ======================================================================================

def _detectar_padrao_base(base_dir):
    """
    Detecta qual padrão de estrutura existe na base_dir.
    Retorna: 'CAUQ' | 'MRAF' | 'FLAT' | 'DESCONHECIDO'
    """
    if not os.path.isdir(base_dir):
        return 'DESCONHECIDO'
    try:
        filhos = os.listdir(base_dir)
    except OSError:
        return 'DESCONHECIDO'

    filhos_up = [f.upper() for f in filhos]

    # Padrão CAUQ: tem sub-pasta de fase diretamente na raiz
    for fase_nome in _SUBFASES_CAUQ['COMPOSICAO'] + _SUBFASES_CAUQ['PROJETO']:
        if fase_nome.upper() in filhos_up:
            return 'CAUQ'

    # Padrão MRAF: tem '004-PROJETOS' na raiz
    if '004-PROJETOS' in filhos_up:
        return 'MRAF'

    # Padrão FLAT: tem pastas de ano (4 dígitos) na raiz
    if any(re.match(r'^\d{4}$', f) for f in filhos):
        return 'FLAT'

    # Padrão CAUQ alternativo: tem pastas que contêm "COMPOSICAO" ou "PROJETO"
    for f in filhos_up:
        if 'COMPOSI' in f or 'PIONEIRO' in f or 'PROJETO' in f:
            return 'CAUQ'

    # Último recurso: se tem subpastas que contêm ano → FLAT
    for f in filhos:
        m = _RE_ANO_PASTA.search(f)
        if m and os.path.isdir(os.path.join(base_dir, f)):
            return 'FLAT'

    return 'DESCONHECIDO'


def _encontrar_subfase(base_dir, nomes_candidatos):
    """Retorna o caminho da primeira sub-pasta que existir dentre os candidatos."""
    try:
        filhos = {f.upper(): f for f in os.listdir(base_dir)}
    except OSError:
        return None
    for cand in nomes_candidatos:
        real = filhos.get(cand.upper())
        if real:
            p = os.path.join(base_dir, real)
            if os.path.isdir(p):
                return p
    return None


def _encontrar_pasta_entrega(pasta_projeto):
    """Retorna sub-pasta de entrega ou a própria pasta se nenhuma for encontrada."""
    for cand in _SUBPASTAS_ENTREGA:
        p = os.path.join(pasta_projeto, cand)
        if os.path.isdir(p):
            return p
    return pasta_projeto


def _inspecionar_pasta(pasta):
    """
    Retorna (arquivo_xl, tem_pdf) para a pasta informada.
    arquivo_xl = nome do primeiro .xlsm/.xlsx não temporário, ou None.
    tem_pdf    = True se encontrar qualquer .pdf.
    """
    arquivo_xl = None
    tem_pdf    = False
    try:
        for f in sorted(os.listdir(pasta)):
            if f.startswith('~$'):
                continue
            fl = f.lower()
            if fl.endswith(('.xlsm', '.xlsx')) and arquivo_xl is None:
                arquivo_xl = f
            if fl.endswith('.pdf'):
                tem_pdf = True
    except OSError:
        pass
    return arquivo_xl, tem_pdf


def _tem_pdf_nome(pasta, token='PROJ'):
    """Verifica se há PDF cujo nome contenha o token (case-insensitive)."""
    if not pasta or not os.path.isdir(pasta):
        return False
    tok = str(token).lower()
    try:
        for f in os.listdir(pasta):
            if f.lower().endswith('.pdf') and tok in f.lower():
                return True
    except OSError:
        return False
    return False


def _status_de_pasta(pasta):
    """Retorna 'OK' | 'ANDAMENTO' | 'VAZIO' baseado no conteúdo da pasta de entrega."""
    xl, pdf = _inspecionar_pasta(pasta)
    if pdf:
        return 'OK', xl, pdf
    if xl:
        return 'ANDAMENTO', xl, pdf
    return 'VAZIO', None, False


# ======================================================================================
# LEITURA PROFUNDA DE CÉLULAS EXCEL — SCANNER ENRIQUECIDO
# Extrai PT, pedreira e cliente diretamente dos arquivos .xlsm/.xlsx de cada fase.
#
# COMPOSIÇÃO  → aba 0 (primeira):
#   I8  (row 7, col 8)   = identificação / PT do projeto
#   C10 (row 9, col 2)   = procedência / pedreira
#   B9  (row 8, col 1)   = cliente (às vezes)
#   H9  (row 8, col 7)   = código / mistura
#   I7  (row 6, col 8)   = variante de I8
#   B8  (row 7, col 1)   = variante de cliente
#
# PIONEIRO / PROJETO → aba "007 E1" ou similar:
#   L9  (row 8, col 11)  = código / PT do projeto
#   B8  (row 7, col 1)   = cliente
#   H8  (row 7, col 7)   = mistura / pedreira
#   L7  (row 6, col 11)  = variante de L9
#   E27 (row 26, col 4)  = posição alternativa
# ======================================================================================

_CELULAS_COMP_XL = [
    (7,  8,  'I8'),    # identificação do projeto / PT
    (9,  2,  'C10'),   # procedência / pedreira
    (8,  1,  'B9'),    # cliente
    (8,  7,  'H9'),    # código / mistura
    (6,  8,  'I7'),    # variante de I8
    (7,  1,  'B8'),    # variante de cliente
]

_CELULAS_ENSAIOS_XL = [
    (8,  11, 'L9'),    # PT / código
    (7,  1,  'B8'),    # cliente
    (7,  7,  'H8'),    # mistura / pedreira
    (6,  11, 'L7'),    # variante de L9
    (26,  4, 'E27'),   # posição alternativa
]

_ABA_ENSAIOS_KEYS_XL = ['007', 'ensaio', 'mecân', 'e1', 'mechanical', 'gst_cadastro']

# ======================================================================================
# CAUQ MARSHALL — CÉLULAS E ABAS ESPECÍFICAS
# ======================================================================================
#
# ABA "GST_CADASTRO PROJETO":
#   E11 (row 10, col 4) = número do projeto
#   E27 (row 26, col 4) = pedreira
#   E7  (row  6, col 4) = campo extra
#
# ABA "PROJ_TRAÇO PIONEIRO":
#   E07 (row 6, col 4)  = mesclado PT + nome pedreira
#   L09 (row 8, col 11) = protocolo PT
#
# COMPOSIÇÃO aba 0:
#   B10 (row 9, col 1)  = mesclado PT + pedreira (ex: "CAUQ 001.A.2026 FX C - PED. IBIPORÃ")
#   I8  (row 7, col 8)  = mesclado PT + pedreira (alternativo)
#   C10 (row 9, col 2)  = procedência / pedreira isolada
#
# STATUS COMPOSIÇÃO:
#   PDF com MESMO NOME BASE do Excel = status OK (entregue)
#   Ex: Excel "CAUQ 001.A.2026 FX C.xlsm" + PDF "CAUQ 001.A.2026 FX C.pdf" → OK
#
# PDF ENTREGA PROJETO (A/B/C padrão):
#   "A (Certificados)..." ou "A (" no nome → certificados
#   "B (DPH)..." ou "B (" no nome → DPH
#   "C (ART)..." ou "C (" no nome → ART
#   PDF com código NNN.M.AAAA no nome → documento do projeto
# ======================================================================================

_CELULAS_GST_CADASTRO_XL = [
    (10, 4, 'E11_PROJ_NUM'),   # E11 = número do projeto
    (26, 4, 'E27_PEDREIRA'),   # E27 = pedreira (confirmado pelo usuário)
    (6,  4, 'E7_CAMPO'),       # E7  = campo extra
]

_CELULAS_PROJ_PION_XL = [
    (6,  4, 'E7_PION'),         # E07 mesclado = PT + pedreira
    (8, 11, 'L9_PROTO_PION'),   # L09 = protocolo
]

# Adicionado B10 (mesclado) à lista de células de composição
_CELULAS_COMP_XL_B10 = [
    (9,  1,  'B10'),    # B10 mesclado = PT + pedreira (CAUQ Marshall)
    (7,  8,  'I8'),     # I8  mesclado = PT + pedreira
    (9,  2,  'C10'),    # C10 = procedência / pedreira
    (8,  1,  'B9'),     # B9  = cliente
    (8,  7,  'H9'),     # H9  = mistura
    (7,  1,  'B8'),     # B8  = cliente alternativo
    (6,  8,  'I7'),     # I7  = variante de I8
]

# Chaves para detectar aba GST_CADASTRO PROJETO
_ABA_GST_KEYS = ['gst_cadastro', 'gst cadastro', 'cadastro projeto', 'cadastro_proj']

# Chaves para detectar aba PROJ_TRAÇO PIONEIRO
_ABA_PROJ_PION_KEYS = ['proj_traç', 'proj_traco', 'proj traç', 'projtraç', 'proj_traço']

# Regex para código NNN.M.AAAA (MRAF / CAUQ Marshall projetos)
# Já existe: _RE_COD_MRAF = re.compile(r'(\d{3}\.\d+\.\d{4})')


def _ler_celula_xl(df, row: int, col: int):
    """Lê célula do DataFrame de forma segura. Retorna str ou None."""
    try:
        if len(df) > row and len(df.columns) > col:
            val = df.iloc[row, col]
            if pd.notna(val):
                s = str(val).strip()
                if s and s.lower() not in ('nan', 'none', '-', 'n.i', ''):
                    return s
    except Exception:
        pass
    return None


def _selecionar_aba_ensaios_xl(xl) -> str | None:
    """Retorna nome da aba de ensaios mecânicos ('007 E1', 'Ensaios', etc.)."""
    for nome in xl.sheet_names:
        nl = nome.lower()
        if any(k in nl for k in _ABA_ENSAIOS_KEYS_XL):
            return nome
    return None


def _ler_excel_composicao(caminho_xl: str) -> dict:
    """
    Extrai células de identificação de um arquivo Excel de COMPOSIÇÃO.
    Tenta a aba 0 primeiro, depois percorre todas as abas como fallback.
    Retorna dict com chaves I8, C10, B9, H9, I7, B8.
    """
    resultado = {c: None for _, _, c in _CELULAS_COMP_XL}
    resultado['_ABA'] = None
    resultado['_CAMINHO'] = caminho_xl
    if not caminho_xl or not os.path.isfile(caminho_xl):
        return resultado
    try:
        xl = pd.ExcelFile(caminho_xl, engine='openpyxl')
        aba_principal = xl.sheet_names[0]
        resultado['_ABA'] = aba_principal
        df = pd.read_excel(xl, sheet_name=aba_principal, header=None, nrows=35)
        for row, col, nome in _CELULAS_COMP_XL:
            resultado[nome] = _ler_celula_xl(df, row, col)
        # Scan extra: percorre outras abas quando células-alvo ainda estão vazias
        cells_vazias = [n for _, _, n in _CELULAS_COMP_XL if resultado[n] is None]
        if cells_vazias:
            for aba_extra in xl.sheet_names[1:]:
                try:
                    df2 = pd.read_excel(xl, sheet_name=aba_extra, header=None, nrows=20)
                    for row, col, nome in _CELULAS_COMP_XL:
                        if resultado[nome] is None:
                            resultado[nome] = _ler_celula_xl(df2, row, col)
                except Exception:
                    continue
    except Exception:
        pass
    return resultado


def _ler_excel_ensaios(caminho_xl: str) -> dict:
    """
    Extrai células de identificação de um arquivo Excel de PIONEIRO ou PROJETO.
    Usa a aba com '007' / 'Ensaios' / 'E1' se disponível; fallback: aba 0.
    Retorna dict com chaves L9, B8, H8, L7, E27.
    """
    resultado = {c: None for _, _, c in _CELULAS_ENSAIOS_XL}
    resultado['_ABA'] = None
    resultado['_CAMINHO'] = caminho_xl
    if not caminho_xl or not os.path.isfile(caminho_xl):
        return resultado
    try:
        xl = pd.ExcelFile(caminho_xl, engine='openpyxl')
        aba_alvo = _selecionar_aba_ensaios_xl(xl) or xl.sheet_names[0]
        resultado['_ABA'] = aba_alvo
        df = pd.read_excel(xl, sheet_name=aba_alvo, header=None, nrows=35)
        for row, col, nome in _CELULAS_ENSAIOS_XL:
            resultado[nome] = _ler_celula_xl(df, row, col)
        # Fallback: outras abas quando tudo ainda está vazio
        if all(resultado[n] is None for _, _, n in _CELULAS_ENSAIOS_XL):
            for aba_extra in xl.sheet_names:
                if aba_extra == aba_alvo:
                    continue
                try:
                    df2 = pd.read_excel(xl, sheet_name=aba_extra, header=None, nrows=35)
                    for row, col, nome in _CELULAS_ENSAIOS_XL:
                        if resultado[nome] is None:
                            resultado[nome] = _ler_celula_xl(df2, row, col)
                except Exception:
                    continue
    except Exception:
        pass
    return resultado


def _consolidar_campos_excel(
    d_comp: dict,
    d_pion: dict,
    d_proj: dict,
    nome_pasta: str,
    cliente_pasta: str = '',
) -> dict:
    """
    Consolida PT / pedreira / cliente / mistura extraídos das 3 fases Excel.

    Prioridade de cada campo:
      PT       → PROJETO(L9/L7) > PIONEIRO(L9/L7) > COMPOSIÇÃO(I8/I7/H9) > nome da pasta
      Pedreira → COMPOSIÇÃO(C10) > PROJETO(H8) > PIONEIRO(H8) > nome da pasta
      Cliente  → PROJETO(B8) > PIONEIRO(B8) > COMPOSIÇÃO(B9/B8) > nome da pasta
    """
    def _pega(*vals):
        for v in vals:
            if v and str(v).strip().lower() not in ('none', 'nan', '', '-', 'n.i'):
                # Filtra valores genéricos que não são identificadores úteis
                s = str(v).strip()
                if len(s) >= 2 and not s.isspace():
                    return s
        return ''

    def _pt_norm(v):
        """Normaliza PT/Protocolo: remove prefixo PT, mantém dígitos e / ."""
        if not v:
            return ''
        s = str(v).strip().upper()
        s = re.sub(r'^PT[-\s]*', '', s)
        s = re.sub(r'[^0-9/\.]+', '', s)
        return s.strip('/.-')

    pt = _pega(
        _pt_norm(d_proj.get('L9')), _pt_norm(d_proj.get('L7')),
        _pt_norm(d_pion.get('L9')), _pt_norm(d_pion.get('L7')),
        _pt_norm(d_comp.get('I8')), _pt_norm(d_comp.get('I7')),
        _pt_norm(d_comp.get('H9')), _pt_norm(d_comp.get('B10')),
    ) or _extrair_pt_de_nome(nome_pasta)

    pedreira = _pega(
        d_comp.get('C10'), d_comp.get('E27_PEDREIRA'),
        d_proj.get('H8'), d_pion.get('H8'),
    ) or _extrair_pedreira_de_nome(nome_pasta)

    cliente = _pega(
        d_proj.get('B8'), d_pion.get('B8'),
        d_comp.get('B9'), d_comp.get('B8'),
        cliente_pasta,
    )

    mistura = _pega(d_comp.get('H9'), d_proj.get('H8'))

    return {
        'PT_EXCEL':       pt,
        'PEDREIRA_EXCEL': pedreira,
        'CLIENTE_EXCEL':  cliente,
        'MISTURA_EXCEL':  mistura,
    }


def _localizar_pasta_pedreira(dir_fase_ano: str, pedreira_ref: str, pt_ref: str = '') -> str | None:
    """
    Localiza sub-pasta de pedreira dentro de dir_fase_ano usando:
    1. PT exato no nome (score 95)
    2. Similaridade fuzzy com pedreira_ref (threshold 60)
    Retorna caminho completo ou None.
    """
    if not os.path.isdir(dir_fase_ano):
        return None
    melhor_pasta = None
    melhor_score = 0
    ped_norm = pedreira_ref.upper().replace('PEDREIRA', '').strip()[:25]
    try:
        for p_ped in os.listdir(dir_fase_ano):
            path_p = os.path.join(dir_fase_ano, p_ped)
            if not os.path.isdir(path_p):
                continue
            # Match por PT (alta prioridade)
            if pt_ref and pt_ref in p_ped:
                return path_p
            # Match por pedreira (fuzzy)
            score = fuzz.partial_ratio(ped_norm, p_ped.upper()) if ped_norm else 0
            if score > melhor_score and score > 60:
                melhor_score = score
                melhor_pasta = path_p
    except OSError:
        pass
    return melhor_pasta


# ── Extração de campos do nome da pasta ─────────────────────────────────────────────

def _extrair_campos_pasta_mraf(nome_pasta):
    """
    Extrai {codigo, cliente, pedreira, mistura} de um nome de pasta de projeto.
    Funciona para MRAF e qualquer tipo que use o padrão:
        {COD} - {CLIENTE} - {PED.*} - {MISTURA}    (4 partes)
        {COD} - {CLIENTE} - {DESCRICAO}             (3 partes)
        {COD} - {DESCRICAO}                         (2 partes, sem cliente)

    Para o código aceita:
        NNN.M.AAAA  (MRAF)
        NN/AA, NN.AA, prefixoNN/AA  (CAUQ/outros)
    Retorna dict ou None se não reconhecer.
    """
    nome = nome_pasta.strip()

    # ── Tenta formato 4 partes com pedreira explícita ──────────────────────────────
    m = _RE_PASTA_NOME_COM_CLIENTE_4.match(nome)
    if m:
        cod = m.group(1).strip()
        # Validar que o "código" realmente parece um código de projeto
        if _RE_COD_MRAF.search(cod) or _RE_COD_CAUQ.match(cod):
            return {
                'codigo':   cod,
                'cliente':  m.group(2).strip(),
                'pedreira': m.group(3).strip(),
                'mistura':  m.group(4).strip(),
            }

    # ── Tenta formato 3 partes ──────────────────────────────────────────────────────
    m = _RE_PASTA_NOME_COM_CLIENTE_3.match(nome)
    if m:
        cod = m.group(1).strip()
        if _RE_COD_MRAF.search(cod) or _RE_COD_CAUQ.match(cod):
            return {
                'codigo':   cod,
                'cliente':  m.group(2).strip(),
                'pedreira': '',
                'mistura':  m.group(3).strip(),
            }

    # ── Fallback: só código (MRAF) ──────────────────────────────────────────────────
    mc = _RE_COD_MRAF.search(nome)
    if mc:
        return {'codigo': mc.group(1), 'cliente': '', 'pedreira': '', 'mistura': nome}

    # ── Fallback: código CAUQ no início ────────────────────────────────────────────
    mc = _RE_COD_CAUQ.match(nome)
    if mc:
        resto = nome[mc.end():].lstrip(' -_')
        return {'codigo': f"{mc.group(1)}/{mc.group(2)}", 'cliente': '', 'pedreira': resto[:40], 'mistura': ''}

    return None


def _extrair_pt_de_nome(nome_pasta):
    """
    Extrai número de PT/código do nome da pasta.
    Aceita: NN/AA, NN.AA, NN_AA, prefixoNN/AA (BGS-001.25, SP-12/25, etc.)
    Retorna string normalizada do número (sem prefixo) ou None.
    """
    # Tenta padrão MRAF primeiro
    m = _RE_COD_MRAF.search(nome_pasta)
    if m:
        return m.group(1)
    # Tenta padrão CAUQ com prefixo opcional
    m = _RE_COD_CAUQ.match(nome_pasta.strip())
    if m:
        return f"{m.group(1)}/{m.group(2)}"
    # Fallback: primeiro número de 2–4 dígitos
    m = _RE_PT_GENERICO.search(nome_pasta)
    if m:
        return m.group(1)
    return None


def _extrair_pedreira_de_nome(nome_pasta):
    """Extrai nome da pedreira do nome da pasta (heurística simples)."""
    nome_up = nome_pasta.upper()
    # Remove prefixo de código
    nome_up = re.sub(r'^\d[\d\._/\-]+\s*', '', nome_up).strip()
    # Remove "PEDREIRA" da string para retornar só o nome
    nome_limpo = re.sub(r'\bPEDREIRA\b', '', nome_up).strip()
    # Pega até 50 chars
    return nome_limpo[:50].strip(' -_')


# ======================================================================================
# 4.3. HELPERS ESPECIALIZADOS CAUQ MARSHALL
# ======================================================================================

def _normalizar_pedreira_match(nome):
    """Normaliza pedreira para comparação fuzzy: remove acentos, PED., pontuação."""
    if not nome:
        return ''
    s = unicodedata.normalize('NFKD', str(nome)).encode('ASCII', 'ignore').decode('ASCII')
    s = re.sub(r'\b(PED|PEDREIRA|PEDREIRAS|PED\.)\b\.?', '', s.upper())
    s = re.sub(r'[^\w\s]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def _pedreira_score(ped_ref, ped_cand):
    """Score 0-100 de similaridade entre duas pedreiras."""
    a = _normalizar_pedreira_match(ped_ref)
    b = _normalizar_pedreira_match(ped_cand)
    if not a or not b:
        return 0
    return fuzz.partial_ratio(a, b)


def _encontrar_pasta_ano_generico(dir_fase, ano_alvo):
    """
    Encontra pasta de ano em dir_fase.
    Aceita: "2026", "PIONEIRO 2026", "_2026 PROJETOS MARSHALL", etc.
    """
    if not os.path.isdir(dir_fase):
        return None
    try:
        for f in sorted(os.listdir(dir_fase)):
            path_f = os.path.join(dir_fase, f)
            if not os.path.isdir(path_f):
                continue
            m = _RE_ANO_PASTA.search(f)
            if m and m.group(1) == str(ano_alvo):
                return path_f
    except OSError:
        pass
    return None


def _pdf_tem_nome_similar_excel(pasta, nome_xl):
    """
    Verifica se existe PDF em `pasta` com nome base similar ao do Excel.
    Composição: PDF = mesmo nome base do Excel (sem extensão).
    """
    if not pasta or not nome_xl or not os.path.isdir(pasta):
        return False
    base = os.path.splitext(nome_xl)[0].upper().strip()
    try:
        for f in os.listdir(pasta):
            if f.startswith('~$') or not f.lower().endswith('.pdf'):
                continue
            pdf_base = os.path.splitext(f)[0].upper().strip()
            # Exato ou muito similar (85%+)
            if pdf_base == base or fuzz.ratio(pdf_base, base) >= 85:
                return True
    except OSError:
        pass
    return False


def _inspecionar_projeto_cauq(pasta):
    """
    Inspeciona pasta de entrega de PROJETO CAUQ.
    Detecta PDFs A (Certificados), B (DPH/Deformação), C (ART) e PDF do projeto.

    Retorna dict:
        pdf_a, pdf_b, pdf_c, pdf_projeto : bool
        tem_pdf, entregue                 : bool
        arquivo_xl                        : str|None
        status                            : 'OK' | 'ANDAMENTO' | 'VAZIO'
        pdfs                              : list[str]
    """
    res = {
        'pdf_a': False, 'pdf_b': False, 'pdf_c': False,
        'pdf_projeto': False, 'tem_pdf': False, 'entregue': False,
        'arquivo_xl': None, 'status': 'VAZIO', 'pdfs': []
    }
    if not os.path.isdir(pasta):
        return res
    try:
        for f in sorted(os.listdir(pasta)):
            if f.startswith('~$'):
                continue
            fl = f.lower()
            fu = f.upper()
            if fl.endswith(('.xlsm', '.xlsx')) and not res['arquivo_xl']:
                res['arquivo_xl'] = f
            if fl.endswith('.pdf'):
                res['tem_pdf'] = True
                res['pdfs'].append(f)
                # A (Certificados)
                if re.search(r'\bA\s*\(', fu) or ' A (' in fu:
                    res['pdf_a'] = True
                # B (DPH / Deformação)
                if re.search(r'\bB\s*\(', fu) or ' B (' in fu:
                    res['pdf_b'] = True
                # C (ART)
                if re.search(r'\bC\s*\(', fu) or ' C (' in fu:
                    res['pdf_c'] = True
                # PDF do projeto (tem código NNN.M.AAAA no nome)
                if _RE_COD_MRAF.search(f):
                    res['pdf_projeto'] = True
    except OSError:
        pass

    res['entregue'] = res['pdf_a'] or res['pdf_b'] or res['pdf_c'] or res['pdf_projeto']
    if res['entregue'] or res['tem_pdf']:
        res['status'] = 'OK'
    elif res['arquivo_xl']:
        res['status'] = 'ANDAMENTO'
    return res


def _ler_excel_gst_cadastro(caminho_xl):
    """
    Lê aba 'GST_CADASTRO PROJETO' do Excel.
    Retorna dict {E11_PROJ_NUM, E27_PEDREIRA, E7_CAMPO}.
    """
    res = {c: None for _, _, c in _CELULAS_GST_CADASTRO_XL}
    res['_ABA'] = None
    if not caminho_xl or not os.path.isfile(caminho_xl):
        return res
    try:
        xl = pd.ExcelFile(caminho_xl, engine='openpyxl')
        aba_alvo = None
        for nome in xl.sheet_names:
            nl = nome.lower()
            if any(k in nl for k in _ABA_GST_KEYS):
                aba_alvo = nome
                break
        if not aba_alvo:
            return res
        res['_ABA'] = aba_alvo
        df = pd.read_excel(xl, sheet_name=aba_alvo, header=None, nrows=35)
        for row, col, nome in _CELULAS_GST_CADASTRO_XL:
            res[nome] = _ler_celula_xl(df, row, col)
    except Exception:
        pass
    return res


def _ler_excel_aba_proj_pion(caminho_xl):
    """
    Lê aba 'PROJ_TRAÇO PIONEIRO' do Excel.
    Retorna dict {E7_PION, L9_PROTO_PION}.
    """
    res = {c: None for _, _, c in _CELULAS_PROJ_PION_XL}
    res['_ABA'] = None
    if not caminho_xl or not os.path.isfile(caminho_xl):
        return res
    try:
        xl = pd.ExcelFile(caminho_xl, engine='openpyxl')
        aba_alvo = None
        for nome in xl.sheet_names:
            nl = nome.lower()
            if any(k in nl for k in _ABA_PROJ_PION_KEYS):
                aba_alvo = nome
                break
        if not aba_alvo:
            return res
        res['_ABA'] = aba_alvo
        df = pd.read_excel(xl, sheet_name=aba_alvo, header=None, nrows=35)
        for row, col, nome in _CELULAS_PROJ_PION_XL:
            res[nome] = _ler_celula_xl(df, row, col)
    except Exception:
        pass
    return res


def _ler_excel_composicao_completo(caminho_xl):
    """
    Lê composição CAUQ: inclui B10 (mesclado) além das células padrão.
    Retorna dict com B10, I8, C10, B9, H9, B8, I7 + aba de ensaios (L9, B8, H8).
    """
    resultado = {c: None for _, _, c in _CELULAS_COMP_XL_B10}
    resultado['_ABA'] = None
    resultado['_CAMINHO'] = caminho_xl
    # Também lê células de ensaios (007 E1)
    for _, _, c in _CELULAS_ENSAIOS_XL:
        resultado[c] = None

    if not caminho_xl or not os.path.isfile(caminho_xl):
        return resultado
    try:
        xl = pd.ExcelFile(caminho_xl, engine='openpyxl')
        # Preferir abas que contenham "COMPOS"; fallback para a primeira
        abas_candidatas = [s for s in xl.sheet_names if re.search(r"compos", s, re.IGNORECASE)]
        if not abas_candidatas:
            abas_candidatas = [xl.sheet_names[0]]

        pt_livre = None
        proto_texto = None

        for aba in abas_candidatas:
            resultado['_ABA'] = aba
            df_comp = pd.read_excel(xl, sheet_name=aba, header=None, nrows=40)
            for row, col, nome in _CELULAS_COMP_XL_B10:
                if resultado[nome] is None:
                    resultado[nome] = _ler_celula_xl(df_comp, row, col)
            if pt_livre is None:
                pt_livre = _procurar_pt_generico(df_comp)
            if proto_texto is None:
                proto_texto = _extrair_proto_texto_compos(df_comp)

        # Aba 007 / ensaios → L9 extra
        aba_ens = _selecionar_aba_ensaios_xl(xl)
        if aba_ens:
            df_ens = pd.read_excel(xl, sheet_name=aba_ens, header=None, nrows=35)
            for row, col, nome in _CELULAS_ENSAIOS_XL:
                resultado[nome] = _ler_celula_xl(df_ens, row, col)
            if pt_livre is None:
                pt_livre = _procurar_pt_generico(df_ens)
            if proto_texto is None:
                proto_texto = _extrair_proto_texto_compos(df_ens)

        resultado['PT_LIVRE'] = pt_livre
        resultado['PROTOCOLO_TEXTO'] = proto_texto
    except Exception:
        pass
    return resultado


def _extrair_pt_de_campo_mesclado(valor):
    """
    Extrai PT de um campo mesclado que contém PT + pedreira.
    Ex: "CAUQ 001.A.2026 FX C DER-PR - PED. IBIPORÃ" → "001.A.2026"
    Ex: "PT 042 - PED. IBIPORÃ" → "042"
    """
    if not valor:
        return None
    s = str(valor).strip()
    # Padrão tipo NNN.A.AAAA (CAUQ Marshall: 001.A.2026)
    m = re.search(r'(\d{3}\.[A-Z0-9]+\.\d{4})', s, re.IGNORECASE)
    if m:
        return m.group(1)
    # Padrão NNN.M.AAAA (MRAF)
    m = _RE_COD_MRAF.search(s)
    if m:
        return m.group(1)
    # PT NNN
    m = re.search(r'\bPT\s*(\d+)\b', s, re.IGNORECASE)
    if m:
        return m.group(1)
    # Número isolado
    m = re.search(r'\b(\d{2,4})\b', s)
    if m:
        return m.group(1)
    return None


def _extrair_pedreira_de_campo_mesclado(valor):
    """
    Extrai pedreira de campo mesclado.
    Ex: "CAUQ 001.A.2026 FX C - PED. IBIPORÃ" → "IBIPORÃ"
    """
    if not valor:
        return None
    s = str(valor)
    # Após "PED."
    m = re.search(r'PED[.\s]+(.{3,40})', s, re.IGNORECASE)
    if m:
        return m.group(1).strip().split(' - ')[0].strip()
    # Última parte após " - "
    partes = s.split(' - ')
    if len(partes) > 1:
        return partes[-1].strip()[:40]
    return None


def _procurar_pt_generico(df):
    """Busca por padrões tipo 045/26 ou 045/2026 em toda a aba."""
    try:
        textos = df.fillna("").astype(str).values.flatten().tolist()
    except Exception:
        return None
    padrao = re.compile(r"\b(\d{2,3})\s*/\s*(\d{2,4})\b")
    for t in textos:
        m = padrao.search(t)
        if m:
            a, b = m.group(1), m.group(2)
            if len(b) == 2:
                b = f"20{b}"
            return f"{a}/{b}"
    return None


def _extrair_proto_texto_compos(df):
    """Extrai texto completo de protocolo+procedência em composição (ex.: "PT 045/2026 - PEDREIRA IBIPORÃ")."""
    candidatos = []
    def _limpa(val):
        return str(val).strip()

    for (r, c) in [(7, 8), (9, 1)]:
        try:
            candidatos.append(_limpa(df.iloc[r, c]))
        except Exception:
            pass

    try:
        textos = df.fillna("").astype(str).values.flatten().tolist()
        candidatos.extend(textos)
    except Exception:
        pass

    padrao = re.compile(r"PT\s*([0-9]{2,3}\s*/\s*[0-9]{2,4}).{0,20}?-\s*([^\n]+)", re.IGNORECASE)
    for cand in candidatos:
        m = padrao.search(_limpa(cand))
        if m:
            pt_raw = m.group(1).replace(" ", "")
            proc = m.group(2).strip()
            return f"PT {pt_raw} - {proc}"

    m_gen = re.search(r"\b([0-9]{2,3}\s*/\s*[0-9]{2,4})\b", " ".join(map(_limpa, candidatos)))
    if m_gen:
        return f"PT {m_gen.group(1).replace(' ', '')}"
    return None


def _extrair_campos_pasta_proj_marshall(nome_pasta):
    """
    Extrai campos de pasta PROJETO CAUQ MARSHALL.
    Padrão: "001.2.2026 - FX C DER.PR - PED. IBIPORÃ - CAP BORRACHA"
    Retorna dict {codigo, faixa, pedreira, cap} ou None.
    """
    nome = nome_pasta.strip()
    # Detectar código NNN.M.AAAA ou NNN.A.AAAA
    m_cod = re.match(r'^(\d{3}\.[A-Z0-9]+\.\d{4})', nome, re.IGNORECASE)
    if not m_cod:
        m_cod = _RE_COD_MRAF.match(nome)
    if not m_cod:
        return None
    codigo = m_cod.group(1)
    resto  = nome[m_cod.end():].lstrip(' -')
    partes = [p.strip() for p in resto.split(' - ')]
    faixa    = partes[0] if partes else ''
    pedreira = ''
    cap      = ''
    for p in partes[1:]:
        p_up = p.upper()
        if re.search(r'\bPED[.\s]', p_up) or 'PEDREIRA' in p_up:
            pedreira = re.sub(r'^PED[.\s]*', '', p, flags=re.IGNORECASE).strip()
        elif re.search(r'\bCAP\b|\bBORRACHA\b|\bSBS\b|\bELÁSTOMERO\b', p_up):
            cap = p
    # Fallback: sem PED., usa segunda parte como pedreira
    if not pedreira and len(partes) > 1:
        pedreira = partes[1]
    if not cap and len(partes) > 2:
        cap = partes[-1]
    return {'codigo': codigo, 'faixa': faixa, 'pedreira': pedreira, 'cap': cap}


def _extrair_pedreira_pasta_comp(nome_pasta):
    """
    Extrai pedreira de pasta COMPOSIÇÃO CAUQ.
    Padrão: "01. PED. IBIPORÃ - FAIXA C DER-PR - CAP AB 08"
    """
    # Remove prefixo "NN. "
    s = re.sub(r'^\d+\.\s*', '', nome_pasta.strip())
    # Remove "PED." inicial
    s = re.sub(r'^PED[.\s]*', '', s, flags=re.IGNORECASE)
    # Pega até o primeiro " - "
    partes = s.split(' - ')
    return partes[0].strip()[:60]


def _extrair_pedreira_pasta_pion(nome_pasta):
    """
    Extrai pedreira de pasta PIONEIRO.
    Padrão: "01. FX 9-SPV-16,0mm - Pedreira Central"
    """
    # Remove prefixo "NN. "
    s = re.sub(r'^\d+\.\s*', '', nome_pasta.strip())
    partes = [p.strip() for p in s.split(' - ')]
    # Última parte geralmente é o nome da pedreira
    if len(partes) > 1:
        return partes[-1][:60]
    return partes[0][:60] if partes else ''


def _varrer_composicoes_cauq(dir_comp, ano):
    """
    Varre pasta 003-COMPOSIÇÕES e retorna lista de records.
    Cada item = um arquivo Excel em uma sub-pasta de pedreira.

    Record keys:
        pedreira_pasta, pedreira_norm, pt_excel, pt_raw,
        caminho_xl, arquivo_xl, status,
        pdf_match_excel, tem_pdf,
        d_comp  (dict de células)
    """
    records = []
    dir_ano = _encontrar_pasta_ano_generico(dir_comp, ano)
    if not dir_ano:
        return records
    try:
        subpastas = sorted(os.listdir(dir_ano))
    except OSError:
        return records

    for nome_sub in subpastas:
        path_sub = os.path.join(dir_ano, nome_sub)
        if not os.path.isdir(path_sub):
            continue
        ped_pasta  = _extrair_pedreira_pasta_comp(nome_sub)
        ped_norm   = _normalizar_pedreira_match(ped_pasta)

        arq_xl, tem_pdf = _inspecionar_pasta(path_sub)
        if not arq_xl:
            continue

        caminho_xl = os.path.join(path_sub, arq_xl)
        d_comp     = _ler_excel_composicao_completo(caminho_xl)

        # PT extraído do campo mesclado B10 ou I8; fallback para busca livre
        pt_raw = d_comp.get('B10') or d_comp.get('I8') or d_comp.get('I7') or d_comp.get('PT_LIVRE') or ''
        pt_excel = _extrair_pt_de_campo_mesclado(pt_raw)

        # PDF com mesmo nome do Excel → composição entregue
        pdf_match = _pdf_tem_nome_similar_excel(path_sub, arq_xl)
        status = 'OK' if pdf_match or tem_pdf else 'ANDAMENTO'

        records.append({
            'pedreira_pasta': ped_pasta,
            'pedreira_norm':  ped_norm,
            'pt_raw':         pt_raw,
            'PROTOCOLO_COMP': d_comp.get('PROTOCOLO_TEXTO', ''),
            'pt_livre':       d_comp.get('PT_LIVRE', ''),
            'caminho_xl':     caminho_xl,
            'arquivo_xl':     arq_xl,
            'status':         status,
            'pdf_match_excel': pdf_match,
            'tem_pdf':        tem_pdf,
            'pasta':          path_sub,
            'nome_sub':       nome_sub,
            'd_comp':         d_comp,
        })
    return records


def _varrer_pioneiros_cauq(dir_pion, ano):
    """
    Varre pasta 004-TRAÇOS PIONEIROS e retorna lista de records.
    Detecta pasta de ano genérica: "PIONEIRO 2026", "2026", etc.

    Record keys:
        pedreira_pasta, pedreira_norm, pt_excel,
        caminho_xl, arquivo_xl, status,
        tem_pdf, pdf_com_proj,
        d_pion (dict de células)
    """
    records = []
    dir_ano = _encontrar_pasta_ano_generico(dir_pion, ano)
    if not dir_ano:
        return records
    try:
        subpastas = sorted(os.listdir(dir_ano))
    except OSError:
        return records

    for nome_sub in subpastas:
        path_sub = os.path.join(dir_ano, nome_sub)
        if not os.path.isdir(path_sub):
            continue
        ped_pasta = _extrair_pedreira_pasta_pion(nome_sub)
        ped_norm  = _normalizar_pedreira_match(ped_pasta)

        arq_xl, tem_pdf = _inspecionar_pasta(path_sub)
        if not arq_xl:
            continue

        caminho_xl = os.path.join(path_sub, arq_xl)

        # Ler aba PROJ_TRAÇO PIONEIRO (E07)
        d_proj_pion = _ler_excel_aba_proj_pion(caminho_xl)
        # Ler aba 007 E1 (L09)
        d_ens       = _ler_excel_ensaios(caminho_xl)

        # PT: E07 > L09 do Excel
        pt_raw   = d_proj_pion.get('E7_PION') or d_ens.get('L9') or ''
        pt_excel = _extrair_pt_de_campo_mesclado(pt_raw) or d_ens.get('L9') or ''
        # Pedreira do Excel: E07 ou nome da pasta
        ped_xl = _extrair_pedreira_de_campo_mesclado(d_proj_pion.get('E7_PION') or '') or ped_pasta

        # PDF com "PROJ" no nome → sinalizar conferência
        pdf_com_proj = _tem_pdf_nome(path_sub, 'PROJ')
        status = 'OK' if tem_pdf else 'ANDAMENTO' if arq_xl else 'VAZIO'

        records.append({
            'pedreira_pasta': ped_pasta,
            'pedreira_norm':  ped_norm,
            'pedreira_xl':    ped_xl,
            'pt_excel':       str(pt_excel).strip(),
            'pt_raw':         pt_raw,
            'caminho_xl':     caminho_xl,
            'arquivo_xl':     arq_xl,
            'status':         status,
            'tem_pdf':        tem_pdf,
            'pdf_com_proj':   pdf_com_proj,
            'pasta':          path_sub,
            'nome_sub':       nome_sub,
            'd_pion':         d_ens,
            'd_proj_pion':    d_proj_pion,
        })
    return records


def _varrer_projetos_marshall(dir_proj, ano):
    """
    Varre pasta 006-PROJETOS para CAUQ Marshall.
    Detecta pasta de ano genérica: "_2026 PROJETOS MARSHALL".

    Record keys:
        codigo, faixa, pedreira, cap,
        pedreira_norm, pt_excel,
        caminho_xl, arquivo_xl,
        insp (dict _inspecionar_projeto_cauq),
        status, tem_pdf,
        d_proj (células aba 007), d_gst (células GST_CADASTRO)
    """
    records = []
    dir_ano = _encontrar_pasta_ano_generico(dir_proj, ano)
    if not dir_ano:
        return records
    try:
        subpastas = sorted(os.listdir(dir_ano))
    except OSError:
        return records

    for nome_sub in subpastas:
        path_sub = os.path.join(dir_ano, nome_sub)
        if not os.path.isdir(path_sub):
            continue

        campos = _extrair_campos_pasta_proj_marshall(nome_sub)
        if not campos:
            continue  # ignora pastas sem código NNN.M.AAAA

        codigo   = campos['codigo']
        faixa    = campos.get('faixa', '')
        pedreira = campos.get('pedreira', '')
        cap      = campos.get('cap', '')
        ped_norm = _normalizar_pedreira_match(pedreira)

        # Procura em sub-pasta de entrega ou diretamente
        dir_ent = _encontrar_pasta_entrega(path_sub)
        insp    = _inspecionar_projeto_cauq(dir_ent)

        caminho_xl = os.path.join(dir_ent, insp['arquivo_xl']) if insp['arquivo_xl'] else None

        # Leitura Excel: GST_CADASTRO + aba 007
        d_gst  = _ler_excel_gst_cadastro(caminho_xl) if caminho_xl else {}
        d_proj = _ler_excel_ensaios(caminho_xl)      if caminho_xl else {}

        # PT: L09 aba 007 > E11 GST_CADASTRO > código da pasta
        pt_excel = (d_proj.get('L9') or
                    d_proj.get('L09_PROTO') or
                    d_gst.get('E11_PROJ_NUM') or
                    codigo)

        # Pedreira do Excel: E27 GST > nome da pasta
        ped_xl = d_gst.get('E27_PEDREIRA') or pedreira

        records.append({
            'codigo':         codigo,
            'faixa':          faixa,
            'pedreira':       ped_xl or pedreira,
            'pedreira_norm':  ped_norm,
            'cap':            cap,
            'pt_excel':       str(pt_excel).strip(),
            'caminho_xl':     caminho_xl or '',
            'arquivo_xl':     insp['arquivo_xl'] or '',
            'insp':           insp,
            'status':         insp['status'],
            'tem_pdf':        insp['tem_pdf'],
            'entregue':       insp['entregue'],
            'd_proj':         d_proj,
            'd_gst':          d_gst,
            'pasta':          path_sub,
            'nome_sub':       nome_sub,
        })
    return records


# ======================================================================================
# 4.3. SCANNER POR PADRÃO
# ======================================================================================

def _varrer_cauq(base_dir, tipo_key, tipo_cfg, db_compasa_pts=None):
    """
    Varre estrutura CAUQ (Marshall, Superpave, BGS, Solo Cimento, Camadas Granulares).

    LÓGICA DE ENUMERAÇÃO (fonte primária = DIRETÓRIO):
      1. Localiza as 3 fases: 003-COMPOSIÇÕES, 004-TRAÇOS PIONEIROS, 006-PROJETOS
      2. Se 006-PROJETOS existe com padrão NNN.M.AAAA → usa projetos como âncora
         Para cada projeto:
           a. Parseia: codigo, faixa, pedreira, cap do nome da pasta
           b. Lê Excel: GST_CADASTRO (E11, E27) + aba 007 (L09)
           c. Inspeta PDFs A/B/C → status OK
           d. Fuzzy-localiza COMPOSIÇÃO por pedreira (≥70)
           e. Fuzzy-localiza PIONEIRO por pedreira (≥70)
           f. FLAG_CONFERENCIA se:
              - Pioneiro existe mas composição não encontrada  → "PION_SEM_COMP"
              - PDF na pasta do pioneiro contém "PROJ"         → "PDF_PROJ"
      3. Se 006-PROJETOS não existe → itera COMPOSIÇÕES como âncora (padrão legado)

    CAMPOS ADICIONAIS (CAUQ Marshall):
        FAIXA, CAP, XL_B10_COMP, XL_E7_PION, XL_E11_PROJ, XL_E27_PROJ,
        PDF_A (Certificados), PDF_B (DPH), PDF_C (ART),
        FLAG_CONFERENCIA, MOTIVO_CONFERENCIA

    STATUS DE COMPOSIÇÃO:
        PDF com mesmo nome base do Excel → 'OK' (entregue)
        Só Excel                          → 'ANDAMENTO'
        Vazio                             → 'VAZIO'

    MATCHING PEDREIRA (fuzzy threshold ≥ 70):
        Normaliza: remove acentos, "PED.", pontuação → fuzz.partial_ratio
    """
    registros    = []
    tem_pioneiro = tipo_cfg.get('tem_pioneiro', True)
    ano          = _ANO_PROJETOS_ATIVO

    # ── Localizar sub-fases ────────────────────────────────────────────────────────
    dir_comp = _encontrar_subfase(base_dir, _SUBFASES_CAUQ['COMPOSICAO'])
    dir_pion = _encontrar_subfase(base_dir, _SUBFASES_CAUQ['PIONEIRO']) if tem_pioneiro else None
    dir_proj = _encontrar_subfase(base_dir, _SUBFASES_CAUQ['PROJETO'])

    if not dir_proj and not dir_comp:
        return registros

    # ==========================================================================
    # RAMO A — 006-PROJETOS com padrão NNN.M.AAAA  (CAUQ Marshall)
    #         Projetos são a âncora; composições e pioneiros são cruzados.
    # ==========================================================================
    regs_proj_marshall = []
    if dir_proj:
        regs_proj_marshall = _varrer_projetos_marshall(dir_proj, ano)

    if regs_proj_marshall:
        # Pré-indexar composições e pioneiros por pedreira normalizada
        recs_comp = _varrer_composicoes_cauq(dir_comp, ano) if dir_comp else []
        recs_pion = _varrer_pioneiros_cauq(dir_pion, ano)  if (dir_pion and tem_pioneiro) else []

        def _melhor_match_comp(ped_norm, threshold=70):
            """Retorna o record de composição com melhor score de pedreira."""
            melhor, melhor_sc = None, 0
            for rc in recs_comp:
                sc = _pedreira_score(ped_norm, rc['pedreira_norm'])
                if sc >= threshold and sc > melhor_sc:
                    melhor_sc = sc
                    melhor = rc
            return melhor

        def _melhor_match_pion(ped_norm, threshold=70):
            """Retorna o record de pioneiro com melhor score de pedreira."""
            melhor, melhor_sc = None, 0
            for rp in recs_pion:
                sc = _pedreira_score(ped_norm, rp['pedreira_norm'])
                if sc >= threshold and sc > melhor_sc:
                    melhor_sc = sc
                    melhor = rp
            return melhor

        for rp in regs_proj_marshall:
            codigo   = rp['codigo']
            faixa    = rp.get('faixa', '')
            pedreira = rp.get('pedreira', '')
            cap      = rp.get('cap', '')
            ped_norm = rp.get('pedreira_norm') or _normalizar_pedreira_match(pedreira)
            insp     = rp.get('insp', {})

            # ── Status PROJETO ───────────────────────────────────────────────
            st_proj      = insp.get('status', 'VAZIO')
            arq_proj_nome= insp.get('arquivo_xl', '')
            pdf_proj     = insp.get('tem_pdf', False)
            pdf_a        = insp.get('pdf_a', False)
            pdf_b        = insp.get('pdf_b', False)
            pdf_c        = insp.get('pdf_c', False)
            caminho_xl_proj = rp.get('caminho_xl', '')

            d_proj = rp.get('d_proj', {})
            d_gst  = rp.get('d_gst', {})

            # ── COMPOSIÇÃO — fuzzy match por pedreira ────────────────────────
            rc = _melhor_match_comp(ped_norm)
            st_comp       = rc['status']      if rc else 'VAZIO'
            arq_comp_nome = rc['arquivo_xl']  if rc else ''
            caminho_xl_comp = rc['caminho_xl']if rc else ''
            pdf_comp      = rc['tem_pdf']     if rc else False
            d_comp        = rc['d_comp']      if rc else {}
            pasta_comp    = rc['pasta']       if rc else ''

            # Status composição: PDF nome similar ao Excel = OK
            if rc and rc.get('pdf_match_excel'):
                st_comp = 'OK'
            elif rc and pdf_comp:
                st_comp = 'OK'

            # ── PIONEIRO — fuzzy match por pedreira ──────────────────────────
            rpi = _melhor_match_pion(ped_norm) if tem_pioneiro else None
            st_pion       = rpi['status']      if rpi else 'NAO_APLICAVEL'
            arq_pion_nome = rpi['arquivo_xl']  if rpi else ''
            caminho_xl_pion = rpi['caminho_xl']if rpi else ''
            pdf_pion      = rpi['tem_pdf']     if rpi else False
            d_pion        = rpi['d_pion']      if rpi else {}
            d_proj_pion   = rpi['d_proj_pion'] if rpi else {}
            pasta_pion    = rpi['pasta']        if rpi else ''

            # ── PT final ─────────────────────────────────────────────────────
            # Precedência: L09 aba 007 do proj > E11 GST > L09 pioneiro > código pasta
            pt_num_final = (
                d_proj.get('L9') or d_proj.get('L09_PROTO') or
                d_gst.get('E11_PROJ_NUM') or
                d_pion.get('L9') or
                rpi['pt_excel'] if rpi else None or
                codigo
            )
            pt_num_final = str(pt_num_final or codigo).strip()
            pt_num_clean = re.sub(r'[^0-9]', '', pt_num_final.split('/')[0])

            # ── Pedreira final ────────────────────────────────────────────────
            pedreira_final = (
                d_gst.get('E27_PEDREIRA') or
                d_comp.get('C10') or
                pedreira
            )

            # ── Mistura ──────────────────────────────────────────────────────
            mistura_final = faixa or d_comp.get('H9', '') or ''

            # ── FLAG_CONFERENCIA ─────────────────────────────────────────────
            flag_conf   = False
            motivo_conf = ''
            if tem_pioneiro:
                # 1. Pioneiro sem composição
                if rpi and not rc:
                    flag_conf   = True
                    motivo_conf = 'PION_SEM_COMP'
                # 2. PDF com "PROJ" na pasta do pioneiro
                if rpi and _tem_pdf_nome(pasta_pion, 'PROJ'):
                    flag_conf   = True
                    motivo_conf = ('PION_SEM_COMP+PDF_PROJ' if motivo_conf else 'PDF_PROJ')

            # ── Classificação ─────────────────────────────────────────────────
            cliente_xl = (
                d_proj.get('B8') or d_pion.get('B8') or
                d_comp.get('B9') or d_comp.get('B8') or ''
            )
            if cliente_xl and not _e_compasa(cliente_xl):
                classif = _classificar_cliente_externo(cliente_xl)
                cliente = cliente_xl
            elif db_compasa_pts and pt_num_clean in db_compasa_pts:
                classif = 'COMPASA'
                cliente = 'COMPASA DO BRASIL'
            elif cliente_xl and _e_compasa(cliente_xl):
                classif = 'COMPASA'
                cliente = 'COMPASA DO BRASIL'
            else:
                classif = 'EXTERNO'
                cliente = cliente_xl or ''

            # ── Status geral ──────────────────────────────────────────────────
            tem_pdf = pdf_proj or pdf_comp or pdf_pion or pdf_a or pdf_b or pdf_c
            if st_proj == 'OK' or (pdf_a or pdf_b or pdf_c):
                status_geral = 'OK'
            elif st_comp == 'OK':
                status_geral = 'OK'
            elif st_proj == 'ANDAMENTO' or st_comp == 'ANDAMENTO':
                status_geral = 'ANDAMENTO'
            elif st_pion == 'ANDAMENTO':
                status_geral = 'ANDAMENTO'
            else:
                status_geral = 'VAZIO'

            registros.append({
                # Identidade
                'CODIGO':          codigo,
                'PT_NUM':          pt_num_final,
                'PT_PASTA':        codigo,
                'CLIENTE':         cliente,
                'PEDREIRA':        pedreira_final,
                'MISTURA':         mistura_final,
                'FAIXA':           faixa,
                'CAP':             cap,
                'ANO_PASTA':       ano,
                'TIPO_PROJETO':    tipo_key,
                'TIPO_LABEL':      tipo_cfg['label'],
                'TIPO_ICONE':      tipo_cfg['icone'],
                'TEM_PIONEIRO':    tem_pioneiro,
                # Status por fase
                'STATUS':          status_geral,
                'STATUS_COMP':     st_comp,
                'STATUS_PION':     st_pion,
                'STATUS_PROJ':     st_proj,
                # Arquivos
                'ARQUIVO_COMP':    arq_comp_nome or '',
                'ARQUIVO_PION':    arq_pion_nome or '',
                'ARQUIVO_PROJ':    arq_proj_nome or '',
                'CAMINHO_COMP':    caminho_xl_comp or '',
                'CAMINHO_PION':    caminho_xl_pion or '',
                'CAMINHO_PROJ':    caminho_xl_proj or '',
                'TEM_PDF':         tem_pdf,
                # PDFs específicos A/B/C
                'PDF_CERTIFICADOS': pdf_a,
                'PDF_DPH':          pdf_b,
                'PDF_ART':          pdf_c,
                # Células Excel
                'XL_I8_COMP':      d_comp.get('I8', ''),
                'XL_C10_COMP':     d_comp.get('C10', ''),
                'XL_B9_COMP':      d_comp.get('B9', ''),
                'XL_B10_COMP':     d_comp.get('B10', ''),
                'XL_L9_PROJ':      d_proj.get('L9', ''),
                'XL_B8_PROJ':      d_proj.get('B8', ''),
                'XL_H8_PROJ':      d_proj.get('H8', ''),
                'XL_L9_PION':      d_pion.get('L9', ''),
                'XL_B8_PION':      d_pion.get('B8', ''),
                'XL_E7_PION':      d_proj_pion.get('E7_PION', ''),
                'XL_E11_PROJ':     d_gst.get('E11_PROJ_NUM', ''),
                'XL_E27_PROJ':     d_gst.get('E27_PEDREIRA', ''),
                'XL_L09_PROTO':    d_proj.get('L09_PROTO', '') or d_proj.get('L9', ''),
                # Metadados
                'CLASSIFICACAO':   classif,
                'PASTA':           rp.get('pasta', ''),
                'NOME_PASTA':      rp.get('nome_sub', ''),
                'FLAG_CONFERENCIA':  flag_conf,
                'MOTIVO_CONFERENCIA': motivo_conf,
            })

        # Se projetos Marshall foram encontrados → retornar agora
        if registros:
            return registros

    # ==========================================================================
    # RAMO B — Sem padrão NNN.M.AAAA  (Superpave, BGS, Solo Cimento, etc.)
    #          Itera COMPOSIÇÕES como âncora e cruza com PIONEIROS por pedreira.
    # ==========================================================================
    dir_referencia = dir_comp
    if not dir_referencia:
        return registros

    # Detectar pasta de ano
    pasta_ano_path = _encontrar_pasta_ano_generico(dir_referencia, ano)
    if not pasta_ano_path:
        # Fallback: tenta estrutura legada (pastas numeradas diretamente)
        try:
            for p in os.listdir(dir_referencia):
                m = _RE_ANO_PASTA.search(p)
                if m and m.group(1) == ano:
                    pasta_ano_path = os.path.join(dir_referencia, p)
                    break
        except OSError:
            pass
    if not pasta_ano_path:
        return registros

    try:
        subpastas_comp = sorted(os.listdir(pasta_ano_path))
    except OSError:
        return registros

    # Pré-carregar pioneiros para cruzamento
    recs_pion_b = _varrer_pioneiros_cauq(dir_pion, ano) if (dir_pion and tem_pioneiro) else []

    def _match_pion_b(ped_norm, threshold=70):
        melhor, melhor_sc = None, 0
        for rp in recs_pion_b:
            sc = _pedreira_score(ped_norm, rp['pedreira_norm'])
            if sc >= threshold and sc > melhor_sc:
                melhor_sc = sc
                melhor = rp
        return melhor

    for nome_sub in subpastas_comp:
        path_sub = os.path.join(pasta_ano_path, nome_sub)
        if not os.path.isdir(path_sub):
            continue

        ped_pasta = _extrair_pedreira_pasta_comp(nome_sub)
        ped_norm  = _normalizar_pedreira_match(ped_pasta)

        arq_xl, tem_pdf = _inspecionar_pasta(path_sub)
        if not arq_xl:
            st_comp = 'VAZIO'
            d_comp  = {}
            caminho_xl_comp = None
        else:
            caminho_xl_comp = os.path.join(path_sub, arq_xl)
            d_comp = _ler_excel_composicao_completo(caminho_xl_comp)
            pdf_match = _pdf_tem_nome_similar_excel(path_sub, arq_xl)
            st_comp = 'OK' if (pdf_match or tem_pdf) else 'ANDAMENTO'

        # PT da composição
        pt_raw = d_comp.get('B10') or d_comp.get('I8') or d_comp.get('I7') or ''
        pt_num_final = _extrair_pt_de_campo_mesclado(pt_raw) or _extrair_pt_de_nome(nome_sub) or ''
        pt_num_clean = re.sub(r'[^0-9]', '', str(pt_num_final))

        # Pedreira final da composição
        pedreira_final = d_comp.get('C10') or ped_pasta

        # ── Pioneiro cruzado ─────────────────────────────────────────────────
        rpi = _match_pion_b(ped_norm) if tem_pioneiro else None
        st_pion       = rpi['status']      if rpi else ('NAO_APLICAVEL' if not tem_pioneiro else 'VAZIO')
        arq_pion_nome = rpi['arquivo_xl']  if rpi else ''
        caminho_xl_pion = rpi['caminho_xl']if rpi else ''
        pdf_pion      = rpi['tem_pdf']     if rpi else False
        d_pion        = rpi['d_pion']      if rpi else {}
        d_proj_pion   = rpi['d_proj_pion'] if rpi else {}
        pasta_pion    = rpi['pasta']        if rpi else ''

        # ── FLAG CONFERÊNCIA ─────────────────────────────────────────────────
        flag_conf   = False
        motivo_conf = ''
        if tem_pioneiro:
            if rpi and st_comp == 'VAZIO':
                flag_conf   = True
                motivo_conf = 'PION_SEM_COMP'
            if rpi and _tem_pdf_nome(pasta_pion, 'PROJ'):
                flag_conf   = True
                motivo_conf = ('PION_SEM_COMP+PDF_PROJ' if motivo_conf else 'PDF_PROJ')

        # ── Projeto (entrega) ────────────────────────────────────────────────
        st_proj = 'VAZIO'
        arq_proj_nome = ''
        caminho_xl_proj = ''
        pdf_proj = False
        d_proj: dict = {}
        pdf_a = pdf_b = pdf_c = False

        if dir_proj:
            pasta_ano_proj = _encontrar_pasta_ano_generico(dir_proj, ano)
            if pasta_ano_proj:
                # Localiza pasta de projeto por pedreira + código
                pasta_proj_p = _localizar_pasta_pedreira(pasta_ano_proj, ped_pasta, pt_ref=pt_num_final or '')
                if pasta_proj_p:
                    dir_ent  = _encontrar_pasta_entrega(pasta_proj_p)
                    insp_p   = _inspecionar_projeto_cauq(dir_ent)
                    st_proj  = insp_p['status']
                    arq_proj_nome = insp_p['arquivo_xl'] or ''
                    pdf_proj = insp_p['tem_pdf']
                    pdf_a    = insp_p['pdf_a']
                    pdf_b    = insp_p['pdf_b']
                    pdf_c    = insp_p['pdf_c']
                    caminho_xl_proj = os.path.join(dir_ent, arq_proj_nome) if arq_proj_nome else ''
                    d_proj = _ler_excel_ensaios(caminho_xl_proj) if caminho_xl_proj else {}

        # ── Classificação ─────────────────────────────────────────────────────
        cliente_xl = d_comp.get('B9') or d_comp.get('B8') or d_proj.get('B8') or ''
        if cliente_xl and not _e_compasa(cliente_xl):
            classif = _classificar_cliente_externo(cliente_xl)
            cliente = cliente_xl
        elif db_compasa_pts and pt_num_clean in db_compasa_pts:
            classif = 'COMPASA'
            cliente = 'COMPASA DO BRASIL'
        elif cliente_xl and _e_compasa(cliente_xl):
            classif = 'COMPASA'
            cliente = 'COMPASA DO BRASIL'
        else:
            classif = 'EXTERNO'
            cliente = cliente_xl or ''

        # ── Status geral ──────────────────────────────────────────────────────
        tem_pdf = tem_pdf or pdf_pion or pdf_proj or pdf_a or pdf_b or pdf_c
        if st_proj == 'OK' or st_comp == 'OK' or pdf_a or pdf_b or pdf_c:
            status_geral = 'OK'
        elif st_proj == 'ANDAMENTO' or st_comp == 'ANDAMENTO':
            status_geral = 'ANDAMENTO'
        elif st_pion not in ('NAO_APLICAVEL', 'VAZIO'):
            status_geral = 'ANDAMENTO'
        else:
            status_geral = 'VAZIO'

        registros.append({
            # Identidade
            'CODIGO':          pt_num_final or nome_sub[:20],
            'PT_NUM':          pt_num_final,
            'PT_PASTA':        _extrair_pt_de_nome(nome_sub) or '',
            'CLIENTE':         cliente,
            'PEDREIRA':        pedreira_final,
            'MISTURA':         d_comp.get('H9', ''),
            'FAIXA':           '',
            'CAP':             '',
            'ANO_PASTA':       ano,
            'TIPO_PROJETO':    tipo_key,
            'TIPO_LABEL':      tipo_cfg['label'],
            'TIPO_ICONE':      tipo_cfg['icone'],
            'TEM_PIONEIRO':    tem_pioneiro,
            # Status por fase
            'STATUS':          status_geral,
            'STATUS_COMP':     st_comp,
            'STATUS_PION':     st_pion,
            'STATUS_PROJ':     st_proj,
            # Arquivos
            'ARQUIVO_COMP':    arq_xl or '',
            'ARQUIVO_PION':    arq_pion_nome or '',
            'ARQUIVO_PROJ':    arq_proj_nome or '',
            'CAMINHO_COMP':    caminho_xl_comp or '',
            'CAMINHO_PION':    caminho_xl_pion or '',
            'CAMINHO_PROJ':    caminho_xl_proj or '',
            'TEM_PDF':         tem_pdf,
            # PDFs específicos A/B/C
            'PDF_CERTIFICADOS': pdf_a,
            'PDF_DPH':          pdf_b,
            'PDF_ART':          pdf_c,
            # Células Excel
            'XL_I8_COMP':      d_comp.get('I8', ''),
            'XL_C10_COMP':     d_comp.get('C10', ''),
            'XL_B9_COMP':      d_comp.get('B9', ''),
            'XL_B10_COMP':     d_comp.get('B10', ''),
            'XL_L9_PROJ':      d_proj.get('L9', ''),
            'XL_B8_PROJ':      d_proj.get('B8', ''),
            'XL_H8_PROJ':      d_proj.get('H8', ''),
            'XL_L9_PION':      d_pion.get('L9', ''),
            'XL_B8_PION':      d_pion.get('B8', ''),
            'XL_E7_PION':      d_proj_pion.get('E7_PION', ''),
            'XL_E11_PROJ':     '',
            'XL_E27_PROJ':     '',
            'XL_L09_PROTO':    d_proj.get('L09_PROTO', '') or d_proj.get('L9', ''),
            # Metadados
            'CLASSIFICACAO':   classif,
            'PASTA':           path_sub,
            'NOME_PASTA':      nome_sub,
            'FLAG_CONFERENCIA':   flag_conf,
            'MOTIVO_CONFERENCIA': motivo_conf,
        })

    return registros


def _varrer_mraf(base_dir, tipo_key, tipo_cfg, db_compasa_pts=None):
    """
    Varre estrutura MRAF: BASE/004-PROJETOS/_AAAA PROJETOS MRAF/{COD} - {CLI} - ...

    Extrai cliente e pedreira diretamente do nome da pasta (padrão confirmado):
        {NNN.M.AAAA} - {CLIENTE} - {PED.*} - {MISTURA}

    ENRIQUECIMENTO EXCEL:
        Após localizar o arquivo .xlsm na pasta 004-ENTREGA, lê as células:
          Aba de ensaios → L9 (PT), B8 (cliente), H8 (mistura)
        O cliente do Excel valida / substitui o do nome da pasta quando divergem.
    """
    registros = []
    projetos_dir_root = os.path.join(base_dir, '004-PROJETOS')
    comp_dir_root     = os.path.join(base_dir, '003-COMPOSIÇÕES')
    if not os.path.isdir(projetos_dir_root):
        return registros

    try:
        pastas_ano = [
            p for p in os.listdir(projetos_dir_root)
            if re.match(r'_?\d{4}', p) and os.path.isdir(os.path.join(projetos_dir_root, p))
        ]
    except OSError:
        return registros

    for pasta_ano_nome in pastas_ano:
        m_ano = _RE_ANO_PASTA.search(pasta_ano_nome)
        ano_str = m_ano.group(1) if m_ano else ''
        if ano_str != _ANO_PROJETOS_ATIVO:
            continue
        pasta_ano_path = os.path.join(projetos_dir_root, pasta_ano_nome)
        pasta_ano_comp = None
        if os.path.isdir(comp_dir_root):
            pasta_ano_comp = _encontrar_pasta_ano_generico(comp_dir_root, ano_str) or os.path.join(comp_dir_root, f"{ano_str}")

        try:
            nomes_proj = sorted(os.listdir(pasta_ano_path))
        except OSError:
            continue

        for nome_proj in nomes_proj:
            pasta_proj_path = os.path.join(pasta_ano_path, nome_proj)
            if not os.path.isdir(pasta_proj_path):
                continue

            campos = _extrair_campos_pasta_mraf(nome_proj)
            if not campos or not campos.get('codigo'):
                continue

            codigo          = campos['codigo']
            cliente_pasta   = campos.get('cliente', '')
            pedreira_pasta  = campos.get('pedreira', '')
            mistura_pasta   = campos.get('mistura', '')

            # ── Status e arquivo da pasta de entrega ─────────────────────────────
            dir_ent   = _encontrar_pasta_entrega(pasta_proj_path)
            status, arq_nome, tem_pdf = _status_de_pasta(dir_ent)
            caminho_xl = os.path.join(dir_ent, arq_nome) if arq_nome else None

            # ── Leitura profunda do Excel de entrega ─────────────────────────────
            d_proj: dict = {}
            if caminho_xl:
                d_proj = _ler_excel_ensaios(caminho_xl)

            # ── Buscar composição associada (pasta 003-COMPOSIÇÕES/ANO) ─────────
            d_comp: dict = {}
            st_comp = 'NAO_APLICAVEL'
            arq_comp_nome = ''
            caminho_xl_comp = ''
            tem_pdf_comp = False
            if pasta_ano_comp and os.path.isdir(pasta_ano_comp):
                try:
                    sub_comps = sorted(os.listdir(pasta_ano_comp))
                except OSError:
                    sub_comps = []
                melhor = None
                melhor_sc = 0
                ped_hint = pedreira_pasta.upper().replace('PED.', '').replace('PEDREIRA', '').strip()
                for subc in sub_comps:
                    path_c = os.path.join(pasta_ano_comp, subc)
                    if not os.path.isdir(path_c):
                        continue
                    sc = fuzz.partial_ratio(ped_hint, subc.upper()) if ped_hint else 0
                    if sc >= 70 and sc > melhor_sc:
                        melhor_sc = sc
                        melhor = path_c
                if melhor:
                    arq_comp_nome, tem_pdf_comp = _inspecionar_pasta(melhor)
                    if arq_comp_nome:
                        caminho_xl_comp = os.path.join(melhor, arq_comp_nome)
                        d_comp = _ler_excel_composicao_completo(caminho_xl_comp)
                        st_comp = 'OK' if tem_pdf_comp else 'ANDAMENTO'
                    else:
                        st_comp = 'VAZIO'

            # ── Consolidar: Excel tem precedência sobre nome da pasta ─────────────
            campos_xl = _consolidar_campos_excel(
                d_comp=d_comp, d_pion={}, d_proj=d_proj,
                nome_pasta=nome_proj,
                cliente_pasta=cliente_pasta,
            )

            # PT: usa código do nome (MRAF tem formato proprio) mas valida com Excel
            pt_final      = campos_xl['PT_EXCEL'] or codigo
            pedreira_final = campos_xl['PEDREIRA_EXCEL'] or pedreira_pasta
            mistura_final  = campos_xl['MISTURA_EXCEL']  or mistura_pasta

            # Cliente: Excel > nome da pasta
            cliente_xl = campos_xl['CLIENTE_EXCEL']
            if cliente_xl:
                cliente = cliente_xl
            elif cliente_pasta:
                cliente = cliente_pasta
            else:
                cliente = ''

            classif = _classificar_cliente_externo(cliente)
            _cliente_up = str(cliente).upper()
            if 'COMPASA' in _cliente_up:
                classif = 'COMPASA'
                cliente = 'COMPASA DO BRASIL'

            # Status geral combinando composição (se houver) e entrega
            status_comp = st_comp if 'st_comp' in locals() else status
            if status == 'OK' or tem_pdf:
                status_geral = 'OK'
            elif status_comp == 'OK':
                status_geral = 'OK'
            elif status == 'ANDAMENTO' or status_comp == 'ANDAMENTO':
                status_geral = 'ANDAMENTO'
            else:
                status_geral = status

            registros.append({
                'CODIGO':         codigo,
                'PT_NUM':         pt_final,
                'PT_PASTA':       codigo,
                'CLIENTE':        cliente,
                'PEDREIRA':       pedreira_final,
                'MISTURA':        mistura_final,
                'ANO_PASTA':      ano_str,
                'TIPO_PROJETO':   tipo_key,
                'TIPO_LABEL':     tipo_cfg['label'],
                'TIPO_ICONE':     tipo_cfg['icone'],
                'TEM_PIONEIRO':   False,      # MRAF não tem Pioneiro separado
                # Status
                'STATUS':         status_geral,
                'STATUS_COMP':    status_comp,
                'STATUS_PION':    'NAO_APLICAVEL',
                'STATUS_PROJ':    status,
                # Arquivos
                'ARQUIVO_COMP':   arq_comp_nome or arq_nome or '',
                'ARQUIVO_PION':   '',
                'ARQUIVO_PROJ':   arq_nome or '',
                'CAMINHO_COMP':   caminho_xl_comp or '',
                'CAMINHO_PION':   '',
                'CAMINHO_PROJ':   caminho_xl or '',
                'TEM_PDF':        tem_pdf,
                # Células Excel
                'XL_I8_COMP':     '',
                'XL_C10_COMP':    pedreira_final,
                'XL_B9_COMP':     cliente,
                'XL_L9_PROJ':     d_proj.get('L9', ''),
                'XL_B8_PROJ':     d_proj.get('B8', ''),
                'XL_H8_PROJ':     d_proj.get('H8', ''),
                'XL_L9_PION':     '',
                'XL_B8_PION':     '',
                # Metadados
                'CLASSIFICACAO':  classif,
                'PASTA':          pasta_proj_path,
                'NOME_PASTA':     nome_proj,
            })

    return registros


def _varrer_flat(base_dir, tipo_key, tipo_cfg, db_compasa_pts=None):
    """
    Fallback para estruturas flat: BASE/{ANO}/{PT}_{PEDREIRA}/

    ENRIQUECIMENTO EXCEL:
        Lê células de entrega para extrair cliente real, classificando
        projetos COMPASA e EXTERNOS corretamente mesmo sem padrão de pasta definido.
    """
    registros = []
    try:
        filhos = os.listdir(base_dir)
    except OSError:
        return registros

    tem_pioneiro = tipo_cfg.get('tem_pioneiro', False)

    for pasta_ano_nome in filhos:
        m_ano = re.match(r'^(\d{4})$', pasta_ano_nome)
        if not m_ano:
            continue
        ano_str = m_ano.group(1)
        if ano_str != _ANO_PROJETOS_ATIVO:
            continue
        pasta_ano_path = os.path.join(base_dir, pasta_ano_nome)

        try:
            projetos = sorted([
                p for p in os.listdir(pasta_ano_path)
                if os.path.isdir(os.path.join(pasta_ano_path, p))
            ])
        except OSError:
            continue

        for nome_proj in projetos:
            pt_codigo_pasta = _extrair_pt_de_nome(nome_proj)
            pedreira_pasta  = _extrair_pedreira_de_nome(nome_proj)
            pasta_proj_path = os.path.join(pasta_ano_path, nome_proj)

            dir_ent   = _encontrar_pasta_entrega(pasta_proj_path)
            status, arq_nome, tem_pdf = _status_de_pasta(dir_ent)
            caminho_xl = os.path.join(dir_ent, arq_nome) if arq_nome else None

            # ── Leitura profunda do Excel ─────────────────────────────────────────
            d_xl: dict = {}
            if caminho_xl:
                d_xl = _ler_excel_ensaios(caminho_xl)

            campos_xl = _consolidar_campos_excel(
                d_comp={}, d_pion={}, d_proj=d_xl,
                nome_pasta=nome_proj,
                cliente_pasta='',
            )

            pt_num_final   = campos_xl['PT_EXCEL'] or pt_codigo_pasta or ''
            pedreira_final = campos_xl['PEDREIRA_EXCEL'] or pedreira_pasta
            mistura_final  = campos_xl['MISTURA_EXCEL']
            cliente_xl     = campos_xl['CLIENTE_EXCEL']

            # Classificação
            pt_num_clean = re.sub(r'[^0-9]', '', pt_num_final.split('/')[0])
            if cliente_xl:
                classif = _classificar_cliente_externo(cliente_xl)
                cliente = cliente_xl
            elif db_compasa_pts and pt_num_clean in db_compasa_pts:
                classif = 'COMPASA'
                cliente = 'COMPASA DO BRASIL'
            else:
                classif = 'EXTERNO'
                cliente = ''

            registros.append({
                'CODIGO':         pt_num_final or nome_proj[:20],
                'PT_NUM':         pt_num_final,
                'PT_PASTA':       pt_codigo_pasta or '',
                'CLIENTE':        cliente,
                'PEDREIRA':       pedreira_final,
                'MISTURA':        mistura_final,
                'ANO_PASTA':      ano_str,
                'TIPO_PROJETO':   tipo_key,
                'TIPO_LABEL':     tipo_cfg['label'],
                'TIPO_ICONE':     tipo_cfg['icone'],
                'TEM_PIONEIRO':   tem_pioneiro,
                'STATUS':         status,
                'STATUS_COMP':    status,
                'STATUS_PION':    'NAO_APLICAVEL',
                'STATUS_PROJ':    status,
                'ARQUIVO_COMP':   '',
                'ARQUIVO_PION':   '',
                'ARQUIVO_PROJ':   arq_nome or '',
                'CAMINHO_COMP':   '',
                'CAMINHO_PION':   '',
                'CAMINHO_PROJ':   caminho_xl or '',
                'TEM_PDF':        tem_pdf,
                'XL_I8_COMP':     '',
                'XL_C10_COMP':    pedreira_final,
                'XL_B9_COMP':     cliente,
                'XL_L9_PROJ':     d_xl.get('L9', ''),
                'XL_B8_PROJ':     d_xl.get('B8', ''),
                'XL_H8_PROJ':     d_xl.get('H8', ''),
                'XL_L9_PION':     '',
                'XL_B8_PION':     '',
                'CLASSIFICACAO':  classif,
                'PASTA':          pasta_proj_path,
                'NOME_PASTA':     nome_proj,
            })

    return registros


def escanear_tipo_projeto(tipo_key, tipo_cfg, db_compasa_pts=None):
    """
    Escaneia um tipo de projeto detectando automaticamente o padrão de pasta.
    Retorna lista de registros (dicts) com todos os projetos encontrados.
    """
    base_dir = tipo_cfg['base_dir']
    padrao   = _detectar_padrao_base(base_dir)

    if padrao == 'MRAF':
        return _varrer_mraf(base_dir, tipo_key, tipo_cfg, db_compasa_pts)
    elif padrao == 'CAUQ':
        return _varrer_cauq(base_dir, tipo_key, tipo_cfg, db_compasa_pts)
    elif padrao == 'FLAT':
        return _varrer_flat(base_dir, tipo_key, tipo_cfg, db_compasa_pts)
    else:
        return []


# ======================================================================================
# 4.4. FUNÇÕES PÚBLICAS — COMPASA e EXTERNOS
# ======================================================================================

def _localizar_pasta_por_pt(tipo_key, tipo_cfg, pt_num, pedreira='', codigo_obs=None):
    """
    Tenta localizar a pasta de um projeto na rede para obter STATUS_PASTA.

    Estratégia por tipo:
      CAUQ_MARSHALL / SOLO_CIMENTO / CAMADAS_GRANULARES / BGS:
        → busca pasta cujo nome começa com PT (ex: "62_IBIPORA" ou "062.25_...")
      CAUQ_SUPERPAVE:
        → busca pasta cujo nome contém CODIGO_OBS ("004/2026" → "004")
      MRAF:
        → busca pasta com cliente extraível (já coberto pelo _varrer_mraf)

    Retorna dict com: STATUS_COMP, STATUS_PION, STATUS_PROJ, ARQUIVO_COMP,
                      ARQUIVO_PION, ARQUIVO_PROJ, TEM_PDF, PASTA
    """
    vazio = {
        'STATUS_COMP': 'VAZIO', 'STATUS_PION': 'VAZIO', 'STATUS_PROJ': 'VAZIO',
        'ARQUIVO_COMP': '', 'ARQUIVO_PION': '', 'ARQUIVO_PROJ': '',
        'TEM_PDF': False, 'PASTA': '',
    }

    base_dir = tipo_cfg.get('base_dir', '')
    if not os.path.isdir(base_dir):
        return vazio

    padrao = _detectar_padrao_base(base_dir)
    tem_pioneiro = tipo_cfg.get('tem_pioneiro', True)

    # ── Determine o token de busca para este projeto ─────────────────────────────────
    # Para SUPERPAVE: usar código da OBS (ex: "004" de "004/2026")
    # Para os demais: usar PT numérico (ex: "62")
    if tipo_key == 'CAUQ_SUPERPAVE' and codigo_obs:
        m = _RE_PT_GENERICO.search(str(codigo_obs))
        token_busca = m.group(1).lstrip('0') if m else str(pt_num).lstrip('0')
    else:
        token_busca = str(pt_num).lstrip('0') or '0'

    # ── Para MRAF: delegamos ao _varrer_mraf (já extrai cliente do nome) ─────────────
    if tipo_key == 'MRAF' or padrao == 'MRAF':
        regs = _varrer_mraf(base_dir, tipo_key, tipo_cfg)
        for r in regs:
            cod = str(r.get('CODIGO', ''))
            if token_busca in cod or pt_num == cod:
                return {
                    'STATUS_COMP': r.get('STATUS', 'VAZIO'),
                    'STATUS_PION': 'NAO_APLICAVEL',
                    'STATUS_PROJ': r.get('STATUS', 'VAZIO'),
                    'ARQUIVO_COMP': r.get('ARQUIVO_PROJ', ''),
                    'ARQUIVO_PION': '',
                    'ARQUIVO_PROJ': r.get('ARQUIVO_PROJ', ''),
                    'TEM_PDF': r.get('TEM_PDF', False),
                    'PASTA': r.get('PASTA', ''),
                    'CLIENTE_PASTA': r.get('CLIENTE', ''),
                    'MISTURA':       r.get('MISTURA', ''),
                }
        return vazio

    # ── Padrão CAUQ: busca nas sub-pastas de fase ────────────────────────────────────
    if padrao in ('CAUQ', 'DESCONHECIDO', 'FLAT'):
        resultado = dict(vazio)
        resultado['PASTA'] = ''

        for fase_key, subfases, resultado_key_status, resultado_key_arq in [
            ('COMPOSICAO', _SUBFASES_CAUQ['COMPOSICAO'], 'STATUS_COMP', 'ARQUIVO_COMP'),
            ('PIONEIRO',   _SUBFASES_CAUQ['PIONEIRO'],   'STATUS_PION', 'ARQUIVO_PION'),
            ('PROJETO',    _SUBFASES_CAUQ['PROJETO'],    'STATUS_PROJ', 'ARQUIVO_PROJ'),
        ]:
            if fase_key == 'PIONEIRO' and not tem_pioneiro:
                resultado['STATUS_PION'] = 'NAO_APLICAVEL'
                continue

            dir_fase = _encontrar_subfase(base_dir, subfases)
            if not dir_fase:
                continue

            # Encontra pasta de ano ativa
            dir_ano = None
            try:
                for p in os.listdir(dir_fase):
                    m = _RE_ANO_PASTA.search(p)
                    if m and m.group(1) == _ANO_PROJETOS_ATIVO:
                        dir_ano = os.path.join(dir_fase, p)
                        break
            except OSError:
                continue
            if not dir_ano:
                continue

            # Encontra pasta do projeto pelo token (PT ou código OBS)
            pasta_proj = None
            melhor_score = 0
            ped_norm = pedreira.upper().replace('PEDREIRA', '').replace('PED.', '').strip()
            try:
                for nome_p in os.listdir(dir_ano):
                    path_p = os.path.join(dir_ano, nome_p)
                    if not os.path.isdir(path_p):
                        continue
                    nome_up = nome_p.upper()
                    # Match pelo token de busca (PT ou código)
                    if nome_up.startswith(token_busca) or f'_{token_busca}_' in nome_up or nome_up.startswith(token_busca.zfill(3)):
                        # Score adicional por pedreira
                        score = 80
                        if ped_norm and len(ped_norm) > 3:
                            score += fuzz.partial_ratio(ped_norm, nome_up) // 4
                        if score > melhor_score:
                            melhor_score = score
                            pasta_proj = path_p
            except OSError:
                continue

            if pasta_proj:
                dir_ent = _encontrar_pasta_entrega(pasta_proj)
                st, arq, pdf = _status_de_pasta(dir_ent)
                resultado[resultado_key_status] = st
                resultado[resultado_key_arq]    = arq or ''
                if pdf:
                    resultado['TEM_PDF'] = True
                if not resultado['PASTA']:
                    resultado['PASTA'] = pasta_proj

        return resultado

    return vazio


@st.cache_data(ttl=300)
def escanear_todos_projetos(df_compasa_db=None):
    """
    VERSÃO ROBUSTA — DIRETÓRIO É A FONTE PRIMÁRIA.
    CLOUD: carrega de todos_projetos.parquet (cache estático).

    Fluxo:
      1ª ETAPA → Varredura de todos os 6 diretórios de rede
                 Cada pasta encontrada = 1 projeto com STATUS real e
                 células Excel extraídas (I8/C10/B9 para COMP, L9/B8/H8 para PROJ/PION).

      2ª ETAPA → Leitura do FORM 022A (Coluna M = Finalidade)
                 Filtra apenas linhas com "PROJETO *" reconhecido.

      3ª ETAPA → CONCATENAÇÃO INTELIGENTE:
                 • Projeto encontrado no diretório
                   → exibe fases reais com dados Excel extraídos
                   → enriquece com FORM 022A se houver cruzamento por PT
                 • Projeto no FORM 022A MAS SEM pasta na rede
                   → STATUS = 'A_INICIAR'

    Campos de células Excel propagados para os cards:
        XL_I8_COMP   → Composição: célula I8 (identificação do projeto)
        XL_C10_COMP  → Composição: célula C10 (procedência/pedreira)
        XL_B9_COMP   → Composição: célula B9 (cliente)
        XL_L9_PROJ   → Projeto: célula L9 (PT/código)
        XL_B8_PROJ   → Projeto: célula B8 (cliente)
        XL_H8_PROJ   → Projeto: célula H8 (mistura)
        XL_L9_PION   → Pioneiro: célula L9 (PT/código)
        XL_B8_PION   → Pioneiro: célula B8 (cliente)

    Parâmetro df_compasa_db: mantido por retrocompatibilidade.
    """

    # ── CLOUD GUARD ──────────────────────────────────────────────────────────
    if bridge.is_cloud:
        from cloud_config import carregar_parquet_cache
        df = carregar_parquet_cache("todos_projetos")
        if not df.empty:
            return df
        return pd.DataFrame()
    # ─────────────────────────────────────────────────────────────────────────

    # =========================================================================
    # ETAPA 1 — VARREDURA DE TODOS OS DIRETÓRIOS (com leitura Excel profunda)
    # =========================================================================
    registros_dir = []
    for tipo_key, tipo_cfg in TIPOS_PROJETO_CONFIG.items():
        regs = escanear_tipo_projeto(tipo_key, tipo_cfg, db_compasa_pts=None)
        registros_dir.extend(regs)

    # =========================================================================
    # ETAPA 2 — LEITURA DO FORM 022A
    # =========================================================================
    df_form = carregar_projetos_form022a(ano=_ANO_PROJETOS_ATIVO)

    # =========================================================================
    # ETAPA 3 — CONCATENAÇÃO
    # =========================================================================

    # Campos padrão de células Excel — garante que colunas existam mesmo sem leitura
    _CAMPOS_XL_PADRAO = {
        'XL_I8_COMP': '', 'XL_C10_COMP': '', 'XL_B9_COMP': '',
        'XL_L9_PROJ': '', 'XL_B8_PROJ': '', 'XL_H8_PROJ': '',
        'XL_L9_PION': '', 'XL_B8_PION': '',
    }

    # Índice de PT × TIPO encontrados nos diretórios
    # Chave: (pt_normalizado, tipo_key)  → normalizado = apenas dígitos do primeiro bloco
    def _norm_pt(pt_raw):
        s = str(pt_raw or '').strip().lstrip('0') or '0'
        return s

    pts_no_dir: dict[tuple, dict] = {}
    for reg in registros_dir:
        pt  = _norm_pt(reg.get('PT_NUM') or reg.get('CODIGO') or '')
        tk  = reg.get('TIPO_PROJETO', '')
        # Mantém o primeiro encontrado (ordem alfabética da pasta → mais recente)
        if (pt, tk) not in pts_no_dir:
            pts_no_dir[(pt, tk)] = reg

    # Índice do FORM 022A: (pt_num, tipo_key) → linha
    form_index: dict[tuple, dict] = {}
    if not df_form.empty:
        for _, frow in df_form.iterrows():
            pt_f = _norm_pt(frow.get('PT', ''))
            tk_f = frow.get('TIPO_PROJETO', '')
            form_index[(pt_f, tk_f)] = frow.to_dict()

    registros_finais = []

    # ── A. Projetos do DIRETÓRIO (enriquecidos com FORM 022A) ──────────────────────
    for (pt, tk), reg in pts_no_dir.items():
        tipo_cfg = TIPOS_PROJETO_CONFIG.get(tk, {})
        frow     = form_index.get((pt, tk), {})

        # Status das fases (vem do diretório)
        st_comp = reg.get('STATUS_COMP', 'VAZIO')
        st_pion = reg.get('STATUS_PION', 'VAZIO')
        st_proj = reg.get('STATUS_PROJ', 'VAZIO')
        tem_pdf = reg.get('TEM_PDF', False)

        # Status geral
        if tem_pdf or st_proj == 'OK' or st_comp == 'OK':
            status_geral = 'OK'
        elif st_proj == 'ANDAMENTO' or st_comp == 'ANDAMENTO':
            status_geral = 'ANDAMENTO'
        else:
            st_form_aux = frow.get('STATUS_FORM', 'DESCONHECIDO')
            if st_form_aux == 'CONCLUIDO':
                status_geral = 'OK'
            elif st_form_aux == 'AGUARDANDO':
                status_geral = 'AGUARDANDO'
            elif st_form_aux == 'ANDAMENTO':
                status_geral = 'ANDAMENTO'
            else:
                status_geral = 'VAZIO'

        # Cliente: MRAF/Excel têm na pasta; outros preferem FORM 022A
        cliente_pasta = reg.get('CLIENTE', '')
        cliente_form  = frow.get('CLIENTE', '')
        # Preferência: Excel extraído > FORM 022A > pasta
        cliente_final = cliente_pasta if cliente_pasta else cliente_form
        classif_final = _classificar_cliente_externo(cliente_final)

        # Se o cliente ou a pedreira contiver "COMPASA", força classificação COMPASA
        _cliente_up = str(cliente_final).upper()
        if 'COMPASA' in _cliente_up:
            classif_final = 'COMPASA'
            cliente_final = 'COMPASA DO BRASIL'

        # Pedreira: Excel > FORM > pasta
        pedreira_reg  = reg.get('PEDREIRA', '')
        pedreira_form = frow.get('PEDREIRA', '') if frow else ''
        pedreira_final = pedreira_reg or pedreira_form

        # Campos Excel — propagados diretamente do scanner
        campos_xl = {k: reg.get(k, '') for k in _CAMPOS_XL_PADRAO}

        # Protocolo: usar PT do FORM 022A (col A) + procedência do FORM (col E) quando disponível
        proto_proc = None
        pt_form = str(frow.get('PT', '')).strip() if frow else ''
        ped_form = str(frow.get('PEDREIRA', '')).strip() if frow else ''
        if pt_form and ped_form:
            proto_proc = f"{pt_form} - {ped_form}"
        elif pt_form and pedreira_final:
            proto_proc = f"{pt_form} - {pedreira_final}"
        elif ped_form and pt:
            proto_proc = f"{pt} - {ped_form}"
        elif reg.get('PROTOCOLO_COMP'):
            proto_proc = reg.get('PROTOCOLO_COMP')

        registros_finais.append({
            # Identidade
            'CODIGO':         reg.get('CODIGO') or pt,
            'PT_NUM':         pt,
            'PT_PASTA':       reg.get('PT_PASTA', ''),
            'CLIENTE':        cliente_final,
            'CLASSIFICACAO':  classif_final,
            'TIPO_PROJETO':   tk,
            'TIPO_LABEL':     tipo_cfg.get('label', tk),
            'TIPO_ICONE':     tipo_cfg.get('icone', '📁'),
            'TEM_PIONEIRO':   tipo_cfg.get('tem_pioneiro', False),
            'ANO_PASTA':      reg.get('ANO_PASTA', _ANO_PROJETOS_ATIVO),
            # Localização
            'PEDREIRA':       pedreira_final,
            'MISTURA':        reg.get('MISTURA', '') or frow.get('MISTURA', ''),
            'OBS':            str(frow.get('OBS', '') or ''),
            'CODIGO_OBS':     frow.get('CODIGO_OBS', '') or '',
            'NOME_PASTA':     reg.get('NOME_PASTA', ''),
            # Dados FORM 022A
            'FINALIDADE_RAW': frow.get('FINALIDADE_RAW', ''),
            'STATUS_FORM':    frow.get('STATUS_FORM', ''),
            'N_MATERIAIS':    frow.get('N_MATERIAIS', 0),
            'MATERIAIS':      frow.get('MATERIAIS', ''),
            'DATA_RECEBIMENTO': frow.get('DATA_RECEBIMENTO'),
            # Status combinado
            'STATUS':         status_geral,
            'STATUS_COMP':    st_comp,
            'STATUS_PION':    st_pion,
            'STATUS_PROJ':    st_proj,
            # Arquivos
            'ARQUIVO_COMP':   reg.get('ARQUIVO_COMP', ''),
            'ARQUIVO_PION':   reg.get('ARQUIVO_PION', ''),
            'ARQUIVO_PROJ':   reg.get('ARQUIVO_PROJ', ''),
            'CAMINHO_COMP':   reg.get('CAMINHO_COMP', ''),
            'CAMINHO_PION':   reg.get('CAMINHO_PION', ''),
            'CAMINHO_PROJ':   reg.get('CAMINHO_PROJ', ''),
            'TEM_PDF':        tem_pdf,
            'PASTA':          reg.get('PASTA', ''),
            'PROTOCOLO_PROC': proto_proc or '',
            # Células Excel extraídas ── usadas nos cards dos painéis
            **campos_xl,
            # Origem
            'ORIGEM_DADOS':   'DIRETORIO',
        })

    # ── B. Projetos SOMENTE NO FORM 022A (pasta não existe na rede) ───────────────
    if not df_form.empty:
        for (pt_f, tk_f), frow in form_index.items():
            if (pt_f, tk_f) in pts_no_dir:
                continue   # Já processado acima

            tipo_cfg = TIPOS_PROJETO_CONFIG.get(tk_f, {})
            tem_pion = tipo_cfg.get('tem_pioneiro', False)
            cliente  = frow.get('CLIENTE', '')
            classif  = _classificar_cliente_externo(cliente)

            registros_finais.append({
                # Identidade
                'CODIGO':         frow.get('PT', pt_f),
                'PT_NUM':         pt_f,
                'PT_PASTA':       '',
                'CLIENTE':        cliente,
                'CLASSIFICACAO':  classif,
                'TIPO_PROJETO':   tk_f,
                'TIPO_LABEL':     tipo_cfg.get('label', tk_f),
                'TIPO_ICONE':     tipo_cfg.get('icone', '📁'),
                'TEM_PIONEIRO':   tem_pion,
                'ANO_PASTA':      _ANO_PROJETOS_ATIVO,
                # Localização
                'PEDREIRA':       frow.get('PEDREIRA', ''),
                'MISTURA':        '',
                'OBS':            str(frow.get('OBS', '') or ''),
                'CODIGO_OBS':     frow.get('CODIGO_OBS', '') or '',
                'NOME_PASTA':     '',
                # Dados FORM 022A
                'FINALIDADE_RAW': frow.get('FINALIDADE_RAW', ''),
                'STATUS_FORM':    frow.get('STATUS_FORM', ''),
                'N_MATERIAIS':    frow.get('N_MATERIAIS', 0),
                'MATERIAIS':      frow.get('MATERIAIS', ''),
                'DATA_RECEBIMENTO': frow.get('DATA_RECEBIMENTO'),
                # Status → A_INICIAR (pasta não encontrada)
                'STATUS':         'A_INICIAR',
                'STATUS_COMP':    'A_INICIAR',
                'STATUS_PION':    'A_INICIAR' if tem_pion else 'NAO_APLICAVEL',
                'STATUS_PROJ':    'A_INICIAR',
                # Arquivos
                'ARQUIVO_COMP':   '', 'ARQUIVO_PION':   '', 'ARQUIVO_PROJ':   '',
                'CAMINHO_COMP':   '', 'CAMINHO_PION':   '', 'CAMINHO_PROJ':   '',
                'TEM_PDF':        False,
                'PASTA':          '',
                # Células Excel — vazias (sem pasta)
                **{k: '' for k in _CAMPOS_XL_PADRAO},
                # Origem
                'ORIGEM_DADOS':   'FORM_022A_SOMENTE',
            })

    if not registros_finais:
        return pd.DataFrame()

    df = pd.DataFrame(registros_finais)
    _ord_map = {'COMPASA': 0, 'CC': 1, 'EXTERNO': 2}
    df['_ord'] = df['CLASSIFICACAO'].map(_ord_map).fillna(9)
    df = (df.sort_values(['_ord', 'TIPO_PROJETO', 'CODIGO'])
            .drop(columns='_ord')
            .reset_index(drop=True))
    return df



@st.cache_data(ttl=300)
def rastrear_projetos_compasa_completo(df_compasa):
    """
    Retorna apenas projetos classificados como COMPASA (retrocompatível).
    Filtra o resultado de escanear_todos_projetos pela coluna CLASSIFICACAO == 'COMPASA'.
    Mantém a estrutura de colunas esperada pelo dashboard.
    """
    df_todos = escanear_todos_projetos(df_compasa)
    if df_todos.empty:
        return pd.DataFrame()

    # Filtra Compasa
    df_comp = df_todos[df_todos['CLASSIFICACAO'] == 'COMPASA'].copy()
    if df_comp.empty:
        return pd.DataFrame()

    # Rebuilda colunas DADOS_* no formato legado esperado pelo dashboard
    def _make_dado(row, fase):
        status_col = f'STATUS_{fase}'
        arq_col    = f'ARQUIVO_{fase}'
        st_val     = row.get(status_col, 'VAZIO')
        arq_val    = row.get(arq_col, '')
        # Mapear VAZIO → NAO_INICIADO para compatibilidade
        st_norm = {
            'OK': 'OK', 'ANDAMENTO': 'ANDAMENTO',
            'VAZIO': 'NAO_INICIADO', 'NAO_APLICAVEL': 'NAO_APLICAVEL',
        }.get(st_val, 'NAO_INICIADO')
        return {
            'STATUS': st_norm,
            'ARQUIVO': arq_val or None,
            'DETALHES': {'pedreira': row.get('PEDREIRA', ''), 'mistura': row.get('MISTURA', '')},
            'TEM_PDF': row.get('TEM_PDF', False),
        }

    registros = []
    for _, row in df_comp.iterrows():
        registros.append({
            'PT':               row['CODIGO'],
            'PEDREIRA':         row['PEDREIRA'],
            'CLIENTE_PASTA':    row['CLIENTE'],
            'STATUS_GERAL':     row['STATUS'],
            'TIPO_PROJETO':     row['TIPO_PROJETO'],
            'TIPO_LABEL':       row['TIPO_LABEL'],
            'TIPO_ICONE':       row['TIPO_ICONE'],
            'TEM_PIONEIRO':     row['TEM_PIONEIRO'],
            'DADOS_COMPOSICAO': _make_dado(row, 'COMP'),
            'DADOS_PIONEIRO':   _make_dado(row, 'PION'),
            'DADOS_PROJETO':    _make_dado(row, 'PROJ'),
        })

    return pd.DataFrame(registros)


@st.cache_data(ttl=300)
def rastrear_projetos_externos():
    """
    Retorna projetos NÃO COMPASA encontrados nas pastas de rede.
    Inclui tanto EXTERNOS puros quanto CCs internos (EPR, CBB, etc.).
    Usa escanear_todos_projetos sem filtro de df_compasa.
    """
    df_todos = escanear_todos_projetos(df_compasa=None)
    if df_todos.empty:
        return pd.DataFrame()
    return df_todos[df_todos['CLASSIFICACAO'] != 'COMPASA'].reset_index(drop=True)


# Mantidos por compatibilidade (wrappers finos)
def _normalizar_codigo_pt(pt_raw):
    """Normaliza código de PT extraindo o primeiro número relevante."""
    if not pt_raw:
        return None
    s = str(pt_raw).strip()
    m = _RE_COD_MRAF.search(s)
    if m:
        return m.group(1)
    m = _RE_COD_CAUQ.search(s)
    if m:
        return m.group(1)
    return None


def localizar_pasta_mraf(codigo_mraf, ano_hint=None):
    """Localiza pasta MRAF pelo código. Wrapper sobre o scanner universal."""
    base = BASE_DIR_PROJETOS_MRAF
    if not os.path.isdir(base):
        return None, None
    regs = _varrer_mraf(base, 'MRAF', TIPOS_PROJETO_CONFIG['MRAF'])
    for r in regs:
        if r['CODIGO'] == codigo_mraf:
            return r['PASTA'], {
                'codigo':   r['CODIGO'],
                'cliente':  r['CLIENTE'],
                'pedreira': r['PEDREIRA'],
                'mistura':  r['MISTURA'],
            }
    return None, None


def verificar_fase_mraf(pt_alvo, ano_hint=None):
    """Verifica status do projeto MRAF. Wrapper sobre o scanner universal."""
    codigo = _normalizar_codigo_pt(pt_alvo)
    retorno = {'STATUS': 'NAO_INICIADO', 'ARQUIVO': None, 'DETALHES': {}, 'TEM_PDF': False, 'CLIENTE_PASTA': ''}
    if not codigo:
        return retorno
    regs = _varrer_mraf(BASE_DIR_PROJETOS_MRAF, 'MRAF', TIPOS_PROJETO_CONFIG['MRAF'])
    for r in regs:
        if r['CODIGO'] == codigo:
            st_map = {'OK': 'OK', 'ANDAMENTO': 'ANDAMENTO', 'VAZIO': 'NAO_INICIADO'}
            retorno.update({
                'STATUS':        st_map.get(r['STATUS'], 'NAO_INICIADO'),
                'ARQUIVO':       r.get('ARQUIVO_PROJ') or None,
                'TEM_PDF':       r['TEM_PDF'],
                'CLIENTE_PASTA': r['CLIENTE'],
                'DETALHES': {
                    'cliente':  r['CLIENTE'],
                    'pedreira': r['PEDREIRA'],
                    'mistura':  r['MISTURA'],
                },
            })
            return retorno
    return retorno


def verificar_fase_compasa_detalhada(tipo_fase, pt_alvo, pedreira, ano, base_dir=None):
    """
    Verifica fase individual de projeto Compasa.
    Wrapper retrocompatível — usa o scanner universal internamente.
    """
    _tipo_key  = 'CAUQ_MARSHALL'
    _base      = base_dir or DIRETORIO_BASE_CAUQ
    _cfg_local = {
        'label': 'CAUQ Marshall', 'base_dir': _base,
        'tem_pioneiro': True, 'keywords': [], 'icone': '🛣️',
    }
    retorno = {'STATUS': 'NAO_INICIADO', 'ARQUIVO': None, 'DETALHES': {}, 'TEM_PDF': False}
    if not pedreira:
        return retorno

    padrao = _detectar_padrao_base(_base)

    if padrao in ('CAUQ', 'FLAT', 'DESCONHECIDO'):
        # Usar lógica CAUQ original para esta fase específica
        sub_map = {
            'COMPOSICAO': _SUBFASES_CAUQ['COMPOSICAO'],
            'PIONEIRO':   _SUBFASES_CAUQ['PIONEIRO'],
            'PROJETO':    _SUBFASES_CAUQ['PROJETO'],
        }
        dir_fase = _encontrar_subfase(_base, sub_map.get(tipo_fase, []))
        if not dir_fase:
            return retorno

        # Encontrar pasta de ano
        try:
            filhos_fase = os.listdir(dir_fase)
        except OSError:
            return retorno
        dir_ano = None
        for p in filhos_fase:
            if str(ano) in p and os.path.isdir(os.path.join(dir_fase, p)):
                dir_ano = os.path.join(dir_fase, p)
                break
        if not dir_ano:
            return retorno

        # Encontrar pasta de projeto por pedreira (fuzzy)
        nome_ped = str(pedreira).upper().replace('PEDREIRA', '').strip()
        melhor_score = 0
        pasta_proj = None
        pt_nums = re.findall(r'\d+', str(pt_alvo))
        pt_clean = pt_nums[0] if pt_nums else ''
        try:
            for p in os.listdir(dir_ano):
                path = os.path.join(dir_ano, p)
                if not os.path.isdir(path):
                    continue
                score = fuzz.partial_ratio(nome_ped, p.upper())
                if score > 80 and score > melhor_score:
                    melhor_score = score
                    pasta_proj = path
                if tipo_fase == 'PROJETO' and pt_clean:
                    if p.startswith(pt_clean) or f' {pt_clean} ' in p:
                        pasta_proj = path
                        break
        except OSError:
            return retorno

        if not pasta_proj:
            return retorno

        dir_busca = _encontrar_pasta_entrega(pasta_proj)
        xl, pdf = _inspecionar_pasta(dir_busca)
        if xl:
            retorno['ARQUIVO'] = xl
            retorno['TEM_PDF'] = pdf
            retorno['STATUS']  = 'OK' if pdf else 'ANDAMENTO'

    elif padrao == 'MRAF':
        r = verificar_fase_mraf(pt_alvo, ano)
        if r['STATUS'] != 'NAO_INICIADO':
            retorno.update(r)

    return retorno



# ======================================================================================
# 5. SINCRONIZAÇÃO
# ======================================================================================

def get_lista_cc_from_excel():
    cc_names = set()
    path = FILES_CONFIG['certificados_2025']['local_path']
    if os.path.exists(path):
        try:
            df = pd.read_excel(path, sheet_name='CLIENTES CAD', header=None, skiprows=1)
            for i, row in df.iterrows():
                nome = str(row.iloc[10]).strip().upper()
                contrato = str(row.iloc[16]).strip()
                if nome and contrato and contrato.upper() not in ['NAN', 'NONE', '']:
                    cc_names.add(nome)
        except: pass
    return cc_names

def sync_recebimento():
    # ── CLOUD GUARD ──────────────────────────────────────────────────────────
    if bridge.is_cloud:
        return  # No cloud, dados vêm do parquet cache estático
    # ─────────────────────────────────────────────────────────────────────────
    todos_dados = []
    keys = [k for k in FILES_CONFIG if FILES_CONFIG[k]['tipo'] == 'recebimento']
    data_corte = pd.Timestamp('2025-12-01')
    cc_list = get_lista_cc_from_excel()
    
    for key in keys:
        source = bridge.get_file_content(key)
        if not source: continue
        try:
            df = pd.read_excel(source, header=None, engine='openpyxl')
            ano = FILES_CONFIG[key]['ano']
            for i in range(6, len(df)):
                try:
                    row = df.iloc[i]
                    empresa_base = str(row.iloc[1]).strip().upper()
                    if len(empresa_base) < 3 or empresa_base in ['EMPRESA','CLIENTE']: continue
                    
                    # --- LÓGICA DE SUBDIVISÃO CBB / ASFALTEC ---
                    classificacao_manual = str(row.iloc[16]).strip().upper()
                    local_servico = str(row.iloc[12]).strip().upper()

                    if "ASFALTEC" in classificacao_manual or ("CBB" in empresa_base and any(x in local_servico for x in ["DEFORMAÇÃO", "DPH", "WHEEL"])):
                        empresa = "CC ASFALTEC"
                        is_manual_cc = True
                    elif "CBB" in empresa_base or "CBB" in classificacao_manual:
                        empresa = "CBB ASFALTOS"
                        is_manual_cc = True
                    else:
                        empresa = empresa_base
                        is_manual_cc = False
                    
                    dt_rec = pd.to_datetime(row.iloc[2], errors='coerce')
                    if pd.notna(dt_rec) and dt_rec < data_corte: continue
                    if pd.isna(dt_rec) and ano == '2025': continue
                    
                    pt_col_a = str(row.iloc[0]).strip()
                    prop_raw = str(row.iloc[13]).strip()
                    status_raw = str(row.iloc[17]).strip()
                    
                    # --- NOVO MAPEAMENTO DA OBSERVAÇÃO RECEBIMENTO ---
                    # Geralmente Coluna L (index 11) ou J (index 9) ou K (index 10)
                    # Baseado na sua planilha: PROTOCOLO(A/0), CLIENTE(B/1), RECEB(C/2), ID(D/3), PROC(E/4), QTD(F/5)
                    # PESO(G/6), RESP_ENT(H/7), OBS_ENT(I/8), OBS_REC(J/9 - OU COLUNA L SE TIVER OCULTAS)
                    
                    # Tentativa de pegar Obs. Recebimento (Coluna 10 ou 11)
                    # Se baseando no padrão da planilha enviada, parece ser a 11ª coluna visual
                    obs_rec = str(row.iloc[10]).strip() # Tentativa padrão (Coluna K)
                    
                    # Se estiver vazio, tenta a próxima (Coluna L)
                    if not obs_rec or obs_rec == 'nan':
                         obs_rec = str(row.iloc[11]).strip()

                    is_cc = (is_manual_cc or empresa.upper() in cc_list or any(x in empresa.upper() for x in ['EPR', 'STRATA', 'COMPASA']))
                    
                    tipo = 'CC' if is_cc else 'PC'
                    num_prop = pt_col_a if tipo == 'CC' else prop_raw
                    
                    if 'COMPASA' in empresa.upper(): empresa = 'COMPASA DO BRASIL'
                    
                    status = normalizar_status(status_raw)
                    if status == 'FINALIZADO': tem_mat = True
                    else:
                        tem_mat = True if (pd.notna(dt_rec) or tipo=='CC') else False
                        status = 'EM ANDAMENTO' if tem_mat else 'AGUARDANDO RECEBIMENTO'
                    
                    # Quantidade (Coluna H / Index 7)
                    qtd_raw = row.iloc[7] if len(row) > 7 else 1
                    try: quantidade = int(float(qtd_raw)) if pd.notna(qtd_raw) else 1
                    except: quantidade = 1
                    
                    todos_dados.append({
                        'NUMERO_PROPOSTA': num_prop, 'CLIENTE': empresa,
                        'DATA_RECEBIMENTO': dt_rec, 'STATUS': status,
                        'TIPO_PROPOSTA': tipo, 'TEM_MATERIAL': tem_mat,
                        'MATERIAL': str(row.iloc[3]), 'PEDREIRA': str(row.iloc[4]),
                        'ANO': ano, 'PT_COLUNA_A': pt_col_a,
                        'QUANTIDADE': quantidade, 'E_CONTRATO_CONTINUO': 1 if is_cc else 0,
                        'OBS_RECEBIMENTO': obs_rec  # <--- NOVA COLUNA SALVA
                    })
                except: continue
        except: continue
        
    if todos_dados:
        df_final = pd.DataFrame(todos_dados)
        for c in ['DATA_RECEBIMENTO']:
             df_final[c] = df_final[c].apply(lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) else None)
        
        df_final = df_final.sort_values(['TIPO_PROPOSTA', 'DATA_RECEBIMENTO', 'STATUS'], ascending=[False, False, True])
        
        # Deduplicação mantida para o DF principal (mas o RAW usará todos)
        df_final = df_final.drop_duplicates(subset=['NUMERO_PROPOSTA', 'PT_COLUNA_A'], keep='first')
        
        with bridge.get_db_conn() as conn:
            df_final.to_sql('recebimentos', conn, if_exists='replace', index=False)

def sync_certificados_067():
    """Sincroniza dados dos certificados 067 com o SQLite"""
    # ── CLOUD GUARD ──────────────────────────────────────────────────────────
    if bridge.is_cloud:
        return  # No cloud, dados vêm do parquet cache estático
    # ─────────────────────────────────────────────────────────────────────────
    todos = []
    for ano in ['2026', '2025']:
        path = FILES_CONFIG[f'certificados_{ano}']['local_path']
        if not os.path.exists(path): continue
        try:
            df = pd.read_excel(path, header=None)
            for i in range(8, len(df)):
                try:
                    row = df.iloc[i]
                    pt = str(row.iloc[5])
                    if len(pt) < 2: continue
                    cliente = str(row.iloc[1]).strip()
                    if len(cliente) < 3: continue
                    todos.append({
                        'PT': pt,
                        'PT_NORMALIZADO': normalizar_pt(pt),
                        'CLIENTE': cliente,
                        'ENSAIO': str(row.iloc[2]).strip(),
                        'NORMA': str(row.iloc[3]).strip(),
                        'QUANTIDADE': float(row.iloc[4]) if pd.notna(row.iloc[4]) else 1,
                        'ENSAIO_CONCLUIDO': 1 if pd.notna(row.iloc[13]) else 0,
                        'RELATORIO_VINCULADO': str(row.iloc[10]) if pd.notna(row.iloc[10]) else "",
                        'ANO': ano
                    })
                except: continue
        except: continue
    
    if todos:
        with bridge.get_db_conn() as conn:
            pd.DataFrame(todos).to_sql('certificados_067', conn, if_exists='replace', index=False)

def sync_propostas():
    pass

def sync_all_data():
    # No cloud, sync não faz nada (dados vêm do cache estático)
    if bridge.is_cloud:
        return
    for func in dir(sys.modules[__name__]):
        if hasattr(getattr(sys.modules[__name__], func), 'clear_cache'):
            getattr(sys.modules[__name__], func).clear_cache()
    try: st.cache_data.clear()
    except: pass
    sync_recebimento()
    sync_certificados_067()
    sync_propostas()

@st.cache_data(ttl=300, show_spinner=False)
def carregar_dados_consolidados_sql():
    if bridge.is_cloud:
        from cloud_config import carregar_parquet_cache
        return carregar_parquet_cache("db_recebimentos")
    with bridge.get_db_conn() as conn:
        try: return pd.read_sql_query("SELECT * FROM recebimentos", conn)
        except: return pd.DataFrame()

@st.cache_data(ttl=600, show_spinner=False)
def carregar_dados_epr_raw(mes_filtro=None, cliente_filtro=None):
    """
    Carrega dados brutos (SEM DEDUPLICAÇÃO) para contagem exata de amostras.
    Inclui a coluna OBS_RECEBIMENTO.
    """
    if bridge.is_cloud:
        from cloud_config import carregar_parquet_cache
        # Usa epr_raw_all (dados brutos SEM dedup) — NÃO db_recebimentos (que é deduplicated)
        df = carregar_parquet_cache("epr_raw_all")
        if not df.empty:
            if 'DATA_RECEBIMENTO' in df.columns:
                df['DATA_RECEBIMENTO'] = pd.to_datetime(df['DATA_RECEBIMENTO'], errors='coerce')
            if cliente_filtro and 'CLIENTE' in df.columns:
                df = df[df['CLIENTE'].astype(str).str.upper().str.contains(cliente_filtro.upper(), na=False)]
        return df
    todos_dados = []
    keys = [k for k in FILES_CONFIG if FILES_CONFIG[k]['tipo'] == 'recebimento']
    data_corte = pd.Timestamp('2025-12-01')
    
    for key in keys:
        source = bridge.get_file_content(key)
        if not source: continue
        try:
            df = pd.read_excel(source, header=None, engine='openpyxl')
            ano = FILES_CONFIG[key]['ano']
            for i in range(6, len(df)):
                try:
                    row = df.iloc[i]
                    empresa = str(row.iloc[1]).strip()
                    if len(empresa) < 3 or empresa.upper() in ['EMPRESA', 'CLIENTE']: continue
                    
                    # Filtrar por cliente (EPR, STRATA, COMPASA, CBB, ASFALTEC)
                    empresa_upper = empresa.upper()
                    if cliente_filtro:
                        if cliente_filtro.upper() not in empresa_upper: continue
                    else:
                        # LISTA ATUALIZADA DE TARGETS PARA O DASHBOARD CC
                        if not any(x in empresa_upper for x in ['EPR', 'STRATA', 'COMPASA', 'CBB', 'ASFALTEC', 'EIXO SP']): continue
                    
                    dt_rec = pd.to_datetime(row.iloc[2], errors='coerce')
                    if pd.notna(dt_rec) and dt_rec < data_corte: continue
                    if pd.isna(dt_rec) and ano == '2025': continue
                    
                    # Aplicar filtro de mês se especificado
                    if mes_filtro is not None and pd.notna(dt_rec):
                        if dt_rec.to_period('M') != mes_filtro:
                            continue
                    
                    # Coluna H (index 7) = QUANTIDADE
                    qtd_raw = row.iloc[7] if len(row) > 7 else 1
                    try:
                        quantidade = int(float(qtd_raw)) if pd.notna(qtd_raw) else 1
                    except:
                        quantidade = 1

                    # Captura OBS_RECEBIMENTO (Tentativa índice 10 ou 11)
                    obs_rec = str(row.iloc[10]).strip()
                    if not obs_rec or obs_rec == 'nan': obs_rec = str(row.iloc[11]).strip()

                    todos_dados.append({
                        'CLIENTE': empresa,
                        'MATERIAL': str(row.iloc[3]),
                        'QUANTIDADE': quantidade,
                        'PT_COLUNA_A': str(row.iloc[0]).strip(),
                        'DATA_RECEBIMENTO': dt_rec,
                        'ANO': ano,
                        'OBS_RECEBIMENTO': obs_rec # Adicionado ao RAW também
                    })
                except: continue
        except: continue
    
    return pd.DataFrame(todos_dados) if todos_dados else pd.DataFrame()

def exportar_dados_csv(df, n): return df.to_csv(index=False).encode('utf-8')
def formatar_numero(val): return str(val) if pd.notna(val) else "0"
def formatar_data(val): return val.strftime('%d/%m/%Y') if pd.notna(val) and not isinstance(val, str) else str(val) if pd.notna(val) else "-"
def formatar_protocolo(pt): return str(pt).strip() if pd.notna(pt) else "-"

@st.cache_data(ttl=600, show_spinner=False)
def carregar_empresa_finalidade_raw():
    """
    Carrega dados brutos de EMPRESA (Col B) × FINALIDADE (Col M) do FORM 022A.
    Retorna DataFrame com colunas: EMPRESA, FINALIDADE, DATA_RECEBIMENTO, QUANTIDADE.
    Usado no gráfico de Quantitativo por Empresa × Finalidade.
    CLOUD: carrega de empresa_finalidade.parquet (cache estático).
    """
    # ── CLOUD GUARD ──────────────────────────────────────────────────────────
    if bridge.is_cloud:
        from cloud_config import carregar_parquet_cache
        df = carregar_parquet_cache("empresa_finalidade")
        if not df.empty:
            return df
        return pd.DataFrame()
    # ─────────────────────────────────────────────────────────────────────────
    todos_dados = []
    keys = [k for k in FILES_CONFIG if FILES_CONFIG[k]['tipo'] == 'recebimento']
    data_corte = pd.Timestamp('2025-12-01')

    for key in keys:
        source = bridge.get_file_content(key)
        if not source:
            continue
        try:
            df = pd.read_excel(source, header=None, engine='openpyxl')
            ano = FILES_CONFIG[key]['ano']
            for i in range(6, len(df)):
                try:
                    row = df.iloc[i]
                    empresa = str(row.iloc[1]).strip()
                    if len(empresa) < 3 or empresa.upper() in ['EMPRESA', 'CLIENTE']:
                        continue

                    dt_rec = pd.to_datetime(row.iloc[2], errors='coerce')
                    if pd.notna(dt_rec) and dt_rec < data_corte:
                        continue
                    if pd.isna(dt_rec) and ano == '2025':
                        continue

                    # Col M (índice 12) = FINALIDADE / LOCAL / SERVIÇO
                    finalidade = str(row.iloc[12]).strip() if len(row) > 12 else ''
                    if not finalidade or finalidade.lower() in ['nan', 'none', '']:
                        continue

                    # Col H (índice 7) = QUANTIDADE
                    qtd_raw = row.iloc[7] if len(row) > 7 else 1
                    try:
                        quantidade = int(float(qtd_raw)) if pd.notna(qtd_raw) else 1
                    except Exception:
                        quantidade = 1

                    todos_dados.append({
                        'EMPRESA':          empresa,
                        'FINALIDADE':       finalidade,
                        'DATA_RECEBIMENTO': dt_rec,
                        'QUANTIDADE':       quantidade,
                        'ANO':              ano,
                    })
                except Exception:
                    continue
        except Exception:
            continue

    return pd.DataFrame(todos_dados) if todos_dados else pd.DataFrame()
def calcular_fas_total(df_aprovadas, df_em_execucao):
    """
    Gera uma lista única de PCs (Propostas Comerciais) para buscar o FORM 045.
    Combina:
    1. Propostas recém aprovadas (do FORM 044)
    2. Propostas em execução ativa (do FORM 022A/Banco de Dados)
    """
    lista_pcs = []

    # 1. Adicionar do FORM 044 (FAS Recebidas/Aprovadas)
    if not df_aprovadas.empty and 'FAS' in df_aprovadas.columns:
        # Filtra apenas o que parece ser PC (contém dígitos e não começa com CC se possível)
        for _, row in df_aprovadas.iterrows():
            pc = str(row['FAS']).strip()
            cliente = str(row.get('EMPRESA', ''))
            if len(pc) > 3 and 'PC' in pc.upper() or '/' in pc:
                lista_pcs.append({'PC': pc, 'EMPRESA': cliente, 'ORIGEM': 'FORM 044'})

    # 2. Adicionar do FORM 022A (Em Execução)
    if not df_em_execucao.empty:
        # Filtra onde TIPO_PROPOSTA é 'PC'
        if 'TIPO_PROPOSTA' in df_em_execucao.columns:
            df_pcs_ativas = df_em_execucao[df_em_execucao['TIPO_PROPOSTA'] == 'PC']
        else:
            # Fallback se não tiver coluna TIPO: tenta identificar pelo numero
            df_pcs_ativas = df_em_execucao[df_em_execucao['NUMERO_PROPOSTA'].astype(str).str.contains('/', na=False)]
        
        for _, row in df_pcs_ativas.iterrows():
            pc = str(row['NUMERO_PROPOSTA']).strip()
            cliente = str(row.get('CLIENTE', ''))
            # Evita duplicar se já adicionou pelo FORM 044
            if pc not in [x['PC'] for x in lista_pcs]:
                lista_pcs.append({'PC': pc, 'EMPRESA': cliente, 'ORIGEM': 'EM EXECUÇÃO'})

    # Retorna DataFrame com a lista consolidada
    if lista_pcs:
        return pd.DataFrame(lista_pcs).drop_duplicates(subset=['PC'])
    else:
        return pd.DataFrame()

@st.cache_data(ttl=600, show_spinner=False)
def carregar_form044():
    """
    Carrega dados do FORM 044 - Controle de Propostas.
    Retorna DataFrame com propostas aprovadas desde 01/12/2025.

    Mapeamento de colunas:
    - COL A (0) = ANO
    - COL C (2) = FAS (número da proposta)
    - COL G (6) = EMPRESA (nome da empresa)
    - COL H (7) = STATUS (deve ser "Aprovada")
    - COL P (15) = DATA_ACEITE (deve ser > 01/12/2025)
    """
    # ── CLOUD GUARD ──────────────────────────────────────────────────────────
    if bridge.is_cloud:
        from cloud_config import carregar_parquet_cache
        df = carregar_parquet_cache("propostas_form044")
        if not df.empty:
            logger.info(f"[CLOUD] carregar_form044 → parquet: {len(df)} registros")
            return df
        return pd.DataFrame()
    # ─────────────────────────────────────────────────────────────────────────
    path = FILES_CONFIG['propostas_comerciais']['local_path']
    if not os.path.exists(path):
        logger.warning(f"FORM 044 não encontrado: {path}")
        return pd.DataFrame()
    
    try:
        # Carregar aba principal (primeira aba), pular cabeçalho
        df = pd.read_excel(path, sheet_name=0, header=None, skiprows=2)
        
        # Mapear colunas conforme estrutura real do FORM 044
        # A=0, B=1, C=2, D=3, E=4, F=5, G=6, H=7, ..., P=15
        df_result = pd.DataFrame()
        
        if len(df.columns) > 15:
            df_result['ANO'] = df.iloc[:, 0]           # COL A
            df_result['FAS'] = df.iloc[:, 2]           # COL C
            df_result['EMPRESA'] = df.iloc[:, 6]       # COL G
            df_result['STATUS'] = df.iloc[:, 7]        # COL H
            df_result['DATA_ACEITE'] = df.iloc[:, 15]  # COL P
        if len(df.columns) > 16:
            df_result['COL_Q'] = df.iloc[:, 16]        # COL Q
        
        # Converter DATA_ACEITE para datetime
        df_result['DATA_ACEITE'] = parse_date_safe(df_result['DATA_ACEITE'])
        
        if 'COL_Q' in df_result.columns:
            df_result['COL_Q'] = parse_date_safe(df_result['COL_Q'])
        
        # Filtrar: STATUS = "Aprovada" E DATA_ACEITE > 01/12/2025
        data_corte = pd.Timestamp('2025-12-01')
        
        mask_aprovada = df_result['STATUS'].astype(str).str.strip().str.upper() == 'APROVADA'
        mask_data = df_result['DATA_ACEITE'] > data_corte
        
        df_filtrado = df_result[mask_aprovada & mask_data].copy()
        
        # Limpar linhas com FAS ou EMPRESA vazios
        df_filtrado = df_filtrado[
            df_filtrado['FAS'].notna() & 
            df_filtrado['EMPRESA'].notna() &
            (df_filtrado['EMPRESA'].astype(str).str.len() > 2)
        ]
        
        logger.info(f"FORM 044 carregado: {len(df_filtrado)} propostas aprovadas desde Dez/2025")
        return df_filtrado
        
    except Exception as e:
        logger.error(f"Erro ao carregar FORM 044: {e}")
        return pd.DataFrame()

def normalizar_entrada_pc(entrada):
    """
    Normaliza entrada de PC para garantir que comece com 'PC '.
    
    Args:
        entrada: String representando o PC, e.g., '150.25' ou 'PC 150.25'
    
    Returns:
        String normalizada, e.g., 'PC 150.25'
    """
    # Remove espaços em branco extras
    entrada = str(entrada).strip()
    
    # Verificar se começa com "PC" (case insensitive)
    entrada_upper = entrada.upper()
    if entrada_upper.startswith("PC"):
        # Se começa com "PC" mas não tem espaço, adicionar espaço
        if len(entrada) > 2 and entrada[2] != " ":
            pc_normalizada = "PC " + entrada[2:]
        else:
            pc_normalizada = entrada
    else:
        # Se não começa com "PC", adicionar o prefixo
        pc_normalizada = f"PC {entrada}"
        
    return pc_normalizada.upper()

@st.cache_data(ttl=600, show_spinner=False)
def carregar_form045(pc_number):
    """
    Carrega dados do FORM 045 para um PC específico.
    Busca na estrutura de diretórios baseada no ano do PC.
    
    Args:
        pc_number: Número do PC, e.g., '123.25' ou 'PC 123.25'
    
    Returns:
        DataFrame com colunas ['Serviço', 'NORMA', 'QUANTIDADE'] ou vazio se não encontrado.
    """
    # Normalizar entrada para garantir que comece com "PC "
    pc_number = normalizar_entrada_pc(pc_number)
    
    base_dir = r"G:\.shortcut-targets-by-id\1JbWwLDR6PaShh0-_xJZLFAvEXQKn65V1\008 - Comercial\000 - Comercial\01 - Propostas Comerciais"
    
    # Extrair ano do pc_number (assume formato XXX.YY, onde YY é ano)
    if '.' in pc_number:
        year_suffix = pc_number.split('.')[-1]
        year = '20' + year_suffix
    else:
        year = '2025'  # default
    
    year_dir = os.path.join(base_dir, year)
    if not os.path.exists(year_dir):
        logger.warning(f"Diretório do ano {year} não encontrado: {year_dir}")
        return pd.DataFrame()
    
    # Encontrar pasta começando com pc_number (já normalizado para "PC XXX.XX")
    pc_folder = None
    for item in os.listdir(year_dir):
        if os.path.isdir(os.path.join(year_dir, item)) and item.startswith(pc_number):
            pc_folder = item
            break
    
    if not pc_folder:
        logger.warning(f"Pasta PC para {pc_number} não encontrada em {year_dir}")
        return pd.DataFrame()
    
    pc_path = os.path.join(year_dir, pc_folder)
    
    # Encontrar arquivo Excel contendo "FORM 045"
    excel_file = None
    for file in os.listdir(pc_path):
        if 'FORM 045' in file and (file.endswith('.xlsx') or file.endswith('.xls') or file.endswith('.xlsm')):
            excel_file = file
            break
    
    if not excel_file:
        logger.warning(f"Arquivo FORM 045 não encontrado em {pc_path}")
        return pd.DataFrame()
    
    excel_path = os.path.join(pc_path, excel_file)
    
    try:
        # Carregar Excel, primeira aba
        df = pd.read_excel(excel_path, header=None, engine='openpyxl')
        
        # Extrair linhas 20 a 42 (índices 19 a 41), colunas E F G (4,5,6)
        data = df.iloc[19:42, [4,5,6]].dropna(how='all')
        data.columns = ['Serviço', 'NORMA', 'QUANTIDADE']
        
        # Limpeza: Remove linhas vazias ou com quantidade zero
        data = data.dropna(subset=['Serviço', 'QUANTIDADE'])
        data['QUANTIDADE'] = pd.to_numeric(data['QUANTIDADE'], errors='coerce').fillna(0)
        data = data[data['QUANTIDADE'] > 0]
        
        # Adiciona colunas de identificação
        data['PC_FAS'] = pc_number
        data['EMPRESA'] = ''  # To be filled later if needed
        data['EMPRESA_PC'] = f"PC {pc_number}"
        
        logger.info(f"FORM 045 carregado para PC {pc_number}: {len(data)} linhas de dados")
        return data
        
    except Exception as e:
        logger.error(f"Erro ao carregar FORM 045 para PC {pc_number}: {e}")
        return pd.DataFrame()

@st.cache_data(ttl=300, show_spinner=False)
def consolidar_fas_totais(df_recebimentos, data_inicio='2025-12-01'):
    """Combina FORM 022A e FORM 045 para gerar visão detalhada das FAS."""
    # ── CLOUD GUARD ──────────────────────────────────────────────────────────
    if bridge.is_cloud:
        from cloud_config import carregar_parquet_cache
        df = carregar_parquet_cache("fas_consolidadas")
        if not df.empty:
            logger.info(f"[CLOUD] consolidar_fas_totais → parquet: {len(df)} registros")
            return df
        return pd.DataFrame()
    # ─────────────────────────────────────────────────────────────────────────
    data_corte = pd.Timestamp(data_inicio)

    # --- Tratamento dos dados do FORM 022A (recebimentos) ---
    if df_recebimentos is not None and not df_recebimentos.empty:
        df_execucao = df_recebimentos.copy()
        df_execucao['DATA_RECEBIMENTO_DT'] = pd.to_datetime(
            df_execucao.get('DATA_RECEBIMENTO'), dayfirst=True, errors='coerce'
        )
        df_execucao['DATA_ENTREGA_DT'] = pd.to_datetime(
            df_execucao.get('DATA_ENTREGA'), dayfirst=True, errors='coerce'
        )
        df_execucao['ANO_INT'] = pd.to_numeric(df_execucao.get('ANO'), errors='coerce')
        filtro_periodo = (
            (df_execucao['DATA_RECEBIMENTO_DT'] >= data_corte) |
            (df_execucao['DATA_ENTREGA_DT'] >= data_corte) |
            (df_execucao['ANO_INT'] >= data_corte.year)
        )
        df_execucao = df_execucao[filtro_periodo.fillna(False)].copy()
        if 'NUMERO_PROPOSTA' in df_execucao.columns:
            df_execucao['PC_KEY'] = df_execucao['NUMERO_PROPOSTA'].apply(normalizar_pc)
        else:
            df_execucao['PC_KEY'] = None
    else:
        df_execucao = pd.DataFrame(columns=['PC_KEY'])

    # --- Metadados do FORM 044 (aprovadas) ---
    df_form044 = carregar_form044()
    fas_meta = {}
    if not df_form044.empty:
        df_form044['PC_KEY'] = df_form044['FAS'].apply(normalizar_pc)
        fas_meta = (
            df_form044.dropna(subset=['PC_KEY'])
            .set_index('PC_KEY')[['DATA_ACEITE', 'EMPRESA']]
            .to_dict('index')
        )

    df_lista_pcs = calcular_fas_total(df_form044, df_execucao)
    if df_lista_pcs.empty:
        return pd.DataFrame()

    registros = []
    for _, info_pc in df_lista_pcs.iterrows():
        pc_bruto = info_pc['PC']
        pc_display = normalizar_entrada_pc(pc_bruto)
        pc_key = normalizar_pc(pc_display)
        if not pc_key:
            continue

        origem = info_pc.get('ORIGEM', 'DESCONHECIDO')
        df_status = df_execucao[df_execucao['PC_KEY'] == pc_key]

        cliente = str(info_pc.get('EMPRESA', '')).strip()
        if not df_status.empty:
            clientes_status = df_status['CLIENTE'].dropna()
            if not clientes_status.empty:
                cliente = clientes_status.iloc[0]
        if not cliente and pc_key in fas_meta:
            cliente = fas_meta[pc_key].get('EMPRESA', '')

        status = 'EM ANDAMENTO'
        obs_rec = ''
        material_022 = ''
        data_receb = None
        data_entrega = None

        if not df_status.empty:
            status_series = df_status['STATUS'].astype(str).str.upper()
            if status_series.str.contains('CONCLU|FINALIZ').any():
                status = 'FINALIZADO'
            elif status_series.str.contains('ANDAMENTO|EXECU').any():
                status = 'EM ANDAMENTO'
            elif status_series.str.contains('AGUARD', na=False).any():
                status = 'AGUARDANDO MATERIAL'
            else:
                status = df_status['STATUS'].dropna().iloc[0]

            if 'DATA_RECEBIMENTO_DT' in df_status.columns:
                data_receb = df_status['DATA_RECEBIMENTO_DT'].dropna().min()
            if 'DATA_ENTREGA_DT' in df_status.columns:
                data_entrega = df_status['DATA_ENTREGA_DT'].dropna().max()

            obs_vals = df_status.get('OBS_RECEBIMENTO')
            if obs_vals is not None:
                obs_list = [str(x).strip() for x in obs_vals.dropna().unique() if str(x).strip()]
                if obs_list:
                    obs_rec = '; '.join(obs_list[:3])

            mat_vals = df_status.get('MATERIAL')
            if mat_vals is not None:
                mat_list = [str(x).strip() for x in mat_vals.dropna().unique() if str(x).strip()]
                if mat_list:
                    material_022 = '; '.join(mat_list[:3])
        else:
            status = 'EM ANDAMENTO' if origem == 'EM EXECUÇÃO' else 'AGUARDANDO MATERIAL'

        dados_form044 = fas_meta.get(pc_key, {})
        data_aceite = dados_form044.get('DATA_ACEITE')

        df_servicos = carregar_form045(pc_display)
        if df_servicos.empty:
            registros.append({
                'PC': pc_display,
                'PC_KEY': pc_key,
                'CLIENTE': cliente,
                'STATUS_FAS': status,
                'ORIGEM': origem,
                'SERVICO': 'SERVIÇOS NÃO LOCALIZADOS NO FORM 045',
                'NORMA': '-',
                'QUANTIDADE': 0.0,
                'DATA_ACEITE': data_aceite,
                'DATA_RECEBIMENTO': data_receb,
                'DATA_ENTREGA': data_entrega,
                'OBS_RECEBIMENTO': obs_rec,
                'MATERIAL_022A': material_022
            })
            continue

        df_servicos = df_servicos.rename(columns={'Servico': 'Serviço'})
        for _, linha_servico in df_servicos.iterrows():
            servico_nome = str(linha_servico.get('Serviço') or '').strip()
            if not servico_nome:
                continue
            registros.append({
                'PC': pc_display,
                'PC_KEY': pc_key,
                'CLIENTE': cliente,
                'STATUS_FAS': status,
                'ORIGEM': origem,
                'SERVICO': servico_nome,
                'NORMA': str(linha_servico.get('NORMA') or '').strip() or '-',
                'QUANTIDADE': float(linha_servico.get('QUANTIDADE') or 0),
                'DATA_ACEITE': data_aceite,
                'DATA_RECEBIMENTO': data_receb,
                'DATA_ENTREGA': data_entrega,
                'OBS_RECEBIMENTO': obs_rec,
                'MATERIAL_022A': material_022
            })

    if not registros:
        return pd.DataFrame()

    df_final = pd.DataFrame(registros)
    df_final['DATA_ACEITE'] = pd.to_datetime(df_final['DATA_ACEITE'], errors='coerce')
    return df_final.sort_values(['STATUS_FAS', 'PC', 'SERVICO']).reset_index(drop=True)

def _contar_cc_por_cliente(df_recebimentos):
    df_cc = df_recebimentos[df_recebimentos.get('E_CONTRATO_CONTINUO') == True].copy()
    if df_cc.empty:
        return {}
    df_cc['CLIENTE_NORM'] = df_cc['CLIENTE'].apply(normalizar_texto)
    agrupado = df_cc.groupby('CLIENTE_NORM')['NUMERO_PROPOSTA'].count().to_dict()
    mapeado = {}
    for cliente_raw in df_cc['CLIENTE'].dropna().unique():
        norm = normalizar_texto(cliente_raw)
        mapeado[norm] = cliente_raw
    resultado = {}
    for norm, total in agrupado.items():
        nome = mapeado.get(norm, norm.title())
        resultado[nome] = {
            'unidades': int(total),
            'descricao': 'Ensaios (CC)'
        }
    return resultado

def _contar_compasa(df_recebimentos):
    df_compasa = df_recebimentos[df_recebimentos['CLIENTE'].astype(str).str.contains('COMPASA', case=False, na=False)].copy()
    if df_compasa.empty:
        return {}
    df_compasa['PEDREIRA_NORMALIZADA'] = df_compasa['PEDREIRA'].fillna('SEM PEDREIRA')
    # Sinalizar PROJETO CAUQ (col FINALIDADE)
    if 'FINALIDADE' in df_compasa.columns:
        df_compasa['E_PROJETO_CAUQ'] = df_compasa['FINALIDADE'].astype(str).str.contains('PROJETO', case=False, na=False)
    else:
        df_compasa['E_PROJETO_CAUQ'] = False

    contagem_total = df_compasa.groupby('PEDREIRA_NORMALIZADA')['NUMERO_PROPOSTA'].nunique().to_dict()
    contagem_proj = df_compasa[df_compasa['E_PROJETO_CAUQ']].groupby('PEDREIRA_NORMALIZADA')['NUMERO_PROPOSTA'].nunique().to_dict()

    # Outros ensaios: soma QUANTIDADE (col H) por pedreira (fallback: contar linhas)
    if 'QUANTIDADE' in df_compasa.columns:
        df_compasa['QUANTIDADE_INT'] = pd.to_numeric(df_compasa['QUANTIDADE'], errors='coerce').fillna(0)
    else:
        df_compasa['QUANTIDADE_INT'] = 1
    contagem_outros = (
        df_compasa[~df_compasa['E_PROJETO_CAUQ']]
        .groupby('PEDREIRA_NORMALIZADA')['QUANTIDADE_INT'].sum()
        .to_dict()
    )

    resultado = {}

    # Bloco geral (todos os PTs por pedreira)
    total = 0
    detalhes = []
    for pedreira, qtd in contagem_total.items():
        total += qtd
        detalhes.append(f"{pedreira}: {qtd} PTs")
    if total > 0:
        resultado['COMPASA DO BRASIL'] = {
            'unidades': int(total),
            'descricao': 'PTs',
            'detalhes': detalhes
        }

    # Projetos CAUQ
    total_proj = sum(contagem_proj.values())
    if total_proj > 0:
        detalhes_proj = [f"{p}: {q} PTs PROJETO" for p, q in contagem_proj.items()]
        resultado['COMPASA - PROJETOS CAUQ'] = {
            'unidades': int(total_proj),
            'descricao': 'PTs (PROJETO CAUQ)',
            'detalhes': detalhes_proj
        }

    # Outros ensaios (barra horizontal)
    total_outros = int(sum(contagem_outros.values()))
    if total_outros > 0:
        detalhes_outros = [f"{p}: {int(q)} ensaios" for p, q in contagem_outros.items()]
        resultado['COMPASA - OUTROS ENSAIOS'] = {
            'unidades': total_outros,
            'descricao': 'Ensaios (soma QUANTIDADE col H)',
            'detalhes': detalhes_outros
        }

    return resultado

def _contar_empresa_pc(df_fas):
    if df_fas.empty:
        return {}
    df_tmp = df_fas.copy()
    df_tmp['EMPRESA_NORMALIZADA'] = df_tmp['CLIENTE'].apply(normalizar_texto)
    nome_original = df_tmp.groupby('EMPRESA_NORMALIZADA')['CLIENTE'].agg(lambda x: x.dropna().iloc[0] if not x.dropna().empty else '').to_dict()

    # Unidade = soma de QUANTIDADE (Serviços/Ensaios das FAS); fallback: contagem de SERVICO
    if 'QUANTIDADE' in df_tmp.columns:
        contagem = df_tmp.groupby('EMPRESA_NORMALIZADA')['QUANTIDADE'].sum()
    else:
        contagem = df_tmp.groupby('EMPRESA_NORMALIZADA')['SERVICO'].count()

    resultado = {}
    for norm, total in contagem.to_dict().items():
        nome = nome_original.get(norm, norm.title())
        resultado[f"{nome}"] = {
            'unidades': int(total),          # <- Serviços/Ensaios (unidade principal)
            'descricao': 'Serviços/Ensaios (PC)'
        }
    return resultado

def _contar_cbb_asfaltec(df_raw_cbb):
    if df_raw_cbb.empty:
        return {}
    df_target = df_raw_cbb[df_raw_cbb['CLIENTE'].str.contains('CBB|ASFALTEC', na=False, case=False)].copy()
    if df_target.empty:
        return {}
    df_target['QTD'] = pd.to_numeric(df_target.get('QUANTIDADE', 1), errors='coerce').fillna(1)

    # Excluir materiais de ligante/CAP/AB08 (similaridade ≥ 70%)
    lista_excluir = ['LIGANTE', 'CAP', 'AB08']
    if 'MATERIAL' in df_target.columns:
        def _eh_ligante(texto):
            if not isinstance(texto, str):
                return False
            t = texto.upper()
            for termo in lista_excluir:
                if fuzz.ratio(t, termo) >= 70 or termo in t:
                    return True
            return False
        df_target = df_target[~df_target['MATERIAL'].apply(_eh_ligante)].copy()
        if df_target.empty:
            return {}
    def grupo(cliente):
        cliente = str(cliente).upper()
        if 'ASFALTEC' in cliente:
            return 'ASFALTEC'
        if 'CBB' in cliente:
            return 'CBB ASFALTOS'
        return None
    df_target['GRUPO'] = df_target['CLIENTE'].apply(grupo)
    df_target = df_target.dropna(subset=['GRUPO'])

    # Coluna de PT: pode ser NUMERO_PROPOSTA (SQL) ou PT_COLUNA_A (raw)
    col_pt = 'NUMERO_PROPOSTA' if 'NUMERO_PROPOSTA' in df_target.columns else 'PT_COLUNA_A'

    # Unidade = número de PTs distintos por grupo
    contagem_pts = df_target.groupby('GRUPO')[col_pt].nunique().to_dict()
    # Amostras = soma de QUANTIDADE por grupo (informação complementar)
    contagem_amostras = df_target.groupby('GRUPO')['QTD'].sum().to_dict()

    resultado = {}
    for grupo_nome, total_pts in contagem_pts.items():
        amostras = int(contagem_amostras.get(grupo_nome, 0))
        resultado[grupo_nome] = {
            'unidades': int(total_pts),      # <- PTs (unidade principal)
            'descricao': 'PTs',
            'amostras': amostras,            # <- amostras (informação extra)
            'descricao_amostras': 'Amostras'
        }
    return resultado

def gerar_quantitativos_empresas(df_recebimentos, df_fas, df_raw_cbb):
    """Gera dicionário consolidado com os valores de cada dashboard por empresa."""
    consolidado = {}

    # 1. Contratos Contínuos (CC)
    cc_totais = _contar_cc_por_cliente(df_recebimentos)
    consolidado.update(cc_totais)

    # 2. Compasa do Brasil (pedreiras)
    compasa_totais = _contar_compasa(df_recebimentos)
    consolidado.update(compasa_totais)

    # 3. Empresa-PC (FAS consolidada)
    empresa_pc = _contar_empresa_pc(df_fas)
    consolidado.update(empresa_pc)

    # 4. CBB / Asfaltec
    cbb_totais = _contar_cbb_asfaltec(df_raw_cbb)
    consolidado.update(cbb_totais)

    return consolidado

def buscar_e_extrair_form045(pc_numero, empresa_nome=""):
    """
    Versão com tratamento de erro melhorado e log para depuração.
    """
    # ── CLOUD GUARD ──────────────────────────────────────────────────────────
    if bridge.is_cloud:
        return pd.DataFrame()  # FORM 045 não disponível no cloud
    # ─────────────────────────────────────────────────────────────────────────
    # 1. Normalização (Ex: 041/25 -> 041.25)
    pc_original = str(pc_numero).strip().upper()
    pc_clean = pc_original.replace('/', '.')
    
    # Se o PC não tiver ponto (ex: 04125), tentar inferir ou pular
    if '.' not in pc_clean and len(pc_clean) > 4:
        # Tenta inserir ponto antes dos ultimos 2 digitos (ex 04125 -> 041.25)
        pc_clean = f"{pc_clean[:-2]}.{pc_clean[-2:]}"

    # 2. Identificação do Ano
    if ".25" in pc_clean: ano = "2025"
    elif ".26" in pc_clean: ano = "2026"
    elif ".24" in pc_clean: ano = "2024"
    else: return pd.DataFrame() # Sem ano, impossível achar a pasta
    
    caminho_ano = os.path.join(BASE_DIR_PROPOSTAS, ano)
    if not os.path.exists(caminho_ano):
        # st.warning(f"Diretório de ano não encontrado: {caminho_ano}") # Descomente para debug
        return pd.DataFrame()
    
    # 3. Localizar a pasta da PC (Busca parcial)
    # Procura por "041.25" dentro dos nomes das pastas
    try:
        pastas = os.listdir(caminho_ano)
    except OSError:
        return pd.DataFrame()

    caminho_pc = ""
    # Busca exata do código dentro do nome da pasta
    termo_busca = pc_clean.replace("PC", "").strip() # Busca por "041.25"
    
    for p in pastas:
        if termo_busca in p:
            caminho_pc = os.path.join(caminho_ano, p)
            break
            
    if not caminho_pc: 
        return pd.DataFrame()
    
    # 4. Localizar o arquivo Excel FORM 045
    # Procura qualquer Excel que tenha "045" no nome
    try:
        arquivos = glob.glob(os.path.join(caminho_pc, "*.xl*")) # Pega xls, xlsx, xlsm
    except:
        return pd.DataFrame()

    arquivo_alvo = ""
    for f in arquivos:
        nome_f = os.path.basename(f).upper()
        if "045" in nome_f and "~$" not in nome_f: # Ignora arquivos temporários
            arquivo_alvo = f
            break
            
    if not arquivo_alvo: 
        return pd.DataFrame()
    
    # 5. Extração dos Dados
    try:
        # Tenta ler o range específico do FORM 045 padrão
        df_temp = pd.read_excel(arquivo_alvo, skiprows=19, nrows=30, usecols="E:G", header=None)
        df_temp.columns = ['Servico', 'Norma', 'Quantidade']
        
        df_temp = df_temp.dropna(subset=['Servico'])
        df_temp['Quantidade'] = pd.to_numeric(df_temp['Quantidade'], errors='coerce').fillna(0)
        df_temp = df_temp[df_temp['Quantidade'] > 0]
        
        if not df_temp.empty:
            df_temp['PC_FAS'] = pc_original
            df_temp['EMPRESA'] = empresa_nome
            df_temp['EMPRESA_PC'] = f"{empresa_nome} - {pc_original}"
            return df_temp
            
    except Exception:
        pass
        
    return pd.DataFrame()

# =====================================================================================
# PROTOCOLO VIA ARQUIVO DE PROJETO (COMPOSIÇÃO!I8)
# =====================================================================================
def extrair_protocolo_composicao(path_excel):
    """
    Lê COMPOSIÇÃO!I8 de um arquivo de projeto e normaliza para o formato "PT-XXX".

    Retorna:
        str ou None
    """
    try:
        import openpyxl  # lazy import para não carregar se não usado
        import re
        wb = openpyxl.load_workbook(path_excel, data_only=True, read_only=True)
        # Tentar nomes comuns da aba
        alvo = None
        for nome in wb.sheetnames:
            if nome.strip().upper().startswith('COMPOSIÇÃO'):
                alvo = nome
                break
        if not alvo:
            return None
        ws = wb[alvo]
        val = ws['I8'].value if 'I8' in ws else None
        if not val:
            return None
        texto = str(val).upper().replace('PT', '').replace('-', ' ').replace('_', ' ')
        # Procurar padrão 000.0.0000
        m = re.search(r"(\d{3}[\.-]\d[\.-]\d{4})", texto)
        if not m:
            # tentar capturar números separados por espaços
            m = re.search(r"(\d{3})\s*(\d)\s*(\d{4})", texto)
            if m:
                texto_norm = f"{m.group(1)}.{m.group(2)}.{m.group(3)}"
                return f"PT-{texto_norm}"
            return None
        protocolo = m.group(1).replace('-', '.').strip()
        return f"PT-{protocolo}"
    except Exception:
        return None

if __name__ == "__main__": pass