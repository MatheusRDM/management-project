"""
=========================================================================
CAUQ SCANNER - Leitura e Geocodificação de Projetos CAUQ Marshall
=========================================================================
Suporta dois formatos de planilha:
  Formato A (2019-2020): .xls, abas "CADASTRO PROJETO CAUQ" + "PROJETO FINAL"
  Formato B (2021-2025): .xlsm, abas "GST_CADASTRO PROJETO" + "PROJ_DOSAGEM MARSHALL"
A leitura é DINÂMICA: campos localizados por cabeçalho/label, sem posição fixa.
=========================================================================
"""

import os
import re
import json
import time
import logging
import unicodedata
import warnings
import xlrd
import pandas as pd
from datetime import datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

logging.basicConfig(level=logging.WARNING)
logger = logging.getLogger(__name__)

# ======================================================================================
# CONFIGURAÇÕES
# ======================================================================================

CAUQ_BASE_DIR = r"G:\.shortcut-targets-by-id\1NUJ7pNAqedohSrLjiwFErFrtVqVZkrjk\006 - Lab. Central\0.2 PROJETOS CAUQ MARSHALL\006-PROJETOS"

# Detecta se estamos no modo cloud (sem acesso ao Google Drive)
_CLOUD_MODE = not os.path.isdir(CAUQ_BASE_DIR)

GEOCODE_CACHE_FILE = os.path.join(
    os.path.dirname(__file__), "..", "cache_certificados", "cauq_geocode_cache.json"
)

YEAR_DIRS = {
    2019: "_2019 PROJETOS MARSHALL",
    2020: "_2020 PROJETOS MARSHALL",
    2021: "_2021 PROJETOS MARSHALL",
    2022: "_2022 PROJETOS MARSHALL",
    2023: "_2023 PROJETOS MARSHALL",
    2024: "_2024 PROJETOS MARSHALL",
    2025: "_2025 PROJETOS MARSHALL",
    2026: "_2026 PROJETOS MARSHALL",
}

NORMA_CORES = {
    "DEINFRA": "purple",
    "DER-PR":  "blue",
    "DER":     "blue",
    "DNIT":    "red",
    "DNER":    "red",
    "PMC":     "orange",
    "OUTRO":   "gray",
}

# ── Limites de especificação por norma ──────────────────────────────────────────
# Referências: DER/PR ES-PA 15/23 e 21/23, DNIT 031/2006-ES, DEINFRA-SC ES-P 05/16
SPEC_LIMITS = {
    "DER-PR": {
        "abrasao_la":          {"max": 40, "unit": "%", "ref": "DNER-ME 035"},
        "equivalente_areia":   {"min": 55, "unit": "%", "ref": "DNER-ME 054"},
        "durabilidade_graudo": {"max": 18, "unit": "%", "ref": "DNER-ME 089"},
        "durabilidade_miudo":  {"max": 18, "unit": "%", "ref": "DNER-ME 089"},
        "indice_forma":        {"max": 0.5, "unit": "", "ref": "DNIT-ME 424/20"},
        "lamelaridade":        {"max": 10, "unit": "%", "ref": "AG-01/DAER"},
        "volume_vazios":       {"min": 3, "max": 5, "unit": "%"},
        "rbv":                 {"min": 65, "max": 82, "unit": "%"},
        "vam":                 {"min": 15, "unit": "%"},
        "dui":                 {"min": 70, "unit": "%", "ref": "DNIT-ME 180/18"},
        "deformacao_permanente": {"max": 10, "unit": "%", "ref": "LCPC"},
    },
    "DNIT": {
        "abrasao_la":          {"max": 50, "unit": "%", "ref": "DNER-ME 035"},
        "equivalente_areia":   {"min": 55, "unit": "%", "ref": "DNER-ME 054"},
        "durabilidade_graudo": {"max": 18, "unit": "%", "ref": "DNER-ME 089"},
        "durabilidade_miudo":  {"max": 18, "unit": "%", "ref": "DNER-ME 089"},
        "volume_vazios":       {"min": 3, "max": 5, "unit": "%"},
        "rbv":                 {"min": 65, "max": 82, "unit": "%"},
        "vam":                 {"min": 15, "unit": "%"},
        "dui":                 {"min": 70, "unit": "%", "ref": "DNIT-ME 180/18"},
    },
    "DEINFRA": {
        "abrasao_la":          {"max": 40, "unit": "%", "ref": "DNER-ME 035"},
        "equivalente_areia":   {"min": 55, "unit": "%", "ref": "DNER-ME 054"},
        "durabilidade_graudo": {"max": 18, "unit": "%", "ref": "DNER-ME 089"},
        "durabilidade_miudo":  {"max": 18, "unit": "%", "ref": "DNER-ME 089"},
        "indice_forma":        {"max": 0.5, "unit": "", "ref": "ABNT NBR 7809"},
        "lamelaridade":        {"max": 10, "unit": "%"},
        "volume_vazios":       {"min": 3, "max": 5, "unit": "%"},
        "rbv":                 {"min": 65, "max": 82, "unit": "%"},
        "vam":                 {"min": 15, "unit": "%"},
        "dui":                 {"min": 70, "unit": "%", "ref": "AASHTO T 283"},
    },
}

# ── Normalização de ligantes (POO) ──────────────────────────────────────────────
# Padrão corporativo usado nos dashboards:
#   CAP 50/70, AMP 55/75, AMP 60/85, AMP 65/90, CAP 30/45, HiMA, CAP BORR.
_LIGANTE_ALIAS = {
    # CAP convencionais
    "CAP 50/70": "CAP 50/70",
    "CAP 50-70": "CAP 50/70",
    "CAP 30/45": "CAP 30/45",
    "CAP 30-45": "CAP 30/45",
    "CAP 30/45 PREMIUM": "CAP 30/45",

    # AMP (polímero)
    "AMP 55/75": "AMP 55/75",
    "AMP 55/75 E": "AMP 55/75",
    "COMPAFLEX AMP 55/75 E": "AMP 55/75",
    "FLEXPAVE 55/75": "AMP 55/75",
    "SUPERCAP 85": "AMP 55/75",
    "ASFALTO MODIFICADO POR POLIMERO 55/75": "AMP 55/75",
    "MODIFICADO POR POLIMERO AMP 55/75 E": "AMP 55/75",

    "AMP 60/85": "AMP 60/85",
    "AMP 60/85 E": "AMP 60/85",
    "COMPAFLEX AMP 60/85 E": "AMP 60/85",
    "FLEXPAVE 60/85": "AMP 60/85",
    "STYLINK 60/85": "AMP 60/85",
    "STYLINK 60.85": "AMP 60/85",

    "AMP 65/90": "AMP 65/90",

    # HiMA (alto módulo)
    "HIMA": "HiMA",

    # Asfaltos borracha / AB-08 → CAP BORR.
    "ECOFLEX B": "CAP BORR.",
    "CAPFLEX AB08": "CAP BORR.",
    "AB08": "CAP BORR.",
    "AB-08": "CAP BORR.",
    "CAP BORR": "CAP BORR.",
    "CAP BORR.": "CAP BORR.",
    "CAP BORRACHA": "CAP BORR.",
    "AMP BORRACHA": "CAP BORR.",
    "ASFALTO BORRACHA": "CAP BORR.",
    "MODIFICADO POR POLIMERO COMPAFLEX PA/RE 55/75 TIPO BORRACHA": "CAP BORR.",

    # Valores não informativos
    "NAO INFORMADO": "",
    "ANALISE": "",
}


class LiganteNormalizer:
    """Normaliza diferentes formas de escrever o ligante para um conjunto padrão."""

    PADROES = (
        "CAP 50/70",
        "AMP 55/75",
        "AMP 60/85",
        "AMP 65/90",
        "CAP 30/45",
        "HiMA",
        "CAP BORR.",
    )

    def __init__(self, aliases: dict[str, str]):
        # Pré-normaliza as chaves para facilitar matching "contains"
        self._aliases_raw = aliases
        self._aliases_norm = {
            self._norm_key(k): v for k, v in aliases.items() if k
        }

    def _norm_key(self, s: str) -> str:
        """Normaliza texto para comparação: sem acento, upper, espaços simples."""
        base = unicodedata.normalize("NFKD", str(s))
        base = "".join(c for c in base if not unicodedata.combining(c))
        base = base.upper()
        base = re.sub(r"[^A-Z0-9/\. ]", " ", base)
        base = re.sub(r"\s+", " ", base).strip()
        return base

    def _match_alias_in_text(self, texto: str) -> str | None:
        """Retorna ligante padrão se alguma chave de alias aparecer no texto."""
        if not texto:
            return None
        txt_norm = self._norm_key(texto)
        for key_norm, val in self._aliases_norm.items():
            if key_norm and key_norm in txt_norm:
                return val
        return None

    def normalizar(self, ligante_raw: str | None, pasta_projeto: str | None = None) -> str:
        """
        Normaliza o ligante lido do Excel.

        - Se for um código numérico (certificado), tenta inferir pelo nome da pasta.
        - Caso contrário, aplica o dicionário de aliases.
        - Se não encontrar nada, devolve o texto original.
        """
        if ligante_raw is None:
            return ""

        lig = str(ligante_raw).strip()
        if not lig:
            return ""

        # Caso 1: parece número de certificado / código NCM (ex.: 4004.12.03, 5307)
        if re.match(r"^[\d\.\-]+$", lig):
            if pasta_projeto:
                p_up = _sem_acento(os.path.basename(pasta_projeto)).upper()
                # Procura qualquer alias que apareça no nome da pasta
                for alias_key in self._aliases_raw.keys():
                    if not alias_key:
                        continue
                    ak_norm = self._norm_key(alias_key)
                    if ak_norm and ak_norm in self._norm_key(p_up):
                        return self._aliases_raw[alias_key] or lig
            # Não conseguiu inferir → descarta o valor inválido
            return ""

        # Caso 2: nome de ligante textual → procurar alias conhecido
        padrao = self._match_alias_in_text(lig)
        if padrao is not None:
            return padrao

        # Caso 3: já está exatamente como um dos padrões esperados
        lig_norm = self._norm_key(lig)
        for p in self.PADROES:
            if self._norm_key(p) == lig_norm:
                return p

        # Fallback: devolve texto original
        return lig


_LIGANTE_NORMALIZER = LiganteNormalizer(_LIGANTE_ALIAS)

# ── Coordenadas conhecidas de pedreiras no Paraná e SC ──────────────────────────
PEDREIRAS_COORDS = {
    "CIANORTE":       (-23.6543, -52.6118),
    "COMPASA":        (-23.2383, -51.0482),  # Ibiporã-PR
    "COMPASA IBIPOR": (-23.2383, -51.0482),
    "ICA":            (-25.0927, -50.1618),  # Ponta Grossa-PR
    "DELLAI":         (-23.7796, -49.9568),  # Tomazina-PR
    "DELLAI SIQUEIRA":(-23.6883, -49.8336),  # Siqueira Campos-PR
    "DELLAI PITANGA": (-24.7575, -51.7636),
    "GUARICANA":      (-25.7671, -49.7168),  # Lapa-PR
    "EXPRESSA":       (-25.5660, -51.4747),  # Guarapuava-PR
    "TREVO":          (-24.9555, -53.4561),  # Cascavel-PR
    "TREVO CASCAVEL": (-24.9555, -53.4561),
    "TREVO TERRA ROXA":(-24.1786, -54.0884),
    "TREVO ST":       (-24.1786, -54.0884),
    "NORTE SUL":      (-23.4060, -51.4570),  # Arapongas-PR
    "ITAIPU":         (-25.4372, -49.3034),  # Curitiba-PR
    "MORO":           (-23.4060, -51.4570),  # Arapongas
    "AGUA AMARELA":   (-25.8737, -49.4978),  # Quitandinha-PR
    "BOSCARDIN":      (-25.0157, -50.1391),  # Ponta Grossa-PR
    "SAMP":           (-23.7746, -50.2866),  # Ibaiti-PR
    "RIO QUATI":      (-23.6883, -49.8336),  # Siqueira Campos-PR
    "BRITAFOZ":       (-25.5480, -54.5882),  # Foz do Iguaçu-PR
    "DALMINA":        (-24.5579, -54.0542),  # M.C.Rondon-PR
    "DALBA":          (-26.2266, -52.6765),  # Pato Branco-PR
    "KERBER":         (-26.1155, -48.8763),  # Três Barras-SC
    "VALE DAS PEDRAS":(-25.4372, -49.3034),  # Curitiba
    "SERRA DA PRATA": (-25.6644, -49.1748),  # SJP-PR
    "ARTECIPE":       (-25.4372, -49.3034),  # Curitiba
    "SANTIAGO":       (-24.0490, -52.3732),  # Campo Mourão
    "UBIRAT":         (-24.5316, -53.0017),  # Ubiratã-PR
    "CMIX":           (-24.5579, -54.0542),  # M.C.Rondon
    "MARC":           (-25.5660, -51.4747),  # Guarapuava
    "REMANSO":        (-25.5480, -54.5882),  # Foz do Iguaçu
    "KM 254":         (-25.4372, -49.3034),
    "IBAITI":         (-23.7746, -50.2866),
    "GRABOWSKI":      (-25.4787, -50.6453),  # Irati-PR
    "PLANACON":       (-25.5660, -51.4747),
    "GASPAR":         (-27.8081, -50.3610),  # Lages-SC
    "CALOGERAS":      (-25.0157, -50.1391),  # Ponta Grossa
    "BARRAC":         (-26.0730, -53.6337),  # Barracão-PR
    "CASTILHO":       (-23.0649, -50.0555),  # Cornélio Procópio
    "APUCARANA":      (-23.5508, -51.4608),
    "ITAPOR":         (-24.6870, -54.0004),  # Itaporã
    "ITAX":           (-25.5660, -51.4747),  # Guarapuava
    "DALL ROSS":      (-26.2266, -52.6765),  # Pato Branco
    "DAL ROSS":       (-26.2266, -52.6765),
    "COMPENSA":       (-24.0490, -52.3732),  # Campo Mourão
    "INFRASUL":       (-26.2336, -51.0828),  # Porto União
    "KERBER PORTO":   (-26.2336, -51.0828),
    "BRITABAL":       (-25.5480, -54.5882),
    "PEDRA NORTE":    (-23.2383, -51.0482),
    "ZANCANARO":      (-25.9319, -52.7096),  # Coronel Vivida
    "POROLA":         (-23.7795, -53.3967),  # Pérola-PR
    "SANTA MARIA":    (-23.4060, -51.4570),
    "PARAGUA":        (-22.4131, -50.5761),  # Paraguaçu Paulista-SP
    "MINERADORA BRITASUL": (-26.1155, -48.8763),
    "MINERACAO CPV":  (-26.1155, -48.8763),
    "AGRO ROQUE":     (-24.0490, -52.3732),
    "PICCINE":        (-25.0157, -50.1391),
    "TUPY":           (-25.4372, -49.3034),
    "TCE":            (-25.4372, -49.3034),
    "ICORP":          (-25.0157, -50.1391),
    "DRISNER":        (-25.4787, -50.6453),
    "CEGE":           (-25.5660, -51.4747),
    "CASALI":         (-24.9555, -53.4561),
    "COTRAGON":       (-24.5579, -54.0542),
    "CONPASUL":       (-26.2266, -52.6765),
    "FORTUNATO":      (-24.9555, -53.4561),
    "JULIO":          (-23.7746, -50.2866),
    "BARALDI":        (-24.0490, -52.3732),
    "KAROLINE":       (-24.5579, -54.0542),
    "SBM":            (-25.4372, -49.3034),
    "SETEP":          (-26.9201, -48.6476),
    "OESTE":          (-24.9555, -53.4561),
    "IMBAU":          (-24.4489, -50.7573),  # Imbaú-PR
    "BRITAX":         (-23.2383, -51.0482),
}

# Campos cujo valor é percentual e pode ser armazenado como fração decimal no Excel
# (ex: célula formatada como "%" guarda 0.15 para "15%")
_CAMPOS_PERCENTUAL = {
    "abrasao_la", "durabilidade_graudo", "durabilidade_miudo",
    "equivalente_areia", "lamelaridade", "adesividade",
    "teor", "volume_vazios", "rbv", "vam", "dui",
}

# ======================================================================================
# ADAPTADOR UNIVERSAL — xlrd (.xls) + openpyxl (.xlsx / .xlsm)
# ======================================================================================

class _Sheet:
    """Interface uniforme (0-indexed) para xlrd.sheet e openpyxl.Worksheet."""

    def __init__(self, sheet, lib: str):
        self._lib = lib
        if lib == "xlrd":
            self._s    = sheet
            self.nrows = sheet.nrows
            self.ncols = sheet.ncols
        else:
            # Pré-carrega todas as linhas para acesso O(1) (compatível com read_only=True)
            rows = list(sheet.iter_rows(values_only=True))
            self._rows = rows
            self.nrows = len(rows)
            self.ncols = max((len(r) for r in rows), default=0) if rows else 0

    def cell_value(self, row: int, col: int):
        """Retorna o valor da célula (0-indexed). Células vazias → ''."""
        try:
            if self._lib == "xlrd":
                return self._s.cell_value(row, col)
            v = self._rows[row][col]
            return v if v is not None else ""
        except Exception:
            return ""


class _Workbook:
    """Interface uniforme para xlrd.Book e openpyxl.Workbook."""

    def __init__(self, wb, lib: str):
        self._wb  = wb
        self._lib = lib

    def sheet_names(self) -> list:
        if self._lib == "xlrd":
            return self._wb.sheet_names()
        return list(self._wb.sheetnames)

    def sheet_by_name(self, name: str) -> _Sheet:
        if self._lib == "xlrd":
            return _Sheet(self._wb.sheet_by_name(name), self._lib)
        return _Sheet(self._wb[name], self._lib)

    def close(self):
        try:
            self._wb.close()
        except Exception:
            pass


def _open_workbook(path: str):
    """Abre .xls com xlrd ou .xlsx/.xlsm com openpyxl. Retorna None em falha."""
    ext = os.path.splitext(path)[1].lower()
    # Tenta xlrd para .xls
    if ext == ".xls":
        try:
            wb = xlrd.open_workbook(path, formatting_info=False)
            return _Workbook(wb, "xlrd")
        except Exception:
            pass  # fallback para openpyxl (arquivo pode ser xlsx renomeado)
    # openpyxl para .xlsx/.xlsm e como fallback do .xls corrompido
    if ext in (".xls", ".xlsx", ".xlsm", ".xlam", ".xltm", ".xltx"):
        try:
            import openpyxl
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            return _Workbook(wb, "openpyxl")
        except Exception as e:
            logger.warning(f"Erro abrindo {path}: {e}")
    return None

# ======================================================================================
# CACHE DE GEOCODIFICAÇÃO
# ======================================================================================

def _load_geocache() -> dict:
    try:
        with open(GEOCODE_CACHE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _save_geocache(cache: dict):
    try:
        os.makedirs(os.path.dirname(GEOCODE_CACHE_FILE), exist_ok=True)
        with open(GEOCODE_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.warning(f"Não foi possível salvar geocache: {e}")


def geocodificar(localizacao: str) -> tuple:
    """Geocodifica via Nominatim (OpenStreetMap, sem API key)."""
    if not localizacao or localizacao.strip() in ("-", "", "N/A"):
        return None, None

    cache = _load_geocache()
    chave = localizacao.strip().upper()
    if chave in cache:
        c = cache[chave]
        return c.get("lat"), c.get("lon")

    try:
        from geopy.geocoders import Nominatim
        geo = Nominatim(user_agent="afirma_evias_cauq_v1", timeout=10)
        for q in [f"{localizacao}, Brasil", f"{localizacao}, Brazil", localizacao]:
            time.sleep(1.1)
            r = geo.geocode(q)
            if r:
                cache[chave] = {"lat": r.latitude, "lon": r.longitude}
                _save_geocache(cache)
                return r.latitude, r.longitude
        cache[chave] = {"lat": None, "lon": None}
        _save_geocache(cache)
        return None, None
    except Exception as e:
        logger.warning(f"Erro geocodificando '{localizacao}': {e}")
        return None, None

# ======================================================================================
# BUSCA DO ARQUIVO EXCEL
# ======================================================================================

def _encontrar_excel(pasta: str) -> str | None:
    """
    Encontra o arquivo Excel principal de um projeto.
    Suporta .xls (Formato A) e .xlsm (Formato B).

    Prioridade:
      1. Raiz: arquivo com 'CAUQ' no nome (ED > R01 > sem revisão)
      2. Subpasta 005-ENTREGA / 007-RESULTADOS / outras: arquivo com 'CAUQ' no nome
      3. Fallback: maior .xlsm/.xls na raiz (ex: 2021 sem 'CAUQ' no nome)
    """
    EXTS = (".xls", ".xlsx", ".xlsm")

    def _prioridade(nome: str) -> int:
        n = nome.upper()
        if re.search(r"[\s\-_]?ED\.XLS", n, re.I): return 0   # edição final
        if re.search(r"\bR0[1-9]\b", n):            return 1   # revisão
        return 2

    def _listar_cauq(d: str) -> list:
        try:
            arqs = [f for f in os.listdir(d)
                    if any(f.lower().endswith(e) for e in EXTS)
                    and "CAUQ" in f.upper()]
            return sorted(arqs, key=_prioridade)
        except Exception:
            return []

    # 1. Raiz
    arqs = _listar_cauq(pasta)
    if arqs:
        return os.path.join(pasta, arqs[0])

    # 2. Subpastas (1 nível)
    def _prio_sub(s: str) -> int:
        u = s.upper()
        if "ENTREGA"   in u: return 0
        if "RESULTADO" in u: return 1
        if "DOSAGEM"   in u: return 2
        return 3

    try:
        subs = sorted(
            [s for s in os.listdir(pasta) if os.path.isdir(os.path.join(pasta, s))],
            key=_prio_sub,
        )
    except Exception:
        subs = []

    for sub in subs:
        arqs = _listar_cauq(os.path.join(pasta, sub))
        if arqs:
            return os.path.join(pasta, sub, arqs[0])

    # 3. Fallback: maior .xlsm/.xls em subpastas (sem exigir "CAUQ" no nome)
    for sub in subs:
        try:
            arqs = [f for f in os.listdir(os.path.join(pasta, sub))
                    if any(f.lower().endswith(e) for e in EXTS)]
            if arqs:
                arqs.sort(key=lambda f: os.path.getsize(os.path.join(pasta, sub, f)), reverse=True)
                return os.path.join(pasta, sub, arqs[0])
        except Exception:
            continue

    # 4. Fallback final: maior .xlsm/.xls na raiz
    try:
        todos = [f for f in os.listdir(pasta)
                 if any(f.lower().endswith(e) for e in EXTS)]
        if todos:
            todos.sort(key=lambda f: os.path.getsize(os.path.join(pasta, f)), reverse=True)
            return os.path.join(pasta, todos[0])
    except Exception:
        pass

    return None

# ======================================================================================
# UTILITÁRIOS DE LEITURA
# ======================================================================================

def _safe_cell(sheet, row: int, col: int, default="") -> str:
    """Lê célula como string. Sempre retorna string."""
    try:
        val = sheet.cell_value(row, col)
        if val is None or val == "":
            return default
        if isinstance(val, float):
            if val == 0.0:
                return default
            return str(int(val)) if val == int(val) else str(val)
        return str(val).strip() or default
    except Exception:
        return default


def _safe_float(sheet, row: int, col: int) -> float | None:
    """Lê célula como float."""
    try:
        val = sheet.cell_value(row, col)
        if isinstance(val, (int, float)) and val != "":
            return float(val)
        # Tentar converter string para float (ex: "5,5" → 5.5)
        if isinstance(val, str):
            v = val.strip().replace(",", ".")
            if v:
                return float(v)
        return None
    except Exception:
        return None


# Limite máximo razoável (em %) por campo — evita ×100 em valores já em %
_CAMPO_LIMITES_MAX = {
    "durabilidade_graudo": 12.0,   # spec máx DER-PR / DNIT = 12 %
    "durabilidade_miudo":  15.0,   # spec máx DER-PR / DNIT = 15 %
}


def _corrigir_percentuais(dados: dict) -> dict:
    """
    Corrige campos percentuais armazenados como frações decimais pelo Excel.
    Células formatadas como "%" no Excel guardam o valor raw (ex: 15% → 0.15).
    Se o valor lido está em (0, 1], multiplica por 100 — EXCETO se o resultado
    exceder o limite físico do campo (o valor já estava em %).
    """
    for campo in _CAMPOS_PERCENTUAL:
        val = dados.get(campo)
        if val is not None and isinstance(val, (int, float)) and 0 < val <= 1.0:
            limite = _CAMPO_LIMITES_MAX.get(campo)
            if limite is None or (val * 100) <= limite:
                dados[campo] = round(val * 100, 4)
    return dados

# ======================================================================================
# HELPERS DINÂMICOS — LEITURA POR CABEÇALHO
# ======================================================================================

def _sem_acento(texto: str) -> str:
    """Remove acentos: MÉDIA→MEDIA, PROCEDÊNCIA→PROCEDENCIA etc."""
    return "".join(
        c for c in unicodedata.normalize("NFKD", str(texto))
        if not unicodedata.combining(c)
    )


def _mapa_cabecalhos(sheet, row: int) -> dict:
    """Varre linha `row` → {texto_normalizado: coluna_0indexed}."""
    mapa = {}
    for col in range(sheet.ncols):
        try:
            raw = str(sheet.cell_value(row, col)).strip()
            if not raw or raw in ("0", ""):
                continue
            chave = re.sub(r"[°º:]+", "", raw.upper()).strip()
            chave = re.sub(r"\s+", " ", chave)
            if chave:
                mapa[chave] = col
        except Exception:
            pass
    return mapa


def _buscar_col(mapa: dict, *keywords) -> int | None:
    """Retorna a coluna cujo cabeçalho contém QUALQUER keyword (sem acento)."""
    for kw in keywords:
        kw_n = _sem_acento(kw).upper()
        for chave, col in mapa.items():
            if kw_n in _sem_acento(chave).upper():
                return col
    return None


def _encontrar_linha(sheet, *keywords, row_start: int = 0, row_max: int = 600) -> int | None:
    """
    Encontra a PRIMEIRA linha cujo conteúdo total contém TODAS as keywords.
    Normaliza acentos em ambos os lados (MEDIA casa com MÉDIA, etc.).
    """
    kws_norm = [_sem_acento(kw).upper() for kw in keywords]
    for r in range(row_start, min(sheet.nrows, row_max)):
        try:
            linha = _sem_acento(" ".join(
                str(sheet.cell_value(r, c)).upper()
                for c in range(sheet.ncols)
            ))
            if all(kw in linha for kw in kws_norm):
                return r
        except Exception:
            pass
    return None

# ======================================================================================
# EXTRAÇÃO — CADASTRO (Formato A horizontal + Formato B vertical)
# ======================================================================================

def _ler_cadastro(sh_cad) -> dict:
    """
    Lê aba CADASTRO suportando dois formatos:

    Formato A (antigo - CADASTRO PROJETO CAUQ):
      Linha de cabeçalho horizontal → dados na linha seguinte.

    Formato B (novo - GST_CADASTRO PROJETO):
      Cada campo em sua própria linha: col D = label, col E = valor.
      Ex: Row 27: "Procedência:" | "PEDREIRA CIANORTE"
    """
    resultado = {
        "num_projeto": "", "procedencia": "", "localizacao": "",
        "natureza_mineralogica": "", "ligante": "", "faixa_granulometrica": "",
    }

    # ── Formato A: cabeçalho horizontal (primeiras 20 linhas) ────────────────
    row_h = _encontrar_linha(sh_cad, "PROCEDENCIA", "FAIXA", row_max=20)
    if row_h is None:
        row_h = _encontrar_linha(sh_cad, "PROCED", "NATUREZA", row_max=20)

    if row_h is not None:
        row_d = row_h + 1
        mapa  = _mapa_cabecalhos(sh_cad, row_h)

        def _campo(*kws) -> str:
            col = _buscar_col(mapa, *kws)
            return _safe_cell(sh_cad, row_d, col) if col is not None else ""

        resultado["num_projeto"]           = _campo("N PROJETO", "PROJETO")
        resultado["procedencia"]           = _campo("PROCED")
        resultado["localizacao"]           = _campo("LOCALIZA")
        resultado["natureza_mineralogica"] = _campo("NATUREZA", "MINERAL")
        resultado["ligante"]               = _campo("LIGANTE")
        resultado["faixa_granulometrica"]  = _campo("FAIXA")

        # fallback LOCAL DA OBRA
        if not resultado["localizacao"] or resultado["localizacao"] in ("-", ""):
            c2 = _buscar_col(mapa, "LOCAL DA OBRA", "LOCAL OBRA")
            if c2 is not None:
                resultado["localizacao"] = _safe_cell(sh_cad, row_d, c2)

        if not resultado["num_projeto"]:
            resultado["num_projeto"] = _safe_cell(sh_cad, row_d, 0)

        return resultado

    # ── Formato B: leitura vertical (GST_CADASTRO PROJETO) ───────────────────
    # Mapa: campo → lista de keywords que o identificam
    CAMPOS_KWS = {
        "num_projeto":           ["N PROJETO", "NUMERO"],
        "procedencia":           ["PROCED"],
        "localizacao":           ["LOCALIZA", "LOCAL DA OBRA", "LOCAL OBRA"],
        "natureza_mineralogica": ["NATUREZA", "MINERAL"],
        "faixa_granulometrica":  ["FAIXA", "ESPECIFICACAO"],
        "ligante":               ["TIPO DE LIGANTE", "LIGANTE"],
    }
    encontrados: set = set()

    for r in range(min(sh_cad.nrows, 80)):
        if len(encontrados) >= len(CAMPOS_KWS):
            break
        for c in range(min(sh_cad.ncols, 6)):
            raw   = str(sh_cad.cell_value(r, c))
            label = _sem_acento(re.sub(r"[°º:.]+", "", raw.upper()).strip())
            if not label:
                continue
            for campo, kws in CAMPOS_KWS.items():
                if campo in encontrados:
                    continue
                kws_n = [_sem_acento(k).upper() for k in kws]
                if any(kw in label for kw in kws_n):
                    # Próxima célula não-vazia à direita
                    # Para "ligante": varre até 15 células, preferindo texto a código numérico
                    if campo == "ligante":
                        best = None
                        for c2 in range(c + 1, min(sh_cad.ncols, c + 15)):
                            v = _safe_cell(sh_cad, r, c2)
                            if not v or v in ("-", "0"):
                                continue
                            # Tenta normalizar: se reconhecer padrão, é o melhor
                            normed = _LIGANTE_NORMALIZER.normalizar(v)
                            if normed:
                                best = normed
                                break
                            # Guarda o primeiro candidato não-numérico como backup
                            if best is None and not re.match(r"^[\d\.\-]+$", str(v)):
                                best = v
                        if best:
                            resultado[campo] = best
                            encontrados.add(campo)
                    else:
                        for c2 in range(c + 1, min(sh_cad.ncols, c + 8)):
                            v = _safe_cell(sh_cad, r, c2)
                            if v and v not in ("-", "0"):
                                resultado[campo] = v
                                encontrados.add(campo)
                                break
                    break

    return resultado

# ======================================================================================
# EXTRAÇÃO — SEÇÃO DE ENSAIOS (AGREGADOS ou MARSHALL)
# ======================================================================================

def _ler_secao_ensaios(sh_pf, row_start: int, row_max: int) -> dict:
    """
    Localiza seção de ensaios pelo cabeçalho e lê cada campo pelo label da linha.
    Funciona para AMBOS os formatos:
      - Formato A: header "Tipos de Ensaios" | "Resultados"
      - Formato B: header "ENSAIOS" | "RESULTADOS"
    """
    campos: dict = {}

    # Busca cabeçalho com "ENSAIO" + "RESULTADO" (funciona nos dois formatos)
    row_h = _encontrar_linha(sh_pf, "ENSAIO", "RESULTADO",
                             row_start=row_start, row_max=row_max)
    if row_h is None:
        return campos

    mapa_h     = _mapa_cabecalhos(sh_pf, row_h)
    col_tipo   = _buscar_col(mapa_h, "TIPO", "ENSAIO")   # "Tipos de Ensaios" ou "ENSAIOS"
    col_result = _buscar_col(mapa_h, "RESULTADO")

    if col_tipo is None or col_result is None:
        return campos

    vazias = 0
    for r in range(row_h + 1, min(sh_pf.nrows, row_h + 35)):
        try:
            label = _sem_acento(str(sh_pf.cell_value(r, col_tipo))).upper().strip()
            # ignorar sub-cabeçalhos como "Mínimo:" / "Máximo:"
            label = re.sub(r"[°º:.]+", "", label).strip()
        except Exception:
            label = ""

        if not label or label in ("EM ANALISE", "LIM MIN", "LIM MAX", "MINIMO", "MAXIMO"):
            vazias += 1
            if vazias >= 3:
                break
            continue
        vazias = 0

        val = _safe_float(sh_pf, r, col_result)

        # ── Características do Agregado ─────────────────────────────────────
        if "ABRAS" in label and "abrasao_la" not in campos:
            campos["abrasao_la"] = val

        elif "FORMA" in label and "LAMELAR" not in label and "indice_forma" not in campos:
            campos["indice_forma"] = val

        elif "DURABILID" in label and ("GRA" in label or "GRAUD" in label) and "durabilidade_graudo" not in campos:
            campos["durabilidade_graudo"] = val

        elif "DURABILID" in label and ("GRA" not in label and "GRAUD" not in label) and "durabilidade_miudo" not in campos:
            campos["durabilidade_miudo"] = val

        elif "EQUIVALENTE" in label and "AREIA" in label and "equivalente_areia" not in campos:
            campos["equivalente_areia"] = val

        elif "LAMELAR" in label and "lamelaridade" not in campos:
            campos["lamelaridade"] = val

        elif "ADESIV" in label and "adesividade" not in campos:
            campos["adesividade"] = val

        # ── Parâmetros Marshall ──────────────────────────────────────────────
        elif "TEOR" in label and "teor" not in campos:
            campos["teor"] = val

        elif "VOLUME" in label and "VAZIO" in label and "volume_vazios" not in campos:
            campos["volume_vazios"] = val

        elif ("RELACAO" in label or "RBV" in label) and "BETUME" in label and "VAZIO" in label and "rbv" not in campos:
            campos["rbv"] = val

        elif "MINERAL" in label and "VAZIO" in label and "vam" not in campos:
            campos["vam"] = val

        elif "MAXIMA" in label and "MEDIDA" in label and "rice" not in campos:
            campos["rice"] = val

        elif "APARENTE" in label and "densidade_aparente" not in campos:
            campos["densidade_aparente"] = val

        elif (("UMIDADE" in label and "INDUZ" in label) or "DUI" in label) and "dui" not in campos:
            campos["dui"] = val

        elif "FILLER" in label and "filler_betume" not in campos:
            campos["filler_betume"] = val

    return campos

# ======================================================================================
# EXTRAÇÃO — DEFORMAÇÃO PERMANENTE
# ======================================================================================

def _ler_deformacao(sh_def) -> float | None:
    """
    Lê Deformação Permanente dinamicamente:
    Encontra linha com 'DEFORMA' + 'MEDIA' e retorna o 1º float plausível (0.01–50 mm).
    """
    row_m = _encontrar_linha(sh_def, "DEFORMA", "MEDIA", row_max=sh_def.nrows)
    if row_m is None:
        row_m = _encontrar_linha(sh_def, "DEFORMACAO", row_max=sh_def.nrows)
    if row_m is None:
        return None

    for c in range(sh_def.ncols):
        val = _safe_float(sh_def, row_m, c)
        if val is not None and 0.01 < abs(val) < 50:
            return abs(val)
    return None

# ======================================================================================
# EXTRAÇÃO PRINCIPAL DO PROJETO
# ======================================================================================

def _extrair_dados_projeto(pasta: str, ano: int) -> dict | None:
    """
    Extrai todos os dados de um projeto CAUQ de forma dinâmica.
    Suporta Formato A (CADASTRO PROJETO CAUQ / PROJETO FINAL)
              e Formato B (GST_CADASTRO PROJETO / PROJ_DOSAGEM MARSHALL).
    """
    xls_path = _encontrar_excel(pasta)
    if not xls_path:
        return None

    wb = _open_workbook(xls_path)
    if wb is None:
        return None

    dados: dict = {
        "ano":    ano,
        "pasta":  os.path.basename(pasta),
        "arquivo": os.path.basename(xls_path),
        "num_projeto": "", "procedencia": "", "localizacao": "",
        "natureza_mineralogica": "", "ligante": "", "faixa_granulometrica": "",
        "norma": "OUTRO",
        "teor": None, "volume_vazios": None, "rbv": None, "vam": None,
        "rice": None, "densidade_aparente": None, "dui": None, "filler_betume": None,
        "abrasao_la": None, "indice_forma": None, "durabilidade_graudo": None,
        "durabilidade_miudo": None, "equivalente_areia": None,
        "lamelaridade": None, "adesividade": None,
        "deformacao_permanente": None,
        "lat": None, "lon": None,
    }

    nomes_abas = {n.upper(): n for n in wb.sheet_names()}

    # ── 1. CADASTRO ───────────────────────────────────────────────────────────
    # Aceita: "CADASTRO PROJETO CAUQ" (Fmt A) ou "GST_CADASTRO PROJETO" (Fmt B)
    try:
        aba_cad = next((orig for up, orig in nomes_abas.items()
                        if "CADASTRO" in up), None)
        if aba_cad:
            dados.update(_ler_cadastro(wb.sheet_by_name(aba_cad)))
    except Exception as e:
        logger.warning(f"CADASTRO {xls_path}: {e}")

    # ── Detectar norma pela faixa granulométrica ──────────────────────────────
    faixa_up = _sem_acento(dados["faixa_granulometrica"]).upper()
    if "DEINFRA" in faixa_up:
        dados["norma"] = "DEINFRA"
    elif "DER" in faixa_up and "DNER" not in faixa_up:
        dados["norma"] = "DER-PR"
    elif "DNIT" in faixa_up:
        dados["norma"] = "DNIT"
    elif "DNER" in faixa_up:
        dados["norma"] = "DNIT"  # DNER virou DNIT
    elif "PMC" in faixa_up:
        dados["norma"] = "DER-PR"  # PMC Curitiba segue DER-PR

    # Fallback norma/faixa pelo nome da pasta
    if dados["norma"] == "OUTRO" or not dados["faixa_granulometrica"]:
        p_up = _sem_acento(os.path.basename(pasta)).upper()
        # Detectar faixa
        m_f = re.search(r"\bF(?:X|AIXA)\s*([A-Z0-9]+(?:[-_\s]\d+[,.]\d+)?)", p_up)
        # Detectar norma (inclui DNER, PMC)
        m_n = re.search(r"\b(DEINFRA|DER[-\s]?PR|DER\b|DNIT|DNER|PMC)\b", p_up)
        if m_f:
            if m_n:
                norma_str = re.sub(r"\s+", "-", m_n.group(1).strip()).upper()
                if norma_str == "DER":
                    norma_str = "DER-PR"
                elif norma_str == "DNER":
                    norma_str = "DNIT"
                elif norma_str == "PMC":
                    norma_str = "DER-PR"
                faixa_str = f"FAIXA {m_f.group(1).strip()} {norma_str}"
            else:
                faixa_str = f"FAIXA {m_f.group(1).strip()}"
            if not dados["faixa_granulometrica"]:
                dados["faixa_granulometrica"] = faixa_str
            if dados["norma"] == "OUTRO" and m_n:
                if "DEINFRA" in norma_str: dados["norma"] = "DEINFRA"
                elif "DER"    in norma_str: dados["norma"] = "DER-PR"
                elif "DNIT"   in norma_str: dados["norma"] = "DNIT"
        elif m_n and dados["norma"] == "OUTRO":
            norma_str = m_n.group(1).strip().upper()
            if norma_str in ("DER", "DER-PR", "PMC"):
                dados["norma"] = "DER-PR"
            elif norma_str in ("DNIT", "DNER"):
                dados["norma"] = "DNIT"
            elif norma_str == "DEINFRA":
                dados["norma"] = "DEINFRA"

    # Normalizar num_projeto para NN.2.AAAA (extrair da pasta se necessário)
    if not re.match(r'^\d+\.\d+\.\d{4}', dados.get('num_projeto', '')):
        _m_num = re.search(r'(\d+\.2\.\d{4})', os.path.basename(pasta))
        if _m_num:
            dados['num_projeto'] = _m_num.group(1)

    # ── 2. PROJETO FINAL / PROJ_DOSAGEM MARSHALL ──────────────────────────────
    # Aceita: "PROJETO FINAL" (Fmt A) ou "PROJ_DOSAGEM MARSHALL" (Fmt B)
    try:
        aba_pf = next(
            (orig for up, orig in nomes_abas.items()
             if "PROJETO FINAL" in up
             or "PROJ_DOSAGEM"  in up
             or ("DOSAGEM" in up and "MARSHALL" in up)),
            None,
        )
        if aba_pf:
            sh_pf = wb.sheet_by_name(aba_pf)

            # Seção AGREGADOS: localiza título "...AGREGAD..." a partir de row 50
            r_agr = _encontrar_linha(sh_pf, "AGREGAD", row_start=50, row_max=300)
            if r_agr is None:
                r_agr = 100  # fallback
            dados.update(_ler_secao_ensaios(sh_pf, row_start=r_agr, row_max=r_agr + 70))

            # Seção MARSHALL: localiza título "...MARSHALL..." a partir de row 250
            r_mar = _encontrar_linha(sh_pf, "MARSHALL", row_start=250, row_max=sh_pf.nrows)
            if r_mar is None:
                r_mar = 300
            dados.update(_ler_secao_ensaios(sh_pf, row_start=r_mar, row_max=sh_pf.nrows))
    except Exception as e:
        logger.warning(f"PROJETO FINAL {xls_path}: {e}")

    # ── 3. DEFORMAÇÃO PERMANENTE ──────────────────────────────────────────────
    # Aceita: "DEFORMAÇÃO PERMANENTE" (Fmt A) ou "DP" (Fmt B)
    try:
        aba_def = next(
            (orig for up, orig in nomes_abas.items()
             if "DEFORMA" in up or up.strip() == "DP"),
            None,
        )
        if aba_def:
            val = _ler_deformacao(wb.sheet_by_name(aba_def))
            if val is not None:
                dados["deformacao_permanente"] = val
    except Exception as e:
        logger.warning(f"DEFORMAÇÃO {xls_path}: {e}")

    # ── 4. Normalizar ligante (classe LiganteNormalizer) ──────────────────────
    dados["ligante"] = _LIGANTE_NORMALIZER.normalizar(
        dados.get("ligante", ""), pasta
    )

    # ── 5. Geocodificar pela procedência (pedreira) ───────────────────────────
    if dados["lat"] is None and dados["procedencia"]:
        proc_up = _sem_acento(dados["procedencia"]).upper()
        for pedreira_key, (lat, lon) in PEDREIRAS_COORDS.items():
            if _sem_acento(pedreira_key).upper() in proc_up:
                dados["lat"] = lat
                dados["lon"] = lon
                break

    dados = _corrigir_percentuais(dados)
    try:
        wb.close()
    except Exception:
        pass
    return dados

# ======================================================================================
# NORMALIZAÇÃO DE LOCALIZAÇÕES
# ======================================================================================

_ESTADOS_BR = {
    "AC","AL","AP","AM","BA","CE","DF","ES","GO","MA",
    "MT","MS","MG","PA","PB","PR","PE","PI","RJ","RN",
    "RS","RO","RR","SC","SP","SE","TO",
}

_PREFIXOS_REMOVER = [
    "LOCALIZADA EM ", "LOCALIZADO EM ", "LOCALIZACAO: ",
    "LOCALIZACAO: ", "MUNICIPIO DE ", "MUNICIPIO DE ",
]

_NOMES_INVALIDOS = {
    "ENG CIVIL", "ENGENHEIRO CIVIL", "RESPONSAVEL", "RESPONSAVEL",
}

_MAX_CIDADE_LEN = 60

_TYPOS_LOCALIZ = {
    "CASACAVEL": "CASCAVEL",
    "SAO MATHEUS": "SAO MATEUS",
}

_CIDADE_ESTADO_FIXO = {
    "TOLEDO":               "TOLEDO - PR",
    "CAMPO MOURAO":         "CAMPO MOURÃO - PR",
    "GUARAPUAVA":           "GUARAPUAVA - PR",
    "SAO JOAO DO IVAI":     "SÃO JOÃO DO IVAÍ - PR",
    "LAPA":                 "LAPA - PR",
    "SAO JOSE DOS PINHAIS": "SÃO JOSÉ DOS PINHAIS - PR",
    "ARAPONGAS":            "ARAPONGAS - PR",
    "ARAPOTI":              "ARAPOTI - PR",
    "CASCAVEL":             "CASCAVEL - PR",
    "CIANORTE":             "CIANORTE - PR",
    "CORONEL VIVIDA":       "CORONEL VIVIDA - PR",
    "CAMPO LARGO":          "CAMPO LARGO - PR",
    "BUTIA":                "BUTÍA - RS",
    "TRES BARRAS":          "TRÊS BARRAS - SC",
    "PORTO UNIAO":          "PORTO UNIÃO - SC",
    "IBIRAMA":              "IBIRAMA - SC",
    "LAGES":                "LAGES - SC",
    "IRANI":                "IRANI - SC",
}


def _limpar_localizacao(loc: str) -> str:
    if not loc:
        return ""
    loc_n = _sem_acento(loc)
    for p in _PREFIXOS_REMOVER:
        if loc_n.startswith(_sem_acento(p)):
            loc = loc[len(p):].strip()
            loc_n = _sem_acento(loc)
            break
    if len(loc) > _MAX_CIDADE_LEN:
        return ""
    if _sem_acento(loc) in {_sem_acento(x) for x in _NOMES_INVALIDOS}:
        return ""
    palavras_invalidas = ("OBRAS ", "PAVIMENTA", "RODOVIA ", "VINCULADAS", "IMPLANTA",
                          "CONFORME PROPOS", "PROPOSTA COMERCIAL")
    if any(p in loc_n for p in [_sem_acento(x) for x in palavras_invalidas]):
        return ""
    if re.search(r'\d{3}/\d{4}', loc):
        return ""
    if " E " in loc:
        partes = [p.strip() for p in loc.split(" E ")]
        if all(len(p) > 4 for p in partes):
            return ""
    loc = re.sub(r'/([A-Z]{2})$', r' - \1', loc)
    loc = re.sub(r'-([A-Z]{2})$', r' - \1', loc)
    loc_sem = _sem_acento(loc)
    for wrong, right in _TYPOS_LOCALIZ.items():
        if wrong in loc_sem:
            loc = loc_sem.replace(wrong, right)
            break
    return loc


def _normalizar_localizacoes(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    def _prep(x):
        s = str(x).strip().upper() if pd.notna(x) else ""
        return "" if s in ("", "-", "NAN", "0") else _limpar_localizacao(s)

    df["localizacao"] = df["localizacao"].apply(_prep)

    # Aplicar mapeamento fixo para cidades sem estado
    def _add_estado_fixo(loc: str) -> str:
        if not loc or " - " in loc:
            return loc
        return _CIDADE_ESTADO_FIXO.get(_sem_acento(loc), loc)

    df["localizacao"] = df["localizacao"].apply(_add_estado_fixo)

    # Mapa dinâmico cidade → "CIDADE - UF" com base nos dados existentes
    mapa: dict = {}
    for loc in df["localizacao"].dropna().unique():
        if not loc:
            continue
        partes = loc.rsplit(" - ", 1)
        if len(partes) == 2 and partes[1].strip() in _ESTADOS_BR:
            cidade = partes[0].strip()
            if cidade not in mapa or " - " not in mapa[cidade]:
                mapa[cidade] = loc

    def _canonico(loc: str) -> str:
        if not loc:
            return loc
        if " - " in loc:
            p = loc.rsplit(" - ", 1)
            if p[1].strip() in _ESTADOS_BR:
                return loc
        return mapa.get(loc, loc)

    df["localizacao"] = df["localizacao"].apply(_canonico)
    return df

# ======================================================================================
# SCANNER DE DEFORMAÇÃO PERMANENTE — DIRETÓRIOS SEPARADOS (2020-2021)
# ======================================================================================

_DP_BASE_DIR = r"G:\.shortcut-targets-by-id\1NUJ7pNAqedohSrLjiwFErFrtVqVZkrjk\006 - Lab. Central\1.4 OBSOLETOS\0.4 PROJETOS DEFORMAÇÃO PERMANENTE"

_DP_YEAR_DIRS = {
    2020: "003-PROJETOS 2020",
    2021: "004-PROJETOS 2021",
}

DP_CACHE_FILE = os.path.join(
    os.path.dirname(__file__), "..", "cache_certificados", "cauq_dp_cache.json"
)


def _load_dp_cache() -> dict:
    try:
        with open(DP_CACHE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _save_dp_cache(cache: dict):
    try:
        os.makedirs(os.path.dirname(DP_CACHE_FILE), exist_ok=True)
        with open(DP_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False)
    except Exception as e:
        logger.warning(f"Não foi possível salvar DP cache: {e}")


def _extrair_dp_separado(pasta_dp: str, ano: int) -> dict | None:
    """
    Extrai resultado de Deformação Permanente dos projetos em diretórios separados
    (2020-2021). Retorna dict com {num_projeto_cauq, deformacao_permanente, procedencia}.
    """
    import openpyxl as _opx

    # Encontrar Excel DPH na pasta
    xlsx_files = [
        f for f in os.listdir(pasta_dp)
        if f.lower().endswith(('.xlsx', '.xlsm')) and 'DPH' in f.upper()
        and not f.startswith('~$')
    ]
    if not xlsx_files:
        return None

    xlsx_path = os.path.join(pasta_dp, xlsx_files[0])

    try:
        wb = _opx.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception:
        return None

    resultado = None
    try:
        # Procura aba RESULTADO
        sh_name = next((s for s in wb.sheetnames if 'RESULTADO' in s.upper()), None)
        if not sh_name:
            wb.close()
            return None

        sh = wb[sh_name]

        num_projeto_cauq = None
        procedencia = None
        deformacao = None

        for row in sh.iter_rows(min_row=1, max_row=80, max_col=10, values_only=False):
            for cell in row:
                if cell.value is None:
                    continue
                val_str = str(cell.value)

                # Identificação da Amostra: PROJETO CAUQ XXXX.X.YYYY
                if 'PROJETO CAUQ' in val_str.upper():
                    # Extrai o número do projeto CAUQ da mesma célula ou adjacente
                    import re as _re
                    m = _re.search(r'(?:PROJETO\s+CAUQ\s+)(\d{4}\.\d\.\d{4})', val_str, _re.IGNORECASE)
                    if m:
                        num_projeto_cauq = m.group(1)
                    else:
                        # Tenta a próxima célula na mesma linha
                        col_idx = cell.column
                        try:
                            adj_cell = sh.cell(row=cell.row, column=col_idx + 1)
                            if adj_cell.value:
                                m2 = _re.search(r'(\d{4}\.\d\.\d{4})', str(adj_cell.value))
                                if m2:
                                    num_projeto_cauq = m2.group(1)
                        except Exception:
                            pass

                # Procedência
                if 'PROCED' in _sem_acento(val_str).upper():
                    try:
                        adj = sh.cell(row=cell.row, column=cell.column + 3)
                        if adj.value:
                            procedencia = str(adj.value).strip()
                    except Exception:
                        pass

                # Deformação da Amostra (mm) - geralmente em D68
                if 'DEFORMA' in _sem_acento(val_str).upper() and 'AMOSTRA' in val_str.upper():
                    # O valor está na linha seguinte, mesma coluna
                    try:
                        val_cell = sh.cell(row=cell.row + 1, column=cell.column)
                        if val_cell.value is not None:
                            v = float(val_cell.value)
                            if 0.01 < abs(v) < 50:
                                deformacao = abs(v)
                    except Exception:
                        pass

        if deformacao is not None:
            resultado = {
                "num_projeto_cauq": num_projeto_cauq,
                "procedencia": procedencia,
                "deformacao_permanente": deformacao,
                "ano": ano,
                "pasta_dp": os.path.basename(pasta_dp),
            }
    except Exception as e:
        logger.warning(f"Erro lendo DP separado {xlsx_path}: {e}")
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return resultado


def _escanear_dp_separados(anos: list[int] | None = None) -> list[dict]:
    """
    Escaneia os diretórios de Deformação Permanente separados (2020-2021).
    Retorna lista de dicts com resultado DP vinculado ao projeto CAUQ.
    """
    anos = anos or list(_DP_YEAR_DIRS.keys())
    cache = _load_dp_cache()
    resultados = []
    cache_novo = {}

    for ano in anos:
        dir_nome = _DP_YEAR_DIRS.get(ano)
        if not dir_nome:
            continue
        dir_path = os.path.join(_DP_BASE_DIR, dir_nome)
        if not os.path.isdir(dir_path):
            continue

        try:
            subpastas = sorted([
                d for d in os.listdir(dir_path)
                if os.path.isdir(os.path.join(dir_path, d)) and not d.startswith(".")
                and not d.startswith("00") == False  # mantém todas
            ])
        except Exception:
            continue

        for nome in subpastas:
            pasta_path = os.path.join(dir_path, nome)
            chave = f"dp::{ano}::{nome}"

            # Cache: verificar mtime
            try:
                xlsx_files = [
                    f for f in os.listdir(pasta_path)
                    if f.lower().endswith(('.xlsx', '.xlsm')) and 'DPH' in f.upper()
                    and not f.startswith('~$')
                ]
                mtime = os.path.getmtime(os.path.join(pasta_path, xlsx_files[0])) if xlsx_files else 0
            except Exception:
                mtime = 0

            entrada = cache.get(chave)
            if entrada and entrada.get("mtime") == mtime and mtime > 0:
                dados = entrada.get("dados")
                if dados:
                    resultados.append(dados)
                continue

            dados = _extrair_dp_separado(pasta_path, ano)
            if dados and mtime > 0:
                cache_novo[chave] = {"mtime": mtime, "dados": dados}
            if dados:
                resultados.append(dados)

    if cache_novo:
        cache.update(cache_novo)
        _save_dp_cache(cache)

    return resultados


def _vincular_dp_aos_projetos(df: pd.DataFrame, dp_list: list[dict]) -> pd.DataFrame:
    """
    Vincula dados de deformação permanente dos diretórios separados
    aos projetos CAUQ existentes, usando o num_projeto_cauq.
    """
    if not dp_list or df.empty:
        return df

    for dp in dp_list:
        num_cauq = dp.get("num_projeto_cauq")
        if not num_cauq:
            continue

        # Tenta match pelo num_projeto no DataFrame
        mask = df["num_projeto"].astype(str).str.contains(num_cauq, na=False, regex=False)
        if mask.any():
            idx = df[mask].index
            for i in idx:
                if pd.isna(df.at[i, "deformacao_permanente"]) or df.at[i, "deformacao_permanente"] is None:
                    df.at[i, "deformacao_permanente"] = dp["deformacao_permanente"]
        else:
            # Tenta match pela pasta (nome da pasta contém referência)
            pasta_dp = dp.get("pasta_dp", "")
            proc = dp.get("procedencia", "")
            # Busca por fragmentos do nome na pasta
            for idx, row in df.iterrows():
                if row["ano"] != dp["ano"]:
                    continue
                if pd.notna(row.get("deformacao_permanente")):
                    continue
                pasta_proj = str(row.get("pasta", ""))
                proc_proj = str(row.get("procedencia", ""))
                # Match por procedência
                if proc and proc_proj and _sem_acento(proc).upper() in _sem_acento(proc_proj).upper():
                    df.at[idx, "deformacao_permanente"] = dp["deformacao_permanente"]
                    break

    return df


# ======================================================================================
# CACHE DE RESULTADOS DO SCANNER
# ======================================================================================

SCAN_CACHE_FILE = os.path.join(
    os.path.dirname(__file__), "..", "cache_certificados", "cauq_scan_cache.json"
)
PARQUET_CACHE_FILE = os.path.join(
    os.path.dirname(__file__), "..", "cache_certificados", "cauq_projetos.parquet"
)


def _load_scan_cache() -> dict:
    """Carrega cache de resultados anteriores {pasta_key: {dados, mtime}}."""
    try:
        with open(SCAN_CACHE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _save_scan_cache(cache: dict):
    try:
        os.makedirs(os.path.dirname(SCAN_CACHE_FILE), exist_ok=True)
        with open(SCAN_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False)
    except Exception as e:
        logger.warning(f"Não foi possível salvar scan cache: {e}")


def _mtime_excel(pasta: str) -> float:
    """Retorna mtime do Excel do projeto, ou 0 se não encontrado."""
    try:
        xls = _encontrar_excel(pasta)
        return os.path.getmtime(xls) if xls else 0.0
    except Exception:
        return 0.0


# ======================================================================================
# SCANNER PRINCIPAL
# ======================================================================================

# ======================================================================================
# NORMALIZACAO DE FAIXAS GRANULOMETRICAS
# ======================================================================================

_FAIXA_MAPA_FIXO = {
    "FAIXA 3":   "FAIXA III DER-SP ET-DE-P00/027",
    "FAIXA EGL": "FAIXA EGL 16-19",
}


def _normalizar_faixa(faixa: str) -> str:
    if not faixa:
        return faixa
    f = faixa.strip()
    f_sem = _sem_acento(f).upper()
    for k, v in _FAIXA_MAPA_FIXO.items():
        if f_sem == _sem_acento(k).upper():
            return v
    f = re.sub(r'(DER[-/]?PR)\s*-?\s*ES[-/]?PA?\s*\d+[./]\d+',
               r'DER-PR', f, flags=re.IGNORECASE)
    f = re.sub(r'DEINFRA\s*[-/]\s*SC\s*-?\s*ES[-/]?PA?\s*\d+[./]\d+',
               'DEINFRA - SC', f, flags=re.IGNORECASE)
    f = re.sub(r'DEINFRA\s*/\s*SC', 'DEINFRA - SC', f, flags=re.IGNORECASE)
    f = re.sub(r'\bDEINFRA\b(?!\s*-\s*SC)', 'DEINFRA - SC', f, flags=re.IGNORECASE)
    f = re.sub(r'\bDER/PR\b', 'DER-PR', f, flags=re.IGNORECASE)
    f = re.sub(r'\bDER\b(?![-/])', 'DER-PR', f, flags=re.IGNORECASE)
    f = re.sub(r'\bDNIT\s+031/2006[-\s]+ES\b', 'DNIT', f, flags=re.IGNORECASE)
    f = re.sub(r'\bADAPATAD[AO]\b', 'ADAPTADA', f, flags=re.IGNORECASE)
    f = re.sub(r'\bADAPTADO\b', 'ADAPTADA', f, flags=re.IGNORECASE)
    f = re.sub(r"""["'‘’]REPERFILAGEM["'‘’]""", "ADAPTADA À REPERFILAGEM", f, flags=re.IGNORECASE)
    f = re.sub(r'\(REPERFILAGEM\)', 'ADAPTADA À REPERFILAGEM', f, flags=re.IGNORECASE)
    f = re.sub(r'\bPMC\s*[-–]\s*ES\b', 'PMC-ES', f, flags=re.IGNORECASE)
    f = re.sub(r'  +', ' ', f).strip()
    f = f.upper()
    f = re.sub(r'DEINFRA\s*-\s*SC', 'DEINFRA - SC', f)
    return f


def _normalizar_faixas(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["faixa_granulometrica"] = df["faixa_granulometrica"].apply(
        lambda x: _normalizar_faixa(str(x))
        if pd.notna(x) and str(x).strip() not in ("", "-", "0", "NAN")
        else ""
    )
    return df


def escanear_projetos(anos_filtro=None, com_geocode: bool = True,
                       com_geocode_api: bool = False,
                       usar_cache: bool = True,
                       max_workers: int = 4) -> pd.DataFrame:
    """
    Varre todos os diretórios anuais e retorna DataFrame com todos os projetos.

    com_geocode      – usa o cache local de coordenadas (rápido, sem rede)
    com_geocode_api  – chama a API Nominatim para localizações sem cache (lento)
    usar_cache       – reutiliza resultados de projetos não modificados (rápido)
    max_workers      – threads paralelas para leitura dos arquivos Excel

    No modo cloud (sem Google Drive), retorna dados do cache sem re-escanear.
    """
    anos  = anos_filtro or list(YEAR_DIRS.keys())

    # ── Modo cloud: retorna apenas dados do cache JSON ──────────────────────
    if _CLOUD_MODE:
        cache = _load_scan_cache()
        if not cache:
            return pd.DataFrame()
        todos = []
        for chave, entrada in cache.items():
            dados = entrada.get("dados")
            if not dados:
                continue
            ano_cache = dados.get("ano")
            if anos_filtro and ano_cache not in anos_filtro:
                continue
            todos.append(dados)
        if not todos:
            return pd.DataFrame()
        df = pd.DataFrame(todos)
        df = _normalizar_localizacoes(df)
        df = _normalizar_faixas(df)
        if com_geocode:
            geo_cache = _load_geocache()
            for idx, row in df.iterrows():
                loc = str(row.get("localizacao", "")).strip()
                if not loc or loc in ("-", "", "N/A"):
                    continue
                chave_loc = loc.upper()
                if chave_loc in geo_cache:
                    df.at[idx, "lat"] = geo_cache[chave_loc].get("lat")
                    df.at[idx, "lon"] = geo_cache[chave_loc].get("lon")
        return df
    cache = _load_scan_cache() if usar_cache else {}

    # Coletar todas as pastas de projeto
    tarefas: list[tuple[str, int]] = []   # (pasta_path, ano)
    for ano in anos:
        dir_nome = YEAR_DIRS.get(ano)
        if not dir_nome:
            continue
        dir_path = os.path.join(CAUQ_BASE_DIR, dir_nome)
        if not os.path.isdir(dir_path):
            continue
        try:
            subpastas = sorted([
                d for d in os.listdir(dir_path)
                if os.path.isdir(os.path.join(dir_path, d)) and not d.startswith(".")
            ])
        except Exception:
            continue
        for nome in subpastas:
            tarefas.append((os.path.join(dir_path, nome), ano))

    todos: list[dict] = []
    cache_novo: dict = {}

    def _processar(pasta: str, ano: int) -> dict | None:
        chave = f"{ano}::{os.path.basename(pasta)}"
        mtime = _mtime_excel(pasta)
        entrada_cache = cache.get(chave)
        if entrada_cache and entrada_cache.get("mtime") == mtime and mtime > 0:
            return entrada_cache.get("dados")
        dados = _extrair_dados_projeto(pasta, ano)
        if dados and mtime > 0:
            cache_novo[chave] = {"mtime": mtime, "dados": dados}
        return dados

    # Leitura paralela (I/O-bound — especialmente em Google Drive)
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = {ex.submit(_processar, p, a): (p, a) for p, a in tarefas}
        for fut in as_completed(futures):
            try:
                dados = fut.result()
                if dados:
                    todos.append(dados)
            except Exception as e:
                logger.warning(f"Erro em {futures[fut][0]}: {e}")

    # Persistir cache atualizado
    if usar_cache and cache_novo:
        cache.update(cache_novo)
        _save_scan_cache(cache)

    if not todos:
        return pd.DataFrame()

    df = pd.DataFrame(todos)

    # ── Validação: detectar deformação template (valores idênticos por ano) ────
    for ano in df["ano"].unique():
        mask_ano = (df["ano"] == ano) & df["deformacao_permanente"].notna()
        vals = df.loc[mask_ano, "deformacao_permanente"]
        if len(vals) > 3:
            # Se >80% dos valores são idênticos, é template
            moda = vals.mode()
            if not moda.empty:
                n_moda = (vals == moda.iloc[0]).sum()
                if n_moda / len(vals) > 0.8:
                    logger.info(
                        f"Ano {ano}: {n_moda}/{len(vals)} deformacoes identicas "
                        f"({moda.iloc[0]:.4f}) - marcando como template"
                    )
                    df.loc[mask_ano & (df["deformacao_permanente"] == moda.iloc[0]),
                           "deformacao_permanente"] = None

    # ── Vincular deformação permanente de diretórios separados (2020-2021) ───
    dp_anos = [a for a in anos if a in _DP_YEAR_DIRS]
    if dp_anos:
        try:
            dp_list = _escanear_dp_separados(dp_anos)
            if dp_list:
                df = _vincular_dp_aos_projetos(df, dp_list)
                logger.info(f"DP separados: {len(dp_list)} resultados vinculados")
        except Exception as e:
            logger.warning(f"Erro ao vincular DP separados: {e}")

    df = _normalizar_localizacoes(df)
    df = _normalizar_faixas(df)

    if com_geocode:
        cache = _load_geocache()
        for idx, row in df.iterrows():
            loc = str(row.get("localizacao", "")).strip()
            if not loc or loc in ("-", "", "N/A"):
                continue
            chave = loc.upper()
            if chave in cache:
                df.at[idx, "lat"] = cache[chave].get("lat")
                df.at[idx, "lon"] = cache[chave].get("lon")
            elif com_geocode_api:
                # Só chama a API quando explicitamente solicitado
                lat, lon = geocodificar(loc)
                df.at[idx, "lat"] = lat
                df.at[idx, "lon"] = lon
            # Se não tem cache nem API autorizada, lat/lon ficam None

    # ── Salvar cache Parquet para carregamento rapido ──
    try:
        os.makedirs(os.path.dirname(PARQUET_CACHE_FILE), exist_ok=True)
        df.to_parquet(PARQUET_CACHE_FILE, index=False)
        logger.info(f"Parquet cache salvo: {len(df)} projetos")
    except Exception as e:
        logger.warning(f"Nao salvou parquet: {e}")

    return df


def carregar_parquet_cache() -> pd.DataFrame | None:
    """Carrega o cache Parquet pre-computado (instantaneo)."""
    try:
        if os.path.exists(PARQUET_CACHE_FILE):
            return pd.read_parquet(PARQUET_CACHE_FILE)
    except Exception:
        pass
    return None


def geocodificar_pendentes(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """
    Geocodifica via API apenas as linhas sem coordenadas.
    Retorna (df_atualizado, n_novos_geocodificados).
    """
    sem_coords = df[df["lat"].isna() & df["localizacao"].notna()].copy()
    sem_coords = sem_coords[~sem_coords["localizacao"].isin(["", "-", "N/A", "NAN"])]

    n_novos = 0
    for idx, row in sem_coords.iterrows():
        loc = str(row["localizacao"]).strip()
        if not loc:
            continue
        lat, lon = geocodificar(loc)
        df.at[idx, "lat"] = lat
        df.at[idx, "lon"] = lon
        if lat is not None:
            n_novos += 1

    return df, n_novos


def anos_disponiveis() -> list:
    """Retorna lista de anos com diretório existente e não vazio."""
    # Modo cloud: extrai anos do cache
    if _CLOUD_MODE:
        cache = _load_scan_cache()
        anos_set = set()
        for chave, entrada in cache.items():
            dados = entrada.get("dados", {})
            ano = dados.get("ano")
            if ano:
                anos_set.add(int(ano))
        return sorted(anos_set) if anos_set else list(YEAR_DIRS.keys())

    anos = []
    for ano, dir_nome in YEAR_DIRS.items():
        dir_path = os.path.join(CAUQ_BASE_DIR, dir_nome)
        if os.path.isdir(dir_path):
            try:
                if any(os.path.isdir(os.path.join(dir_path, d))
                       for d in os.listdir(dir_path)):
                    anos.append(ano)
            except Exception:
                pass
    return sorted(anos)
