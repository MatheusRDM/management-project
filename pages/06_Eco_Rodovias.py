"""
=============================================================================
ECO RODOVIAS — Gestão de Contrato 6771
=============================================================================
BR-050 (Eco Minas Goiás) + BR-365 (Eco Cerrado)
Checklist APP + Ensaios AEVIAS
=============================================================================
"""
import streamlit as st
import sys
import os
import json
import re as _re
import threading
import requests
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import folium
from streamlit_folium import st_folium
from datetime import datetime, date

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from styles import aplicar_estilos
from page_auth import proteger_pagina

# =============================================================================
st.set_page_config(
    page_title="Eco Rodovias | Afirma E-vias",
    page_icon="Imagens/logo_icon.png",
    layout="wide",
    initial_sidebar_state="expanded"
)
aplicar_estilos()
proteger_pagina("Eco Rodovias")
# =============================================================================

# ── Paleta ────────────────────────────────────────────────────────────────────
COR_PRIMARY  = "#566E3D"
COR_ACCENT   = "#BFCF99"
COR_BG       = "#0D1B2A"
COR_CARD     = "rgba(26, 31, 46, 0.85)"
COR_BORDER   = "rgba(86,110,61,0.35)"
COR_TEXT     = "#E8EFD8"
COR_MUTED    = "#8FA882"
COR_OK       = "#3cb44b"
COR_COBRAR   = "#e6194b"
COR_NE       = "#3a4a5e"
COR_ELAB     = "#4363d8"

PLOTLY_LAYOUT = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font=dict(family="Poppins, sans-serif", color=COR_TEXT, size=12),
    margin=dict(l=10, r=10, t=35, b=10),
    hoverlabel=dict(bgcolor=COR_BG, bordercolor=COR_PRIMARY,
                    font=dict(color=COR_TEXT, size=12, family="Poppins")),
    hovermode="closest",
    dragmode=False,
)
PLOTLY_CONFIG = {"displayModeBar": False, "scrollZoom": False}

# ── CSS ──────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;500;600;700&display=swap');

.eco-header {
    padding: 18px 0 8px 0;
    border-bottom: 2px solid rgba(86,110,61,0.4);
    margin-bottom: 24px;
}
.eco-header h1 {
    font-family: 'Poppins', sans-serif;
    font-size: 1.55rem;
    font-weight: 700;
    color: #BFCF99;
    margin: 0;
}
.eco-header p {
    font-family: 'Poppins', sans-serif;
    font-size: 0.82rem;
    color: #8FA882;
    margin: 4px 0 0 0;
}
.eco-kpi {
    background: rgba(26,31,46,0.85);
    border: 1px solid rgba(86,110,61,0.35);
    border-radius: 10px;
    padding: 14px 18px;
    text-align: center;
    margin-bottom: 10px;
}
.eco-kpi .val {
    font-family: 'Poppins', sans-serif;
    font-size: 2rem;
    font-weight: 700;
    line-height: 1;
}
.eco-kpi .lbl {
    font-family: 'Poppins', sans-serif;
    font-size: 0.72rem;
    color: #8FA882;
    margin-top: 4px;
    text-transform: uppercase;
    letter-spacing: 0.06em;
}
.cal-wrap { overflow-x: auto; width: 100%; }
.cal-table {
    border-collapse: collapse;
    font-family: 'Poppins', sans-serif;
    font-size: 0.68rem;
    width: 100%;
    min-width: 900px;
}
.cal-table th {
    background: rgba(86,110,61,0.25);
    color: #BFCF99;
    padding: 5px 3px;
    text-align: center;
    font-weight: 600;
    border: 1px solid rgba(86,110,61,0.2);
    white-space: nowrap;
    font-size: 0.62rem;
}
.cal-table td {
    padding: 5px 4px;
    border: 1px solid rgba(255,255,255,0.05);
    text-align: center;
    white-space: nowrap;
}
.cal-table td.colab {
    text-align: left;
    font-weight: 500;
    color: #E8EFD8;
    padding-left: 8px;
    min-width: 160px;
    max-width: 200px;
    overflow: hidden;
    text-overflow: ellipsis;
}
.cal-table td.funcao {
    text-align: left;
    color: #8FA882;
    font-size: 0.60rem;
    min-width: 120px;
    max-width: 160px;
    overflow: hidden;
    text-overflow: ellipsis;
}
.status-ok    { background: rgba(60,180,75,0.25);  color: #3cb44b; font-weight:600; border-radius:3px; }
.status-cobrar{ background: rgba(230,25,75,0.25);  color: #ff5577; font-weight:600; border-radius:3px; }
.status-ne    { background: rgba(58,74,94,0.4);    color: #7a90a8; }
.status-elab  { background: rgba(67,99,216,0.25);  color: #6ec6ff; font-weight:600; border-radius:3px; }
.status-vazio { background: transparent; color: #3a4a5e; }
.legend-item { display:inline-flex; align-items:center; gap:6px; margin-right:14px; font-size:0.75rem; font-family:'Poppins',sans-serif; color:#E8EFD8; }
.legend-dot { width:12px; height:12px; border-radius:3px; display:inline-block; }
</style>
""", unsafe_allow_html=True)

# =============================================================================
# CAMINHOS
# =============================================================================
_BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_CACHE_DIR  = os.path.join(_BASE_DIR, "cache_certificados")
_Y_BASE     = "Y:/24-017 ECO 050 e CERRADO - Supervisão de Obras/04. Medição AFIRMA"
_IS_CLOUD   = not os.path.exists(_Y_BASE)

# =============================================================================
# CARREGAMENTO DE DADOS
# =============================================================================

@st.cache_data(ttl=3600, show_spinner=False)
def _carregar_checklist_cache() -> dict:
    p = os.path.join(_CACHE_DIR, "eco_checklist.json")
    if os.path.exists(p):
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    return {}


@st.cache_data(ttl=3600, show_spinner=False)
def _carregar_checklist_y(med_folder: str) -> dict | None:
    """Lê o xlsx de controle diretamente do Y:."""
    import openpyxl
    from glob import glob
    folder = os.path.join(_Y_BASE, med_folder)
    if not os.path.exists(folder):
        return None
    # Busca arquivo xlsx com "controle" e "app" no nome
    hits = [f for f in os.listdir(folder)
            if f.lower().endswith(".xlsx")
            and "controle" in f.lower()
            and not f.startswith("~")]
    if not hits:
        return None
    path = os.path.join(folder, hits[0])
    wb = openpyxl.load_workbook(path)
    result = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue
        header = rows[0]
        dates = []
        for v in header[2:]:
            if isinstance(v, datetime):
                dates.append(v.strftime("%Y-%m-%d"))
            else:
                dates.append(None)
        people = []
        for row in rows[2:]:
            if not row[0]:
                continue
            entry = {
                "colaborador": str(row[0]),
                "funcao": str(row[1]) if row[1] else "",
                "dias": {},
            }
            for i, d in enumerate(dates):
                if d and (i + 2) < len(row):
                    v = row[i + 2]
                    entry["dias"][d] = str(v) if v else None
            people.append(entry)
        result[sn] = people
    return result


@st.cache_data(ttl=3600, show_spinner=False)
def _listar_meds() -> list[str]:
    cache = _carregar_checklist_cache()
    local_meds = list(cache.keys())
    if not _IS_CLOUD and os.path.exists(_Y_BASE):
        y_meds = [d for d in sorted(os.listdir(_Y_BASE)) if d.startswith("MED")]
        # Merge, mantendo cache + novos do Y
        all_meds = sorted(set(local_meds + y_meds))
        return all_meds
    return local_meds


@st.cache_data(ttl=3600, show_spinner=False)
def _carregar_ensaios() -> list[dict]:
    # 1. Tenta Y / desktop local
    local = os.path.join(
        os.path.expanduser("~"),
        "OneDrive", "Área de Trabalho", "Ensaios AEVIAS", "ensaios_dados.json"
    )
    for p in [local, os.path.join(_CACHE_DIR, "eco_ensaios.json")]:
        if os.path.exists(p):
            with open(p, encoding="utf-8") as f:
                return json.load(f)
    return []


# =============================================================================
# COMPONENTES VISUAIS
# =============================================================================

def _kpi_card(val, label, cor="#BFCF99"):
    return f"""
    <div class="eco-kpi">
        <div class="val" style="color:{cor}">{val}</div>
        <div class="lbl">{label}</div>
    </div>"""


def _status_class(v):
    if v is None or v == "":
        return "status-vazio", "·"
    vu = str(v).upper().strip()
    if vu == "OK":
        return "status-ok", "OK"
    if vu in ("COBRAR", "COBRE"):
        return "status-cobrar", "COB"
    if vu in ("N/E", "NE"):
        return "status-ne", "N/E"
    if vu in ("ELAB.", "ELAB"):
        return "status-elab", "ELB"
    return "status-vazio", v[:3] if v else "·"


def _renderizar_calendario(people: list[dict], mes_ref: str):
    """Renderiza tabela HTML de calendário."""
    if not people:
        st.warning("Nenhum colaborador encontrado.")
        return

    # Coleta todas as datas disponíveis
    datas = set()
    for p in people:
        datas.update(p.get("dias", {}).keys())
    datas = sorted(d for d in datas if d)

    if not datas:
        st.info("Sem datas registradas.")
        return

    # Monta cabeçalho de datas
    DAY_ABBR = {0: "SEG", 1: "TER", 2: "QUA", 3: "QUI", 4: "SEX", 5: "SÁB", 6: "DOM"}
    html = ['<div class="cal-wrap"><table class="cal-table">']
    # Linha 1: números dos dias
    html.append("<thead><tr>")
    html.append('<th>Colaborador</th><th>Função</th>')
    for d in datas:
        dt = datetime.strptime(d, "%Y-%m-%d")
        html.append(f'<th>{dt.day:02d}</th>')
    html.append('<th>OK</th><th>COB</th></tr>')
    # Linha 2: siglas dos dias da semana
    html.append("<tr>")
    html.append('<th></th><th></th>')
    for d in datas:
        dt = datetime.strptime(d, "%Y-%m-%d")
        html.append(f'<th style="font-size:0.55rem;color:#8FA882">{DAY_ABBR[dt.weekday()]}</th>')
    html.append('<th></th><th></th></tr></thead>')

    # Linhas por colaborador
    html.append("<tbody>")
    for p in people:
        dias = p.get("dias", {})
        ok_count  = sum(1 for v in dias.values() if v and str(v).upper().strip() == "OK")
        cob_count = sum(1 for v in dias.values() if v and str(v).upper().strip() in ("COBRAR", "COBRE"))
        html.append("<tr>")
        html.append(f'<td class="colab" title="{p["colaborador"]}">{p["colaborador"]}</td>')
        html.append(f'<td class="funcao" title="{p["funcao"]}">{p["funcao"]}</td>')
        for d in datas:
            v = dias.get(d)
            cls, txt = _status_class(v)
            html.append(f'<td class="{cls}">{txt}</td>')
        cor_ok  = "#3cb44b" if ok_count  > 0 else "#7a90a8"
        cor_cob = "#ff5577" if cob_count > 0 else "#7a90a8"
        html.append(f'<td style="color:{cor_ok};font-weight:700">{ok_count}</td>')
        html.append(f'<td style="color:{cor_cob};font-weight:700">{cob_count if cob_count else "—"}</td>')
        html.append("</tr>")
    html.append("</tbody></table></div>")

    # Legenda
    html.append("""
    <div style="margin-top:10px">
        <span class="legend-item"><span class="legend-dot" style="background:rgba(60,180,75,0.35)"></span>OK — Checklist enviado</span>
        <span class="legend-item"><span class="legend-dot" style="background:rgba(230,25,75,0.35)"></span>COBRAR — Pendente</span>
        <span class="legend-item"><span class="legend-dot" style="background:rgba(58,74,94,0.6)"></span>N/E — Não estava em campo</span>
        <span class="legend-item"><span class="legend-dot" style="background:rgba(67,99,216,0.35)"></span>ELAB. — Em elaboração</span>
    </div>""")

    st.markdown("".join(html), unsafe_allow_html=True)


# =============================================================================
# ABAS DE CONTEÚDO
# =============================================================================

def _aba_checklist():
    meds = _listar_meds()
    if not meds:
        st.info("Nenhuma medição encontrada. Verifique o acesso ao servidor Y:")
        return

    med_labels = {m: m for m in reversed(meds)}  # mais recente primeiro
    meds_reversed = list(reversed(meds))

    col_sel, col_info = st.columns([2, 4])
    with col_sel:
        med_escolhida = st.selectbox(
            "📅 Medição:",
            options=meds_reversed,
            format_func=lambda x: x,
            key="eco_med_sel",
        )

    # Carrega dados da medição selecionada
    with st.spinner("Carregando checklist..."):
        sheets = None
        if not _IS_CLOUD:
            sheets = _carregar_checklist_y(med_escolhida)
        if sheets is None:
            cache = _carregar_checklist_cache()
            entry = cache.get(med_escolhida, {})
            sheets = entry.get("sheets", {})

    if not sheets:
        st.warning(f"Nenhum arquivo de controle encontrado para **{med_escolhida}**.")
        return

    # KPIs globais
    total_ok = total_cob = total_ne = 0
    for plist in sheets.values():
        for p in plist:
            for v in p.get("dias", {}).values():
                vu = str(v).upper().strip() if v else ""
                if vu == "OK":
                    total_ok += 1
                elif vu in ("COBRAR", "COBRE"):
                    total_cob += 1
                elif vu in ("N/E", "NE"):
                    total_ne += 1

    c1, c2, c3, c4 = st.columns(4)
    c1.markdown(_kpi_card(total_ok,  "Checklists OK",     COR_OK),    unsafe_allow_html=True)
    c2.markdown(_kpi_card(total_cob, "A Cobrar",          COR_COBRAR), unsafe_allow_html=True)
    c3.markdown(_kpi_card(total_ne,  "Não em Campo (N/E)", COR_MUTED), unsafe_allow_html=True)
    pct = f"{100*total_ok/(total_ok+total_cob):.0f}%" if (total_ok+total_cob) > 0 else "—"
    c4.markdown(_kpi_card(pct, "Taxa de Conformidade", COR_ACCENT), unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Sub-abas por contrato
    sheet_names = list(sheets.keys())
    tab_labels = [f"🛣️ {s}" for s in sheet_names]
    tabs = st.tabs(tab_labels)

    for tab, sn in zip(tabs, sheet_names):
        with tab:
            people = sheets[sn]
            if not people:
                st.info("Nenhum colaborador.")
                continue

            # KPIs por contrato
            ok_c = sum(
                1 for p in people for v in p.get("dias", {}).values()
                if v and str(v).upper().strip() == "OK"
            )
            cob_c = sum(
                1 for p in people for v in p.get("dias", {}).values()
                if v and str(v).upper().strip() in ("COBRAR", "COBRE")
            )
            colab_cob = [
                p["colaborador"].split()[0]
                for p in people
                if any(str(v).upper().strip() in ("COBRAR", "COBRE")
                       for v in p.get("dias", {}).values() if v)
            ]

            if cob_c > 0:
                st.error(
                    f"⚠️ **{cob_c} checklist(s) pendente(s)** — Colaboradores: "
                    + ", ".join(colab_cob[:6])
                    + ("..." if len(colab_cob) > 6 else "")
                )
            else:
                st.success(f"✅ Todos os checklists enviados neste contrato — {ok_c} registros OK")

            _renderizar_calendario(people, med_escolhida)


def _aba_ensaios():
    data = _carregar_ensaios()
    if not data:
        st.info("Nenhum dado de ensaios encontrado. Execute a automação de download.")
        return

    df = pd.DataFrame(data)
    # Normalizar encoding
    df["obra"]         = df["obra"].str.strip()
    df["tipo"]         = df["tipo"].str.strip()
    df["profissional"] = df["profissional"].str.strip()
    df["data_dt"]      = pd.to_datetime(df["data"], format="%d/%m/%Y", errors="coerce")

    # Ordem das obras
    OBRAS_ORDEM = ["SST", "Pavimento", "TOPOGRAFIA", "OAE / Terraplenos",
                   "Ampliações", "ESCRITÓRIO", "Conserva"]
    OBRAS_CORES = {
        "SST":              "#e6194b",
        "Pavimento":        "#3cb44b",
        "TOPOGRAFIA":       "#ffe119",
        "OAE / Terraplenos":"#4363d8",
        "Ampliações":       "#f58231",
        "ESCRITÓRIO":       "#911eb4",
        "Conserva":         "#42d4f4",
    }

    # KPIs por categoria
    st.markdown("#### Ensaios por Categoria")
    cols = st.columns(len(OBRAS_ORDEM))
    for col, obra in zip(cols, OBRAS_ORDEM):
        cnt = len(df[df["obra"] == obra])
        cor = OBRAS_CORES.get(obra, COR_ACCENT)
        col.markdown(
            f'<div class="eco-kpi"><div class="val" style="color:{cor}">{cnt}</div>'
            f'<div class="lbl">{obra}</div></div>',
            unsafe_allow_html=True,
        )

    st.markdown("<br>", unsafe_allow_html=True)

    # Gráficos
    col_a, col_b = st.columns(2)

    with col_a:
        st.markdown("#### Distribuição por Categoria")
        obras_count = df.groupby("obra").size().reset_index(name="qtd")
        obras_count = obras_count.sort_values("qtd", ascending=False)
        cores_list = [OBRAS_CORES.get(o, COR_MUTED) for o in obras_count["obra"]]
        fig = go.Figure(go.Pie(
            labels=obras_count["obra"],
            values=obras_count["qtd"],
            marker_colors=cores_list,
            hole=0.52,
            textinfo="label+value",
            textfont_size=12,
        ))
        fig.update_layout(**PLOTLY_LAYOUT, showlegend=False, height=320)
        st.plotly_chart(fig, use_container_width=True, config=PLOTLY_CONFIG)

    with col_b:
        st.markdown("#### Distribuição por Tipo de Documento")
        tipos_count = df.groupby("tipo").size().reset_index(name="qtd")
        tipos_count = tipos_count.sort_values("qtd", ascending=True)
        fig2 = go.Figure(go.Bar(
            x=tipos_count["qtd"],
            y=tipos_count["tipo"],
            orientation="h",
            marker_color=COR_PRIMARY,
            text=tipos_count["qtd"],
            textposition="outside",
        ))
        fig2.update_layout(**PLOTLY_LAYOUT, height=320,
                           xaxis=dict(showgrid=False, visible=False),
                           yaxis=dict(showgrid=False, tickfont=dict(size=11)))
        st.plotly_chart(fig2, use_container_width=True, config=PLOTLY_CONFIG)

    # Timeline
    if df["data_dt"].notna().any():
        st.markdown("#### Timeline de Ensaios")
        df_time = df.dropna(subset=["data_dt"]).copy()
        df_time["data_str"] = df_time["data_dt"].dt.strftime("%Y-%m-%d")
        timeline = df_time.groupby(["data_str", "obra"]).size().reset_index(name="qtd")
        fig3 = px.bar(
            timeline, x="data_str", y="qtd", color="obra",
            color_discrete_map=OBRAS_CORES,
            labels={"data_str": "Data", "qtd": "Qtd", "obra": "Categoria"},
        )
        fig3.update_layout(**PLOTLY_LAYOUT, height=280,
                           bargap=0.1, showlegend=True,
                           legend=dict(orientation="h", y=-0.18, x=0))
        st.plotly_chart(fig3, use_container_width=True, config=PLOTLY_CONFIG)

    # Filtros e tabela
    st.markdown("#### Tabela de Ensaios")
    f_col1, f_col2, f_col3 = st.columns(3)
    with f_col1:
        f_obra = st.multiselect("Categoria:", sorted(df["obra"].unique()), key="eco_f_obra")
    with f_col2:
        f_tipo = st.multiselect("Tipo:", sorted(df["tipo"].unique()), key="eco_f_tipo")
    with f_col3:
        f_prof = st.multiselect("Profissional/Projeto:", sorted(df["profissional"].unique()), key="eco_f_prof")

    df_view = df.copy()
    if f_obra: df_view = df_view[df_view["obra"].isin(f_obra)]
    if f_tipo: df_view = df_view[df_view["tipo"].isin(f_tipo)]
    if f_prof: df_view = df_view[df_view["profissional"].isin(f_prof)]

    _sort_col = "data_dt" if "data_dt" in df_view.columns else "data"
    df_display = df_view.sort_values(_sort_col, ascending=False)[["data", "obra", "tipo", "profissional"]]
    df_display.columns = ["Data", "Categoria", "Tipo", "Profissional/Projeto"]
    st.dataframe(df_display, use_container_width=True, hide_index=True, height=300)
    st.caption(f"{len(df_display)} registro(s) exibido(s) de {len(df)} total")


# =============================================================================
# SIDEBAR
# =============================================================================

def _sidebar():
    with st.sidebar:
        st.markdown("""
        <style>
        section[data-testid="stSidebar"] { background: #0D1B2A !important; }
        .eco-sidebar-title {
            font-family:'Poppins',sans-serif; font-size:0.78rem;
            color:#8FA882; text-transform:uppercase; letter-spacing:.06em;
            margin: 8px 0 4px 0;
        }
        div[data-testid="stButton"] button {
            background: rgba(86,110,61,0.15) !important;
            border: 1px solid rgba(86,110,61,0.4) !important;
            color: #BFCF99 !important;
            font-family:'Poppins',sans-serif !important;
            font-size:0.78rem !important;
            padding:0.2rem 0.6rem !important;
            border-radius:6px !important;
            margin-bottom:0.5rem !important;
        }
        </style>""", unsafe_allow_html=True)

        if st.button("< Menu Principal", key="back_menu_eco"):
            st.switch_page("app.py")

        try:
            st.image("Imagens/AE - Logo Hor Principal_2.png", use_container_width=True)
        except Exception:
            st.markdown('<h3 style="color:white;text-align:center">AFIRMA E-VIAS</h3>', unsafe_allow_html=True)

        st.markdown("---")
        st.markdown('<div class="eco-sidebar-title">Contrato</div>', unsafe_allow_html=True)
        st.markdown("""
        <div style="font-family:'Poppins',sans-serif; font-size:0.80rem; color:#E8EFD8; line-height:1.6">
            <b style="color:#BFCF99">ECO RODOVIAS 6771</b><br>
            🛣️ BR-050 — Eco Minas Goiás<br>
            🛣️ BR-365 — Eco Cerrado<br>
            <span style="color:#8FA882; font-size:0.72rem">Supervisão de Obras</span>
        </div>""", unsafe_allow_html=True)

        st.markdown("---")
        st.markdown('<div class="eco-sidebar-title">Acesso ao Servidor</div>', unsafe_allow_html=True)
        if _IS_CLOUD:
            st.warning("🌐 Modo Cloud — dados do cache", icon=None)
        else:
            st.success("✅ Servidor Y: conectado", icon=None)


# =============================================================================
# MAIN
# =============================================================================

# =============================================================================
# LOGOS RASTREAMENTO — API (síncrono, via @st.fragment)
# =============================================================================
_LOGOS_BASE = "https://rastrear.logosrastreamento.com.br"
_CORES_VEICULOS = [
    "#FF6B35","#4CC9F0","#F7B731","#7BED9F","#FF4757",
    "#A29BFE","#FD79A8","#00CEC9","#FDCB6E","#6C5CE7",
]


def _logos_login():
    """Autentica no Logos e retorna (sess, idcli)."""
    try:
        usuario = st.secrets["logos_usuario"]
        senha   = st.secrets["logos_senha"]
    except Exception:
        usuario = "matheus.resende@afirmaevias.com.br"
        senha   = "19072019Joaquim*"

    sess = requests.Session()
    sess.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0 Safari/537.36",
    })
    r = sess.get(f"{_LOGOS_BASE}/Identity/Account/Login", timeout=15)
    m = _re.search(r'name="__RequestVerificationToken"[^>]+value="([^"]+)"', r.text)
    token = m.group(1) if m else ""
    sess.post(f"{_LOGOS_BASE}/Identity/Account/Login", data={
        "Input.UserName": usuario,
        "Input.Password": senha,
        "__RequestVerificationToken": token,
    }, timeout=15, allow_redirects=True)
    idcli = next((c.value for c in sess.cookies if c.name == "IDCLI"), None)
    if not idcli:
        raise ValueError("Login falhou — credenciais inválidas ou sessão expirada.")
    return sess, idcli


def _logos_get_eco(sess, idcli):
    """Retorna lista de veículos com ECO no nome (última posição)."""
    r = sess.post(f"{_LOGOS_BASE}/api/ultimaposicao", json={
        "idcliente": int(idcli), "texto": "", "placa": "", "serial": "",
        "descricao": "", "grupoveiculo": "", "idsVeiculos": [],
    }, timeout=20)
    items = r.json() if isinstance(r.json(), list) else r.json().get("data", [])
    return [v for v in items if "ECO" in str(v.get("descricaovel", "")).upper()]


def _logos_get_rota(sess, idveiculo, d_ini, d_fim):
    """Retorna histórico de posições. d_ini/d_fim: 'YYYY-MM-DD HH:MM'"""
    r = sess.post(f"{_LOGOS_BASE}/api/historicoposicao", json={
        "idveiculo": idveiculo, "datainicio": d_ini, "datafinal": d_fim,
    }, timeout=60)
    d = r.json()
    return d if isinstance(d, list) else d.get("data", [])


# Nomes possíveis de campos no historicoposicao (o Logos usa variações)
_HIST_ODO_FIELDS = [
    "pos_odometro", "odometro", "hodometro", "pos_hodometro",
    "km", "quilometragem", "pos_km", "distancia",
]
_HIST_LAT_FIELDS = [
    "pos_coordenada_latitude", "latitude", "lat",
    "coordenada_latitude", "pos_latitude",
]
_HIST_LON_FIELDS = [
    "pos_coordenada_longitude", "longitude", "lon", "lng",
    "coordenada_longitude", "pos_longitude",
]
_HIST_VEL_FIELDS = [
    "pos_velocidade", "velocidade", "speed", "vel",
]
_HIST_IGN_FIELDS = [
    "pos_ignicao", "ignicao", "ignition", "motor",
]
_HIST_DT_FIELDS  = [
    "pos_dt_posicao", "dt_posicao", "data_hora", "dataHora",
    "data", "datetime", "timestamp",
]
_HIST_CID_FIELDS = [
    "pos_end_cidade", "cidade", "municipio", "city",
]
_HIST_UF_FIELDS  = [
    "pos_end_uf", "uf", "estado", "state",
]


def _pick(d, fields, default=None):
    """Retorna o primeiro campo encontrado em d dentre a lista fields."""
    for f in fields:
        v = d.get(f)
        if v is not None:
            return v
    return default


def _km_from_hist(hist):
    """
    Calcula km percorridos no histórico.
    1) Tenta odômetro (max-min): detecta campo automaticamente.
    2) Fallback Haversine sobre coordenadas GPS.
    """
    import math

    # Tenta odômetro com detecção automática de campo
    odos = []
    for p in hist:
        v = _pick(p, _HIST_ODO_FIELDS)
        try:
            iv = int(float(v or 0))
            if iv > 0:
                odos.append(iv)
        except Exception:
            pass
    if odos:
        return max(odos) - min(odos)

    # Fallback: Haversine
    def _hav(lat1, lon1, lat2, lon2):
        R = 6371.0
        dlat = math.radians(lat2 - lat1)
        dlon = math.radians(lon2 - lon1)
        a = (math.sin(dlat/2)**2
             + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2))
             * math.sin(dlon/2)**2)
        return R * 2 * math.asin(math.sqrt(a))

    coords = []
    for p in hist:
        try:
            lt = float(_pick(p, _HIST_LAT_FIELDS) or 0)
            ln = float(_pick(p, _HIST_LON_FIELDS) or 0)
            if lt and ln:
                coords.append((lt, ln))
        except Exception:
            pass

    if len(coords) < 2:
        return 0
    total = sum(_hav(coords[i][0], coords[i][1], coords[i+1][0], coords[i+1][1])
                for i in range(len(coords) - 1))
    return round(total)


def _normalizar_contrato(s: str) -> str:
    """
    Canonicaliza nomes de contrato com variações de separador.
    ECO 050/CERRADO | ECO-050/CERRADO | ECO050/CERRADO → ECO-050/CERRADO
    ECO 135 | ECO-135 | ECO135 → ECO-135
    """
    s = s.strip().upper()
    # ECO[espaço|-|nada]NNN → ECO-NNN
    s = _re.sub(r"ECO[\s\-]?(\d+)", r"ECO-\1", s)
    # remove espaços ao redor da barra
    s = _re.sub(r"\s*/\s*", "/", s)
    return s


def _parse_eco(v, i):
    """Normaliza um veículo ECO para dict de analytics."""
    desc  = v.get("descricaovel", f"V{i}")
    parts = desc.split(" - ", 1)
    contrato  = _normalizar_contrato(parts[0].strip() if len(parts) > 1 else desc)
    motorista = parts[1].strip() if len(parts) > 1 else desc
    dt = str(v.get("pos_dt_posicao", ""))[:10]
    try:
        from datetime import date as _d
        sugest = _d.fromisoformat(dt) if dt else date.today()
    except Exception:
        sugest = date.today()
    return {
        "contrato":       contrato,
        "motorista":      motorista,
        "desc":           desc,
        "placa":          v.get("placavel", "—"),
        "odometro":       int(v.get("pos_odometro") or 0),
        "velocidade":     v.get("pos_velocidade", 0),
        "ignicao":        bool(v.get("pos_ignicao")),
        "uf":             v.get("pos_end_uf", "—"),
        "cidade":         v.get("pos_end_cidade", "—"),
        "localizacao":    v.get("localizacao", "—"),
        "dt_posicao":     str(v.get("pos_dt_posicao", ""))[:16].replace("T", " "),
        "ultima_data":    sugest,
        "tempo_dir_h":    round((v.get("pos_tempo_dirigindo") or 0) / 3600, 1),
        "tempo_par_min":  round((v.get("pos_tempo_parado") or 0) / 60, 1),
        "horimetro":      v.get("pos_horimetro", 0),
        "bateria":        v.get("pos_bateria_externa", 0),
        "lat":            v.get("pos_coordenada_latitude"),
        "lon":            v.get("pos_coordenada_longitude"),
        "idvei":          v.get("pos_idvei"),
        "cor":            _CORES_VEICULOS[i % len(_CORES_VEICULOS)],
    }


def _render_mapa_posicao(itens):
    mapa   = folium.Map(location=[-18.5, -47.5], zoom_start=6, tiles="CartoDB dark_matter")
    bounds = []
    for it in itens:
        if it["lat"] and it["lon"]:
            try:
                lt, ln = float(it["lat"]), float(it["lon"])
                ign = "🟢" if it["ignicao"] else "🔴"
                popup_html = (
                    f"<b style='color:{it['cor']}'>{it['desc']}</b><br>"
                    f"Placa: {it['placa']}<br>Ignição: {ign}<br>"
                    f"Velocidade: {it['velocidade']} km/h<br>"
                    f"Hodômetro: {it['odometro']:,} km<br>{it['localizacao']}"
                )
                folium.CircleMarker(
                    [lt, ln], radius=8, color=it["cor"], fill=True,
                    fill_color=it["cor"], fill_opacity=0.9,
                    tooltip=f"{ign} {it['motorista']} — {it['velocidade']} km/h",
                    popup=folium.Popup(popup_html, max_width=280),
                ).add_to(mapa)
                bounds.append([lt, ln])
            except Exception:
                pass
    if bounds:
        lats = [c[0] for c in bounds]; lons = [c[1] for c in bounds]
        mapa.fit_bounds([[min(lats), min(lons)], [max(lats), max(lons)]])
    st_folium(mapa, width="100%", height=450, key="logos_mapa_pos", returned_objects=[])


def _render_estatisticas(itens):
    import plotly.express as px
    import plotly.graph_objects as go

    # Paleta da aplicação
    _C = {
        "bg":      "rgba(0,0,0,0)",
        "grid":    "rgba(255,255,255,0.06)",
        "text":    "#C8D8A8",
        "eco135":  "#7BBF6A",
        "cerrado": "#4CC9F0",
        "acc1":    "#F7B731",
        "acc2":    "#FF6B6B",
        "seq":     ["#7BBF6A","#4CC9F0","#F7B731","#FF6B6B","#A29BFE","#FD79A8","#00CEC9"],
    }
    _NO_INTERACT = dict(displayModeBar=False, scrollZoom=False)
    _BASE = dict(
        paper_bgcolor=_C["bg"], plot_bgcolor=_C["bg"],
        font=dict(family="Inter, sans-serif", color=_C["text"], size=12),
        margin=dict(l=12, r=12, t=36, b=12),
    )

    df = pd.DataFrame(itens)
    ligados   = int(df["ignicao"].sum())
    deslig    = len(df) - ligados
    total_km  = df["odometro"].sum()
    n_estados = df["uf"].nunique()
    n_cidades = df["cidade"].nunique()

    # ── Cards ─────────────────────────────────────────────────────────────────
    st.markdown(f"""
    <div style="display:flex;gap:12px;flex-wrap:wrap;margin-bottom:20px">
      <div style="flex:1;min-width:120px;background:rgba(123,191,106,0.12);border:1px solid #7BBF6A55;
                  border-radius:10px;padding:16px;text-align:center">
        <div style="font-size:1.8rem;font-weight:700;color:#7BBF6A">{len(df)}</div>
        <div style="color:#C8D8A8;font-size:.8rem">Veículos ECO</div></div>
      <div style="flex:1;min-width:120px;background:rgba(76,201,240,0.12);border:1px solid #4CC9F055;
                  border-radius:10px;padding:16px;text-align:center">
        <div style="font-size:1.8rem;font-weight:700;color:#4CC9F0">{ligados}</div>
        <div style="color:#C8D8A8;font-size:.8rem">🟢 Ligados</div></div>
      <div style="flex:1;min-width:120px;background:rgba(255,107,107,0.12);border:1px solid #FF6B6B55;
                  border-radius:10px;padding:16px;text-align:center">
        <div style="font-size:1.8rem;font-weight:700;color:#FF6B6B">{deslig}</div>
        <div style="color:#C8D8A8;font-size:.8rem">🔴 Desligados</div></div>
      <div style="flex:1;min-width:120px;background:rgba(247,183,49,0.12);border:1px solid #F7B73155;
                  border-radius:10px;padding:16px;text-align:center">
        <div style="font-size:1.8rem;font-weight:700;color:#F7B731">{total_km:,.0f}</div>
        <div style="color:#C8D8A8;font-size:.8rem">Km total (hodômetro)</div></div>
      <div style="flex:1;min-width:120px;background:rgba(162,155,254,0.12);border:1px solid #A29BFE55;
                  border-radius:10px;padding:16px;text-align:center">
        <div style="font-size:1.8rem;font-weight:700;color:#A29BFE">{n_estados} UFs · {n_cidades} cidades</div>
        <div style="color:#C8D8A8;font-size:.8rem">Distribuição geográfica</div></div>
    </div>""", unsafe_allow_html=True)

    # ── Hodômetro por motorista ────────────────────────────────────────────────
    df_odo = df.sort_values("odometro", ascending=True)
    cores_contrato = {c: _C["seq"][i % len(_C["seq"])]
                      for i, c in enumerate(df["contrato"].unique())}
    fig_odo = go.Figure()
    for contrato, grp in df_odo.groupby("contrato"):
        fig_odo.add_trace(go.Bar(
            x=grp["odometro"], y=grp["motorista"], orientation="h",
            name=contrato,
            marker_color=cores_contrato[contrato],
            marker_line_width=0,
            text=[f"{v:,.0f} km" for v in grp["odometro"]],
            textposition="outside",
            textfont=dict(size=10, color=_C["text"]),
            hovertemplate="<b>%{y}</b><br>Hodômetro: %{x:,.0f} km<extra></extra>",
        ))
    fig_odo.update_layout(
        **_BASE,
        title=dict(text="Hodômetro Acumulado por Motorista", font=dict(size=14, color=_C["text"]), x=0),
        barmode="group",
        height=max(420, len(df) * 24),
        xaxis=dict(tickformat=",", gridcolor=_C["grid"], zeroline=False, showline=False),
        yaxis=dict(gridcolor=_C["grid"], tickfont=dict(size=10)),
        legend=dict(orientation="h", y=1.04, x=0, font=dict(size=11)),
        bargap=0.25,
    )
    st.plotly_chart(fig_odo, use_container_width=True, config=_NO_INTERACT)

    col1, col2 = st.columns(2)

    # ── Ignição: donut ─────────────────────────────────────────────────────────
    with col1:
        fig_ign = go.Figure(go.Pie(
            labels=["🟢 Ligados", "🔴 Desligados"],
            values=[ligados, deslig],
            hole=0.62,
            marker=dict(colors=["#7BBF6A", "#FF6B6B"],
                        line=dict(color="#0E1117", width=2)),
            textinfo="label+percent",
            textfont=dict(size=11, color=_C["text"]),
            hovertemplate="<b>%{label}</b>: %{value}<extra></extra>",
        ))
        fig_ign.add_annotation(text=f"<b>{len(df)}</b><br>ECO", x=0.5, y=0.5,
                               font=dict(size=16, color=_C["text"]), showarrow=False)
        fig_ign.update_layout(**_BASE, height=280,
                              title=dict(text="Status de Ignição", font=dict(size=13, color=_C["text"]), x=0),
                              showlegend=False)
        st.plotly_chart(fig_ign, use_container_width=True, config=_NO_INTERACT)

    # ── Por Contrato: donut ────────────────────────────────────────────────────
    with col2:
        cnt_c = df.groupby("contrato").size().reset_index(name="n")
        fig_cont = go.Figure(go.Pie(
            labels=cnt_c["contrato"], values=cnt_c["n"],
            hole=0.62,
            marker=dict(colors=_C["seq"], line=dict(color="#0E1117", width=2)),
            textinfo="label+value",
            textfont=dict(size=11, color=_C["text"]),
            hovertemplate="<b>%{label}</b>: %{value} veículos<extra></extra>",
        ))
        fig_cont.add_annotation(text="<b>Contratos</b>", x=0.5, y=0.5,
                                font=dict(size=13, color=_C["text"]), showarrow=False)
        fig_cont.update_layout(**_BASE, height=280,
                               title=dict(text="Distribuição por Contrato", font=dict(size=13, color=_C["text"]), x=0),
                               showlegend=False)
        st.plotly_chart(fig_cont, use_container_width=True, config=_NO_INTERACT)

    # ── Cidades e estados ──────────────────────────────────────────────────────
    col3, col4 = st.columns(2)

    with col3:
        cnt_uf = df.groupby("uf").size().reset_index(name="n").sort_values("n")
        fig_uf = go.Figure(go.Bar(
            x=cnt_uf["n"], y=cnt_uf["uf"], orientation="h",
            marker=dict(
                color=cnt_uf["n"],
                colorscale=[[0,"#1a3a2a"],[0.5,"#4A8A5A"],[1,"#7BBF6A"]],
                line_width=0,
            ),
            text=cnt_uf["n"], textposition="outside",
            textfont=dict(size=11, color=_C["text"]),
            hovertemplate="<b>%{y}</b>: %{x} veículos<extra></extra>",
        ))
        fig_uf.update_layout(
            **_BASE,
            title=dict(text="Veículos por Estado", font=dict(size=13, color=_C["text"]), x=0),
            height=280,
            xaxis=dict(gridcolor=_C["grid"], zeroline=False, showline=False),
            yaxis=dict(gridcolor="rgba(0,0,0,0)"),
        )
        st.plotly_chart(fig_uf, use_container_width=True, config=_NO_INTERACT)

    with col4:
        cnt_cid = (df.groupby(["cidade","uf"]).size()
                     .reset_index(name="n")
                     .sort_values("n", ascending=False)
                     .head(12))
        cnt_cid["label"] = cnt_cid["cidade"] + " · " + cnt_cid["uf"]
        fig_cid = go.Figure(go.Bar(
            x=cnt_cid["n"], y=cnt_cid["label"].iloc[::-1],
            orientation="h",
            marker=dict(
                color=cnt_cid["n"].iloc[::-1],
                colorscale=[[0,"#132840"],[0.5,"#1E6B9E"],[1,"#4CC9F0"]],
                line_width=0,
            ),
            text=cnt_cid["n"].iloc[::-1], textposition="outside",
            textfont=dict(size=10, color=_C["text"]),
            hovertemplate="<b>%{y}</b>: %{x} veículos<extra></extra>",
        ))
        fig_cid.update_layout(
            **_BASE,
            title=dict(text="Top 12 Cidades", font=dict(size=13, color=_C["text"]), x=0),
            height=280,
            xaxis=dict(gridcolor=_C["grid"], zeroline=False, showline=False),
            yaxis=dict(gridcolor="rgba(0,0,0,0)", tickfont=dict(size=9)),
        )
        st.plotly_chart(fig_cid, use_container_width=True, config=_NO_INTERACT)

    # ── Tabela completa ────────────────────────────────────────────────────────
    st.markdown("#### 📋 Tabela Completa")
    df_tab = df[["contrato","motorista","placa","odometro","velocidade",
                 "tempo_dir_h","cidade","uf","dt_posicao","ignicao"]].copy()
    df_tab["ignicao"] = df_tab["ignicao"].map({True:"🟢 Ligado", False:"🔴 Desligado"})
    df_tab = df_tab.rename(columns={
        "contrato":"Contrato","motorista":"Motorista","placa":"Placa",
        "odometro":"Hodômetro km","velocidade":"Vel. km/h",
        "tempo_dir_h":"Tempo dirigindo (h)","cidade":"Cidade","uf":"UF",
        "dt_posicao":"Última posição","ignicao":"Ignição",
    })
    st.dataframe(df_tab.sort_values("Hodômetro km", ascending=False),
                 use_container_width=True, hide_index=True, height=400)


def _render_rota_individual(itens):
    st.markdown("**Selecione o veículo e o período:**")
    opcoes = {f"{it['motorista']}  [{it['ultima_data']}]": it for it in itens}
    r1, r2, r3, r4 = st.columns([3, 2, 2, 1])
    with r1:
        sel = st.selectbox("Motorista [última data]:", list(opcoes.keys()), key="logos_sel_v")
    it_sel = opcoes[sel]
    with r2:
        d_ini = st.date_input("De:", value=it_sel["ultima_data"], key="logos_r_ini")
    with r3:
        d_fim = st.date_input("Até:", value=it_sel["ultima_data"], key="logos_r_fim")
    with r4:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        ver_rota = st.button("🗺️ Ver Rota", key="logos_btn_rota", use_container_width=True)

    if ver_rota:
        with st.spinner("Buscando rota..."):
            try:
                sess2, _ = _logos_login()
                hist = _logos_get_rota(
                    sess2, it_sel["idvei"],
                    d_ini.strftime("%Y-%m-%d 00:00"),
                    d_fim.strftime("%Y-%m-%d 23:59"),
                )
                if not hist:
                    st.warning(f"Nenhuma posição para {d_ini} – {d_fim}. Tente outras datas.")
                    return
                st.session_state["logos_rota"]     = hist
                st.session_state["logos_rota_sel"] = sel
                st.session_state["logos_rota_cor"] = it_sel["cor"]
            except Exception as e:
                st.error(f"❌ {e}")
                return

    hist = st.session_state.get("logos_rota", [])
    if not hist:
        return

    coords = []
    cidades_rota = []
    for p in hist:
        lt = p.get("pos_coordenada_latitude")
        ln = p.get("pos_coordenada_longitude")
        if lt and ln:
            try:
                coords.append([float(lt), float(ln)])
            except Exception:
                pass
        cidade = p.get("pos_end_cidade", "")
        if cidade and (not cidades_rota or cidades_rota[-1] != cidade):
            cidades_rota.append(cidade)

    if not coords:
        st.warning("Sem coordenadas válidas neste período.")
        return

    cor_rota  = st.session_state.get("logos_rota_cor", "#4CC9F0")
    desc_rota = st.session_state.get("logos_rota_sel", sel)

    # Cards da rota
    km_rota = _km_from_hist(hist)
    c1, c2, c3 = st.columns(3)
    c1.metric("Posições GPS", len(coords))
    c2.metric("Km percorridos", f"{km_rota:,} km")
    c3.metric("Cidades passadas", len(set(cidades_rota)))

    if cidades_rota:
        st.caption("Rota: " + " → ".join(dict.fromkeys(cidades_rota)))

    mapa_r = folium.Map(tiles="CartoDB dark_matter")
    folium.PolyLine(coords, color=cor_rota, weight=4, opacity=0.9,
                    tooltip=desc_rota).add_to(mapa_r)
    folium.CircleMarker(coords[0],  radius=8, color="#00FF00", fill=True,
                        fill_color="#00FF00", tooltip="▶ Início").add_to(mapa_r)
    folium.CircleMarker(coords[-1], radius=8, color="#FF4757", fill=True,
                        fill_color="#FF4757", tooltip="⏹ Fim").add_to(mapa_r)
    lats = [c[0] for c in coords]; lons = [c[1] for c in coords]
    mapa_r.fit_bounds([[min(lats), min(lons)], [max(lats), max(lons)]])
    st_folium(mapa_r, width="100%", height=500, key="logos_mapa_rota", returned_objects=[])


def _render_analise_periodo(itens):
    _C = {
        "bg":      "rgba(0,0,0,0)",
        "grid":    "rgba(255,255,255,0.06)",
        "text":    "#C8D8A8",
        "eco135":  "#7BBF6A",
        "cerrado": "#4CC9F0",
        "acc1":    "#F7B731",
        "acc2":    "#FF6B6B",
        "seq":     ["#7BBF6A","#4CC9F0","#F7B731","#FF6B6B","#A29BFE","#FD79A8","#00CEC9"],
    }
    _NO_INTERACT = dict(displayModeBar=False, scrollZoom=False)
    _BASE = dict(
        paper_bgcolor=_C["bg"], plot_bgcolor=_C["bg"],
        font=dict(family="Inter, sans-serif", color=_C["text"], size=12),
        margin=dict(l=12, r=12, t=36, b=12),
    )

    # Deriva datas padrão a partir dos próprios veículos (evita período sem dados)
    _datas_vei = [it["ultima_data"] for it in itens if it.get("ultima_data")]
    _d_max = max(_datas_vei) if _datas_vei else date.today()
    _d_min = _d_max.replace(day=1)  # primeiro dia do mês do último dado

    st.markdown(
        f"Busca o histórico de todos os veículos ECO em um período e gera estatísticas consolidadas. "
        f"**Última posição disponível: {_d_max.strftime('%d/%m/%Y')}**"
    )
    p1, p2, p3 = st.columns([2, 2, 1])
    with p1:
        pd_ini = st.date_input("Data início:", value=_d_min, key="logos_p_ini")
    with p2:
        pd_fim = st.date_input("Data fim:", value=_d_max, key="logos_p_fim")
    with p3:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        carregar = st.button("📊 Carregar Período", key="logos_btn_periodo", use_container_width=True)

    if carregar:
        resultados = []
        all_pontos = []
        _raw_sample = None   # primeiro ponto bruto para diagnóstico de campos
        prog = st.progress(0, text="Iniciando...")
        with st.spinner("Buscando histórico de todos os veículos ECO..."):
            try:
                sess3, _ = _logos_login()
                for idx, it in enumerate(itens):
                    prog.progress((idx + 1) / len(itens),
                                  text=f"Buscando {it['motorista']} ({idx+1}/{len(itens)})...")
                    try:
                        hist = _logos_get_rota(
                            sess3, it["idvei"],
                            pd_ini.strftime("%Y-%m-%d 00:00"),
                            pd_fim.strftime("%Y-%m-%d 23:59"),
                        )
                    except Exception:
                        hist = []

                    if hist:
                        if _raw_sample is None:
                            _raw_sample = hist[0]  # guarda primeiro ponto para debug
                        km = _km_from_hist(hist)
                        cidades = list(dict.fromkeys(
                            str(_pick(p, _HIST_CID_FIELDS) or "") for p in hist
                            if _pick(p, _HIST_CID_FIELDS)
                        ))
                        ufs = list(dict.fromkeys(
                            str(_pick(p, _HIST_UF_FIELDS) or "") for p in hist
                            if _pick(p, _HIST_UF_FIELDS)
                        ))
                        # Coleta pontos temporais (com detecção automática de campos)
                        for p in hist:
                            dt_str = str(_pick(p, _HIST_DT_FIELDS, ""))
                            try:
                                dt  = pd.to_datetime(dt_str)
                                vel = float(_pick(p, _HIST_VEL_FIELDS) or 0)
                                ign = bool(_pick(p, _HIST_IGN_FIELDS) or False)
                                odo = int(float(_pick(p, _HIST_ODO_FIELDS) or 0))
                                all_pontos.append({
                                    "motorista":  it["motorista"],
                                    "contrato":   it["contrato"],
                                    "hora":       dt.hour,
                                    "dia_semana": dt.dayofweek,
                                    "data":       dt.date(),
                                    "odometro":   odo,
                                    "ignicao":    ign,
                                    "velocidade": vel,
                                    "idle":       ign and vel <= 3,
                                    "cidade":     str(_pick(p, _HIST_CID_FIELDS) or ""),
                                    "uf":         str(_pick(p, _HIST_UF_FIELDS) or ""),
                                })
                            except Exception:
                                pass
                    else:
                        km = 0; cidades = []; ufs = []

                    resultados.append({
                        "contrato":    it["contrato"],
                        "motorista":   it["motorista"],
                        "placa":       it["placa"],
                        "km_periodo":  km,
                        "registros":   len(hist),
                        "cidades":     len(set(cidades)),
                        "estados":     ", ".join(ufs[:3]),
                        "rota_resumo": " → ".join(list(dict.fromkeys(cidades))[:5]),
                    })
                prog.empty()
                st.session_state["logos_periodo_result"] = resultados
                st.session_state["logos_periodo_pontos"] = all_pontos
                st.session_state["logos_periodo_label"]  = f"{pd_ini} a {pd_fim}"
                st.session_state["logos_hist_sample"] = _raw_sample
            except Exception as e:
                st.error(f"❌ {e}")
                return

    res = st.session_state.get("logos_periodo_result", [])
    if not res:
        return

    label_periodo = st.session_state.get("logos_periodo_label", "")
    st.markdown(f"#### Resultados: {label_periodo}")

    df_p  = pd.DataFrame(res)
    pontos = st.session_state.get("logos_periodo_pontos", [])
    df_pt  = pd.DataFrame(pontos) if pontos else pd.DataFrame()

    total_km = df_p["km_periodo"].sum()

    if total_km == 0 and df_pt.empty:
        st.warning(
            "⚠️ Nenhum dado encontrado para este período. "
            "Verifique se as datas selecionadas têm dados disponíveis — "
            f"a última posição conhecida dos veículos é **{_d_max.strftime('%d/%m/%Y')}**."
        )
        return
    # ── Métricas base ──────────────────────────────────────────────────────────
    # ── Debug: campos reais do historicoposicao ────────────────────────────────
    _sample = st.session_state.get("logos_hist_sample")
    if _sample:
        total_km_check = sum(r.get("km_periodo", 0) for r in res)
        if total_km_check == 0:
            st.error("⚠️ Km = 0 para todos os veículos. Exibindo campos reais da API para diagnóstico:")
            with st.expander("🔍 Campos retornados pelo historicoposicao (clique para ver)", expanded=True):
                st.json({k: v for k, v in _sample.items()})
                st.caption("Os campos acima são os REAIS retornados. Se 'pos_odometro' não aparecer, "
                            "o nome correto está nessa lista.")
        else:
            with st.expander("🔍 Debug: campos da API (opcional)"):
                st.json({k: v for k, v in _sample.items()})

    MINS_POR_PONTO = 3  # estimativa Logos: ~1 ponto a cada 3 min
    CUSTO_KM       = 0.50   # R$/km — diesel highway pickup (Geotab benchmark)
    CUSTO_IDLE_H   = 5.09   # R$/h idle — 1L diesel/h × R$5,09/L (preço fev/2026)
    LIMIAR_IDLE_OK = 5      # % idle aceitável (Geotab: alvo ≤5%)
    LIMIAR_IDLE_AL = 10     # % idle alarme (Geotab: >10% = crítico)
    LIMIAR_KM_WA   = 300    # km/dia atenção
    LIMIAR_KM_AL   = 500    # km/dia alarme (≈ 5h30 dirigindo a 90 km/h — Lei 13.103)
    dias_nome      = ["Seg","Ter","Qua","Qui","Sex","Sáb","Dom"]

    ativos   = (df_p["registros"] > 0).sum()
    inativos = len(df_p) - ativos
    n_dias   = max(1, (pd.to_datetime(label_periodo.split(" a ")[-1]) -
                       pd.to_datetime(label_periodo.split(" a ")[0])).days + 1) if " a " in label_periodo else 1

    # ── Agrega KPIs por motorista ──────────────────────────────────────────────
    # h_rastreio = total de pontos GPS × intervalo estimado (≈3 min/ponto)
    # Isso equivale ao tempo TOTAL rastreado (motor ligado ou não).
    # idle = vel ≤ 3 km/h (parado com ou sem ignição)
    if not df_pt.empty:
        agg = {}
        for mot, grp in df_pt.groupby("motorista"):
            n_total  = len(grp)
            n_mov    = int((grp["velocidade"] > 3).sum()) if "velocidade" in grp else n_total
            n_idle   = n_total - n_mov  # parado = vel ≤ 3
            n_fds    = int((grp["dia_semana"] >= 5).sum()) if "dia_semana" in grp else 0
            h_rastreio = n_total * MINS_POR_PONTO / 60   # total horas rastreadas
            h_mov      = n_mov   * MINS_POR_PONTO / 60   # horas em movimento
            h_idle     = n_idle  * MINS_POR_PONTO / 60   # horas parado
            idle_pct   = round(h_idle / max(0.1, h_rastreio) * 100, 1)
            fds_pct    = round(n_fds / max(1, n_total) * 100, 1)
            km         = float(df_p.loc[df_p["motorista"] == mot, "km_periodo"].values[0]) \
                         if mot in df_p["motorista"].values else 0.0
            # Eficiência = km / horas em MOVIMENTO (não total)
            # Cap em 120 km/h (máx razoável rodovia)
            eff_raw    = km / max(0.1, h_mov)
            eff        = round(min(eff_raw, 120.0), 1)
            custo_cb   = round(km * CUSTO_KM, 2)
            custo_id   = round(h_idle * CUSTO_IDLE_H, 2)
            agg[mot] = {
                "motorista":   mot,
                "km_periodo":  km,
                "h_rastreio":  round(h_rastreio, 1),
                "h_mov":       round(h_mov, 1),
                "h_idle":      round(h_idle, 1),
                "idle_pct":    idle_pct,
                "fds_pct":     fds_pct,
                "n_fds":       n_fds,
                "efficiency":  eff,
                "custo_cb":    custo_cb,
                "custo_idle":  custo_id,
                "custo_total": round(custo_cb + custo_id, 2),
            }
        df_mot = pd.DataFrame(list(agg.values()))
    else:
        df_mot = pd.DataFrame()

    total_custo_cb   = df_mot["custo_cb"].sum()   if not df_mot.empty else 0
    total_custo_idle = df_mot["custo_idle"].sum()  if not df_mot.empty else 0
    total_custo      = df_mot["custo_total"].sum() if not df_mot.empty else 0
    km_dia           = total_km / n_dias

    # ═══════════════════════════════════════════════════════════════════════════
    # CARDS RESUMO
    # ═══════════════════════════════════════════════════════════════════════════
    st.markdown(f"""
    <div style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:18px">
      <div style="flex:1;min-width:110px;background:rgba(123,191,106,0.12);border:1px solid #7BBF6A55;
                  border-radius:10px;padding:14px;text-align:center">
        <div style="font-size:1.6rem;font-weight:700;color:#7BBF6A">{total_km:,.0f} km</div>
        <div style="color:#C8D8A8;font-size:.75rem">Total rodado</div></div>
      <div style="flex:1;min-width:110px;background:rgba(76,201,240,0.12);border:1px solid #4CC9F055;
                  border-radius:10px;padding:14px;text-align:center">
        <div style="font-size:1.6rem;font-weight:700;color:#4CC9F0">{km_dia:,.0f} km</div>
        <div style="color:#C8D8A8;font-size:.75rem">Média km/dia</div></div>
      <div style="flex:1;min-width:110px;background:rgba(247,183,49,0.12);border:1px solid #F7B73155;
                  border-radius:10px;padding:14px;text-align:center">
        <div style="font-size:1.6rem;font-weight:700;color:#F7B731">R$ {total_custo_cb:,.0f}</div>
        <div style="color:#C8D8A8;font-size:.75rem">Custo combustível est.</div></div>
      <div style="flex:1;min-width:110px;background:rgba(255,107,107,0.12);border:1px solid #FF6B6B55;
                  border-radius:10px;padding:14px;text-align:center">
        <div style="font-size:1.6rem;font-weight:700;color:#FF6B6B">R$ {total_custo_idle:,.0f}</div>
        <div style="color:#C8D8A8;font-size:.75rem">Custo idle est.</div></div>
      <div style="flex:1;min-width:110px;background:rgba(162,155,254,0.12);border:1px solid #A29BFE55;
                  border-radius:10px;padding:14px;text-align:center">
        <div style="font-size:1.6rem;font-weight:700;color:#A29BFE">R$ {total_custo:,.0f}</div>
        <div style="color:#C8D8A8;font-size:.75rem">Custo total estimado</div></div>
      <div style="flex:1;min-width:110px;background:rgba(123,191,106,0.12);border:1px solid #7BBF6A55;
                  border-radius:10px;padding:14px;text-align:center">
        <div style="font-size:1.6rem;font-weight:700;color:#7BBF6A">{ativos}</div>
        <div style="color:#C8D8A8;font-size:.75rem">Veículos ativos</div></div>
      <div style="flex:1;min-width:110px;background:rgba(255,107,107,0.08);border:1px solid #FF6B6B33;
                  border-radius:10px;padding:14px;text-align:center">
        <div style="font-size:1.6rem;font-weight:700;color:#FF6B6B">{inativos}</div>
        <div style="color:#C8D8A8;font-size:.75rem">Sem dados</div></div>
      <div style="flex:1;min-width:110px;background:rgba(247,183,49,0.08);border:1px solid #F7B73133;
                  border-radius:10px;padding:14px;text-align:center">
        <div style="font-size:1.6rem;font-weight:700;color:#F7B731">{n_dias}</div>
        <div style="color:#C8D8A8;font-size:.75rem">Dias analisados</div></div>
    </div>""", unsafe_allow_html=True)

    # ═══════════════════════════════════════════════════════════════════════════
    # SEÇÃO 1 — ⚠️ ALERTAS CRÍTICOS (nomes específicos)
    # ═══════════════════════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("## ⚠️ Alertas Críticos — Pessoas e Quantidades")
    st.caption("Baseado em: Geotab Fleet KPIs · Lei 13.103/2015 · CTB Art. 61 · preço diesel Fev/2026")

    if not df_mot.empty:
        a1, a2, a3 = st.columns(3)

        # Alerta 1: Idle > limiar
        with a1:
            df_idle_al = df_mot[df_mot["idle_pct"] > LIMIAR_IDLE_AL].sort_values("idle_pct", ascending=False)
            st.markdown(f"""
            <div style="background:rgba(255,70,70,0.12);border:1px solid #FF6B6B88;border-radius:10px;
                        padding:14px;margin-bottom:12px">
              <div style="color:#FF6B6B;font-weight:700;font-size:1rem">
                🔴 Motor Ocioso &gt;{LIMIAR_IDLE_AL}% — {len(df_idle_al)} motorista(s)
              </div>
              <div style="color:#C8D8A8;font-size:.75rem">Benchmark Geotab: alvo ≤5%, alarme &gt;10%<br>
              Custo: ~R$5,09/hora parado</div>
            </div>""", unsafe_allow_html=True)
            if not df_idle_al.empty:
                tab_idle = df_idle_al[["motorista","idle_pct","h_idle","custo_idle"]].copy()
                tab_idle.columns = ["Motorista","Idle %","Horas idle","Custo idle (R$)"]
                st.dataframe(tab_idle, use_container_width=True, hide_index=True)
            else:
                st.success("Nenhum motorista acima do limite.")
            df_idle_wa = df_mot[(df_mot["idle_pct"] > LIMIAR_IDLE_OK) & (df_mot["idle_pct"] <= LIMIAR_IDLE_AL)]
            if not df_idle_wa.empty:
                st.markdown(f"🟡 **Atenção ({LIMIAR_IDLE_OK}–{LIMIAR_IDLE_AL}%):** {', '.join(df_idle_wa['motorista'].tolist())}")

        # Alerta 2: Km/dia excessivo
        with a2:
            if not df_pt.empty and "data" in df_pt.columns and "odometro" in df_pt.columns:
                _dfo = df_pt[df_pt["odometro"] > 0]
                df_daily_v = (_dfo.groupby(["motorista","data"])["odometro"]
                                    .agg(["max","min"]).reset_index())
                df_daily_v["km_dia"] = (df_daily_v["max"] - df_daily_v["min"]).clip(lower=0, upper=1500)
                dias_alarme = df_daily_v[df_daily_v["km_dia"] > LIMIAR_KM_AL].sort_values("km_dia", ascending=False)
                dias_atencao = df_daily_v[(df_daily_v["km_dia"] > LIMIAR_KM_WA) &
                                          (df_daily_v["km_dia"] <= LIMIAR_KM_AL)]
                st.markdown(f"""
                <div style="background:rgba(255,70,70,0.12);border:1px solid #FF6B6B88;border-radius:10px;
                            padding:14px;margin-bottom:12px">
                  <div style="color:#FF6B6B;font-weight:700;font-size:1rem">
                    🔴 Km/dia &gt;{LIMIAR_KM_AL} — {len(dias_alarme)} ocorrência(s)
                  </div>
                  <div style="color:#C8D8A8;font-size:.75rem">Alarme: &gt;500 km/dia ≈ risco Lei 13.103<br>
                  {len(dias_atencao)} ocorrências em zona atenção ({LIMIAR_KM_WA}–{LIMIAR_KM_AL} km)</div>
                </div>""", unsafe_allow_html=True)
                if not dias_alarme.empty:
                    tab_al = dias_alarme[["motorista","data","km_dia"]].copy()
                    tab_al.columns = ["Motorista","Data","Km no dia"]
                    st.dataframe(tab_al, use_container_width=True, hide_index=True)
                else:
                    st.success("Nenhum dia com >500 km.")

        # Alerta 3: Fim de semana
        with a3:
            if not df_mot.empty and "n_fds" in df_mot.columns:
                df_fds_al = df_mot[df_mot["n_fds"] > 0].sort_values("fds_pct", ascending=False)
                st.markdown(f"""
                <div style="background:rgba(255,183,49,0.12);border:1px solid #F7B73188;border-radius:10px;
                            padding:14px;margin-bottom:12px">
                  <div style="color:#F7B731;font-weight:700;font-size:1rem">
                    🟡 Ativos no Fim de Semana — {len(df_fds_al)} motorista(s)
                  </div>
                  <div style="color:#C8D8A8;font-size:.75rem">Qualquer trip FDS sem OS aprovada = custo não justificado<br>
                  Benchmark: 0% FDS sem autorização</div>
                </div>""", unsafe_allow_html=True)
                if not df_fds_al.empty:
                    tab_fds = df_fds_al[["motorista","n_fds","fds_pct","km_periodo"]].copy()
                    tab_fds.columns = ["Motorista","Reg. FDS","% do total","Km total"]
                    st.dataframe(tab_fds, use_container_width=True, hide_index=True)

    # ═══════════════════════════════════════════════════════════════════════════
    # SEÇÃO 2 — 💰 ANÁLISE DE CUSTO POR MOTORISTA
    # ═══════════════════════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("## 💰 Custo Estimado por Motorista")
    st.caption("Combustível: R$0,50/km · Idle: R$5,09/h (1L diesel/h × preço Fev/2026)")

    if not df_mot.empty:
        df_custo = df_mot[df_mot["km_periodo"] > 0].sort_values("custo_total", ascending=True)

        fig_custo = go.Figure()
        fig_custo.add_trace(go.Bar(
            x=df_custo["custo_cb"], y=df_custo["motorista"], orientation="h",
            name="Combustível (R$)", marker_color="#7BBF6A", marker_line_width=0,
            hovertemplate="<b>%{y}</b><br>Combustível: R$ %{x:,.0f}<extra></extra>",
        ))
        fig_custo.add_trace(go.Bar(
            x=df_custo["custo_idle"], y=df_custo["motorista"], orientation="h",
            name="Idle (R$)", marker_color="#FF6B6B", marker_line_width=0,
            hovertemplate="<b>%{y}</b><br>Idle: R$ %{x:,.0f}<extra></extra>",
        ))
        fig_custo.update_layout(
            **_BASE,
            title=dict(text="💰 Custo estimado: combustível + idle por motorista",
                       font=dict(size=14, color=_C["text"]), x=0),
            barmode="stack",
            height=max(400, len(df_custo) * 22),
            xaxis=dict(tickprefix="R$", tickformat=",", gridcolor=_C["grid"], zeroline=False),
            yaxis=dict(gridcolor=_C["grid"], tickfont=dict(size=9)),
            legend=dict(orientation="h", y=1.04, x=0),
            bargap=0.2,
        )
        st.plotly_chart(fig_custo, use_container_width=True, config=_NO_INTERACT)

    # ═══════════════════════════════════════════════════════════════════════════
    # SEÇÃO 3 — 📊 SCORE DE EFICIÊNCIA
    # ═══════════════════════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("## 📊 Score de Eficiência — km por hora em movimento")
    st.caption("km / horas com vel >3 km/h (cap 120 km/h). Baixo = viagens curtas · Alto = rodovia contínua.")

    if not df_mot.empty:
        col_e1, col_e2 = st.columns(2)

        with col_e1:
            df_eff = df_mot[df_mot["km_periodo"] > 0].sort_values("efficiency", ascending=True)
            media_eff = df_eff["efficiency"].mean()
            cores_eff = [_C["acc2"] if v < media_eff * 0.7
                         else (_C["acc1"] if v < media_eff else _C["eco135"])
                         for v in df_eff["efficiency"]]
            fig_eff = go.Figure(go.Bar(
                x=df_eff["efficiency"], y=df_eff["motorista"], orientation="h",
                marker_color=cores_eff, marker_line_width=0,
                text=[f"{v:.1f} km/h" for v in df_eff["efficiency"]],
                textposition="outside",
                textfont=dict(size=10, color=_C["text"]),
                hovertemplate="<b>%{y}</b>: %{x:.1f} km/h em movimento<extra></extra>",
            ))
            fig_eff.add_vline(x=media_eff, line_dash="dash", line_color=_C["acc1"],
                              annotation_text=f"Média: {media_eff:.1f}",
                              annotation_font_color=_C["acc1"])
            fig_eff.update_layout(
                **_BASE,
                title=dict(text="⚡ Eficiência: km por hora em movimento",
                           font=dict(size=13, color=_C["text"]), x=0),
                height=max(400, len(df_eff) * 22),
                xaxis=dict(gridcolor=_C["grid"], zeroline=False),
                yaxis=dict(gridcolor=_C["grid"], tickfont=dict(size=9)),
            )
            st.plotly_chart(fig_eff, use_container_width=True, config=_NO_INTERACT)

        with col_e2:
            # Scatter: km_total vs horas_idle — vilões ficam no canto direito-baixo
            df_sc = df_mot[df_mot["km_periodo"] > 0].copy()
            df_sc["tamanho"] = (df_sc["km_periodo"] / df_sc["km_periodo"].max() * 30 + 8).clip(8, 40)
            df_sc["cor"] = df_sc["idle_pct"].apply(
                lambda x: "#FF6B6B" if x > LIMIAR_IDLE_AL else ("#F7B731" if x > LIMIAR_IDLE_OK else "#7BBF6A")
            )
            fig_sc = go.Figure()
            for label, color, mask in [
                (f"Idle OK (≤{LIMIAR_IDLE_OK}%)", "#7BBF6A", df_sc["idle_pct"] <= LIMIAR_IDLE_OK),
                (f"Atenção ({LIMIAR_IDLE_OK}–{LIMIAR_IDLE_AL}%)", "#F7B731",
                 (df_sc["idle_pct"] > LIMIAR_IDLE_OK) & (df_sc["idle_pct"] <= LIMIAR_IDLE_AL)),
                (f"Alarme (>{LIMIAR_IDLE_AL}%)", "#FF6B6B", df_sc["idle_pct"] > LIMIAR_IDLE_AL),
            ]:
                sub = df_sc[mask]
                if not sub.empty:
                    fig_sc.add_trace(go.Scatter(
                        x=sub["h_idle"], y=sub["km_periodo"],
                        mode="markers+text",
                        name=label,
                        marker=dict(color=color, size=sub["tamanho"], opacity=0.85,
                                    line=dict(color="#0D1B2A", width=1)),
                        text=sub["motorista"].str.split().str[0],
                        textposition="top center",
                        textfont=dict(size=9, color=_C["text"]),
                        hovertemplate="<b>%{text}</b><br>Idle: %{x:.1f}h · Km: %{y:,.0f}<extra></extra>",
                        customdata=sub[["motorista","idle_pct"]].values,
                    ))
            fig_sc.update_layout(
                **_BASE,
                title=dict(text="🎯 Km rodado vs Horas idle (tamanho = km total)",
                           font=dict(size=13, color=_C["text"]), x=0),
                height=420,
                xaxis=dict(title="Horas idle", gridcolor=_C["grid"], zeroline=False),
                yaxis=dict(title="Km rodado", gridcolor=_C["grid"], zeroline=False, tickformat=","),
                legend=dict(orientation="h", y=1.06, x=0, font=dict(size=10)),
            )
            st.plotly_chart(fig_sc, use_container_width=True, config=_NO_INTERACT)

    # ═══════════════════════════════════════════════════════════════════════════
    # SEÇÃO 4 — ⏱️ IDLE TIME DETALHADO
    # ═══════════════════════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("## ⏱️ Tempo Ocioso — Quem mais desperdiça combustível")

    if not df_mot.empty:
        col_i1, col_i2 = st.columns(2)

        with col_i1:
            df_idle_r = df_mot[df_mot["h_rastreio"] > 0].sort_values("idle_pct", ascending=True)
            cores_idle = [
                _C["eco135"] if v <= LIMIAR_IDLE_OK
                else (_C["acc1"] if v <= LIMIAR_IDLE_AL else _C["acc2"])
                for v in df_idle_r["idle_pct"]
            ]
            fig_idle_pct = go.Figure(go.Bar(
                x=df_idle_r["idle_pct"], y=df_idle_r["motorista"], orientation="h",
                marker_color=cores_idle, marker_line_width=0,
                text=[f"{v:.1f}%" for v in df_idle_r["idle_pct"]],
                textposition="outside",
                textfont=dict(size=10, color=_C["text"]),
                hovertemplate="<b>%{y}</b>: %{x:.1f}% idle<extra></extra>",
            ))
            fig_idle_pct.add_vline(x=LIMIAR_IDLE_OK, line_dash="dot", line_color=_C["acc1"],
                                   annotation_text="Alvo 5%", annotation_font_color=_C["acc1"])
            fig_idle_pct.add_vline(x=LIMIAR_IDLE_AL, line_dash="dash", line_color=_C["acc2"],
                                   annotation_text="Alarme 10%", annotation_font_color=_C["acc2"])
            fig_idle_pct.update_layout(
                **_BASE,
                title=dict(text="🔴 % de Idle por Motorista (verde ≤5% · amarelo 5–10% · vermelho >10%)",
                           font=dict(size=12, color=_C["text"]), x=0),
                height=max(400, len(df_idle_r) * 22),
                xaxis=dict(ticksuffix="%", gridcolor=_C["grid"], zeroline=False),
                yaxis=dict(gridcolor=_C["grid"], tickfont=dict(size=9)),
            )
            st.plotly_chart(fig_idle_pct, use_container_width=True, config=_NO_INTERACT)

        with col_i2:
            df_idle_h = df_mot[df_mot["h_idle"] > 0].sort_values("h_idle", ascending=True)
            fig_idle_h = go.Figure(go.Bar(
                x=df_idle_h["h_idle"], y=df_idle_h["motorista"], orientation="h",
                marker=dict(
                    color=df_idle_h["h_idle"],
                    colorscale=[[0,"#1a1a2a"],[0.5,"#8B2020"],[1,"#FF6B6B"]],
                    line_width=0,
                ),
                text=[f"{h:.1f}h · R${c:.0f}" for h, c in zip(df_idle_h["h_idle"], df_idle_h["custo_idle"])],
                textposition="outside",
                textfont=dict(size=10, color=_C["text"]),
                hovertemplate="<b>%{y}</b>: %{x:.1f}h idle<extra></extra>",
            ))
            fig_idle_h.update_layout(
                **_BASE,
                title=dict(text="⏱️ Horas idle absolutas (h · custo est.)",
                           font=dict(size=13, color=_C["text"]), x=0),
                height=max(400, len(df_idle_h) * 22),
                xaxis=dict(ticksuffix="h", gridcolor=_C["grid"], zeroline=False),
                yaxis=dict(gridcolor=_C["grid"], tickfont=dict(size=9)),
            )
            st.plotly_chart(fig_idle_h, use_container_width=True, config=_NO_INTERACT)

    # ═══════════════════════════════════════════════════════════════════════════
    # SEÇÃO 5 — 🏆 RANKING KM + MÉDIA KM/DIA
    # ═══════════════════════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("## 🏆 Quem Mais Rodou no Período")

    col_k1, col_k2 = st.columns(2)

    with col_k1:
        if not df_p.empty:
            cores_c = {c: _C["seq"][i % len(_C["seq"])]
                       for i, c in enumerate(df_p["contrato"].unique())}
            df_km_r = df_p.sort_values("km_periodo", ascending=True)
            fig_km = go.Figure()
            for contrato, grp in df_km_r.groupby("contrato"):
                fig_km.add_trace(go.Bar(
                    x=grp["km_periodo"], y=grp["motorista"], orientation="h",
                    name=contrato, marker_color=cores_c[contrato], marker_line_width=0,
                    text=[f"{v:,.0f} km" for v in grp["km_periodo"]],
                    textposition="outside",
                    textfont=dict(size=10, color=_C["text"]),
                    hovertemplate="<b>%{y}</b><br>%{x:,.0f} km<extra></extra>",
                ))
            fig_km.update_layout(
                **_BASE,
                title=dict(text="Km total por motorista no período",
                           font=dict(size=13, color=_C["text"]), x=0),
                barmode="group", height=max(400, len(df_p) * 22),
                xaxis=dict(tickformat=",", gridcolor=_C["grid"], zeroline=False),
                yaxis=dict(gridcolor=_C["grid"], tickfont=dict(size=9)),
                legend=dict(orientation="h", y=1.04, x=0),
                bargap=0.25,
            )
            st.plotly_chart(fig_km, use_container_width=True, config=_NO_INTERACT)

    with col_k2:
        if not df_pt.empty and "data" in df_pt.columns and "odometro" in df_pt.columns:
            _dfo2 = df_pt[df_pt["odometro"] > 0]
            df_mv = (_dfo2.groupby(["motorista","data"])["odometro"].agg(["max","min"]).reset_index())
            df_mv["km_d"] = (df_mv["max"] - df_mv["min"]).clip(lower=0, upper=1500)
            df_media = (df_mv.groupby("motorista")["km_d"].mean().reset_index()
                              .sort_values("km_d", ascending=True))
            cores_kmdia = [
                _C["acc2"] if v > LIMIAR_KM_AL
                else (_C["acc1"] if v > LIMIAR_KM_WA else _C["eco135"])
                for v in df_media["km_d"]
            ]
            fig_med = go.Figure(go.Bar(
                x=df_media["km_d"], y=df_media["motorista"], orientation="h",
                marker_color=cores_kmdia, marker_line_width=0,
                text=[f"{v:.0f} km/dia" for v in df_media["km_d"]],
                textposition="outside",
                textfont=dict(size=10, color=_C["text"]),
                hovertemplate="<b>%{y}</b>: %{x:.0f} km/dia médio<extra></extra>",
            ))
            fig_med.add_vline(x=LIMIAR_KM_WA, line_dash="dot", line_color=_C["acc1"],
                              annotation_text=f"Atenção {LIMIAR_KM_WA}km",
                              annotation_font_color=_C["acc1"])
            fig_med.add_vline(x=LIMIAR_KM_AL, line_dash="dash", line_color=_C["acc2"],
                              annotation_text=f"Alarme {LIMIAR_KM_AL}km",
                              annotation_font_color=_C["acc2"])
            fig_med.update_layout(
                **_BASE,
                title=dict(text="🚗 Média km/dia (verde OK · amarelo >300 · vermelho >500)",
                           font=dict(size=12, color=_C["text"]), x=0),
                height=max(400, len(df_media) * 22),
                xaxis=dict(gridcolor=_C["grid"], zeroline=False),
                yaxis=dict(gridcolor=_C["grid"], tickfont=dict(size=9)),
            )
            st.plotly_chart(fig_med, use_container_width=True, config=_NO_INTERACT)

    # ── Km por dia — linha temporal ────────────────────────────────────────────
    if not df_pt.empty and "data" in df_pt.columns and "odometro" in df_pt.columns:
        _dfo3 = df_pt[df_pt["odometro"] > 0]
        df_daily2 = (_dfo3.groupby(["data","motorista"])["odometro"].agg(["max","min"]).reset_index())
        df_daily2["km_d"] = (df_daily2["max"] - df_daily2["min"]).clip(lower=0, upper=1500)
        df_agg2 = df_daily2.groupby("data")["km_d"].sum().reset_index().sort_values("data")
        fig_kmd = go.Figure(go.Scatter(
            x=df_agg2["data"].astype(str), y=df_agg2["km_d"],
            mode="lines+markers",
            line=dict(color=_C["acc1"], width=2),
            marker=dict(color=_C["acc1"], size=5),
            fill="tozeroy", fillcolor="rgba(247,183,49,0.10)",
            hovertemplate="<b>%{x}</b>: %{y:,.0f} km (frota)<extra></extra>",
        ))
        fig_kmd.update_layout(
            **_BASE,
            title=dict(text="📈 Km percorrido por dia — frota total",
                       font=dict(size=13, color=_C["text"]), x=0),
            height=280,
            xaxis=dict(gridcolor=_C["grid"], zeroline=False, tickangle=-35, tickfont=dict(size=9)),
            yaxis=dict(gridcolor=_C["grid"], zeroline=False, tickformat=","),
        )
        st.plotly_chart(fig_kmd, use_container_width=True, config=_NO_INTERACT)

    # ═══════════════════════════════════════════════════════════════════════════
    # SEÇÃO 6 — 🚨 FIM DE SEMANA — ANÁLISE COMPLETA
    # ═══════════════════════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("## 🚨 Fim de Semana — Quem e Quanto")

    if not df_pt.empty and "dia_semana" in df_pt.columns:
        col_f1, col_f2 = st.columns(2)

        with col_f1:
            df_fds_r = df_pt[df_pt["dia_semana"] >= 5]
            dias_nome_pt = ["Seg","Ter","Qua","Qui","Sex","Sáb","Dom"]
            dia_cnt = (df_pt.groupby("dia_semana").size()
                            .reindex(range(7), fill_value=0).reset_index())
            dia_cnt.columns = ["dia_semana","n"]
            fds_pct_tot = int(dia_cnt[dia_cnt["dia_semana"] >= 5]["n"].sum() /
                              max(1, dia_cnt["n"].sum()) * 100)
            fig_dia = go.Figure(go.Bar(
                x=[dias_nome_pt[d] for d in dia_cnt["dia_semana"]],
                y=dia_cnt["n"],
                marker_color=[_C["acc2"] if d >= 5 else _C["cerrado"] for d in dia_cnt["dia_semana"]],
                marker_line_width=0,
                text=dia_cnt["n"],
                textposition="outside",
                textfont=dict(size=11, color=_C["text"]),
                hovertemplate="<b>%{x}</b>: %{y} registros<extra></extra>",
            ))
            fig_dia.update_layout(
                **_BASE,
                title=dict(text=f"📅 Atividade por dia — {fds_pct_tot}% dos registros são FDS",
                           font=dict(size=13, color=_C["text"]), x=0),
                height=280,
                xaxis=dict(gridcolor=_C["grid"], zeroline=False),
                yaxis=dict(gridcolor=_C["grid"], zeroline=False),
                bargap=0.2,
            )
            st.plotly_chart(fig_dia, use_container_width=True, config=_NO_INTERACT)

        with col_f2:
            if not df_fds_r.empty:
                fds_mot = (df_fds_r.groupby("motorista").size()
                                    .reset_index(name="reg_fds")
                                    .sort_values("reg_fds", ascending=True))
                tot_por = df_pt.groupby("motorista").size().reindex(fds_mot["motorista"]).fillna(1)
                fds_mot["pct_fds"] = (fds_mot["reg_fds"].values / tot_por.values * 100).round(1)
                fig_fds = go.Figure(go.Bar(
                    x=fds_mot["reg_fds"], y=fds_mot["motorista"], orientation="h",
                    marker=dict(
                        color=fds_mot["pct_fds"],
                        colorscale=[[0,"#3a1020"],[0.4,"#A0304A"],[1,"#FF6B6B"]],
                        line_width=0, colorbar=dict(title="%FDS", thickness=10,
                                                    tickfont=dict(color=_C["text"]),
                                                    title_font=dict(color=_C["text"])),
                    ),
                    text=[f"{r} reg · {p:.0f}%" for r, p in zip(fds_mot["reg_fds"], fds_mot["pct_fds"])],
                    textposition="outside",
                    textfont=dict(size=9, color=_C["text"]),
                    hovertemplate="<b>%{y}</b>: %{x} reg FDS<extra></extra>",
                ))
                fig_fds.update_layout(
                    **_BASE,
                    title=dict(text="🚨 Ranking Fim de Semana por Motorista",
                               font=dict(size=13, color=_C["text"]), x=0),
                    height=max(380, len(fds_mot) * 22),
                    xaxis=dict(gridcolor=_C["grid"], zeroline=False),
                    yaxis=dict(gridcolor=_C["grid"], tickfont=dict(size=9)),
                )
                st.plotly_chart(fig_fds, use_container_width=True, config=_NO_INTERACT)

    # ═══════════════════════════════════════════════════════════════════════════
    # SEÇÃO 7 — 🕐 HORÁRIOS + HEATMAPS
    # ═══════════════════════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("## 🕐 Quando Trabalham — Horários e Padrões")

    if not df_pt.empty and "hora" in df_pt.columns:
        col_h1, col_h2 = st.columns(2)

        with col_h1:
            hora_cnt = (df_pt.groupby("hora").size()
                             .reindex(range(24), fill_value=0).reset_index())
            hora_cnt.columns = ["hora","n"]
            pico = int(hora_cnt.loc[hora_cnt["n"].idxmax(), "hora"])
            fig_hora = go.Figure(go.Bar(
                x=hora_cnt["hora"], y=hora_cnt["n"],
                marker_color=[_C["acc1"] if h == pico else _C["eco135"] for h in hora_cnt["hora"]],
                marker_line_width=0,
                hovertemplate="<b>%{x}h</b>: %{y} registros<extra></extra>",
            ))
            fig_hora.update_layout(
                **_BASE,
                title=dict(text=f"🕐 Horários — pico às {pico}h",
                           font=dict(size=13, color=_C["text"]), x=0),
                height=280,
                xaxis=dict(tickmode="array", tickvals=list(range(0,24,2)),
                           ticktext=[f"{h}h" for h in range(0,24,2)],
                           gridcolor=_C["grid"], zeroline=False),
                yaxis=dict(gridcolor=_C["grid"], zeroline=False),
                bargap=0.1,
            )
            st.plotly_chart(fig_hora, use_container_width=True, config=_NO_INTERACT)

        with col_h2:
            # Heatmap: hora × dia_semana (toda frota)
            heatmap_data = (df_pt.groupby(["dia_semana","hora"]).size()
                                  .unstack(fill_value=0))
            z_mat = heatmap_data.reindex(range(7)).fillna(0).values.tolist()
            fig_hm = go.Figure(go.Heatmap(
                z=z_mat,
                x=[f"{h}h" for h in range(24)],
                y=dias_nome,
                colorscale=[[0,"#0D1B2A"],[0.3,"#1E4D3A"],[0.7,"#4A8A5A"],[1,"#7BBF6A"]],
                hovertemplate="<b>%{y} %{x}</b>: %{z} registros<extra></extra>",
                showscale=True,
                colorbar=dict(title="Registros", thickness=12,
                              tickfont=dict(color=_C["text"]),
                              title_font=dict(color=_C["text"])),
            ))
            fig_hm.update_layout(
                **_BASE,
                title=dict(text="🌡️ Heatmap: hora × dia (intensidade = atividade)",
                           font=dict(size=13, color=_C["text"]), x=0),
                height=280,
                xaxis=dict(tickfont=dict(size=9)),
                yaxis=dict(tickfont=dict(size=10)),
                margin=dict(l=60, r=12, t=36, b=12),
            )
            st.plotly_chart(fig_hm, use_container_width=True, config=_NO_INTERACT)

        # Heatmap: motorista × hora do dia (quem trabalha quando)
        top_mots = (df_pt.groupby("motorista").size().sort_values(ascending=False)
                         .head(20).index.tolist())
        df_top = df_pt[df_pt["motorista"].isin(top_mots)]
        hm2 = (df_top.groupby(["motorista","hora"]).size()
                      .unstack(fill_value=0)
                      .reindex(columns=range(24), fill_value=0))
        fig_hm2 = go.Figure(go.Heatmap(
            z=hm2.values.tolist(),
            x=[f"{h}h" for h in range(24)],
            y=[m.split()[0] if m else m for m in hm2.index.tolist()],
            colorscale=[[0,"#0D1B2A"],[0.3,"#132840"],[0.7,"#1E6B9E"],[1,"#4CC9F0"]],
            hovertemplate="<b>%{y} %{x}</b>: %{z} registros<extra></extra>",
            showscale=True,
            colorbar=dict(title="Reg.", thickness=12,
                          tickfont=dict(color=_C["text"]),
                          title_font=dict(color=_C["text"])),
        ))
        fig_hm2.update_layout(
            **_BASE,
            title=dict(text="👤 Quando cada motorista trabalha (top 20)",
                       font=dict(size=13, color=_C["text"]), x=0),
            height=max(380, len(top_mots) * 22),
            xaxis=dict(tickfont=dict(size=9)),
            yaxis=dict(tickfont=dict(size=9)),
            margin=dict(l=120, r=12, t=36, b=12),
        )
        st.plotly_chart(fig_hm2, use_container_width=True, config=_NO_INTERACT)

    # ═══════════════════════════════════════════════════════════════════════════
    # SEÇÃO 8 — 📍 ONDE ANDAM
    # ═══════════════════════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("## 📍 Onde Andam — Cidades e Regiões")

    if not df_pt.empty and "cidade" in df_pt.columns:
        col_c1, col_c2 = st.columns(2)

        with col_c1:
            cnt_cid = (df_pt[df_pt["cidade"] != ""]
                       .groupby(["cidade","uf"]).size()
                       .reset_index(name="n")
                       .sort_values("n", ascending=False)
                       .head(15))
            if not cnt_cid.empty:
                cnt_cid["label"] = cnt_cid["cidade"] + " · " + cnt_cid["uf"]
                fig_cid = go.Figure(go.Bar(
                    x=cnt_cid["n"].iloc[::-1], y=cnt_cid["label"].iloc[::-1],
                    orientation="h",
                    marker=dict(
                        color=cnt_cid["n"].iloc[::-1],
                        colorscale=[[0,"#132840"],[0.5,"#1E6B9E"],[1,"#4CC9F0"]],
                        line_width=0,
                    ),
                    text=cnt_cid["n"].iloc[::-1], textposition="outside",
                    textfont=dict(size=10, color=_C["text"]),
                    hovertemplate="<b>%{y}</b>: %{x} registros<extra></extra>",
                ))
                fig_cid.update_layout(
                    **_BASE,
                    title=dict(text="📍 Top 15 cidades mais frequentes",
                               font=dict(size=13, color=_C["text"]), x=0),
                    height=380,
                    xaxis=dict(gridcolor=_C["grid"], zeroline=False),
                    yaxis=dict(gridcolor="rgba(0,0,0,0)", tickfont=dict(size=9)),
                )
                st.plotly_chart(fig_cid, use_container_width=True, config=_NO_INTERACT)

        with col_c2:
            # Por estado
            cnt_uf = (df_pt[df_pt["uf"] != ""].groupby("uf").size()
                           .reset_index(name="n").sort_values("n", ascending=True))
            if not cnt_uf.empty:
                fig_uf = go.Figure(go.Bar(
                    x=cnt_uf["n"], y=cnt_uf["uf"], orientation="h",
                    marker=dict(
                        color=cnt_uf["n"],
                        colorscale=[[0,"#1a3a2a"],[0.5,"#4A8A5A"],[1,"#7BBF6A"]],
                        line_width=0,
                    ),
                    text=cnt_uf["n"], textposition="outside",
                    textfont=dict(size=11, color=_C["text"]),
                    hovertemplate="<b>%{y}</b>: %{x} registros<extra></extra>",
                ))
                fig_uf.update_layout(
                    **_BASE,
                    title=dict(text="🗺️ Distribuição por Estado (UF)",
                               font=dict(size=13, color=_C["text"]), x=0),
                    height=300,
                    xaxis=dict(gridcolor=_C["grid"], zeroline=False),
                    yaxis=dict(gridcolor="rgba(0,0,0,0)", tickfont=dict(size=11)),
                )
                st.plotly_chart(fig_uf, use_container_width=True, config=_NO_INTERACT)

    # ═══════════════════════════════════════════════════════════════════════════
    # SEÇÃO 9 — 🛣️ TABELAS COMPLETAS
    # ═══════════════════════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("## 📋 Tabelas Detalhadas")

    tab_t1, tab_t2, tab_t3 = st.tabs(["Resumo por Motorista", "Rotas", "Dias com Excesso de Km"])

    with tab_t1:
        if not df_mot.empty:
            df_tab_mot = df_mot.sort_values("custo_total", ascending=False).copy()
            df_tab_mot["idle_pct"] = df_tab_mot["idle_pct"].apply(lambda x: f"{x:.1f}%")
            df_tab_mot["fds_pct"]  = df_tab_mot["fds_pct"].apply(lambda x: f"{x:.1f}%")
            df_tab_mot = df_tab_mot.rename(columns={
                "motorista":"Motorista","km_periodo":"Km total",
                "h_rastreio":"H rastreio","h_mov":"H movimento","h_idle":"H idle",
                "idle_pct":"Idle %","fds_pct":"FDS %",
                "n_fds":"Reg FDS","efficiency":"km/h mov",
                "custo_cb":"Custo CB (R$)","custo_idle":"Custo Idle (R$)","custo_total":"Custo Total (R$)",
            })
            st.dataframe(df_tab_mot, use_container_width=True, hide_index=True, height=500)

    with tab_t2:
        df_rotas = df_p[df_p["rota_resumo"] != ""].sort_values("km_periodo", ascending=False)[
            ["motorista","placa","km_periodo","rota_resumo","estados","cidades"]
        ].rename(columns={
            "motorista":"Motorista","placa":"Placa","km_periodo":"Km",
            "rota_resumo":"Rota (resumo)","estados":"UFs","cidades":"N° Cidades",
        })
        st.dataframe(df_rotas, use_container_width=True, hide_index=True, height=400)

    with tab_t3:
        if not df_pt.empty and "data" in df_pt.columns and "odometro" in df_pt.columns:
            _dfo4 = df_pt[df_pt["odometro"] > 0]
            df_dv = (_dfo4.groupby(["motorista","data"])["odometro"].agg(["max","min"]).reset_index())
            df_dv["km_dia"] = (df_dv["max"] - df_dv["min"]).clip(lower=0, upper=1500)
            df_exc = df_dv[df_dv["km_dia"] > LIMIAR_KM_WA].sort_values("km_dia", ascending=False).copy()
            df_exc["status"] = df_exc["km_dia"].apply(
                lambda x: "🔴 ALARME >500km" if x > LIMIAR_KM_AL else "🟡 ATENÇÃO >300km"
            )
            df_exc = df_exc[["motorista","data","km_dia","status"]].rename(columns={
                "motorista":"Motorista","data":"Data","km_dia":"Km no dia","status":"Status",
            })
            st.dataframe(df_exc, use_container_width=True, hide_index=True, height=400)
        else:
            st.info("Dados diários insuficientes.")

    # ═══════════════════════════════════════════════════════════════════════════
    # SEÇÃO 10 — 🚀 VELOCIDADE: quem dirige rápido demais
    # ═══════════════════════════════════════════════════════════════════════════
    if not df_pt.empty and "velocidade" in df_pt.columns:
        st.markdown("---")
        st.markdown("## 🚀 Velocidade — Quem Corre Mais")
        st.caption("CTB Art. 61: limite 110 km/h pista dupla · 100 km/h pista simples · >120 km/h infração grave/gravíssima")

        df_vel = df_pt[df_pt["velocidade"] > 0]
        if not df_vel.empty:
            col_v1, col_v2 = st.columns(2)

            with col_v1:
                # Velocidade média por motorista
                vel_media = (df_vel.groupby("motorista")["velocidade"].mean()
                                    .reset_index().sort_values("velocidade", ascending=True))
                vel_media["velocidade"] = vel_media["velocidade"].round(1)
                cores_vel = [
                    _C["acc2"] if v > 110 else (_C["acc1"] if v > 90 else _C["eco135"])
                    for v in vel_media["velocidade"]
                ]
                fig_vel = go.Figure(go.Bar(
                    x=vel_media["velocidade"], y=vel_media["motorista"], orientation="h",
                    marker_color=cores_vel, marker_line_width=0,
                    text=[f"{v:.0f} km/h" for v in vel_media["velocidade"]],
                    textposition="outside",
                    textfont=dict(size=10, color=_C["text"]),
                    hovertemplate="<b>%{y}</b>: %{x:.1f} km/h média<extra></extra>",
                ))
                fig_vel.add_vline(x=110, line_dash="dash", line_color=_C["acc2"],
                                  annotation_text="110 km/h (pista dupla)",
                                  annotation_font_color=_C["acc2"])
                fig_vel.update_layout(
                    **_BASE,
                    title=dict(text="Velocidade Média por Motorista",
                               font=dict(size=13, color=_C["text"]), x=0),
                    height=max(400, len(vel_media) * 22),
                    xaxis=dict(gridcolor=_C["grid"], zeroline=False, ticksuffix=" km/h"),
                    yaxis=dict(gridcolor=_C["grid"], tickfont=dict(size=9)),
                )
                st.plotly_chart(fig_vel, use_container_width=True, config=_NO_INTERACT)

            with col_v2:
                # Registros acima de 110 km/h por motorista
                df_exc_vel = df_vel[df_vel["velocidade"] > 110]
                if not df_exc_vel.empty:
                    exc_cnt = (df_exc_vel.groupby("motorista")
                                         .agg(n=("velocidade","size"), vel_max=("velocidade","max"))
                                         .reset_index().sort_values("n", ascending=True))
                    exc_cnt["vel_max"] = exc_cnt["vel_max"].round(0).astype(int)
                    fig_exc = go.Figure(go.Bar(
                        x=exc_cnt["n"], y=exc_cnt["motorista"], orientation="h",
                        marker=dict(
                            color=exc_cnt["vel_max"],
                            colorscale=[[0,"#F7B731"],[0.5,"#FF6B6B"],[1,"#FF0000"]],
                            line_width=0,
                            colorbar=dict(title="Vel máx", thickness=10,
                                          tickfont=dict(color=_C["text"]),
                                          title_font=dict(color=_C["text"])),
                        ),
                        text=[f"{n}x · máx {m} km/h" for n, m in zip(exc_cnt["n"], exc_cnt["vel_max"])],
                        textposition="outside",
                        textfont=dict(size=9, color=_C["text"]),
                        hovertemplate="<b>%{y}</b>: %{x} reg >110 km/h<extra></extra>",
                    ))
                    fig_exc.update_layout(
                        **_BASE,
                        title=dict(text=f"🚨 Excesso >110 km/h — {len(df_exc_vel)} registros total",
                                   font=dict(size=13, color=_C["text"]), x=0),
                        height=max(320, len(exc_cnt) * 24),
                        xaxis=dict(gridcolor=_C["grid"], zeroline=False),
                        yaxis=dict(gridcolor=_C["grid"], tickfont=dict(size=9)),
                    )
                    st.plotly_chart(fig_exc, use_container_width=True, config=_NO_INTERACT)
                else:
                    st.success("Nenhum registro acima de 110 km/h no período.")

            # Histograma de velocidade da frota
            fig_hist_vel = go.Figure(go.Histogram(
                x=df_vel["velocidade"],
                nbinsx=50,
                marker_color=_C["cerrado"],
                marker_line_width=0,
                hovertemplate="%{x:.0f} km/h: %{y} registros<extra></extra>",
            ))
            fig_hist_vel.add_vline(x=110, line_dash="dash", line_color=_C["acc2"],
                                   annotation_text="Limite 110", annotation_font_color=_C["acc2"])
            fig_hist_vel.add_vline(x=80, line_dash="dot", line_color=_C["eco135"],
                                   annotation_text="Econômico 80", annotation_font_color=_C["eco135"])
            fig_hist_vel.update_layout(
                **_BASE,
                title=dict(text="📊 Distribuição de velocidade (toda frota)",
                           font=dict(size=13, color=_C["text"]), x=0),
                height=260,
                xaxis=dict(title="km/h", gridcolor=_C["grid"], zeroline=False),
                yaxis=dict(title="Registros", gridcolor=_C["grid"], zeroline=False),
            )
            st.plotly_chart(fig_hist_vel, use_container_width=True, config=_NO_INTERACT)

    # ═══════════════════════════════════════════════════════════════════════════
    # SEÇÃO 11 — 🌙 FORA DO HORÁRIO COMERCIAL (22h–6h)
    # ═══════════════════════════════════════════════════════════════════════════
    if not df_pt.empty and "hora" in df_pt.columns:
        df_noturno = df_pt[(df_pt["hora"] >= 22) | (df_pt["hora"] < 6)]
        if not df_noturno.empty:
            st.markdown("---")
            st.markdown("## 🌙 Atividade Noturna (22h–6h)")
            st.caption("Registros com ignição fora do horário comercial — possível uso não autorizado")

            col_n1, col_n2 = st.columns(2)

            with col_n1:
                not_mot = (df_noturno.groupby("motorista").size()
                                      .reset_index(name="reg_noturno")
                                      .sort_values("reg_noturno", ascending=True))
                tot_tot = df_pt.groupby("motorista").size()
                not_mot["pct_noturno"] = (not_mot["reg_noturno"].values /
                                          tot_tot.reindex(not_mot["motorista"]).fillna(1).values * 100).round(1)
                fig_not = go.Figure(go.Bar(
                    x=not_mot["reg_noturno"], y=not_mot["motorista"], orientation="h",
                    marker=dict(
                        color=not_mot["pct_noturno"],
                        colorscale=[[0,"#1a1a3a"],[0.5,"#4A3A8A"],[1,"#A29BFE"]],
                        line_width=0,
                        colorbar=dict(title="%", thickness=10,
                                      tickfont=dict(color=_C["text"]),
                                      title_font=dict(color=_C["text"])),
                    ),
                    text=[f"{n} reg · {p:.0f}%" for n, p in zip(not_mot["reg_noturno"], not_mot["pct_noturno"])],
                    textposition="outside",
                    textfont=dict(size=9, color=_C["text"]),
                    hovertemplate="<b>%{y}</b>: %{x} reg noturno<extra></extra>",
                ))
                fig_not.update_layout(
                    **_BASE,
                    title=dict(text=f"🌙 Ranking atividade noturna — {len(df_noturno)} reg total",
                               font=dict(size=13, color=_C["text"]), x=0),
                    height=max(320, len(not_mot) * 22),
                    xaxis=dict(gridcolor=_C["grid"], zeroline=False),
                    yaxis=dict(gridcolor=_C["grid"], tickfont=dict(size=9)),
                )
                st.plotly_chart(fig_not, use_container_width=True, config=_NO_INTERACT)

            with col_n2:
                # Distribuição horária do uso noturno
                hora_not = (df_noturno.groupby("hora").size()
                                       .reindex(list(range(22,24))+list(range(0,6)), fill_value=0)
                                       .reset_index())
                hora_not.columns = ["hora","n"]
                fig_hnot = go.Figure(go.Bar(
                    x=[f"{h}h" for h in hora_not["hora"]], y=hora_not["n"],
                    marker_color="#A29BFE", marker_line_width=0,
                    hovertemplate="<b>%{x}</b>: %{y} reg<extra></extra>",
                ))
                fig_hnot.update_layout(
                    **_BASE,
                    title=dict(text="🕐 Distribuição horária noturna",
                               font=dict(size=13, color=_C["text"]), x=0),
                    height=260,
                    xaxis=dict(gridcolor=_C["grid"], zeroline=False),
                    yaxis=dict(gridcolor=_C["grid"], zeroline=False),
                    bargap=0.15,
                )
                st.plotly_chart(fig_hnot, use_container_width=True, config=_NO_INTERACT)

    # ═══════════════════════════════════════════════════════════════════════════
    # SEÇÃO 12 — 🗺️ IDLE POR CIDADE: onde ficam parados com motor ligado
    # ═══════════════════════════════════════════════════════════════════════════
    if not df_pt.empty and "idle" in df_pt.columns and "cidade" in df_pt.columns:
        df_idle_cid = df_pt[(df_pt["idle"] == True) & (df_pt["cidade"] != "")]
        if not df_idle_cid.empty:
            st.markdown("---")
            st.markdown("## 🗺️ Onde Ficam Parados com Motor Ligado")
            st.caption("Pontos GPS com ignição ON + velocidade ≤3 km/h — cada ponto ≈3 min de idle")

            col_ic1, col_ic2 = st.columns(2)

            with col_ic1:
                idle_cid = (df_idle_cid.groupby(["cidade","uf"]).size()
                                        .reset_index(name="idle_pontos")
                                        .sort_values("idle_pontos", ascending=False)
                                        .head(15))
                idle_cid["horas_idle"] = (idle_cid["idle_pontos"] * MINS_POR_PONTO / 60).round(1)
                idle_cid["custo"] = (idle_cid["horas_idle"] * CUSTO_IDLE_H).round(0).astype(int)
                idle_cid["label"] = idle_cid["cidade"] + " · " + idle_cid["uf"]
                fig_icid = go.Figure(go.Bar(
                    x=idle_cid["horas_idle"].iloc[::-1],
                    y=idle_cid["label"].iloc[::-1],
                    orientation="h",
                    marker=dict(
                        color=idle_cid["horas_idle"].iloc[::-1],
                        colorscale=[[0,"#1a1a2a"],[0.5,"#8B2020"],[1,"#FF6B6B"]],
                        line_width=0,
                    ),
                    text=[f"{h:.0f}h · R${c}" for h, c in zip(
                        idle_cid["horas_idle"].iloc[::-1], idle_cid["custo"].iloc[::-1])],
                    textposition="outside",
                    textfont=dict(size=9, color=_C["text"]),
                    hovertemplate="<b>%{y}</b>: %{x:.1f}h idle<extra></extra>",
                ))
                fig_icid.update_layout(
                    **_BASE,
                    title=dict(text="🏙️ Top 15 cidades com mais idle (horas · custo est.)",
                               font=dict(size=12, color=_C["text"]), x=0),
                    height=380,
                    xaxis=dict(ticksuffix="h", gridcolor=_C["grid"], zeroline=False),
                    yaxis=dict(gridcolor="rgba(0,0,0,0)", tickfont=dict(size=9)),
                )
                st.plotly_chart(fig_icid, use_container_width=True, config=_NO_INTERACT)

            with col_ic2:
                # Idle por motorista × cidade (quem fica parado onde)
                idle_mc = (df_idle_cid.groupby(["motorista","cidade"]).size()
                                       .reset_index(name="n")
                                       .sort_values("n", ascending=False)
                                       .head(20))
                idle_mc["horas"] = (idle_mc["n"] * MINS_POR_PONTO / 60).round(1)
                idle_mc["label"] = idle_mc["motorista"].str.split().str[0] + " — " + idle_mc["cidade"]
                fig_imc = go.Figure(go.Bar(
                    x=idle_mc["horas"].iloc[::-1],
                    y=idle_mc["label"].iloc[::-1],
                    orientation="h",
                    marker=dict(
                        color=idle_mc["horas"].iloc[::-1],
                        colorscale=[[0,"#132840"],[0.5,"#8B4020"],[1,"#FF6B6B"]],
                        line_width=0,
                    ),
                    text=[f"{h:.1f}h" for h in idle_mc["horas"].iloc[::-1]],
                    textposition="outside",
                    textfont=dict(size=9, color=_C["text"]),
                    hovertemplate="<b>%{y}</b>: %{x:.1f}h idle<extra></extra>",
                ))
                fig_imc.update_layout(
                    **_BASE,
                    title=dict(text="👤🏙️ Motorista × Cidade com mais idle (Top 20)",
                               font=dict(size=12, color=_C["text"]), x=0),
                    height=max(380, len(idle_mc) * 22),
                    xaxis=dict(ticksuffix="h", gridcolor=_C["grid"], zeroline=False),
                    yaxis=dict(gridcolor="rgba(0,0,0,0)", tickfont=dict(size=8)),
                )
                st.plotly_chart(fig_imc, use_container_width=True, config=_NO_INTERACT)

    # ═══════════════════════════════════════════════════════════════════════════
    # SEÇÃO 13 — 🎯 RANKING FINAL: TOP 10 MOTORISTAS PROBLEMÁTICOS
    # ═══════════════════════════════════════════════════════════════════════════
    if not df_mot.empty:
        st.markdown("---")
        st.markdown("## 🎯 Ranking Final — Motoristas que Precisam de Atenção")
        st.caption(
            "Score de risco composto: (idle% × 3) + (fds% × 2) + (1 se km/dia > 300) + (excesso velocidade). "
            "Quanto maior o score, mais urgente a intervenção."
        )

        # Compute risk score
        df_risk = df_mot.copy()
        # idle peso 3
        df_risk["risk_idle"] = df_risk["idle_pct"].clip(upper=100) * 3
        # fds peso 2
        df_risk["risk_fds"]  = df_risk["fds_pct"] * 2
        # km excesso
        if not df_pt.empty and "data" in df_pt.columns and "odometro" in df_pt.columns:
            _dfoR = df_pt[df_pt["odometro"] > 0]
            if not _dfoR.empty:
                _dailyR = (_dfoR.groupby(["motorista","data"])["odometro"]
                                 .agg(["max","min"]).reset_index())
                _dailyR["km_d"] = (_dailyR["max"] - _dailyR["min"]).clip(lower=0, upper=1500)
                _avgR = _dailyR.groupby("motorista")["km_d"].mean().reset_index()
                _avgR.columns = ["motorista","km_dia_avg"]
                df_risk = df_risk.merge(_avgR, on="motorista", how="left")
                df_risk["km_dia_avg"] = df_risk["km_dia_avg"].fillna(0)
                df_risk["risk_km"] = (df_risk["km_dia_avg"] > LIMIAR_KM_WA).astype(float) * 50
            else:
                df_risk["km_dia_avg"] = 0.0
                df_risk["risk_km"]    = 0.0
        else:
            df_risk["km_dia_avg"] = 0.0
            df_risk["risk_km"]    = 0.0

        # velocidade excesso
        if not df_pt.empty and "velocidade" in df_pt.columns:
            _velR = (df_pt[df_pt["velocidade"] > 110]
                          .groupby("motorista").size()
                          .reset_index(name="n_excesso_vel"))
            df_risk = df_risk.merge(_velR, on="motorista", how="left")
            df_risk["n_excesso_vel"] = df_risk["n_excesso_vel"].fillna(0)
            df_risk["risk_vel"] = df_risk["n_excesso_vel"].clip(upper=50) * 2
        else:
            df_risk["n_excesso_vel"] = 0
            df_risk["risk_vel"]      = 0.0

        # noturno
        if not df_pt.empty and "hora" in df_pt.columns:
            _notR = (df_pt[(df_pt["hora"] >= 22) | (df_pt["hora"] < 6)]
                          .groupby("motorista").size()
                          .reset_index(name="n_noturno"))
            df_risk = df_risk.merge(_notR, on="motorista", how="left")
            df_risk["n_noturno"] = df_risk["n_noturno"].fillna(0)
            df_risk["risk_not"] = df_risk["n_noturno"].clip(upper=50)
        else:
            df_risk["n_noturno"] = 0
            df_risk["risk_not"]  = 0.0

        df_risk["risk_score"] = (df_risk["risk_idle"] + df_risk["risk_fds"]
                                 + df_risk["risk_km"] + df_risk["risk_vel"]
                                 + df_risk["risk_not"])
        df_risk = df_risk.sort_values("risk_score", ascending=False)

        # Top 10
        top10 = df_risk.head(10).copy()
        fig_risk = go.Figure(go.Bar(
            x=top10["risk_score"].iloc[::-1],
            y=top10["motorista"].iloc[::-1],
            orientation="h",
            marker=dict(
                color=top10["risk_score"].iloc[::-1],
                colorscale=[[0,"#1a3a2a"],[0.3,"#F7B731"],[0.6,"#FF6B6B"],[1,"#FF0000"]],
                line_width=0,
            ),
            hovertemplate="<b>%{y}</b>: score %{x:.0f}<extra></extra>",
        ))
        fig_risk.update_layout(
            **_BASE,
            title=dict(text="🎯 Top 10 — Score de Risco Composto",
                       font=dict(size=14, color=_C["text"]), x=0),
            height=max(350, len(top10) * 30),
            xaxis=dict(title="Score de risco", gridcolor=_C["grid"], zeroline=False),
            yaxis=dict(gridcolor=_C["grid"], tickfont=dict(size=10)),
        )
        st.plotly_chart(fig_risk, use_container_width=True, config=_NO_INTERACT)

        # Detalhes dos top 10
        st.markdown("##### Detalhamento Top 10")
        tab_risk = top10[[
            "motorista","risk_score","km_periodo","idle_pct","h_idle",
            "fds_pct","n_fds","km_dia_avg","n_excesso_vel","n_noturno",
            "custo_total"
        ]].copy()
        tab_risk = tab_risk.rename(columns={
            "motorista":"Motorista","risk_score":"Score","km_periodo":"Km total",
            "idle_pct":"Idle %","h_idle":"H idle","fds_pct":"FDS %",
            "n_fds":"Reg FDS","km_dia_avg":"Km/dia médio",
            "n_excesso_vel":"Reg >110km/h","n_noturno":"Reg noturno",
            "custo_total":"Custo Est. (R$)",
        })
        tab_risk["Km/dia médio"] = tab_risk["Km/dia médio"].round(0)
        tab_risk["Score"] = tab_risk["Score"].round(0)
        st.dataframe(tab_risk, use_container_width=True, hide_index=True)

        # Cards dos 3 piores
        st.markdown("##### Os 3 que mais precisam de atenção:")
        worst3 = top10.head(3)
        cols_w = st.columns(3)
        for i, (_, row) in enumerate(worst3.iterrows()):
            with cols_w[i]:
                problemas = []
                if row["idle_pct"] > LIMIAR_IDLE_AL:
                    problemas.append(f"🔴 Idle {row['idle_pct']:.0f}%")
                if row["fds_pct"] > 0:
                    problemas.append(f"📅 FDS {row['fds_pct']:.0f}%")
                if row.get("km_dia_avg", 0) > LIMIAR_KM_WA:
                    problemas.append(f"🚗 {row['km_dia_avg']:.0f} km/dia")
                if row.get("n_excesso_vel", 0) > 0:
                    problemas.append(f"🚀 {int(row['n_excesso_vel'])}× >110km/h")
                if row.get("n_noturno", 0) > 0:
                    problemas.append(f"🌙 {int(row['n_noturno'])} reg noturno")
                prob_html = "<br>".join(problemas) if problemas else "Sem infrações graves"
                medal = ["🥇","🥈","🥉"][i]
                st.markdown(f"""
                <div style="background:rgba(255,70,70,0.10);border:1px solid #FF6B6B55;
                            border-radius:10px;padding:14px;min-height:180px">
                  <div style="font-size:1.1rem;font-weight:700;color:#FF6B6B">
                    {medal} {row['motorista']}</div>
                  <div style="font-size:.85rem;color:#F7B731;margin:4px 0">
                    Score: {row['risk_score']:.0f} · Custo: R$ {row['custo_total']:,.0f}</div>
                  <div style="font-size:.8rem;color:#C8D8A8;line-height:1.5">{prob_html}</div>
                </div>""", unsafe_allow_html=True)


@st.fragment
def _aba_rastreamento():
    atu = st.session_state.get("logos_ultima_atualizacao")
    c1, c2 = st.columns([5, 1])
    with c1:
        if atu:
            n = len(st.session_state.get("logos_veiculos", []))
            st.caption(f"✅ {n} veículos ECO · Atualizado: {atu}")
    with c2:
        atualizar = st.button("🔄 Atualizar", key="logos_btn", use_container_width=True)

    if atualizar:
        with st.spinner("Conectando ao Logos..."):
            try:
                sess, idcli = _logos_login()
                veiculos = _logos_get_eco(sess, idcli)
                if not veiculos:
                    st.warning("Nenhum veículo ECO encontrado.")
                    return
                st.session_state["logos_veiculos"]           = veiculos
                st.session_state["logos_ultima_atualizacao"] = datetime.now().strftime("%d/%m/%Y %H:%M")
                st.session_state.pop("logos_rota", None)
                st.session_state.pop("logos_periodo_result", None)
            except Exception as e:
                st.error(f"❌ {e}")
                return

    veiculos = st.session_state.get("logos_veiculos", [])
    if not veiculos:
        st.info("Clique em **🔄 Atualizar** para buscar os veículos ECO do Logos Rastreamento.")
        return

    itens = [_parse_eco(v, i) for i, v in enumerate(veiculos)]

    tab_pos, tab_stats, tab_rota, tab_periodo = st.tabs([
        "📍 Posição Atual",
        "📊 Estatísticas",
        "🛣️ Rota Individual",
        "📅 Análise de Período",
    ])
    with tab_pos:
        _render_mapa_posicao(itens)
    with tab_stats:
        _render_estatisticas(itens)
    with tab_rota:
        _render_rota_individual(itens)
    with tab_periodo:
        _render_analise_periodo(itens)


# =============================================================================
def main():
    _sidebar()

    st.markdown("""
    <div class="eco-header">
        <h1>🛣️ Eco Rodovias — Contrato 6771</h1>
        <p>BR-050 (Eco Minas Goiás) · BR-365 (Eco Cerrado) · Supervisão de Obras AFIRMA E-VIAS</p>
    </div>""", unsafe_allow_html=True)

    tab_checklist, tab_ensaios, tab_rastr = st.tabs([
        "📋 Checklist APP",
        "🔬 Ensaios AEVIAS",
        "🛰️ Rastreamento",
    ])

    with tab_checklist:
        _aba_checklist()

    with tab_ensaios:
        _aba_ensaios()

    with tab_rastr:
        _aba_rastreamento()


if __name__ == "__main__" or True:
    main()
