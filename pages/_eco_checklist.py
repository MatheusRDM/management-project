"""
_eco_checklist.py — checklist tab for ECO Rodovias.
"""
import sys
import os

_PAGES_DIR = os.path.dirname(os.path.abspath(__file__))
_ROOT_DIR  = os.path.dirname(_PAGES_DIR)
if _PAGES_DIR not in sys.path: sys.path.insert(0, _PAGES_DIR)
if _ROOT_DIR  not in sys.path: sys.path.insert(0, _ROOT_DIR)

import streamlit as st
import json
import time
from datetime import datetime, date

from _eco_shared import (
    COR_PRIMARY, COR_ACCENT, COR_BG, COR_CARD, COR_BORDER,
    COR_TEXT, COR_MUTED, COR_OK, COR_COBRAR, COR_NE, COR_ELAB,
    PLOTLY_LAYOUT, PLOTLY_CONFIG,
    _BASE_DIR, _CACHE_DIR, _Y_BASE, _IS_CLOUD,
)


# =============================================================================
# CARREGAMENTO DE DADOS
# =============================================================================

@st.cache_data(ttl=3600, show_spinner=False)
def _carregar_checklist_cache() -> dict:
    p = os.path.join(_CACHE_DIR, "eco_checklist.json")
    if os.path.exists(p):
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    return {}


@st.cache_data(ttl=3600, show_spinner=False)
def _carregar_checklist_y(med_folder: str) -> dict | None:
    """Lê o xlsx de controle diretamente do Y:."""
    import openpyxl
    from glob import glob
    folder = os.path.join(_Y_BASE, med_folder)
    if not os.path.exists(folder):
        return None
    # Busca arquivo xlsx com "controle" e "app" no nome
    hits = [f for f in os.listdir(folder)
            if f.lower().endswith(".xlsx")
            and "controle" in f.lower()
            and not f.startswith("~")]
    if not hits:
        return None
    path = os.path.join(folder, hits[0])
    wb = openpyxl.load_workbook(path)
    result = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue
        header = rows[0]
        dates = []
        for v in header[2:]:
            if isinstance(v, datetime):
                dates.append(v.strftime("%Y-%m-%d"))
            else:
                dates.append(None)
        people = []
        for row in rows[2:]:
            if not row[0]:
                continue
            entry = {
                "colaborador": str(row[0]),
                "funcao": str(row[1]) if row[1] else "",
                "dias": {},
            }
            for i, d in enumerate(dates):
                if d and (i + 2) < len(row):
                    v = row[i + 2]
                    entry["dias"][d] = str(v) if v else None
            people.append(entry)
        result[sn] = people
    return result


@st.cache_data(ttl=3600, show_spinner=False)
def _listar_meds() -> list[str]:
    cache = _carregar_checklist_cache()
    local_meds = list(cache.keys())
    if not _IS_CLOUD and os.path.exists(_Y_BASE):
        y_meds = [d for d in sorted(os.listdir(_Y_BASE)) if d.startswith("MED")]
        # Merge, mantendo cache + novos do Y
        all_meds = sorted(set(local_meds + y_meds))
        return all_meds
    return local_meds


# =============================================================================
# COMPONENTES VISUAIS
# =============================================================================

def _kpi_card(val, label, cor="#BFCF99"):
    return f"""
    <div class="eco-kpi">
        <div class="val" style="color:{cor}">{val}</div>
        <div class="lbl">{label}</div>
    </div>"""


import unicodedata as _ucd

def _norm(s: str) -> str:
    """Lowercase + remove acentos + strip — para comparação robusta."""
    s = _ucd.normalize("NFD", s.lower().strip())
    return "".join(c for c in s if _ucd.category(c) != "Mn")

# Funções isentas de checklist de campo (sem acentos — comparado via _norm)
_FUNCOES_ISENTAS_NORM = {
    "assistente de engenharia",
    "encarregado sala tecnica",
    "encarregado de sala tecnica",
    "desenhista",
    "engenheiro sala tecnica",
    "engenheiro de sala tecnica",
}
# Substrings que indicam isenção (sem acentos)
_ISENTOS_SUBSTR = ("auxiliar",)

def _isento_checklist(funcao: str) -> bool:
    """True se o cargo não exige checklist de campo."""
    f = _norm(funcao)
    if f in _FUNCOES_ISENTAS_NORM:
        return True
    if any(sub in f for sub in _ISENTOS_SUBSTR):
        return True
    return False


def _status_class(v):
    if v is None or v == "":
        return "status-vazio", "·"
    vu = str(v).upper().strip()
    if vu == "OK":
        return "status-ok", "OK"
    if vu in ("COBRAR", "COBRE"):
        return "status-cobrar", "COB"
    if vu in ("N/E", "NE"):
        return "status-ne", "N/E"
    if vu in ("ELAB.", "ELAB"):
        return "status-elab", "ELB"
    return "status-vazio", v[:3] if v else "·"


_CSS_CARDS = """
<style>
@keyframes pulse-dot {
  0%,100%{opacity:1;transform:scale(1)}
  50%{opacity:.35;transform:scale(.8)}
}
.ck-wrap{padding:0 2px}
.ck-grid{display:flex;flex-direction:column;gap:10px}
@media(min-width:620px){
  .ck-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(320px,1fr));gap:12px}
}
.ck-card{background:rgba(13,27,42,.75);border:1px solid rgba(86,110,61,.3);
  border-radius:12px;padding:14px 14px 12px;border-left:4px solid #566E3D;
  transition:border-color .15s}
.ck-card.card-cob{border-left-color:#e6194b}
.ck-card.card-ok{border-left-color:#3cb44b}
.ck-card.card-ne{border-left-color:#3a4a5e}
.ck-name{font-size:.92rem;font-weight:700;color:#E8EFD8;
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.ck-role{font-size:.7rem;color:#8FA882;margin-bottom:10px;
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.ck-days{display:flex;gap:5px;flex-wrap:nowrap;overflow-x:auto;
  padding-bottom:2px;scrollbar-width:none}
.ck-days::-webkit-scrollbar{display:none}
.ck-pill{display:flex;flex-direction:column;align-items:center;gap:2px;
  min-width:34px;flex-shrink:0}
.ck-dd{font-size:.6rem;color:#8FA882;font-weight:500}
.ck-badge{display:inline-flex;align-items:center;justify-content:center;
  width:34px;height:26px;border-radius:6px;font-size:.65rem;font-weight:700;
  letter-spacing:.02em}
.ck-badge.b-ok{background:rgba(60,180,75,.25);color:#3cb44b;border:1px solid rgba(60,180,75,.5)}
.ck-badge.b-cob{background:rgba(230,25,75,.25);color:#e6194b;border:1px solid rgba(230,25,75,.5)}
.ck-badge.b-ne{background:rgba(58,74,94,.5);color:#7a90a8;border:1px solid rgba(58,74,94,.7)}
.ck-badge.b-elab{background:rgba(67,99,216,.25);color:#4363d8;border:1px solid rgba(67,99,216,.5)}
.ck-badge.b-vazio{background:rgba(255,255,255,.04);color:#4a5a6a;border:1px dashed #2D3748}
.ck-badge.b-hoje-ok{background:rgba(60,180,75,.35);color:#3cb44b;
  border:2px solid #3cb44b;box-shadow:0 0 6px rgba(60,180,75,.4)}
.ck-badge.b-hoje-cob{background:rgba(230,25,75,.35);color:#e6194b;
  border:2px solid #e6194b;box-shadow:0 0 6px rgba(230,25,75,.4)}
.ck-badge.b-hoje-sem{background:rgba(247,183,49,.12);color:#F7B731;
  border:2px dashed #F7B731}
.ck-dot{width:7px;height:7px;border-radius:50%;background:#F7B731;
  animation:pulse-dot 1.4s ease-in-out infinite;margin:0 auto}
.ck-foot{display:flex;gap:8px;margin-top:9px;flex-wrap:wrap}
.ck-stat{font-size:.68rem;padding:2px 8px;border-radius:999px;font-weight:600}
.ck-stat.s-ok{background:rgba(60,180,75,.2);color:#3cb44b}
.ck-stat.s-cob{background:rgba(230,25,75,.2);color:#e6194b}
.ck-stat.s-ne{background:rgba(58,74,94,.5);color:#7a90a8}
.ck-sep{height:1px;background:rgba(255,255,255,.05);margin:14px 0}
</style>
"""

_DAY_ABBR = {0:"SEG",1:"TER",2:"QUA",3:"QUI",4:"SEX",5:"SÁB",6:"DOM"}

def _badge_class(v: str | None, is_hoje: bool) -> tuple[str, str]:
    """Retorna (css_class, texto) para um badge de status."""
    if v is None or str(v).strip() == "":
        if is_hoje:
            return "b-hoje-sem", "?"
        return "b-vazio", "·"
    vu = str(v).upper().strip()
    if vu == "OK":
        return ("b-hoje-ok" if is_hoje else "b-ok"), "OK"
    if vu in ("COBRAR","COBRE"):
        return ("b-hoje-cob" if is_hoje else "b-cob"), "COB"
    if vu in ("N/E","NE"):
        return "b-ne", "N/E"
    if vu in ("ELAB.","ELAB"):
        return "b-elab", "ELB"
    return "b-vazio", v[:3] if v else "·"


def _renderizar_cards(people: list[dict], datas_janela: list[str]):
    """Renderiza cards mobile-first com janela de 7 dias."""
    today_str = date.today().strftime("%Y-%m-%d")

    def _urgency(p):
        dias = p.get("dias", {})
        v_hoje = dias.get(today_str)
        vu = str(v_hoje).upper().strip() if v_hoje else ""
        if vu in ("COBRAR","COBRE"): return 0   # pior: cobrança hoje
        if not v_hoje:               return 1   # sem registro hoje
        if vu in ("N/E","NE"):       return 2
        return 3                                # OK = último

    def _tem_dado_no_periodo(p):
        """True se a pessoa tem QUALQUER registro não-nulo na janela."""
        dias = p.get("dias", {})
        return any(v for d, v in dias.items() if d in datas_janela and v)

    # Separa: exige checklist × isento por cargo × sem dados no período
    isentos_cargo  = [p for p in people if _isento_checklist(p.get("funcao",""))]
    com_checklist  = [p for p in people if not _isento_checklist(p.get("funcao",""))]
    sem_dados      = [p for p in com_checklist if not _tem_dado_no_periodo(p)]
    obrigados      = [p for p in com_checklist if _tem_dado_no_periodo(p)]
    isentos        = isentos_cargo  # para o bloco visual abaixo
    people_sorted  = sorted(obrigados, key=_urgency)

    cards_html = ['<div class="ck-wrap">', _CSS_CARDS, '<div class="ck-grid">']

    for p in people_sorted:
        dias      = p.get("dias", {})
        ok_count  = sum(1 for d,v in dias.items()
                        if v and str(v).upper().strip()=="OK" and d in datas_janela)
        cob_count = sum(1 for d,v in dias.items()
                        if v and str(v).upper().strip() in ("COBRAR","COBRE"))
        ne_count  = sum(1 for d,v in dias.items()
                        if v and str(v).upper().strip() in ("N/E","NE") and d in datas_janela)
        v_hoje    = dias.get(today_str)
        vu_hoje   = str(v_hoje).upper().strip() if v_hoje else ""

        card_cls = "card-cob" if cob_count > 0 else ("card-ok" if ok_count > 0 else "card-ne")

        # Pills dos dias
        pills_html = ""
        for d in datas_janela:
            try:
                dt_obj = datetime.strptime(d, "%Y-%m-%d")
            except Exception:
                continue
            is_hoje = (d == today_str)
            v       = dias.get(d)
            b_cls, b_txt = _badge_class(v, is_hoje)

            dot = ""
            if is_hoje and not v:
                dot = '<div class="ck-dot"></div>'

            dd_label = ("HOJE" if is_hoje
                        else f"{_DAY_ABBR[dt_obj.weekday()]} {dt_obj.day:02d}")
            dd_style = ("font-weight:700;color:#F7B731" if is_hoje else "")

            pills_html += (
                f'<div class="ck-pill">'
                f'<span class="ck-dd" style="{dd_style}">{dd_label}</span>'
                f'<span class="ck-badge {b_cls}">{b_txt}</span>'
                f'{dot}'
                f'</div>'
            )

        # Footer stats
        stats = f'<span class="ck-stat s-ok">✓ {ok_count} OK</span>'
        if cob_count:
            stats += f'<span class="ck-stat s-cob">⚠ {cob_count} pend.</span>'
        if ne_count:
            stats += f'<span class="ck-stat s-ne">N/E {ne_count}</span>'

        nome   = p.get("colaborador","—")
        funcao = p.get("funcao","—")

        cards_html.append(f"""
        <div class="ck-card {card_cls}">
          <div class="ck-name">{nome}</div>
          <div class="ck-role">{funcao}</div>
          <div class="ck-days">{pills_html}</div>
          <div class="ck-foot">{stats}</div>
        </div>""")

    cards_html.append("</div></div>")
    st.markdown("".join(cards_html), unsafe_allow_html=True)

    # ── Rodapé: isentos por cargo ─────────────────────────────────────────────
    if isentos_cargo:
        nomes_isentos = " · ".join(
            f"{p.get('colaborador','').split()[0]}"
            f"<span style='color:#3a4a5e;font-size:.6rem'> ({p.get('funcao','—')})</span>"
            for p in sorted(isentos_cargo, key=lambda x: x.get("colaborador",""))
        )
        st.markdown(
            f'<div style="margin-top:6px;padding:7px 12px;'
            f'background:rgba(58,74,94,.2);border-radius:8px;'
            f'border-left:3px solid #3a4a5e;color:#5a6a7e;font-size:.75rem">'
            f'<span style="font-weight:600;color:#8FA882">🚫 Sem checklist de campo:</span> '
            f'{nomes_isentos}</div>',
            unsafe_allow_html=True
        )

    # ── Rodapé: sem dados no período (ex: afastados, novos) ──────────────────
    if sem_dados:
        nomes_sd = " · ".join(
            p.get("colaborador","").split()[0]
            for p in sorted(sem_dados, key=lambda x: x.get("colaborador",""))
        )
        st.markdown(
            f'<div style="margin-top:4px;padding:6px 12px;'
            f'background:rgba(40,40,40,.2);border-radius:8px;'
            f'border-left:3px solid #2D3748;color:#4a5568;font-size:.72rem">'
            f'<span style="font-weight:600;color:#4a5568">⚫ Sem registro no período ({len(sem_dados)}):</span> '
            f'{nomes_sd}</div>',
            unsafe_allow_html=True
        )


def _renderizar_calendario(people: list[dict], mes_ref: str):
    """View principal: cards mobile-first (7 dias) + tabela completa em expander."""
    if not people:
        st.warning("Nenhum colaborador encontrado.")
        return

    today      = date.today()
    today_str  = today.strftime("%Y-%m-%d")

    # ── Filtra isentos e sem dados ANTES de qualquer renderização ─────────────
    isentos_cargo_cal = [p for p in people if     _isento_checklist(p.get("funcao",""))]
    campo_people      = [p for p in people if not _isento_checklist(p.get("funcao",""))]

    # Coleta datas do mês que já passaram (sem futuro)
    datas_mes = sorted(
        d for p in campo_people for d in p.get("dias", {}).keys()
        if d and d <= today_str
    )
    datas_mes = sorted(set(datas_mes))

    if not datas_mes:
        st.info("Sem datas registradas até hoje.")
        return

    # Pessoas com pelo menos 1 registro no mês
    def _tem_dado_mes(p):
        return any(v for d,v in p.get("dias",{}).items() if d in datas_mes and v)

    sem_dados_cal = [p for p in campo_people if not _tem_dado_mes(p)]
    people_ativos = [p for p in campo_people if _tem_dado_mes(p)]

    # Janela: últimos 7 dias disponíveis (incluindo hoje se existir)
    datas_janela = datas_mes[-7:]

    # ── View rápida: cards 7 dias (só ativos) ────────────────────────────────
    _renderizar_cards(people_ativos, datas_janela)

    # ── Rodapé compacto ───────────────────────────────────────────────────────
    rodape = []
    if isentos_cargo_cal:
        nomes = " · ".join(
            f"<span style='color:#5a6a7e'>{p['colaborador'].split()[0]}"
            f"<span style='font-size:.6rem;color:#3a4a5e'> ({p.get('funcao','—')})</span></span>"
            for p in sorted(isentos_cargo_cal, key=lambda x: x.get("colaborador",""))
        )
        rodape.append(f"🚫 <b style='color:#8FA882'>Sem checklist:</b> {nomes}")
    if sem_dados_cal:
        nomes = " · ".join(
            f"<span style='color:#4a5568'>{p['colaborador'].split()[0]}</span>"
            for p in sorted(sem_dados_cal, key=lambda x: x.get("colaborador",""))
        )
        rodape.append(f"⚫ <b style='color:#4a5568'>Sem registro ({len(sem_dados_cal)}):</b> {nomes}")
    if rodape:
        st.markdown(
            '<div style="margin:6px 0;padding:7px 12px;background:rgba(30,30,40,.3);'
            'border-radius:8px;font-size:.72rem;line-height:1.8">'
            + "<br>".join(rodape) + "</div>",
            unsafe_allow_html=True
        )

    st.markdown('<div class="ck-sep"></div>', unsafe_allow_html=True)

    # ── Expander: tabela completa do mês (só ativos de campo) ────────────────
    n_ativos = len(people_ativos)
    with st.expander(f"📊 Calendário completo — {n_ativos} pessoas em campo · {len(datas_mes)} dias", expanded=False):
        DAY_ABBR = _DAY_ABBR
        html = ['<div class="cal-wrap"><table class="cal-table"><thead><tr>']
        html.append('<th>Colaborador</th><th>Função</th>')
        for d in datas_mes:
            dt = datetime.strptime(d, "%Y-%m-%d")
            is_hj = (d == today_str)
            style = "color:#F7B731;font-weight:700" if is_hj else ""
            html.append(f'<th style="{style}">{dt.day:02d}</th>')
        html.append('<th>OK</th><th>COB</th></tr><tr><th></th><th></th>')
        for d in datas_mes:
            dt = datetime.strptime(d, "%Y-%m-%d")
            html.append(f'<th style="font-size:.55rem;color:#8FA882">'
                        f'{DAY_ABBR[dt.weekday()]}</th>')
        html.append('<th></th><th></th></tr></thead><tbody>')

        for p in people_ativos:
            dias      = p.get("dias", {})
            ok_count  = sum(1 for d,v in dias.items()
                            if v and str(v).upper().strip()=="OK" and d in datas_mes)
            cob_count = sum(1 for d,v in dias.items()
                            if v and str(v).upper().strip() in ("COBRAR","COBRE"))
            html.append("<tr>")
            html.append(f'<td class="colab">{p["colaborador"]}</td>')
            html.append(f'<td class="funcao">{p["funcao"]}</td>')
            for d in datas_mes:
                v = dias.get(d)
                cls, txt = _status_class(v)
                sty = "outline:2px solid #F7B731" if d==today_str else ""
                html.append(f'<td class="{cls}" style="{sty}">{txt}</td>')
            c1 = "#3cb44b" if ok_count  > 0 else "#7a90a8"
            c2 = "#e6194b" if cob_count > 0 else "#7a90a8"
            html.append(f'<td style="color:{c1};font-weight:700">{ok_count}</td>')
            html.append(f'<td style="color:{c2};font-weight:700">'
                        f'{cob_count if cob_count else "—"}</td>')
            html.append("</tr>")
        html.append("</tbody></table></div>")
        st.markdown("".join(html), unsafe_allow_html=True)


# =============================================================================
# ABA CHECKLIST
# =============================================================================

# =============================================================================
# ENSAIOS AEVIAS — fonte: aevias-controle.base44.app (cache local)
# =============================================================================

_ENSAIOS_PATH  = os.path.join(_CACHE_DIR, "ensaios_aevias.json")
_BASE44_URL    = "https://aevias-controle.base44.app"

# Cores por obra
_COR_OBRA = {
    "SST":              "#F7B731",
    "Pavimento":        "#7BBF6A",
    "TOPOGRAFIA":       "#4CC9F0",
    "OAE / Terraplenos":"#A29BFE",
    "Ampliações":       "#FD79A8",
    "Conserva":         "#00CEC9",
    "ESCRITÓRIO":       "#FDCB6E",
}
# Ícones por tipo
_ICON_TIPO = {
    "Diário de Obra":        "📋",
    "Checklist de Usina":    "🏭",
    "Checklist de Aplicação":"🚧",
    "Checklist de MRAF":     "🔩",
    "Ensaio de CAUQ":        "🧪",
}


@st.cache_data(ttl=300, show_spinner=False)  # auto-refresh a cada 5 min
def _carregar_ensaios() -> tuple[list, float]:
    """Carrega ensaios_aevias.json. Retorna (lista, timestamp_modificação)."""
    if not os.path.exists(_ENSAIOS_PATH):
        return [], 0.0
    mtime = os.path.getmtime(_ENSAIOS_PATH)
    with open(_ENSAIOS_PATH, encoding="utf-8") as f:
        dados = json.load(f)
    return dados, mtime


_CSS_PROD = """
<style>
.prod-wrap{padding:0 2px}
.prod-grid{display:flex;flex-direction:column;gap:10px}
@media(min-width:640px){
  .prod-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(340px,1fr));gap:12px}
}
.prod-card{background:rgba(13,27,42,.8);border:1px solid rgba(86,110,61,.3);
  border-radius:12px;padding:14px;border-left:4px solid #566E3D}
.prod-card.pc-ok{border-left-color:#3cb44b}
.prod-card.pc-pend{border-left-color:#F7B731}
.prod-card.pc-rep{border-left-color:#e6194b}
.prod-name{font-size:.92rem;font-weight:700;color:#E8EFD8;margin-bottom:2px}
.prod-sub{font-size:.7rem;color:#8FA882;margin-bottom:10px}
.prod-days{display:flex;gap:4px;flex-wrap:nowrap;overflow-x:auto;
  padding-bottom:4px;scrollbar-width:none}
.prod-days::-webkit-scrollbar{display:none}
.prod-pill{display:flex;flex-direction:column;align-items:center;gap:2px;
  min-width:44px;flex-shrink:0;cursor:default}
.prod-dd{font-size:.58rem;color:#8FA882;font-weight:500;text-align:center}
.prod-cell{width:44px;min-height:28px;border-radius:6px;font-size:.6rem;
  font-weight:700;display:flex;flex-direction:column;align-items:center;
  justify-content:center;gap:1px;padding:2px 1px;text-align:center;line-height:1.1}
.prod-cell.c-ok{background:rgba(60,180,75,.25);color:#3cb44b;border:1px solid rgba(60,180,75,.4)}
.prod-cell.c-pend{background:rgba(247,183,49,.2);color:#F7B731;border:1px solid rgba(247,183,49,.4)}
.prod-cell.c-rep{background:rgba(230,25,75,.2);color:#e6194b;border:1px solid rgba(230,25,75,.4)}
.prod-cell.c-exec{background:rgba(67,99,216,.2);color:#4363d8;border:1px solid rgba(67,99,216,.4)}
.prod-cell.c-vazio{background:rgba(255,255,255,.03);color:#3a4a5e;border:1px dashed #2D3748}
.prod-cell.c-hoje{outline:2px solid #F7B731;outline-offset:1px}
.prod-emp{font-size:.55rem;color:inherit;opacity:.8;overflow:hidden;
  max-width:42px;text-overflow:ellipsis;white-space:nowrap}
.prod-foot{display:flex;gap:6px;flex-wrap:wrap;margin-top:8px}
.prod-tag{font-size:.65rem;padding:2px 7px;border-radius:999px;font-weight:600}
.prod-tag.t-ok{background:rgba(60,180,75,.2);color:#3cb44b}
.prod-tag.t-pend{background:rgba(247,183,49,.2);color:#F7B731}
.prod-tag.t-rep{background:rgba(230,25,75,.2);color:#e6194b}
</style>
"""

# Mapeamento obra → categoria de exibição
_CATEGORIAS_PROD = {
    "Diário de Obra": {
        "obras":  None,          # todas as obras
        "tipos":  {"Diário de Obra"},
        "icon":   "📋",
        "cor":    "#4CC9F0",
    },
    "SST": {
        "obras":  {"SST"},
        "tipos":  None,
        "icon":   "🦺",
        "cor":    "#F7B731",
    },
    "Pavimento": {
        "obras":  {"Pavimento"},
        "tipos":  None,
        "icon":   "🏗️",
        "cor":    "#7BBF6A",
    },
    "OAE / Terraplenos": {
        "obras":  {"OAE / Terraplenos"},
        "tipos":  None,
        "icon":   "🏛️",
        "cor":    "#A29BFE",
    },
    "Topografia": {
        "obras":  {"TOPOGRAFIA"},
        "tipos":  None,
        "icon":   "📐",
        "cor":    "#4CC9F0",
    },
}


def _render_produtividade(dados: list, datas_janela: list[str]):
    """
    View de Produtividade por Laboratorista — similar ao BASE44.
    Dividida em 5 categorias: Diário de Obra / SST / Pavimento / OAE / Topografia.
    Usa campo 'lab' (novo) ou 'profissional' (retrocompatível).
    """
    today_str = date.today().strftime("%Y-%m-%d")
    DAY_ABBR_P = {0:"SEG",1:"TER",2:"QUA",3:"QUI",4:"SEX",5:"SÁB",6:"DOM"}

    # Nomes que são contratos/grupos, não pessoas individuais
    _GRUPOS = {"eco cerrado", "eco minas goiás", "eco minas goias",
               "ecl minas goiás", "eco-cerrado", "eco-minas goiás"}

    # Normaliza: usa 'lab' se disponível (dados novos após re-scraping)
    # Se 'lab' vazio e 'profissional' é contrato → mostra como grupo
    dados_sem_nome = False
    for e in dados:
        lab_raw  = (e.get("lab") or "").strip()
        prof_raw = (e.get("profissional") or "").strip()
        if lab_raw:
            e["lab"] = lab_raw
        elif prof_raw.lower() in _GRUPOS:
            e["lab"] = f"📁 {prof_raw} (grupo)"
            dados_sem_nome = True
        else:
            e["lab"] = prof_raw or "—"

    if dados_sem_nome:
        st.warning(
            "⚠️ Alguns registros mostram o contrato (ex: 'Eco Cerrado') "
            "em vez do nome individual do laboratorista. "
            "**Execute `baixar_ensaios.py` novamente** para capturar os nomes reais.",
            icon="👤"
        )
        # Normaliza data para YYYY-MM-DD
        try:
            e["_d"] = datetime.strptime(e["data"], "%d/%m/%Y").strftime("%Y-%m-%d")
        except Exception:
            e["_d"] = ""
        # Normaliza status
        s = str(e.get("status","")).lower()
        if "aprovado" in s:      e["_status"] = "ok"
        elif "pendente" in s:    e["_status"] = "pend"
        elif "reprovado" in s:   e["_status"] = "rep"
        elif "execu" in s:       e["_status"] = "exec"
        else:                    e["_status"] = "ok"   # sem status = aprovado implícito

    # Apenas datas da janela (sem futuro)
    dados_janela = [e for e in dados if e.get("_d") in datas_janela]
    if not dados_janela:
        st.info("Sem registros no período exibido.")
        return

    from collections import defaultdict

    def _filtrar_categoria(cat_cfg: dict) -> list:
        """Filtra registros para uma categoria."""
        r = dados_janela
        if cat_cfg["obras"]:
            r = [e for e in r if e.get("obra","") in cat_cfg["obras"]]
        if cat_cfg["tipos"]:
            r = [e for e in r if e.get("tipo","") in cat_cfg["tipos"]]
        return r

    def _render_grid_categoria(registros: list, cor: str):
        """Renderiza cards de uma categoria."""
        por_lab: dict = defaultdict(lambda: defaultdict(list))
        for e in registros:
            por_lab[e["lab"]][e["_d"]].append(e)

        def _urg(lab):
            ts = [e for d in datas_janela for e in por_lab[lab].get(d,[])]
            if any(e["_status"]=="rep"  for e in ts): return 0
            if any(e["_status"]=="pend" for e in ts): return 1
            return 2

        labs_sorted = sorted(por_lab.keys(), key=_urg)
        cards = [f'<div class="prod-grid">']

        for lab in labs_sorted:
            por_data = por_lab[lab]
            todos    = [e for d in datas_janela for e in por_data.get(d,[])]
            has_rep  = any(e["_status"]=="rep"  for e in todos)
            has_pend = any(e["_status"]=="pend" for e in todos)
            card_cls = "pc-rep" if has_rep else ("pc-pend" if has_pend else "pc-ok")

            obras_lab = list(dict.fromkeys(e.get("obra","") for e in todos if e.get("obra")))
            pills = ""
            for d in datas_janela:
                try:
                    dt_obj = datetime.strptime(d, "%Y-%m-%d")
                except Exception:
                    continue
                is_hoje = (d == today_str)
                ens_dia = por_data.get(d, [])
                if not ens_dia:
                    cls = "c-vazio" + (" c-hoje" if is_hoje else "")
                    txt = "HOJE" if is_hoje else "—"
                    emp = ""
                else:
                    statuses = [e["_status"] for e in ens_dia]
                    cls = ("c-rep"  if "rep"  in statuses else
                           "c-pend" if "pend" in statuses else
                           "c-exec" if "exec" in statuses else "c-ok")
                    if is_hoje: cls += " c-hoje"
                    n   = len(ens_dia)
                    emp = (ens_dia[0].get("empreiteira","") or
                           ens_dia[0].get("obra",""))[:8]
                    lbl = {"rep":"REP","pend":"PND","exec":"EXE"}.get(statuses[0],"OK")
                    txt = f"{lbl}×{n}" if n > 1 else lbl

                dd_label = "HOJE" if is_hoje else f"{DAY_ABBR_P[dt_obj.weekday()]} {dt_obj.day:02d}"
                dd_style = "font-weight:700;color:#F7B731" if is_hoje else ""
                pills += (
                    f'<div class="prod-pill">'
                    f'<span class="prod-dd" style="{dd_style}">{dd_label}</span>'
                    f'<div class="prod-cell {cls}">'
                    f'<span>{txt}</span>'
                    f'{"" if not ens_dia else f"<span class=prod-emp>{emp}</span>"}'
                    f'</div></div>'
                )

            n_ok   = sum(1 for e in todos if e["_status"]=="ok")
            n_pend = sum(1 for e in todos if e["_status"]=="pend")
            n_rep  = sum(1 for e in todos if e["_status"]=="rep")
            tags = f'<span class="prod-tag t-ok">✓ {n_ok}</span>'
            if n_pend: tags += f'<span class="prod-tag t-pend">⏳ {n_pend} pend.</span>'
            if n_rep:  tags += f'<span class="prod-tag t-rep">✗ {n_rep} reprov.</span>'

            cards.append(
                f'<div class="prod-card {card_cls}" style="border-left-color:{cor}">'
                f'<div class="prod-name">{lab}</div>'
                f'<div class="prod-sub">{" · ".join(obras_lab[:3])}</div>'
                f'<div class="prod-days">{pills}</div>'
                f'<div class="prod-foot">{tags}</div></div>'
            )
        cards.append("</div>")
        st.markdown("".join(cards), unsafe_allow_html=True)

    # ── Renderiza cada categoria como sub-tab ─────────────────────────────
    cat_names  = list(_CATEGORIAS_PROD.keys())
    cat_tabs   = st.tabs([
        f"{_CATEGORIAS_PROD[c]['icon']} {c}" for c in cat_names
    ])

    for tab, cat_name in zip(cat_tabs, cat_names):
        with tab:
            cfg      = _CATEGORIAS_PROD[cat_name]
            regs     = _filtrar_categoria(cfg)
            if not regs:
                st.info(f"Sem registros de **{cat_name}** no período.")
                continue
            st.caption(f"{len(regs)} registros · últimos {len(datas_janela)} dias")
            st.markdown(_CSS_PROD, unsafe_allow_html=True)
            _render_grid_categoria(regs, cfg["cor"])
    return   # impede cair no código antigo abaixo

    # ── Código antigo (não alcançado) ─────────────────────────────────────
    por_lab = defaultdict(lambda: defaultdict(list))
    for e in dados_janela:
        por_lab[e["lab"]][e["_d"]].append(e)

    def _urgency_lab(lab):
        ensaios = [e for d in datas_janela for e in por_lab[lab].get(d,[])]
        if any(e["_status"]=="rep"  for e in ensaios): return 0
        if any(e["_status"]=="pend" for e in ensaios): return 1
        return 2

    labs_sorted = sorted(por_lab.keys(), key=_urgency_lab)

    cards = ['<div class="prod-wrap">', _CSS_PROD, '<div class="prod-grid">']

    for lab in labs_sorted:
        por_data = por_lab[lab]

        # Status geral do card
        todos = [e for d in datas_janela for e in por_data.get(d,[])]
        has_rep  = any(e["_status"]=="rep"  for e in todos)
        has_pend = any(e["_status"]=="pend" for e in todos)
        card_cls = "pc-rep" if has_rep else ("pc-pend" if has_pend else "pc-ok")

        # Obras únicas do período
        obras_lab = list(dict.fromkeys(e.get("obra","") for e in todos if e.get("obra")))

        # Pills por dia
        pills = ""
        for d in datas_janela:
            try:
                dt_obj = datetime.strptime(d, "%Y-%m-%d")
            except Exception:
                continue
            is_hoje = (d == today_str)
            ensaios_dia = por_data.get(d, [])

            if not ensaios_dia:
                cls = "c-vazio" + (" c-hoje" if is_hoje else "")
                txt = "HOJE" if is_hoje else "—"
                emp = ""
            else:
                # Pior status do dia
                statuses = [e["_status"] for e in ensaios_dia]
                if "rep"  in statuses: cls = "c-rep"
                elif "pend" in statuses: cls = "c-pend"
                elif "exec" in statuses: cls = "c-exec"
                else:                    cls = "c-ok"
                if is_hoje: cls += " c-hoje"

                n   = len(ensaios_dia)
                emp = ensaios_dia[0].get("empreiteira","") or ensaios_dia[0].get("obra","")
                emp = emp[:8] if emp else ""
                txt = f"OK×{n}" if n > 1 else "OK"
                if "rep"  in statuses: txt = f"REP×{n}" if n>1 else "REP"
                elif "pend" in statuses: txt = f"PND×{n}" if n>1 else "PND"

            dd_label = "HOJE" if is_hoje else f"{DAY_ABBR_P[dt_obj.weekday()]} {dt_obj.day:02d}"
            dd_style = "font-weight:700;color:#F7B731" if is_hoje else ""

            pills += (
                f'<div class="prod-pill">'
                f'<span class="prod-dd" style="{dd_style}">{dd_label}</span>'
                f'<div class="prod-cell {cls}">'
                f'<span>{txt}</span>'
                f'{"" if not ensaios_dia else f"<span class=prod-emp>{emp}</span>"}'
                f'</div>'
                f'</div>'
            )

        # Tags resumo
        n_ok   = sum(1 for e in todos if e["_status"]=="ok")
        n_pend = sum(1 for e in todos if e["_status"]=="pend")
        n_rep  = sum(1 for e in todos if e["_status"]=="rep")
        tags = f'<span class="prod-tag t-ok">✓ {n_ok}</span>'
        if n_pend: tags += f'<span class="prod-tag t-pend">⏳ {n_pend} pend.</span>'
        if n_rep:  tags += f'<span class="prod-tag t-rep">✗ {n_rep} reprov.</span>'

        sub = " · ".join(obras_lab[:3])

        cards.append(f"""
        <div class="prod-card {card_cls}">
          <div class="prod-name">{lab}</div>
          <div class="prod-sub">{sub}</div>
          <div class="prod-days">{pills}</div>
          <div class="prod-foot">{tags}</div>
        </div>""")

    cards.append("</div></div>")
    st.markdown("".join(cards), unsafe_allow_html=True)


_TIPO_COR = {
    "Diário de Obra":         ("#4CC9F0", "📋"),
    "Checklist de Usina":     ("#7BBF6A", "🏭"),
    "Checklist de Aplicação": ("#F7B731", "🚧"),
    "Checklist de MRAF":      ("#A29BFE", "🔩"),
    "Ensaio de CAUQ":         ("#FD79A8", "🧪"),
}

_GRUPOS_NORM = {"eco cerrado", "eco minas goias", "eco minas goiás",
                "ecl minas goiás", "eco-cerrado", "eco-minas goiás"}

_CSS_ENSAIOS = """
<style>
.ea-kpi-row{display:flex;gap:10px;flex-wrap:wrap;margin:10px 0 16px}
.ea-kpi{background:rgba(13,27,42,.75);border:1px solid rgba(86,110,61,.3);
  border-radius:10px;padding:10px 16px;min-width:90px;text-align:center}
.ea-kpi .val{font-size:1.5rem;font-weight:700}
.ea-kpi .lbl{font-size:.65rem;color:#8FA882;margin-top:2px}
.ea-tipo-chip{display:inline-block;font-size:.62rem;font-weight:700;
  padding:2px 8px;border-radius:999px;margin:2px 3px 2px 0;border:1px solid}
.ea-day-row{display:flex;gap:4px;flex-wrap:nowrap;overflow-x:auto;
  padding-bottom:4px;scrollbar-width:none;margin:6px 0}
.ea-day-row::-webkit-scrollbar{display:none}
.ea-day-pill{display:flex;flex-direction:column;align-items:center;gap:2px;
  min-width:38px;flex-shrink:0}
.ea-dd{font-size:.57rem;color:#8FA882;font-weight:500;text-align:center}
.ea-cell{width:38px;height:26px;border-radius:6px;font-size:.6rem;font-weight:700;
  display:flex;align-items:center;justify-content:center;text-align:center}
.ea-cell.c-ok{background:rgba(60,180,75,.25);color:#3cb44b;border:1px solid rgba(60,180,75,.4)}
.ea-cell.c-pend{background:rgba(247,183,49,.2);color:#F7B731;border:1px solid rgba(247,183,49,.4)}
.ea-cell.c-rep{background:rgba(230,25,75,.2);color:#e6194b;border:1px solid rgba(230,25,75,.4)}
.ea-cell.c-vazio{background:rgba(255,255,255,.03);color:#3a4a5e;border:1px dashed #2D3748}
.ea-cell.c-warn{background:rgba(247,183,49,.1);color:#F7B731;border:2px dashed #F7B731}
.ea-cell.c-hoje{outline:2px solid #F7B731;outline-offset:1px}
.ea-cobrar{display:inline-block;font-size:.65rem;font-weight:700;
  color:#FF6B6B;background:rgba(255,107,107,.15);border:1px solid rgba(255,107,107,.4);
  border-radius:6px;padding:2px 8px;margin-left:8px}
.ea-report-link{color:#4CC9F0;text-decoration:none;font-size:.7rem}
.ea-report-link:hover{text-decoration:underline}
.ea-status-badge{display:inline-block;font-size:.6rem;padding:1px 6px;
  border-radius:999px;border:1px solid;font-weight:600;margin-right:4px}
.ea-entry-row{display:flex;align-items:center;gap:8px;flex-wrap:wrap;
  padding:5px 0;border-bottom:1px solid rgba(255,255,255,.04)}
.ea-entry-row:last-child{border-bottom:none}
</style>
"""


def _render_ensaios_aevias():
    """Seção de ensaios — view per-person limpa com cobrar tracking."""
    from collections import defaultdict

    st.markdown("## 🧪 Ensaios & Relatórios — AEVIAS Controle")

    # ── Header: info + refresh ────────────────────────────────────────────────
    col_info, col_btn = st.columns([5, 1])

    dados, mtime = _carregar_ensaios()

    with col_info:
        if mtime:
            dt_mod = datetime.fromtimestamp(mtime).strftime("%d/%m/%Y %H:%M")
            st.caption(
                f"Fonte: `ensaios_aevias.json` · Atualizado: **{dt_mod}** · "
                f"{len(dados)} registros · "
                f"[Abrir site ↗]({_BASE44_URL}/MeusEnsaios)"
            )
        else:
            st.warning(
                "`cache_certificados/ensaios_aevias.json` não encontrado. "
                "Execute `baixar_ensaios.py` para gerar o cache."
            )
            return
    with col_btn:
        if st.button("🔄 Refresh", key="btn_refresh_ensaios", use_container_width=True):
            st.cache_data.clear()
            st.rerun()

    if not dados:
        st.info("Nenhum ensaio carregado.")
        return

    # ── Period selector ───────────────────────────────────────────────────────
    periodo = st.radio(
        "Período:",
        options=["Hoje", "7 dias", "30 dias"],
        index=1,
        horizontal=True,
        key="ens_periodo_sel",
    )

    today = date.today()
    today_str = today.strftime("%Y-%m-%d")
    if periodo == "Hoje":
        delta_days = 0
    elif periodo == "7 dias":
        delta_days = 6
    else:
        delta_days = 29

    # ── Parse e enriquece cada ensaio ────────────────────────────────────────
    for e in dados:
        try:
            e["_date"] = datetime.strptime(e["data"], "%d/%m/%Y").date()
            e["_dstr"] = e["_date"].strftime("%Y-%m-%d")
        except Exception:
            e["_date"] = None
            e["_dstr"] = ""

        # Normaliza nome do profissional
        lab_raw  = (e.get("lab") or "").strip()
        prof_raw = (e.get("profissional") or "").strip()
        if lab_raw:
            e["_prof"] = lab_raw
            e["_is_grupo"] = False
        elif _norm(prof_raw) in _GRUPOS_NORM:
            e["_prof"] = prof_raw  # mantém original para label
            e["_is_grupo"] = True
        else:
            e["_prof"] = prof_raw or "—"
            e["_is_grupo"] = False

        # Normaliza status
        s = str(e.get("status", "")).lower()
        if "reprovado" in s:
            e["_status"] = "rep"
        elif "pendente" in s:
            e["_status"] = "pend"
        elif "aprovado" in s:
            e["_status"] = "ok"
        else:
            e["_status"] = "ok"

    # ── Filtra janela (sem futuro) ────────────────────────────────────────────
    from datetime import timedelta
    date_inicio = today - timedelta(days=delta_days)
    dados_periodo = [
        e for e in dados
        if e["_dstr"] and e["_dstr"] <= today_str and e["_date"] >= date_inicio
    ]

    # ── KPIs globais ─────────────────────────────────────────────────────────
    n_total    = len(dados_periodo)
    n_pend     = sum(1 for e in dados_periodo if e["_status"] == "pend")
    n_rep      = sum(1 for e in dados_periodo if e["_status"] == "rep")
    n_cobrar   = n_pend + n_rep

    st.markdown(_CSS_ENSAIOS, unsafe_allow_html=True)
    st.markdown(
        f'<div class="ea-kpi-row">'
        f'<div class="ea-kpi"><div class="val" style="color:#C8D8A8">{n_total}</div>'
        f'<div class="lbl">Submissões</div></div>'
        f'<div class="ea-kpi"><div class="val" style="color:#F7B731">{n_pend}</div>'
        f'<div class="lbl">Pendentes</div></div>'
        f'<div class="ea-kpi"><div class="val" style="color:#FF6B6B">{n_rep}</div>'
        f'<div class="lbl">Reprovados</div></div>'
        f'<div class="ea-kpi"><div class="val" style="color:#{"FF6B6B" if n_cobrar else "7BBF6A"}">'
        f'{"🚨 " if n_cobrar else "✓ "}{n_cobrar}</div>'
        f'<div class="lbl">A Cobrar</div></div>'
        f'</div>',
        unsafe_allow_html=True,
    )

    if not dados_periodo:
        st.info(f"Sem registros no período selecionado ({periodo}).")
        return

    # ── Dias com pelo menos 1 registro no período ─────────────────────────────
    datas_com_dados = sorted(
        {e["_dstr"] for e in dados_periodo if e["_dstr"]}
    )

    # ── Agrupa por profissional ───────────────────────────────────────────────
    por_prof: dict = defaultdict(list)
    for e in dados_periodo:
        por_prof[e["_prof"]].append(e)

    # Separa grupos dos indivíduos
    grupos  = {p: v for p, v in por_prof.items()
               if any(e["_is_grupo"] for e in v)}
    indivs  = {p: v for p, v in por_prof.items()
               if not any(e["_is_grupo"] for e in v)}

    DAY_ABBR_E = {0: "SEG", 1: "TER", 2: "QUA", 3: "QUI",
                  4: "SEX", 5: "SÁB", 6: "DOM"}

    def _urgencia_prof(registros):
        if any(e["_status"] == "rep"  for e in registros): return 0
        if any(e["_status"] == "pend" for e in registros): return 1
        return 2

    def _render_person_expander(nome: str, registros: list, label_prefix: str = ""):
        """Renderiza um expander por pessoa com todos os dados."""
        a_cobrar = any(e["_status"] in ("pend", "rep") for e in registros)
        has_rep  = any(e["_status"] == "rep" for e in registros)

        expander_label = f"{label_prefix}{nome}"
        if a_cobrar:
            expander_label += "  🚨 A COBRAR"

        with st.expander(expander_label, expanded=a_cobrar):
            n_reg   = len(registros)
            n_p_loc = sum(1 for e in registros if e["_status"] == "pend")
            n_r_loc = sum(1 for e in registros if e["_status"] == "rep")

            # Mini KPI row
            kpi_color = "#FF6B6B" if (n_p_loc + n_r_loc) else "#7BBF6A"
            st.markdown(
                f'<div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:8px">'
                f'<span style="color:#C8D8A8;font-size:.8rem;font-weight:700">'
                f'{n_reg} submissão(ões)</span>'
                f'{"<span class=ea-cobrar>🚨 A COBRAR</span>" if a_cobrar else ""}'
                f'</div>',
                unsafe_allow_html=True,
            )

            # Breakdown por tipo (chips)
            from collections import Counter
            cnt_tipo = Counter(e.get("tipo", "—") for e in registros)
            chips_html = ""
            for tipo, cnt in cnt_tipo.most_common():
                cor, icon = _TIPO_COR.get(tipo, ("#8FA882", "📄"))
                chips_html += (
                    f'<span class="ea-tipo-chip" '
                    f'style="color:{cor};border-color:{cor}55;background:{cor}18">'
                    f'{icon} {tipo} ({cnt})</span>'
                )
            if chips_html:
                st.markdown(chips_html, unsafe_allow_html=True)

            # Dias com / sem submissão (compact calendar row)
            por_data_p: dict = defaultdict(list)
            for e in registros:
                por_data_p[e["_dstr"]].append(e)

            pills_html = ""
            for d in datas_com_dados:
                try:
                    dt_obj = datetime.strptime(d, "%Y-%m-%d")
                except Exception:
                    continue
                is_hoje  = (d == today_str)
                ens_dia  = por_data_p.get(d, [])
                dd_label = ("HOJE" if is_hoje
                            else f"{DAY_ABBR_E[dt_obj.weekday()]} {dt_obj.day:02d}")
                dd_style = "font-weight:700;color:#F7B731" if is_hoje else ""

                if not ens_dia:
                    cls = "c-vazio" + (" c-hoje" if is_hoje else "")
                    txt = "—"
                else:
                    statuses = [e["_status"] for e in ens_dia]
                    cls = ("c-rep"  if "rep"  in statuses else
                           "c-pend" if "pend" in statuses else "c-ok")
                    if is_hoje:
                        cls += " c-hoje"
                    n = len(ens_dia)
                    txt = f"×{n}" if n > 1 else ("REP" if "rep" in statuses
                                                  else "PND" if "pend" in statuses
                                                  else "OK")

                pills_html += (
                    f'<div class="ea-day-pill">'
                    f'<span class="ea-dd" style="{dd_style}">{dd_label}</span>'
                    f'<div class="ea-cell {cls}">{txt}</div>'
                    f'</div>'
                )

            if pills_html:
                st.markdown(
                    f'<div class="ea-day-row">{pills_html}</div>',
                    unsafe_allow_html=True,
                )

            # Lista de entradas individuais com status badge + link
            st.markdown(
                '<div style="margin-top:8px;font-size:.7rem;color:#8FA882;font-weight:600">'
                'Registros:</div>',
                unsafe_allow_html=True,
            )
            rows_html = ""
            for e in sorted(registros, key=lambda x: x["_dstr"], reverse=True):
                tipo   = e.get("tipo", "—")
                obra   = e.get("obra", "—")
                local  = e.get("local", "")
                emp    = e.get("empreiteira", "") or ""
                url    = e.get("reportUrl", "")
                status = e.get("status", "")
                dstr   = e.get("data", "")
                s_norm = e["_status"]
                cor_s  = ("#7BBF6A" if s_norm == "ok"
                           else "#F7B731" if s_norm == "pend"
                           else "#FF6B6B")
                s_badge = (
                    f'<span class="ea-status-badge" '
                    f'style="color:{cor_s};border-color:{cor_s}55">{status}</span>'
                    if status else ""
                )
                ico_tipo = _TIPO_COR.get(tipo, ("#8FA882", "📄"))[1]
                link_html = (
                    f'<a class="ea-report-link" href="{_BASE44_URL}{url}" '
                    f'target="_blank">🔗 Relatório</a>'
                    if url else ""
                )
                detail = " · ".join(filter(None, [obra, emp, local]))
                rows_html += (
                    f'<div class="ea-entry-row">'
                    f'<span style="color:#8FA882;min-width:70px;font-size:.65rem">{dstr}</span>'
                    f'<span style="color:#C8D8A8">{ico_tipo} {tipo}</span>'
                    f'<span style="color:#8FA882;font-size:.65rem;flex:1">{detail}</span>'
                    f'{s_badge}{link_html}'
                    f'</div>'
                )
            st.markdown(
                f'<div style="background:rgba(0,0,0,.15);border-radius:8px;padding:6px 10px">'
                f'{rows_html}</div>',
                unsafe_allow_html=True,
            )

    # ── Views: Por Pessoa (expanders) | Por Categoria (grid) ─────────────────
    tab_pessoa, tab_cat = st.tabs(["👤 Por Pessoa", "📊 Por Categoria"])

    with tab_pessoa:
        # Render: indivíduos (ordenados por urgência)
        indivs_sorted = sorted(
            indivs.items(),
            key=lambda kv: (_urgencia_prof(kv[1]), kv[0])
        )
        for nome, regs in indivs_sorted:
            _render_person_expander(nome, regs, label_prefix="👤 ")

        # Render: grupos (Eco Cerrado, ECO Minas Goiás etc.) — se ainda existirem
        if grupos:
            st.markdown(
                '<div style="margin:12px 0 4px;font-size:.75rem;color:#8FA882;font-weight:600">'
                'Grupos / Contratos (nome individual não disponível — execute o scraper):</div>',
                unsafe_allow_html=True,
            )
            grupos_sorted = sorted(
                grupos.items(),
                key=lambda kv: (_urgencia_prof(kv[1]), kv[0])
            )
            for nome, regs in grupos_sorted:
                _render_person_expander(nome, regs, label_prefix="📁 ")

    with tab_cat:
        _render_produtividade(dados_periodo, datas_com_dados)


def _aba_checklist():
    meds = _listar_meds()
    if not meds:
        st.info("Nenhuma medição encontrada. Verifique o acesso ao servidor Y:")
        return

    med_labels = {m: m for m in reversed(meds)}  # mais recente primeiro
    meds_reversed = list(reversed(meds))

    col_sel, col_info = st.columns([2, 4])
    with col_sel:
        med_escolhida = st.selectbox(
            "📅 Medição:",
            options=meds_reversed,
            format_func=lambda x: x,
            key="eco_med_sel",
        )

    # Carrega dados da medição selecionada
    with st.spinner("Carregando checklist..."):
        sheets = None
        if not _IS_CLOUD:
            sheets = _carregar_checklist_y(med_escolhida)
        if sheets is None:
            cache = _carregar_checklist_cache()
            entry = cache.get(med_escolhida, {})
            sheets = entry.get("sheets", {})

    if not sheets:
        st.warning(f"Nenhum arquivo de controle encontrado para **{med_escolhida}**.")
        return

    # KPIs globais
    total_ok = total_cob = total_ne = 0
    for plist in sheets.values():
        for p in plist:
            for v in p.get("dias", {}).values():
                vu = str(v).upper().strip() if v else ""
                if vu == "OK":
                    total_ok += 1
                elif vu in ("COBRAR", "COBRE"):
                    total_cob += 1
                elif vu in ("N/E", "NE"):
                    total_ne += 1

    c1, c2, c3, c4 = st.columns(4)
    c1.markdown(_kpi_card(total_ok,  "Checklists OK",     COR_OK),    unsafe_allow_html=True)
    c2.markdown(_kpi_card(total_cob, "A Cobrar",          COR_COBRAR), unsafe_allow_html=True)
    c3.markdown(_kpi_card(total_ne,  "Não em Campo (N/E)", COR_MUTED), unsafe_allow_html=True)
    pct = f"{100*total_ok/(total_ok+total_cob):.0f}%" if (total_ok+total_cob) > 0 else "—"
    c4.markdown(_kpi_card(pct, "Taxa de Conformidade", COR_ACCENT), unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Sub-abas por contrato
    sheet_names = list(sheets.keys())
    tab_labels = [f"🛣️ {s}" for s in sheet_names]
    tabs = st.tabs(tab_labels)

    for tab, sn in zip(tabs, sheet_names):
        with tab:
            people = sheets[sn]
            if not people:
                st.info("Nenhum colaborador.")
                continue

            # KPIs por contrato
            ok_c = sum(
                1 for p in people for v in p.get("dias", {}).values()
                if v and str(v).upper().strip() == "OK"
            )
            cob_c = sum(
                1 for p in people for v in p.get("dias", {}).values()
                if v and str(v).upper().strip() in ("COBRAR", "COBRE")
            )
            colab_cob = [
                p["colaborador"].split()[0]
                for p in people
                if any(str(v).upper().strip() in ("COBRAR", "COBRE")
                       for v in p.get("dias", {}).values() if v)
            ]

            if cob_c > 0:
                st.error(
                    f"⚠️ **{cob_c} checklist(s) pendente(s)** — Colaboradores: "
                    + ", ".join(colab_cob[:6])
                    + ("..." if len(colab_cob) > 6 else "")
                )
            else:
                st.success(f"✅ Todos os checklists enviados neste contrato — {ok_c} registros OK")

            _renderizar_calendario(people, med_escolhida)

    # ── Ensaios AEVIAS — dados do aevias-controle.base44.app ─────────────────
    st.markdown("---")
    _render_ensaios_aevias()
