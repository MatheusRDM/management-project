"""
=========================================================================
EPR LITORAL PIONEIRO — Utilitários de Dados (FORM 022A + FORM 103C)
=========================================================================
Carrega, filtra e processa os dados do FORM 022A especificamente para
o cliente EPR Litoral Pioneiro.

Campos extraídos da COL L (OBS_RECEBIMENTO) por tipo de material:
  • CP_CONCRETO → DATA_MOLDAGEM (60%)
  • CAUQ_PISTA  → DATA_EXECUCAO (70%), NUMERO_CP, TRECHO, LOCALIZACAO
  • CAUQ_MASSA  → DATA_MOLDAGEM (60%), PROJETO_NUM, LOCALIZACAO,
                   TIPO_SERVICO, MATERIAL_OBS

FORM 103C (Rompimento de Concreto):
  • Células-chave por aba PT_XXXX:
    M9=PT Nº, D21=Rodovia, F21=Estaca, I26=Dt.Moldagem
    J26=Dt.Rupt.7d, L26=Dt.Rupt.28d
    (E30,N32), (E34,N37), (E38,N40) → blocos dias/resultado MPa
=========================================================================
"""

import glob
import re
import os
import sys
import pandas as pd
import streamlit as st

_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

from thefuzz import fuzz
from utils_certificados import FILES_CONFIG, bridge, DB_NAME

# =============================================================================
# CONFIGURAÇÕES
# =============================================================================

EPR_CLIENTE_ALVO     = "EPR LITORAL PIONEIRO"
EPR_SIMILARIDADE_MIN = 70

TABELA_EPR = "epr_form022a"

# Pasta com os arquivos FORM 103C de rompimento de concreto EPR
FORM103C_PASTA = (
    r"G:\.shortcut-targets-by-id\1NUJ7pNAqedohSrLjiwFErFrtVqVZkrjk"
    r"\006 - Lab. Central\0.1 RELATÓRIOS TÉCNICOS\003-PROJETOS\2026"
    r"\001.1.2026 - EPR - ENSAIOS GERAIS - CC EPR\003-RESULTADOS\1. Concreto"
)
FORM103C_ABAS_IGNORAR = {"CAD TRAÇO", "RESUMO", "CAD TRACO", "TEMPLATE"}

MATERIAL_GRUPOS = {
    "CP_CONCRETO": [
        "CP - Concreto (Cilíndrico)", "CP - BGTC", "CP - CCR",
        "CP - Concreto (Prismático)", "CP - Solo-cimento",
        "CP CONCRETO", "CONCRETO CILINDRICO", "CONCRETO PRISMATICO",
        "BGTC", "CCR", "SOLO CIMENTO", "SOLO-CIMENTO",
    ],
    "CAUQ_PISTA": [
        "CP - CAUQ (Pista)", "CAUQ (Pista)", "CAUQ PISTA", "CP CAUQ PISTA",
    ],
    "CAUQ_MASSA": [
        "CAUQ (Massa Usinada)", "CAUQ MASSA USINADA",
        "CAUQ MASSA", "MASSA USINADA", "CAUQ USINADA",
    ],
}

_REGRAS_GRUPOS = {
    "CP_CONCRETO": [
        {"campo": "DATA DA COLETA",   "threshold": 60, "dest": "DATA_MOLDAGEM"},
        {"campo": "DATA DE MOLDAGEM", "threshold": 60, "dest": "DATA_MOLDAGEM"},
        {"campo": "DATA MOLDAGEM",    "threshold": 60, "dest": "DATA_MOLDAGEM"},
    ],
    "CAUQ_PISTA": [
        {"campo": "DATA DE EXECUCAO", "threshold": 70, "dest": "DATA_EXECUCAO"},
        {"campo": "DATA DA EXECUCAO", "threshold": 70, "dest": "DATA_EXECUCAO"},
        {"campo": "DATA EXECUCAO",    "threshold": 70, "dest": "DATA_EXECUCAO"},
        {"campo": "DATA",             "threshold": 85, "dest": "DATA_EXECUCAO"},
    ],
    "CAUQ_MASSA": [
        {"campo": "DATA DA COLETA",   "threshold": 60, "dest": "DATA_MOLDAGEM"},
        {"campo": "DATA DE COLETA",   "threshold": 60, "dest": "DATA_MOLDAGEM"},
        {"campo": "DATA COLETA",      "threshold": 65, "dest": "DATA_MOLDAGEM"},
    ],
}

_LABELS_GRUPOS = {
    "CP_CONCRETO": "CP Concreto / BGTC / CCR / Solo-cimento",
    "CAUQ_PISTA":  "CAUQ (Pista)",
    "CAUQ_MASSA":  "CAUQ (Massa Usinada)",
    "OUTROS":      "Outros materiais",
}

_ICONES_GRUPOS = {
    "CP_CONCRETO": "🏗️",
    "CAUQ_PISTA":  "🛣️",
    "CAUQ_MASSA":  "♨️",
    "OUTROS":      "📦",
}

# Palavras-chave de serviços para CAUQ Massa
_KEYWORDS_SERVICO = [
    "FRESAGEM", "RECOMPOSICAO", "RECOMPOSIÇÃO", "IMPLANTACAO", "IMPLANTAÇÃO",
    "PAVIMENTACAO", "PAVIMENTAÇÃO", "REMENDO", "RECAPEAMENTO", "MICRORREVESTIMENTO",
    "REJUVENESCIMENTO", "BASE", "SUB-BASE", "REGULARIZACAO", "REGULARIZAÇÃO",
]

# Padrões de localização: KM, rodovias, avenidas, etc.
_RE_LOCALIZACAO = re.compile(
    r"(?:"
    r"KM\s*\d+[\+\-\.]?\d*"          # KM 02+680
    r"|(?:AV|RUA|ROD|BR|PR|SP|RS|SC|MG|BA|PE|RJ|ES|GO|MT|MS|PA|AM)\b[^\n\t]*"
    r"|\bFX\s+\w"                      # FX A / FX C
    r"|\b(?:LESTE|OESTE|NORTE|SUL|DER|ESQ|DIR|CENTRAL)\b"
    r")",
    re.IGNORECASE,
)

# Padrão de projeto: "PROJETO 113/2025" ou "PC 091.26" ou "PC091.26"
_RE_PROJETO = re.compile(r"\bPROJETO\s+[\w\./\-]+|\bPC\s*\d+[\./]\d+\b", re.IGNORECASE)


# =============================================================================
# FUNÇÕES DE PARSING DA COL L
# =============================================================================

def _classificar_material(material_str: str) -> str:
    if not material_str or str(material_str).strip().lower() in ("nan", "none", ""):
        return "OUTROS"
    mat_up = str(material_str).upper().strip()

    # 1ª passagem: correspondência EXATA (case-insensitive) em todos os grupos
    for grupo, candidatos in MATERIAL_GRUPOS.items():
        for cand in candidatos:
            if cand.upper() == mat_up:
                return grupo

    # 2ª passagem: fuzzy — CAUQ verificado ANTES de CP_CONCRETO para evitar
    # falsos positivos com materiais "CP - CAUQ..." vs candidatos "CP - ..."
    _ORDEM_FUZZY = ["CAUQ_PISTA", "CAUQ_MASSA", "CP_CONCRETO", "OUTROS"]
    for grupo in _ORDEM_FUZZY:
        candidatos = MATERIAL_GRUPOS.get(grupo, [])
        for cand in candidatos:
            if fuzz.partial_ratio(cand.upper(), mat_up) >= 70:
                return grupo

    return "OUTROS"


def _parse_obs_campo(obs_text: str, campo_alvo: str, threshold: int = 60):
    """
    Extrai valor do campo_alvo dentro do texto OBS (Col L).
    Estratégia 1: split por \\t/\\n → compara chave com campo_alvo.
    Estratégia 2: regex no texto corrido (linha única sem quebras).
    """
    if not obs_text or str(obs_text).strip().lower() in ("nan", "none", "-", ""):
        return None
    obs_text = str(obs_text).strip()

    # Estratégia 1
    tokens = re.split(r"[\t\n\r]+", obs_text)
    for token in tokens:
        token = token.strip()
        if not token or ":" not in token:
            continue
        chave, _, valor = token.partition(":")
        chave_norm = re.sub(r"\s+", " ", chave.strip().upper())
        campo_norm = re.sub(r"[ÁÀÂÃ]", "A",
                    re.sub(r"[ÉÈÊ]", "E",
                    re.sub(r"[ÍÌÎ]", "I",
                    re.sub(r"[ÓÒÔÕ]", "O",
                    re.sub(r"[ÚÙÛ]", "U", campo_alvo.upper())))))
        score = fuzz.partial_ratio(campo_norm, chave_norm)
        if score >= threshold:
            v = valor.strip()
            return v if v else None

    # Estratégia 2: regex no texto corrido
    palavras = campo_alvo.split()
    pattern = r"\b" + r"\s+".join(re.escape(p) for p in palavras) + r"\s*[:\-]\s*([^\t\n\r,;]+)"
    m = re.search(pattern, obs_text, re.IGNORECASE)
    if m:
        v = m.group(1).strip()
        return v if v else None

    return None


def _extrair_numero_cp(obs_text: str) -> str:
    if not obs_text or str(obs_text).strip().lower() in ("nan", "none", "-", ""):
        return "-"
    m = re.search(r"\bCP\s*(\d+)\b", str(obs_text).upper())
    return f"CP {m.group(1)}" if m else "-"


def _extrair_localizacao(obs_text: str) -> str:
    """
    Extrai a linha de localização da OBS:
    - Para CAUQ Pista: tudo após o número do CP até 'DATA'
      Ex: 'CP 501 FX C DER/PR DATA...' → 'FX C DER/PR'
    - Para CAUQ Massa: linha com KM, AV, ROD, etc.
    Retorna a melhor localização encontrada ou '-'.
    """
    if not obs_text or str(obs_text).strip().lower() in ("nan", "none", "-", ""):
        return "-"
    obs = str(obs_text).strip()

    # Tenta encontrar linha com KM ou rodovia/avenida
    linhas = re.split(r"[\t\n\r]+", obs)
    for linha in linhas:
        if _RE_LOCALIZACAO.search(linha):
            return linha.strip()

    # Para texto corrido (CAUQ Pista): extrai entre CP NNN e DATA
    m = re.search(r"\bCP\s*\d+\s+(.*?)(?:DATA|$)", obs, re.IGNORECASE)
    if m:
        trecho = m.group(1).strip().rstrip(",;")
        if len(trecho) > 2:
            return trecho

    return "-"


def _extrair_trecho(obs_text: str) -> str:
    """
    Extrai informação de faixa/trecho: 'FX C DER/PR', 'FX 02 LESTE', etc.
    """
    if not obs_text:
        return "-"
    m = re.search(r"\bFX\s+\w[\w\s/]*(?:DER|ESQ|DIR|LESTE|OESTE|NORTE|SUL)?", str(obs_text), re.IGNORECASE)
    return m.group(0).strip() if m else "-"


def _extrair_projeto(obs_text: str) -> str:
    """Extrai número de projeto: 'PROJETO 113/2025', 'PC 091.26', etc."""
    if not obs_text or str(obs_text).strip().lower() in ("nan", "none", "-", ""):
        return "-"
    m = _RE_PROJETO.search(str(obs_text))
    return m.group(0).strip() if m else "-"


def _extrair_tipo_servico(obs_text: str) -> str:
    """Extrai tipo de serviço executado (FRESAGEM, RECOMPOSIÇÃO, etc.)."""
    if not obs_text:
        return "-"
    obs_up = str(obs_text).upper()
    for kw in _KEYWORDS_SERVICO:
        # Busca linha que contém a palavra-chave
        for linha in re.split(r"[\t\n\r]+", obs_up):
            if kw.upper() in linha:
                return linha.strip().title()
    return "-"


def _extrair_material_obs(obs_text: str) -> str:
    """
    Extrai o tipo específico de material mencionado na OBS.
    Ex: 'CBUQ FX C BORRACHA', 'BGTC', 'CCR'
    """
    if not obs_text:
        return "-"
    # Padrão: linha com CBUQ, BGTC, CCR, CAP, etc.
    for linha in re.split(r"[\t\n\r]+", str(obs_text)):
        linha_up = linha.strip().upper()
        if any(x in linha_up for x in ["CBUQ", "CAUQ", "BGTC", "CCR", "CAP", "PEN", "PMF", "DNER"]):
            return linha.strip()
    return "-"


def _normalizar_status(status_raw: str) -> str:
    if not status_raw or str(status_raw).strip().lower() in ("nan", "none", ""):
        return "AGUARDANDO"
    s = str(status_raw).upper().strip()
    if any(x in s for x in ("CONCLU", "FINALIZ", "ENTREGUE", "OK")):
        return "CONCLUIDO"
    if any(x in s for x in ("ANDAMENTO", "EXECU", "PROCESSANDO", "EM PROG")):
        return "EM ANDAMENTO"
    return "AGUARDANDO"


# =============================================================================
# PROCESSAMENTO DE LINHA
# =============================================================================

def _processar_linha_epr(row, ano: str) -> dict | None:
    try:
        empresa_raw = str(row.iloc[1]).strip() if len(row) > 1 else ""
        if len(empresa_raw) < 3 or empresa_raw.upper() in ("EMPRESA", "CLIENTE", "NAN", "NONE", ""):
            return None
        if fuzz.partial_ratio(EPR_CLIENTE_ALVO.upper(), empresa_raw.upper()) < EPR_SIMILARIDADE_MIN:
            return None

        # Protocolo
        pt_raw = row.iloc[0] if len(row) > 0 else None
        try:
            pt = str(int(float(str(pt_raw).strip())))
        except Exception:
            pt = str(pt_raw).strip() if pt_raw else "-"

        # Datas
        dt_rec = pd.to_datetime(row.iloc[2], errors="coerce") if len(row) > 2 else pd.NaT

        # Material
        material_raw = str(row.iloc[3]).strip() if len(row) > 3 else ""
        if material_raw.lower() in ("nan", "none", ""):
            material_raw = "-"

        # Quantidade
        qtd_raw = row.iloc[7] if len(row) > 7 else 1
        try:
            quantidade = int(float(qtd_raw)) if pd.notna(qtd_raw) else 1
        except Exception:
            quantidade = 1

        # Pedreira / Procedência (Col E, idx 4)
        pedreira_raw = str(row.iloc[4]).strip() if len(row) > 4 else ""
        pedreira = "" if pedreira_raw.lower() in ("nan", "none", "ni", "n.i", "-", "") else pedreira_raw

        # OBS / Col L (idx 11)
        obs_raw = str(row.iloc[11]).strip() if len(row) > 11 else ""
        if obs_raw.lower() in ("nan", "none"):
            obs_raw = ""

        # Status
        status_raw = str(row.iloc[17]).strip() if len(row) > 17 else ""
        status = _normalizar_status(status_raw)

        # Classificar material
        grupo = _classificar_material(material_raw)

        # Campos extraídos da OBS
        data_moldagem  = None
        data_execucao  = None
        numero_cp      = "-"
        localizacao    = "-"
        trecho         = "-"
        projeto_num    = "-"
        tipo_servico   = "-"
        material_obs   = "-"

        # Extração por regras do grupo
        for regra in _REGRAS_GRUPOS.get(grupo, []):
            valor = _parse_obs_campo(obs_raw, regra["campo"], regra["threshold"])
            if valor:
                if regra["dest"] == "DATA_MOLDAGEM":
                    data_moldagem = valor
                elif regra["dest"] == "DATA_EXECUCAO":
                    data_execucao = valor
                break

        # Campos comuns a todos os grupos
        localizacao = _extrair_localizacao(obs_raw)
        trecho      = _extrair_trecho(obs_raw)
        projeto_num = _extrair_projeto(obs_raw)

        # Campos específicos por grupo
        if grupo == "CAUQ_PISTA":
            numero_cp = _extrair_numero_cp(obs_raw)

        if grupo in ("CAUQ_MASSA", "CAUQ_PISTA"):
            tipo_servico = _extrair_tipo_servico(obs_raw)
            material_obs = _extrair_material_obs(obs_raw)

        return {
            "PT":               pt,
            "CLIENTE":          empresa_raw,
            "DATA_RECEBIMENTO": dt_rec.strftime("%d/%m/%Y") if pd.notna(dt_rec) else "-",
            "MES_ANO":          dt_rec.strftime("%m/%Y") if pd.notna(dt_rec) else "-",
            "ANO":              str(dt_rec.year) if pd.notna(dt_rec) else ano,
            "MATERIAL":         material_raw,
            "MATERIAL_GRUPO":   grupo,
            "PEDREIRA":         pedreira,
            "OBS_RAW":          obs_raw,
            # Campos extraídos
            "DATA_MOLDAGEM":    data_moldagem or "-",
            "DATA_EXECUCAO":    data_execucao or "-",
            "NUMERO_CP":        numero_cp,
            "LOCALIZACAO":      localizacao,
            "TRECHO":           trecho,
            "PROJETO_NUM":      projeto_num,
            "TIPO_SERVICO":     tipo_servico,
            "MATERIAL_OBS":     material_obs,
            "QUANTIDADE":       quantidade,
            "STATUS":           status,
        }

    except Exception:
        return None


# =============================================================================
# FORM 103C — ROMPIMENTO DE CONCRETO
# =============================================================================

def _normalizar_pt(pt_str: str) -> str:
    """
    Normaliza número de protocolo para comparação cross-form.
    Ex: "1166/25" → "1166", "PT 1166" → "1166", "1166" → "1166"
    """
    s = str(pt_str).strip().upper()
    s = s.replace("PT", "").replace(" ", "")
    return s.split("/")[0].strip()


@st.cache_data(ttl=300, show_spinner=False)
def carregar_form103c() -> pd.DataFrame:
    """
    Lê todos os arquivos FORM 103C (Rompimento Concreto) na pasta EPR.
    Para cada aba PT_XXXX extrai: PT, rodovia, estaca, datas, resultados Fc7/Fc28.
    Retorna DataFrame com colunas de resultado para merge com FORM 022A.
    """
    # ── CLOUD GUARD ──────────────────────────────────────────────────────────
    from cloud_config import IS_CLOUD
    if IS_CLOUD:
        return pd.DataFrame()  # FORM 103C não disponível no cloud
    # ─────────────────────────────────────────────────────────────────────────
    import openpyxl

    arquivos = glob.glob(os.path.join(FORM103C_PASTA, "FORM 103*.xlsx"))
    if not arquivos:
        return pd.DataFrame()

    registros = []
    for arq in arquivos:
        try:
            wb = openpyxl.load_workbook(arq, data_only=True)
        except Exception:
            continue

        for sh_name in wb.sheetnames:
            if sh_name.strip().upper() in FORM103C_ABAS_IGNORAR:
                continue

            ws = wb[sh_name]

            def cel(ref):
                try:
                    return ws[ref].value
                except Exception:
                    return None

            pt_raw = cel("M9")
            if pt_raw is None:
                continue

            def fmt_data(d):
                if d is None:
                    return None
                if hasattr(d, "strftime"):
                    return d.strftime("%d/%m/%Y")
                s = str(d).strip()
                return s if s.lower() not in ("nan", "none", "") else None

            rodovia = cel("D21")
            estaca  = cel("F21")
            dt_mold  = fmt_data(cel("I26"))
            dt_7d    = fmt_data(cel("J26"))
            dt_28d   = fmt_data(cel("L26"))

            # Blocos: (célula_dias, célula_resultado)
            blocos = [
                (cel("E30"), cel("N32")),
                (cel("E34"), cel("N37")),
                (cel("E38"), cel("N40")),
            ]

            res_7d = res_28d = dias_7d = dias_28d = None
            for dias_val, res_val in blocos:
                if res_val is None:
                    continue
                try:
                    d = int(float(str(dias_val)))
                    r = float(str(res_val))
                    if d <= 7 and res_7d is None:
                        res_7d, dias_7d = round(r, 2), d
                    elif d >= 14:
                        res_28d, dias_28d = round(r, 2), d
                except Exception:
                    pass

            registros.append({
                "PT_NUM":              str(pt_raw).strip(),
                "PT_NORM":             _normalizar_pt(str(pt_raw)),
                "RODOVIA_103":         str(rodovia).strip() if rodovia else None,
                "ESTACA_103":          str(estaca).strip() if estaca else None,
                "DATA_MOLDAGEM_103":   dt_mold,
                "DATA_ROMPIMENTO_7D":  dt_7d,
                "DATA_ROMPIMENTO_28D": dt_28d,
                "RESULTADO_7D_MPA":    res_7d,
                "DIAS_7D":             dias_7d,
                "RESULTADO_28D_MPA":   res_28d,
                "DIAS_28D":            dias_28d,
                "ABA_FORM103":         sh_name,
            })

    return pd.DataFrame(registros) if registros else pd.DataFrame()


# =============================================================================
# CARREGAMENTO PRINCIPAL
# =============================================================================

@st.cache_data(ttl=600, show_spinner=False)
def carregar_dados_epr_form022a() -> pd.DataFrame:
    # ── CLOUD GUARD ──────────────────────────────────────────────────────────
    from cloud_config import IS_CLOUD
    if IS_CLOUD:
        from cloud_config import carregar_parquet_cache
        df = carregar_parquet_cache("db_epr_form022a")
        return df if not df.empty else pd.DataFrame()
    # ─────────────────────────────────────────────────────────────────────────
    keys_recebimento = [k for k in FILES_CONFIG if FILES_CONFIG[k].get("tipo") == "recebimento"]
    if not keys_recebimento:
        return pd.DataFrame()

    registros = []
    for key in sorted(keys_recebimento):
        ano = FILES_CONFIG[key].get("ano", "????")
        source = bridge.get_file_content(key)
        if not source:
            continue
        try:
            df_raw = pd.read_excel(source, header=None, engine="openpyxl")
        except Exception:
            continue

        for i in range(6, len(df_raw)):
            row = df_raw.iloc[i]
            resultado = _processar_linha_epr(row, ano)
            if resultado:
                registros.append(resultado)

    if not registros:
        return pd.DataFrame()

    df = pd.DataFrame(registros)
    df = df.drop_duplicates(subset=["PT", "MATERIAL", "DATA_RECEBIMENTO"])
    df = df.reset_index(drop=True)

    # ── Enriquecer com resultados do FORM 103C (Fc7/Fc28) ──────────────────
    _colunas_103 = [
        "RODOVIA_103", "ESTACA_103",
        "DATA_ROMPIMENTO_7D", "DATA_ROMPIMENTO_28D",
        "RESULTADO_7D_MPA", "RESULTADO_28D_MPA",
        "DIAS_7D", "DIAS_28D",
    ]
    try:
        df_103 = carregar_form103c()
        if not df_103.empty:
            df["PT_NORM"] = df["PT"].apply(_normalizar_pt)
            cols_merge = ["PT_NORM"] + _colunas_103
            df = df.merge(df_103[cols_merge], on="PT_NORM", how="left")
            df.drop(columns=["PT_NORM"], inplace=True, errors="ignore")
        else:
            for c in _colunas_103:
                df[c] = None
    except Exception:
        for c in _colunas_103:
            df[c] = None

    return df


# =============================================================================
# PERSISTÊNCIA
# =============================================================================

def sincronizar_epr() -> bool:
    from cloud_config import IS_CLOUD
    if IS_CLOUD:
        return False  # No cloud, dados vêm do cache estático
    import sqlite3
    st.cache_data.clear()
    df = carregar_dados_epr_form022a()
    if df.empty:
        return False
    db_path = os.path.join(_ROOT, DB_NAME)
    try:
        with sqlite3.connect(db_path) as conn:
            df.to_sql(TABELA_EPR, conn, if_exists="replace", index=False)
        return True
    except Exception:
        return False


def carregar_do_db() -> pd.DataFrame:
    from cloud_config import IS_CLOUD
    if IS_CLOUD:
        from cloud_config import carregar_parquet_cache
        return carregar_parquet_cache("db_epr_form022a")
    import sqlite3
    db_path = os.path.join(_ROOT, DB_NAME)
    try:
        with sqlite3.connect(db_path) as conn:
            return pd.read_sql_query(f"SELECT * FROM {TABELA_EPR}", conn)
    except Exception:
        return pd.DataFrame()


def carregar_dados(forcar_excel: bool = False) -> pd.DataFrame:
    from cloud_config import IS_CLOUD
    if IS_CLOUD:
        return carregar_do_db()  # Sempre do cache no cloud
    if not forcar_excel:
        df = carregar_do_db()
        if not df.empty:
            return df
    return carregar_dados_epr_form022a()


# =============================================================================
# UTILITÁRIOS
# =============================================================================

def formatar_numero(valor) -> str:
    try:
        return f"{int(valor):,}".replace(",", ".")
    except Exception:
        return str(valor)

def get_label_grupo(grupo: str) -> str:
    return _LABELS_GRUPOS.get(grupo, grupo)

def get_icone_grupo(grupo: str) -> str:
    return _ICONES_GRUPOS.get(grupo, "📦")

def get_todos_grupos() -> list:
    return list(_LABELS_GRUPOS.keys())

def calcular_kpis(df: pd.DataFrame) -> dict:
    if df.empty:
        return {k: 0 for k in ["total_amostras", "concluidos", "em_andamento", "aguardando", "pct_concluido"]}
    total        = int(df["QUANTIDADE"].sum()) if "QUANTIDADE" in df.columns else len(df)
    concluidos   = int(df[df["STATUS"] == "CONCLUIDO"]["QUANTIDADE"].sum()) if total > 0 else 0
    em_andamento = int(df[df["STATUS"] == "EM ANDAMENTO"]["QUANTIDADE"].sum()) if total > 0 else 0
    aguardando   = int(df[df["STATUS"] == "AGUARDANDO"]["QUANTIDADE"].sum()) if total > 0 else 0
    pct          = (concluidos / total * 100) if total > 0 else 0.0
    return {
        "total_amostras":  total,
        "concluidos":      concluidos,
        "em_andamento":    em_andamento,
        "aguardando":      aguardando,
        "pct_concluido":   round(pct, 1),
    }
